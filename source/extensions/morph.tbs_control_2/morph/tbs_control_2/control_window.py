# SPDX-FileCopyrightText: Copyright (c) 2024 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: LicenseRef-NvidiaProprietary

"""
control_window.py — TBS 제어창 UI 및 이벤트 핸들러

【역할】
- build_control_window(ext): "TBS 제어창" 창. 최상단 화면 옵션(기본 메뉴/패널 숨기기),
  USD Load(TbsUsdWindow — extension.py),
  USD 타임라인(수동/자동), 가상 시그널 샘플,
  XML 제너레이터(6종 시퀀스 콤보·입력 필드), 우선 표시 접두사, prim 목록.
- refresh_object_list(ext): 드롭다운/목록 갱신.
- on_play_usd_animation / on_play_generator_sample / on_refresh_prim_list 등 버튼·콤보 핸들러.

【수정 포인트】
- USD 재생 UI: build_control_window() 상단 ~ "USD 애니메이션 정지" 버튼 근처.
- XML 제너레이터 UI: "XML 제너레이터 생성기" Frame 블록.
  · 콤보 항목 추가/순서 변경: ext._xml_seq_combo = ui.ComboBox(0, ...) 인자 목록.
  · 하단 입력 전환: on_xml_seq_changed — FROM/TO 보일지(ext._xml_ab_inputs_frame), PORT만 보일지(ext._xml_port_inputs_frame).
    → xml_generator.FROM_TO_SEQS / PORT_ID_ONLY_SEQS 와 동일한 규칙 유지.
  · OK/역파싱: on_xml_ok_clicked, on_xml_run_clicked — 내부 seqs 리스트는 콤보 순서와 반드시 일치.
- 애니메이션 버튼(예: 이동/포물선/회전): 파일 하단 on_* 및 SAMPLE_GENERATOR_JSON.

【이벤트→XML→역파싱→애니메이션(JSON) 상태 기반 룰 매핑 유지보수 가이드】
- 목적: 같은 이벤트라도 from/to/포트 점유 상태가 다르면 다른 JSON을 실행할 수 있게 한다.
- 규칙 파일(우선순위):
  1) `config/event_animation_rules.json`  ← 권장(상태 기반)
  2) `config/event_animation_map.json`    ← 기본 fallback(이벤트 단순 매핑)
- 포트별 LOT prim 가시성: `config/port_lot_prim_paths.json` — `port_lot_visibility.apply_port_lot_prim_visibility` (시뮬 이벤트마다)
- rules 형식(요약):
  · 리스트 항목: {"name","priority","when","use"}
  · when:
    - sequence: 정식 시퀀스명(예: EAPEIS_PORT_MOVE_TRANSFERING)
    - from_port / to_port / port: 문자열 일치
    - ports_occupancy: {"EP2":"FULL","BP3":"EMPTY"} 같은 상태 조건
      (값은 FULL/EMPTY 또는 특정 LOT_ID)
  · use:
    - json: 실행할 시퀀스 JSON 경로
    - runner / description: 부가 메타
- 호출 흐름(현재 구현의 주 경로):
  1) simulation_engine._emit_event()에서 payload + ports_occupancy 전달
  2) on_sim_start_clicked()의 on_event 콜백이 post_sim_anim_event(...) → 큐 SimUiQueueKind.ANIM_EVENT
  3) _drain_sim_log_queue() → _sim_ui_sink_anim_event → handle_sim_event_for_animation(ext, payload)
  4) handle_sim_event_for_animation():
     - payload를 canonical sequence(EAPEIS_PORT_*)로 정규화
     - xml_generator.build_xml_string(...)로 XML 생성
     - xml_generator.parse_xml_string(...)으로 역파싱
     - 역파싱 결과(sequence_name/from/to/port)를 rules/map 입력 payload로 표준화
  5) _resolve_event_animation_entry(seq, payload):
     - 먼저 rules에서 조건 매칭(우선순위 높은 규칙 우선)
     - 없으면 event_animation_map fallback
  6) _execute_mapped_sequence_stub(...)에서 파일 존재/파싱 검증 후 SequenceRunner로 즉시 실행 시도
- 유지보수 체크포인트:
  · XML 생성/역파싱 규칙 수정: `xml_generator.py` (상수/빌더/파서)
  · 이벤트 별칭(canonical 변환) 수정: `SIM_SEQ_ALIAS`
  · rules 조건 필드 확장: `_resolve_rule_entry`, `_matches_occupancy_rule`
  · JSON 실행 연결/예외처리 수정: `_execute_mapped_sequence_stub`
  · 최종 분기 로그/표시 메시지 수정: `handle_sim_event_for_animation`
- 시퀀스 편집기 JSON 연결 방법:
  1) 시퀀스 편집기에서 JSON 저장
  2) 파일을 extension 내부 경로(예: data/sim_sequences/*.json)에 배치
  3) rules 또는 map의 use.json 경로에 등록
  4) 시뮬레이션 이벤트 발생 시 자동 매칭/검증 로그 확인
 - 표시모드(SimLogPanelMode): 콤보 인덱스와 `_drain_sim_log_queue`의 이력 스킵 여부가 연동된다.
  · "둘다": 진행현황 + 이력로그
  · "진행현황": 진행현황만
  · "이력로그": 스토리/시뮬 이력만
 - 시뮬 UI 큐 라우팅: `SimUiQueueKind` + `_dispatch_sim_ui_queue_item` + `_sim_ui_sink_*`.
   새 공정 텍스트 로그는 `post_sim_history_line(ext, line)`(시뮬 스레드)만 쓰면 이력 창으로 간다.
 - 시뮬레이션 종료 시 `_export_sim_logs_to_xlsx()`가 자동 호출되어
   `data/sim_logs/sim_logs_YYYYmmdd_HHMMSS.xlsx`에 **JSON 타임테이블만** 저장한다.
   (시트는 한 장(`타임테이블`)으로 통합, 1열에 JSON 라인을 시간순으로 한 줄씩 기록.
   화면 구분은 각 라인의 `"screen"` 키로 한다.)

【시뮬레이션 이벤트→애니메이션(JSON) 매핑 요약 (요구사항 반영)】
주의:
- 실제 실행 우선순위는 `EVENT_JSON_CASE_MAP` → rules(`config/event_animation_rules.json`) → map(`config/event_animation_map.json`) 순이다.
- 아래 목록은 코드 내 기본 테이블 `EVENT_JSON_CASE_MAP` 기준이며, payload의 키는 seq + (from_port_id/to_port_id 또는 port_id)이다.
- 포트 ID: **INOUT**(IN/OUT), 버퍼 **BP1~BP4**(EP3 구성일 때만 BP4 사용), **EP1~EP3**.

1) 생성/투입(OHT 운반) — 애니는 ARRIVED에서만
- **이벤트(sequence_name)**: `EAPEIS_PORT_ARRIVED`
- **조건·케이스맵 키**
  - OHT→EP 직접 투입: `from_port_id="OHT"`, `to_port_id="EP1|EP2|EP3"` → 키 `OHT->EPn`
  - OHT→IN/OUT 안착: `port_id="INOUT"` 만 전달(from/to 없음) → 키 `INOUT`
- **JSON**
  - `arrived_inout.json` — IN/OUT 안착
  - `arrived_ep1.json` … `arrived_ep3.json` — EP 직접 투입

2) IN/OUT → 버퍼(BP) 이동 — 애니 실행
- **이벤트(sequence_name)**: `EAPEIS_PORT_MOVE_TRANSFERING`
- **조건**: `from_port_id="INOUT"`, `to_port_id="BP1|BP2|BP3|BP4"`
- **JSON**: `move_inout_bp1.json` … `move_inout_bp4.json` (도착 버퍼별)

3) 버퍼 → EP(공정포트) 이동 — 애니 실행
- **이벤트(sequence_name)**: `EISEAP_PORT_MOVE_REQ`
- **조건**: `from_port_id="BP1|BP2|BP3|BP4"`, `to_port_id="EP1|EP2|EP3"`
- **JSON**: `move_{버퍼}_{EP}.json` (예: `move_bp1_ep1.json`, `move_bp4_ep3.json`)

4) 회수 우선 실행 — 애니는 REMOVED에서만
- **이벤트(sequence_name)**: `EAPEIS_PORT_REMOVED`
- **조건**: `port_id="EP1|EP2|EP3"`
- **JSON**
  - EP1: `data/sim_sequences/removed_ep1.json`
  - EP2: `data/sim_sequences/removed_ep2.json`
  - EP3: `data/sim_sequences/removed_ep3.json`

5) 애니 없는 이벤트(상태/큐 의미만)
- `EAPEIS_PORT_READYTOLOAD` (생성/수신 준비)
- `EAPEIS_PORT_READYTOUNLOAD` (회수 요청 큐 적재)

【XML 시퀀스와 UI 필드】 (로직·상수는 xml_generator.py)
- FROM_PORT_ID + TO_PORT_ID: MOVE_TRANSFERING, MOVE, MOVE_REQ
- PORT_ID만: READYTOLOAD, ARRIVED, READYTOUNLOAD, REMOVED
새 종류 추가 시: xml_generator 수정 + 이 파일의 ComboBox·seqs 3곳 + 필요 시 IntField/모델 추가.

【주요 함수 색인(빠른 참조)】
- 경로·규칙 파일: _extension_root_dir(확장 루트), _event_animation_map_path / _event_animation_rules_path(JSON 경로),
  _load_event_animation_map·_load_event_animation_rules(mtime 캐시·로드)
- 규칙 매칭: _matches_occupancy_rule(ports_occupancy 조건), _resolve_rule_entry(rules.json 우선순위),
  _resolve_event_case_map_entry(EVENT_JSON_CASE_MAP), _resolve_event_animation_entry(통합: case→rules→map),
  _normalize_json_path(상대→절대)
- 실행: _execute_mapped_sequence_stub(매핑된 JSON 검증·SequenceRunner 실행), _estimate_step_duration_sec_for_log,
  _estimate_sequence_total_duration_sec_for_log, _estimate_anim_duration_for_gate_payload(게이트 대기 시간 추정)
- 이벤트 처리: handle_sim_event_for_animation(시뮬 payload→XML→역파싱→룰→JSON 실행), _on_sim_event(로그용 래퍼)
- 시뮬 UI 큐(스레드→UI): post_sim_history_line, post_sim_anim_event, post_sim_progress_update,
  _enqueue_sim_log·_enqueue_anim_event·_enqueue_control_action·_enqueue_gate_request·_enqueue_sim_progress,
  _drain_sim_log_queue(메인 스레드에서 소비), _dispatch_sim_ui_queue_item, _coerce_sim_ui_queue_kind,
  _sim_ui_sink_progress·_sim_ui_sink_anim_event·_sim_ui_sink_history_line·_sim_ui_sink_action·_sim_ui_sink_gate
- 게이트: _show_sim_gate_dialog, _close_sim_gate_dialog, on_sim_start_clicked 내부 on_gate 연동
- 진행·로그 UI: _append_sim_log, _format_history_line·_with_history_color_icon, _append_anim_history_log(노옵),
  _render_pending_dots,
  _update_sim_progress, _is_progress_only_mode, on_copy_sim_progress
- 포트 패널: _port_cell_text, _compact_cell_value, _sync_ep3_port_cell_visibility, _set_port_box_style, _update_port_occupancy_panel
- 시뮬 제어: on_sim_start_clicked·on_sim_stop_clicked·on_sim_reset_clicked, _detach_sim_update,
  on_sim_log_view_changed, on_sim_ep_count_changed, _export_sim_logs_to_xlsx
- XML UI: on_xml_seq_changed, on_xml_ok_clicked, on_xml_run_clicked
- 포트 문자열: _parse_port_num, _port_kind, _normalize_port_text_from_xml
- Prim 목록: on_refresh_prim_list, refresh_object_list, build_object_panel, on_button_0/1/2
- 가상 시그널: receive_signal_data, run_generator_from_parsed
- 창: build_control_window(전체 UI 조립)

【멀티 뷰포트 분할·화면별 시뮼 독립 진행 (유지보수 가이드)】
개요:
- **뷰포트 분할(2~4)**: `sim_multi_view.apply_sim_viewport_split_layout` 이 Kit Viewport/보조 창·USD 컨텍스트를 구성한다.
  제어창은 `ext._sim_viewport_split_count` 만 “실제 적용된 분할 수”로 유지하고, 체크박스·모니터 UI와 동기화한다.
- **시뮼 엔진**: 분할 수가 N>1 이면 `TBSSimulationEngine` 인스턴스를 N개 만들고 `ext._sim_engines` 에 넣는다. 각 엔진은
  `event_tags={"tbs_sim_screen": "1".."N"}` 로 로그·진행·게이트 payload에 화면 번호를 붙인다.
- **독립 tick**: N>1 일 때 **화면마다 별도 스레드**(`_sim_multi_engine_tick_worker`)에서 `sim.tick()` 만 호출한다.
  한 화면이 공정 확인(`on_gate` → `done_evt.wait()`)에 블로킹돼도 다른 화면 스레드는 계속 진행한다.
- **애니·pause와의 관계**: JSON `SequenceRunner`·Kit translate/rotate/curve 가 돌 때 `_sim_tick_pause_event` 가 켜질 수 있다.
  멀티일 때는 `_multi_tick_should_skip_for_screen` 으로 **해당 화면만** tick 을 잠시 건너뛰어, 다른 화면 sim 시간은 진행시킨다.
  애니 job 에는 `tbs_sim_screen` 이 들어가 `_sim_active_anim_owner_screen` 이 “어느 화면용 pause 인지”를 판별한다.
- **애니 재생 경로**: `_sim_ui_sink_anim_event` → `handle_sim_event_for_animation` → JSON `SequenceRunner`
  (`usd_context_name` 으로 분할 보조 스테이지에 MOVE 등 적용).

주요 ext 필드:
- `_sim_viewport_split_count` : 1~4, sim_multi_view 와 제어창 체크박스의 단일 소스.
- `_sim_engines` / `_sim_engine` : 멀티 시 엔진 리스트·호환용 [0] 참조.
- `_sim_tick_threads` : 멀티 시 worker 스레드들; `_detach_sim_update` 에서 모두 join.
- `_sim_thread` / `_sim_thread_stop` : 단일 채널용 기존 tick 스레드; 멀티 시 `_sim_thread` 는 None 일 수 있음.
- `_sim_per_screen_snapshots` : 화면별 시뮼 설정 스냅샷(저장/미저장); `_on_save_sim_settings_to_screen` 으로 갱신.
- `_sim_monitor_channels` : 분할 모니터(포트/진행/로그) 컬럼 dict 리스트; `_rebuild_sim_monitor_split_ui` 가 조립.
- `_sim_multi_export_done` / `_sim_multi_tick_shutdown` : 멀티 전체 종료·엑셀 export 한 번만.

함수별 역할 (분할·멀티 시뮼 직접 관련):
- `_refresh_sim_per_screen_status_labels` : 스냅샷 유무에 따라 화면별 “(저장됨)/(미저장)” 라벨 갱신.
- `_refresh_sim_per_screen_rows` : 분할 수에 맞춰 화면별 설정 행(HStack) visible 처리.
- `_on_save_sim_settings_to_screen` : 현재 제어창 값을 해당 화면 스냅샷에 저장·EP3 셀·HUD 스케줄.
- `_auto_fill_per_screen_snapshots_on_start` : 시뮼 시작 시 비어 있는 화면 슬롯에 현재 설정 복사.
- `_fault_ports_from_snapshot` / `_timing_and_init_from_snapshot` : 스냅샷→엔진용 고장 집합·Timing/InitConfig.
- `_sync_sim_multi_split_row_visibility` : USD 로드 후에만 분할 설정 행 표시.
- `_sync_sim_split_checkboxes_from_ext_count` : `ext` 분할 수와 1~4 체크박스 모델 동기화(재진입 가드).
- `_force_sim_split_to_default` : 분할 불가·스테이지 없을 때 1화면·스냅샷 초기화·레이아웃 1로 롤백.
- `_on_sim_split_choice_changed` : 사용자가 분할 체크박스 변경 시 상호 배타 처리 후 `apply_sim_viewport_split_layout`.
- `_usd_context_name_for_sim_screen` : 애니/가시성용 보조 USD 컨텍스트 이름(`morph_tbs_split_aux_*`).
- `_sim_monitor_channel_count` : 모니터 열 개수(분할 수와 동일).
- `_snapshot_monitor_channel_texts` / `_create_sim_monitor_channel_column` / `_rebuild_sim_monitor_split_ui` :
  분할 모니터 UI 재구축 및 이전 채널 텍스트 보존.
- `_sim_active_anim_owner_screen` : `_sim_anim_active["tbs_sim_screen"]` 기반 pause 대상 화면(1-based).
- `_multi_tick_should_skip_for_screen` : pause 이벤트 시 **이번 프레임에 tick 을 건너뛸 화면** 판정.
- `_sim_multi_engine_tick_worker` : 한 엔진 전용 tick 루프 스레드; 전 엔진 종료 시 export 액션 enqueue.
- `_detach_sim_update` : tick 스레드(복수 우선) 정지·join, 로그 UI 구독 해제.
- `on_sim_start_clicked` : N채널 엔진 생성·콜백 연결·N>1 이면 worker 스레드 기동, N==1 이면 기존 단일 `_tick_loop`.
- `_tick_loop` (내부 함수) : 단일/구조상 단일 스레드에서 모든 엔진 순차 tick(레거시 멀티 경로는 worker로 대체됨).
- `_execute_mapped_sequence_stub` : 이벤트→JSON 실행; job 에 `tbs_sim_screen` 포함(멀티 pause 판별용).
- `_sim_ui_sink_anim_event` : 큐에서 소비; 화면별 가시성 + 이벤트→JSON(화면별 USD 컨텍스트).
- `_update_sim_progress` : `tbs_sim_screen` 으로 멀티 진행현황 라벨 라우팅.

사용처: extension.py on_startup → build_control_window(self)
  · 재호출 시 기존 TBS 제어창은 destroy 후 재생성(확장 리로드 등으로 위젯 이중 생성 방지).
"""

from enum import Enum
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
import copy
import random
import asyncio
import threading
import time
import queue
import json
import re
from datetime import datetime
from pathlib import Path

import omni.kit.app as app
import omni.ui as ui
from pxr import Gf

from . import usd_animation_control
from . import xml_generator
from .curve_animation import (
    is_curve_animation_running,
    make_parabolic_path,
    run_prim_curve_animation,
    stop_all_curve_animations,
    stop_prim_curve_animation,
)
from .kit_chrome_visibility import KIT_CHROME_HIDE_DEFAULT_ON_LAUNCH, apply_kit_chrome_hidden
from .port_lot_visibility import (
    apply_port_lot_prim_visibility,
    apply_port_lot_prim_visibility_for_context,
    clear_port_lot_authoring_cache,
    foup_proc_y_lift,
    sync_port_lot_positions_after_visibility,
)
from .prim_info import get_prim_display_name, safe_str
from .prim_utils import (
    collect_prim_paths_safe,
    find_all_prim_paths_by_name,
    get_prim_local_translate,
    get_stage,
    is_usd_file_stage_loaded,
    set_prim_translate_only,
)
from .rotate_animation import (
    is_rotate_animation_running,
    run_prim_rotate_animation,
    stop_all_rotate_animations,
    stop_prim_rotate_animation,
)
from .selection_overlay import show_prim_info_in_viewport
from .signal_parser import parse_signal
from .sequence_engine import SequenceRunner
from . import sim_multi_view
from .sim_control_defaults import SIM_CONTROL_DEFAULTS as _SIM_DEF
from .simulation_engine import (
    Lot,
    SimulationInitConfig,
    SimulationLogConfig,
    SimulationTimingConfig,
    TBSSimulationEngine,
)
from .translate_animation import (
    is_translate_animation_running,
    run_prim_translate_animation,
    stop_all_translate_animations,
    stop_prim_translate_animation,
)

from .control_sim_bar_graph import (
    BAR_STATE_DOWN,
    BAR_STATE_EMPTY,
    BAR_STATE_LOAD,
    BAR_STATE_PROC,
    EpBarPrecomputed,
    allocate_bar_segment_pixels,
    bar_graph_row_order,
    normalize_bar_graph_row_order,
    bar_state_color,
    bar_state_from_seg,
    build_ep_bar_from_playback_schedule,
    build_ep_bar_from_progress_items,
    build_bar_graph_copy_document,
    build_prerun_export_document,
    build_prerun_export_document_web_slim,
    format_row_state_duration_summary,
    merge_bar_row_segments,
    overlay_bar_rows_tip_from_occ,
    truncate_bar_rows_at_t,
    write_prerun_export_json,
)
from .control_sim_bar_graph import _aggregate_all_ep_state  # noqa: PLC2701 — UI live 막대 ALL_EP 집계
from .control_sim_fix_proc_ui import build_fix_proc_window
from .sim_lot_fix_proc import lot_id_from_payload, read_lot_fix_proc_at_start
from .control_sim_prerun_playback import (
    PlaybackEngine,
    SimPreRunResult,
    SimTimelinePlayer,
    _progress_event_affects_ep,
    build_seek_snapshots_by_item_index,
    build_timetable_row_metas,
    effective_ports_occupancy_at_t,
    interval_occ_parts,
    prerun_engine_to_timeline,
    resolve_seek_through_index,
)
from .ebs_control_panel_ui import (
    build_ebs_control_panel_content,
    get_sim_ep_count_idx,
    init_ebs_control_models,
    sync_aux_kit_window_visibility,
)
from .control_sim_playback_gate import (
    can_emit_timeline_event,
    clear_playback_gate_state,
    clear_proc_gates,
    compute_json_effective_speed,
    is_screen_runner_busy,
    json_wall_duration_sec,
    set_json_wall_busy,
    set_proc_gate_end,
    try_release_json_wall_when_idle,
)
from .control_sim_playback_speed import (
    clear_playback_step_speed_locks,
    ensure_step_speed_locked,
    lock_playback_step_speed,
    unlock_playback_step_speed,
)
from .progress_step_state import (
    apply_engine_progress_payload,
    bind_linked_anim_on_dispatch,
    build_payload_from_step,
    build_playback_tick_payload,
    clear_progress_step_state,
    format_progress_anim_footer,
    notify_anim_finished,
    notify_anim_queued,
    notify_anim_started,
    progress_dedupe_extra,
    sync_anim_runtime_from_ext,
)
from .control_sim_multi_playback import (
    add_playback_sessions_after_prerun,
    bootstrap_playback_after_prerun,
    get_playback_runtime,
    get_sim_playback_player,
    is_multi_playback_instances,
    iter_sim_playback_players,
    set_sim_playback_active,
    sim_log_ui_drain_limit,
    sim_log_ui_history_drain_limit,
    stop_playback_for_screen,
    stop_playback_runtime,
)
from .control_sim_timetable_ui import (
    build_timetable_column_ui,
    mount_interactive_timetable,
    refresh_all_timetable_highlights,
    refresh_timetable_row_highlight,
    reset_timetable_channel_to_idle,
    set_timetable_busy_label,
    timetable_rows_locked,
    unlock_timetable_rows,
)

MAX_PRIMS_DISPLAY = 80
DEFAULT_PRIORITY_NAME_PREFIX = "Mesh_"
CHECKBOX_WHITE_STYLE = {
    "color": 0xFF000000,
    "background_color": 0xFFEEEEEE,
}


class SimUiQueueKind(str, Enum):
    """
    `_sim_log_queue` 튜플 (kind, payload) 의 kind.
    payload가 도달하는 UI 영역은 아래 sink와 1:1에 가깝게 대응한다.
    """

    PROGRESS = "progress"  # → 진행현황 패널 (_update_sim_progress)
    HISTORY_LINE = "log"  # → 이력 로그 패널 (_append_sim_log). 값 "log"는 기존 큐 호환 유지.
    ANIM_EVENT = "anim_event"  # → 포트 상태 + 시퀀스 실행 (handle_sim_event_for_animation)
    ACTION = "action"  # → 제어 액션 (예: xlsx보내기)
    GATE = "gate"  # → 공정 확인 창


class SimUiControlAction(str, Enum):
    """SimUiQueueKind.ACTION 의 payload 로 허용되는 값."""

    EXPORT_XLSX = "export_xlsx"


class SimLogPanelMode(int, Enum):
    """표시모드 콤보 인덱스 (`on_sim_log_view_changed` 와 동일)."""

    ALL = 0
    PROGRESS_ONLY = 1
    HISTORY_ONLY = 2


SIM_SEQ_ALIAS = {
    "READYTOLOAD": xml_generator.SEQ_READYTOLOAD,
    "ARRIVED": xml_generator.SEQ_ARRIVED,
    "MOVE_TRANSFERING": xml_generator.SEQ_MOVE_TRANSFERING,
    "MOVE": xml_generator.SEQ_MOVE,
    "MOVE_REQ": xml_generator.SEQ_MOVE_REQ,
    "READYTOUNLOAD": xml_generator.SEQ_READYTOUNLOAD,
    "REMOVED": xml_generator.SEQ_REMOVED,
}
# 이벤트별/포트별 JSON 매핑(최상단 일원화)
# - 운영 중 수정은 이 테이블을 우선 수정한다.
# - key 규칙:
#   * READYTOLOAD/READYTOUNLOAD: 애니 없음(매핑 비워둠)
#   * ARRIVED: OHT 이동 애니만 실행 → key="FROM->TO"(직접 EP) 또는 port만 있으면 key=port(INOUT 안착)
#   * EAPEIS_PORT_MOVE_TRANSFERING: INOUT->BPx 이동 애니만 실행 → key="FROM->TO"
#   * EISEAP_PORT_MOVE_REQ: BPx->EPy 이동 애니만 실행 → key="FROM->TO"
#   * REMOVED: 회수 우선순위가 되었을 때 회수 애니 실행 → key="PORT" (EP1/2/3)
EVENT_JSON_CASE_MAP: Dict[str, Dict[str, str]] = {
    # IN/OUT → 버퍼: 시뮬 payload from=INOUT, to=BP1..BP4
    xml_generator.SEQ_MOVE_TRANSFERING: {
        "INOUT->BP1": "data/sim_sequences/move_inout_bp1.json",
        "INOUT->BP2": "data/sim_sequences/move_inout_bp2.json",
        "INOUT->BP3": "data/sim_sequences/move_inout_bp3.json",
        "INOUT->BP4": "data/sim_sequences/move_inout_bp4.json",
    },
    xml_generator.SEQ_ARRIVED: {
        # OHT→EP 직접: FROM->TO / OHT→IN/OUT 안착: port_id만 → 키 INOUT
        "INOUT": "data/sim_sequences/arrived_inout.json",
        "OHT->EP1": "data/sim_sequences/arrived_ep1.json",
        "OHT->EP2": "data/sim_sequences/arrived_ep2.json",
        "OHT->EP3": "data/sim_sequences/arrived_ep3.json",
    },
    # 요구사항: READYTOLOAD / READYTOUNLOAD 는 애니 실행 안함(빈 dict 유지)
    xml_generator.SEQ_READYTOLOAD: {},
    xml_generator.SEQ_READYTOUNLOAD: {},
    # 요구사항: BP->EP 이동 애니는 EISEAP_PORT_MOVE_REQ 에서만 실행
    xml_generator.SEQ_MOVE_REQ: {
        "BP1->EP1": "data/sim_sequences/move_bp1_ep1.json",
        "BP1->EP2": "data/sim_sequences/move_bp1_ep2.json",
        "BP1->EP3": "data/sim_sequences/move_bp1_ep3.json",
        "BP2->EP1": "data/sim_sequences/move_bp2_ep1.json",
        "BP2->EP2": "data/sim_sequences/move_bp2_ep2.json",
        "BP2->EP3": "data/sim_sequences/move_bp2_ep3.json",
        "BP3->EP1": "data/sim_sequences/move_bp3_ep1.json",
        "BP3->EP2": "data/sim_sequences/move_bp3_ep2.json",
        "BP3->EP3": "data/sim_sequences/move_bp3_ep3.json",
        "BP4->EP1": "data/sim_sequences/move_bp4_ep1.json",
        "BP4->EP2": "data/sim_sequences/move_bp4_ep2.json",
        "BP4->EP3": "data/sim_sequences/move_bp4_ep3.json",
    },
    xml_generator.SEQ_REMOVED: {
        "EP1": "data/sim_sequences/removed_ep1.json",  # EP1에서 LOT/Foup 회수 완료 연출
        "EP2": "data/sim_sequences/removed_ep2.json",  # EP2에서 LOT/Foup 회수 완료 연출
        "EP3": "data/sim_sequences/removed_ep3.json",  # EP3에서 LOT/Foup 회수 완료 연출
    },
}

def _case_map_port_token(p: str) -> str:
    """이벤트 payload 포트 문자열을 케이스맵 키용으로 정규화(대문자·트림)."""
    return str(p or "").strip().upper()
SAMPLE_GENERATOR_JSON = """{
  "objects": ["Mesh_308", "Mesh_561", "WalkwayEndA_01"],
  "animation": {
    "segments": [
      {"duration": 1.0, "delta": [100, 0, 0]},
      {"duration": 1.0, "delta": [0, 100, 0]},
      {"duration": 2.0, "delta": [-100, -100, 0]}
    ]
  }
}"""

_EVENT_ANIM_MAP_CACHE: Optional[Dict[str, Any]] = None
_EVENT_ANIM_MAP_MTIME: Optional[float] = None
_EVENT_ANIM_RULES_CACHE: Optional[List[Dict[str, Any]]] = None
_EVENT_ANIM_RULES_MTIME: Optional[float] = None


def _extension_root_dir() -> Path:
    """이 파일 기준 확장 루트(.../morph.tbs_control_2). config·data 경로 계산에 사용."""
    # .../source/extensions/morph.tbs_control_2
    return Path(__file__).resolve().parents[2]


def _sequence_json_search_roots() -> Tuple[Path, ...]:
    """
    ``data/sim_sequences/*.json`` 을 찾을 때 쓸 확장 패키지 루트 후보(앞쪽 우선).

    Kit 가 확장을 빌드 산출물/캐시에서만 로드하고 ``data`` 폴더는 소스 트리에만 둔 경우,
    ``__file__`` 기준 단일 루트로는 JSON 이 "없음"으로 판정될 수 있다.
    """
    roots: List[Path] = []
    seen: Set[str] = set()

    def _add(path: Optional[Path]) -> None:
        if path is None:
            return
        try:
            rp = path.resolve()
        except Exception:
            rp = Path(path)
        key = str(rp).lower()
        if key in seen:
            return
        seen.add(key)
        roots.append(rp)

    try:
        import carb

        rt = carb.tokens.get_tokens_interface().resolve("${root}")
        if rt:
            src = Path(rt) / "source" / "extensions" / "morph.tbs_control_2"
            if (src / "data" / "sim_sequences").is_dir():
                _add(src)
    except Exception:
        pass

    _add(Path(__file__).resolve().parents[2])

    try:
        em = app.get_app().get_extension_manager()
        for ext in em.get_extensions() or []:
            if not isinstance(ext, dict) or not ext.get("enabled"):
                continue
            eid = str(ext.get("id", "") or "")
            if eid != "morph.tbs_control_2" and not eid.startswith("morph.tbs_control_2-"):
                continue
            epath = em.get_extension_path(eid)
            if not epath:
                continue
            pp = Path(epath)
            if (pp / "data" / "sim_sequences").is_dir():
                _add(pp)
    except Exception:
        pass

    return tuple(roots)


def _event_animation_map_path() -> Path:
    """이벤트 seq → JSON 단순 매핑 파일 경로."""
    return _extension_root_dir() / "config" / "event_animation_map.json"


def _event_animation_rules_path() -> Path:
    """상태 기반 애니메이션 규칙(우선순위 리스트) 파일 경로."""
    return _extension_root_dir() / "config" / "event_animation_rules.json"


def _load_event_animation_map() -> Dict[str, Any]:
    """event_animation_map.json을 읽어 dict로 반환. mtime이 같으면 캐시 재사용."""
    global _EVENT_ANIM_MAP_CACHE, _EVENT_ANIM_MAP_MTIME
    p = _event_animation_map_path()
    if not p.exists():
        _EVENT_ANIM_MAP_CACHE = {}
        _EVENT_ANIM_MAP_MTIME = None
        return {}
    try:
        mtime = p.stat().st_mtime
        if _EVENT_ANIM_MAP_CACHE is not None and _EVENT_ANIM_MAP_MTIME == mtime:
            return _EVENT_ANIM_MAP_CACHE
        data = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            data = {}
        _EVENT_ANIM_MAP_CACHE = data
        _EVENT_ANIM_MAP_MTIME = mtime
        return data
    except Exception as e:
        print(f"[ANIM MAP] 매핑 파일 로드 실패: {p} err={e}", flush=True)
        _EVENT_ANIM_MAP_CACHE = {}
        _EVENT_ANIM_MAP_MTIME = None
        return {}


def _load_event_animation_rules() -> List[Dict[str, Any]]:
    """event_animation_rules.json을 읽어 규칙 리스트로 반환(priority 오름차순 정렬, mtime 캐시)."""
    global _EVENT_ANIM_RULES_CACHE, _EVENT_ANIM_RULES_MTIME
    p = _event_animation_rules_path()
    if not p.exists():
        _EVENT_ANIM_RULES_CACHE = []
        _EVENT_ANIM_RULES_MTIME = None
        return []
    try:
        mtime = p.stat().st_mtime
        if _EVENT_ANIM_RULES_CACHE is not None and _EVENT_ANIM_RULES_MTIME == mtime:
            return _EVENT_ANIM_RULES_CACHE
        data = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            data = []
        norm: List[Dict[str, Any]] = [x for x in data if isinstance(x, dict)]
        norm.sort(key=lambda r: int(r.get("priority", 1000)))
        _EVENT_ANIM_RULES_CACHE = norm
        _EVENT_ANIM_RULES_MTIME = mtime
        return norm
    except Exception as e:
        print(f"[ANIM RULES] 규칙 파일 로드 실패: {p} err={e}", flush=True)
        _EVENT_ANIM_RULES_CACHE = []
        _EVENT_ANIM_RULES_MTIME = None
        return []


def _matches_occupancy_rule(rule_occ: Dict[str, Any], occ: Dict[str, Any]) -> bool:
    """규칙의 ports_occupancy(포트→FULL/EMPTY/LOT_ID)가 현재 occ 스냅샷과 일치하는지."""
    for port, expected in (rule_occ or {}).items():
        p = str(port).strip().upper()
        got = str((occ or {}).get(p, "") or "")
        exp = str(expected or "").strip()
        if exp.upper() == "FULL":
            if not got:
                return False
        elif exp.upper() == "EMPTY":
            if got:
                return False
        else:
            if got != exp:
                return False
    return True


def _resolve_rule_entry(seq: str, payload: Dict[str, str]) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str]]:
    rules = _load_event_animation_rules()
    if not rules:
        return (None, None, None)
    p_from = str(payload.get("from_port_id", "") or "")
    p_to = str(payload.get("to_port_id", "") or "")
    p_port = str(payload.get("port_id", "") or "")
    occ = payload.get("ports_occupancy", {}) if isinstance(payload.get("ports_occupancy", {}), dict) else {}

    for r in rules:
        when = r.get("when", {}) if isinstance(r.get("when", {}), dict) else {}
        if str(when.get("sequence", "") or "").strip() not in ("", seq):
            continue
        if str(when.get("from_port", "") or "").strip() not in ("", p_from):
            continue
        if str(when.get("to_port", "") or "").strip() not in ("", p_to):
            continue
        if str(when.get("port", "") or "").strip() not in ("", p_port):
            continue
        rule_occ = when.get("ports_occupancy", {})
        if isinstance(rule_occ, dict) and not _matches_occupancy_rule(rule_occ, occ):
            continue

        use = r.get("use", {}) if isinstance(r.get("use", {}), dict) else {}
        j = use.get("json")
        if isinstance(j, str) and j.strip():
            return (j.strip(), use, str(r.get("name", "")))
    return (None, None, None)


def _resolve_event_case_map_entry(seq: str, payload: Dict[str, str]) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str]]:
    table = EVENT_JSON_CASE_MAP.get(seq, {})
    if not isinstance(table, dict) or not table:
        return (None, None, None)
    p_from = _case_map_port_token(str(payload.get("from_port_id", "") or ""))
    p_to = _case_map_port_token(str(payload.get("to_port_id", "") or ""))
    p_port = _case_map_port_token(str(payload.get("port_id", "") or ""))
    key = f"{p_from}->{p_to}" if p_from and p_to else p_port
    if not key:
        return (None, None, None)
    j = table.get(key)
    if not isinstance(j, str) or not j.strip():
        return (None, None, None)
    meta = {
        "runner": "sequence_editor",
        "description": f"top-case-map:{key}",
    }
    return (j.strip(), meta, f"top_case_map:{seq}:{key}")


def _resolve_event_animation_entry(seq: str, payload: Optional[Dict[str, str]] = None) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str], str]:
    """
    반환:
    - json_path_str: 실제 JSON 경로 문자열(없으면 None)
    - meta: runner/description 등 부가정보
    """
    # 0) 파일 최상단 케이스 매핑(운영 우선)
    j_case, meta_case, case_name = _resolve_event_case_map_entry(seq, payload or {})
    if j_case:
        return (j_case, meta_case, case_name, "top_case_map")

    # 1) 상태 기반 rules 우선
    j_rule, meta_rule, rule_name = _resolve_rule_entry(seq, payload or {})
    if j_rule:
        return (j_rule, meta_rule, rule_name or "(unnamed-rule)", "rules")

    # 2) 기존 단순 map fallback
    m = _load_event_animation_map()
    if not m:
        return (None, None, None, "")

    # 키 우선순위:
    # 1) 정식 seq (EAPEIS_PORT_...)
    # 2) 별칭(READYTOLOAD 등)
    # 3) raw 값
    raw_alias = None
    for alias, canonical in SIM_SEQ_ALIAS.items():
        if canonical == seq:
            raw_alias = alias
            break

    cand = [seq]
    if raw_alias:
        cand.append(raw_alias)

    for key in cand:
        if key not in m:
            continue
        v = m.get(key)
        if isinstance(v, str):
            return (v, {"runner": "sequence_editor", "description": ""}, None, "map")
        if isinstance(v, dict):
            j = v.get("json")
            if isinstance(j, str) and j.strip():
                return (j.strip(), v, None, "map")
    return (None, None, None, "")


def _normalize_json_path(path_text: str) -> Path:
    """
    시퀀스 JSON 상대 경로를 확장 루트 기준 절대 Path 로.

    후보 루트(``_sequence_json_search_roots``)를 순서대로 보며 ``.is_file()`` 인 첫 경로를 반환한다.
    없으면 기존과 같이 ``_extension_root_dir()`` 기준 경로를 반환(오류 메시지·로그용).
    """
    raw = str(path_text or "").strip()
    p = Path(raw)
    if not raw:
        return p
    if p.is_absolute():
        try:
            return p.resolve()
        except Exception:
            return p
    rel = p
    roots = _sequence_json_search_roots()
    last: Optional[Path] = None
    for root in roots:
        try:
            cand = (root / rel).resolve()
        except Exception:
            cand = root / rel
        last = cand
        try:
            if cand.is_file():
                return cand
        except Exception:
            pass
    if last is not None:
        return last
    base = roots[0] if roots else _extension_root_dir()
    try:
        return (base / rel).resolve()
    except Exception:
        return base / rel


def _estimate_step_duration_sec_for_log(step: Dict[str, Any], *, speed_scale: float = 1.0) -> Optional[float]:
    """
    애니메이션 실행이력용 "예상 길이" 계산(보수적).
    - MOVE/ROTATE: duration_max가 있으면 max, 없으면 duration.
    - DELAY: duration
    - USD_TIMELINE / TIMESAMPLES_REPLAY: 프레임 범위(start/end)로 추정
    - PRIM_VISIBILITY / SET_PRIM_VISIBILITY / PRIM_HIDE / PRIM_SHOW: duration (기본 0.02, tbs_lam_sequence_engine 과 동일)
    """
    # NOTE(정책):
    # - "표시/예상 시간"은 **1배속 기준(콘텐츠 기준)** 으로 유지한다.
    # - 시뮬 배속(ext._sim_speed_model)은 "재생/진행 속도"만 바꾸고, 여기의 표기 시간에는 반영하지 않는다.
    # - 단, USD_TIMELINE / TIMESAMPLES_REPLAY 의 per-step 배속(step["speed_scale"])은
    #   "그 스텝 자체를 빠르게 재생"하므로 표기에도 반영한다.
    try:
        from .sequence_renewal import is_renewal_marker

        if is_renewal_marker(step or {}):
            return 0.0
    except Exception:
        pass
    try:
        t = str((step or {}).get("type") or "").upper()
    except Exception:
        return None
    try:
        if t in ("MOVE", "ROTATE"):
            if "duration_max" in (step or {}):
                return max(0.0, float((step or {}).get("duration_max", (step or {}).get("duration", 0.0))))
            return max(0.0, float((step or {}).get("duration", 0.0)))
        if t == "DELAY":
            return max(0.0, float((step or {}).get("duration", 0.0)))
        if t in ("PRIM_VISIBILITY", "SET_PRIM_VISIBILITY", "PRIM_HIDE", "PRIM_SHOW"):
            # sticky visibility tail — _start_set_prim_visibility 와 동일 (1배속 표기).
            return max(0.0, float((step or {}).get("duration", 0.02) or 0.02))
        if t in ("USD_TIMELINE", "TIMESAMPLES_REPLAY"):
            play = (step or {}).get("play") or {}
            if not isinstance(play, dict):
                play = {}

            def _frame_val(key: str, default: int = 0) -> int:
                if key in play and play[key] is not None:
                    return int(play[key])
                if key in (step or {}) and (step or {}).get(key) is not None:
                    return int((step or {}).get(key))
                return default

            start = _frame_val("start_frame", 0)
            end = _frame_val("end_frame", 0)
            if end <= start:
                return 0.0
            # 정책: 기본 30fps(TPS) 기반 환산 + 배속 반영 (tbs_lam_sequence_engine 과 동일 schema)
            try:
                if "speed_scale" in play and play["speed_scale"] is not None:
                    step_sp = float(play["speed_scale"])
                else:
                    step_sp = float((step or {}).get("speed_scale", 1.0))
            except Exception:
                step_sp = 1.0
            step_sp = max(0.01, float(step_sp))
            base = float(usd_animation_control.frame_to_time(float(end - start)))
            return max(0.0, base / float(step_sp))
    except Exception:
        return None
    return None


def _estimate_sequence_total_duration_sec_for_log(steps: List[Dict[str, Any]], *, speed_scale: float = 1.0) -> Optional[float]:
    """
    SequenceRunner의 그룹/지연 규칙을 단순화해서 "예상 총 길이"를 계산한다.
    - 병렬 그룹: 리더 시작 시각 기준으로 (offset + duration)의 최대값을 그룹 종료로 본다.
    - 다음 그룹 시작: engine과 동일하게 anchor_end + next.step_delay_ms 를 사용하되,
      그룹 시작(t0)보다 앞당기지는 않는다.
    """
    if not steps:
        return 0.0
    # 첫 스텝의 step_delay_ms는 시퀀스 시작 전 지연으로 해석
    try:
        t_cursor = max(0.0, int((steps[0] or {}).get("step_delay_ms", 0)) / 1000.0)
    except Exception:
        t_cursor = 0.0
    last_finish = t_cursor

    i = 0
    while i < len(steps):
        try:
            g_end = _group_end_index(steps, i)
        except Exception:
            g_end = i
        t0 = t_cursor

        # 그룹 내 예상 종료(병렬 최대)
        group_finish = t0
        for j in range(i, g_end + 1):
            st = steps[j] if isinstance(steps[j], dict) else {}
            off = 0.0
            if j != i:
                try:
                    off = max(0.0, int((st or {}).get("step_delay_ms", 0)) / 1000.0)
                except Exception:
                    off = 0.0
            # 표기/예상 시간은 1배속 기준(콘텐츠 기준).
            # (USD_TIMELINE step["speed_scale"]만 _estimate_step_duration_sec_for_log 내부에서 반영)
            dur = _estimate_step_duration_sec_for_log(st, speed_scale=1.0)
            if dur is None:
                # 미지원 타입은 해당 스텝만 0초로 간주(시퀀스 전체 추정은 계속).
                dur = 0.0
            group_finish = max(group_finish, t0 + off + float(dur))
        last_finish = max(last_finish, group_finish)

        next_idx = g_end + 1
        if next_idx >= len(steps):
            break

        # anchor 종료 시각(앵커 스텝은 그룹 마지막)
        anchor_step = steps[g_end] if isinstance(steps[g_end], dict) else {}
        anchor_off = 0.0
        if g_end > i:
            try:
                anchor_off = max(0.0, int((anchor_step or {}).get("step_delay_ms", 0)) / 1000.0)
            except Exception:
                anchor_off = 0.0
        anchor_dur = _estimate_step_duration_sec_for_log(anchor_step, speed_scale=1.0)
        if anchor_dur is None:
            anchor_dur = 0.0
        anchor_end = t0 + anchor_off + float(anchor_dur)

        try:
            delay_next = int((steps[next_idx] or {}).get("step_delay_ms", 0)) / 1000.0
        except Exception:
            delay_next = 0.0
        t_cursor = max(t0, anchor_end + float(delay_next))
        i = next_idx

    return max(0.0, float(last_finish))


def _execute_mapped_sequence_stub(
    ext: Any,
    seq: str,
    payload: Dict[str, str],
    json_path_text: str,
    meta: Optional[Dict[str, Any]],
    rule_name: Optional[str],
    verbose: bool,
) -> None:
    """
    rules/map이 가리키는 JSON을 검증한 뒤 SequenceRunner.run()으로 실제 재생한다.

    - 시뮬 tick: 배속>1 등에서 ``_sim_tick_pause_event`` 로 잠시 맞출 수 있다. 멀티 뷰에서는 job 의
      ``tbs_sim_screen``(``payload`` 의 ``tbs_sim_screen``)으로 **어느 화면의 tick 만** 멈출지 판별한다.
    - JSON 시퀀스(SequenceRunner) 재생 중에는 단일 스레드 모드에서는 tick 을 막지 않아 공정과 병행될 수 있다.
    - 동시 ``run()`` 방지를 위해 ``_sim_anim_pending`` 큐로 직렬화한다.
    """
    p = _normalize_json_path(json_path_text)
    runner = str((meta or {}).get("runner", "sequence_editor"))
    desc = str((meta or {}).get("description", ""))
    from_port = str(payload.get("from_port_id", "")).strip()
    to_port = str(payload.get("to_port_id", "")).strip()
    port = str(payload.get("port_id", "")).strip()
    if from_port and to_port:
        route = f"{from_port}->{to_port}"
    elif to_port:
        route = f"to={to_port}"
    elif from_port:
        route = f"from={from_port}"
    elif port:
        route = f"port={port}"
    else:
        route = "port=미상"
    base_desc = desc if desc else "동작설명 없음"
    lot_id = lot_id_from_payload(payload) or "-"
    action_text = f"{base_desc} ({route} | lot={lot_id})"
    sim_time = str(payload.get("sim_time", "")).strip()
    # 스토리-JSON 요약(애니 실행이력 창) 기록: "실행/스킵/실패" 모두 남겨야 누적이 끊기지 않는다.
    def _push_story_json(status: str) -> None:
        """(애니 실행이력 UI 제거) 스토리-JSON 요약 누적은 더 이상 사용하지 않는다."""
        return

    if not p.is_file():
        _push_story_json("SKIP (MISSING)")
        _append_anim_history_log(
            ext,
            f"[ANIM] 파일없음 | event={seq} | action={action_text} | need={p.name} | path={p}",
        )
        if verbose:
            print(
                f"[ANIM MAP] 이벤트={seq} -> JSON 파일 없음: {p} "
                f"(runner={runner}, rule={rule_name or '-'}, desc={desc})",
                flush=True,
            )
        return
    try:
        # 파일 유효성 확인(실행 전 파싱 검증)
        parsed = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(parsed, list):
            raise ValueError("시퀀스 JSON 루트는 list여야 합니다.")
    except Exception as e:
        _push_story_json("SKIP (PARSE_FAIL)")
        _append_anim_history_log(ext, f"[ANIM] JSON 파싱실패 | event={seq} | action={action_text} | file={p.name} | err={e}")
        if verbose:
            print(f"[ANIM MAP] JSON 파싱 실패: {p} err={e}", flush=True)
        return

    # 시뮬레이션 중에는 "빈 JSON([])"을 실행(run)하면 runner 초기화/복원 경로를 타면서
    # 포즈/숨김이 튀는 원인이 된다. 빈 시퀀스는 실행하지 않고 스킵한다.
    if not parsed:
        _push_story_json("SKIP (EMPTY)")
        _append_anim_history_log(
            ext,
            f"[ANIM] 스킵(EMPTY JSON) | t={sim_time or '-'} | event={seq} | action={action_text} | file={p.name}",
        )
        if verbose:
            print(f"[ANIM MAP] 빈 JSON([]) 스킵: {p} event={seq}", flush=True)
        return

    try:
        from .json_playback_timing import renewal_info_from_steps

        _has_renewal_parsed, _ = renewal_info_from_steps(parsed)
        try:
            _scr_defer = int(str(payload.get("tbs_sim_screen", "1") or "1").strip() or "1")
        except Exception:
            _scr_defer = 1
        _set_renewal_port_defer(ext, _scr_defer, bool(_has_renewal_parsed))
        _set_renewal_json_guard(ext, _scr_defer, bool(_has_renewal_parsed))
    except Exception:
        pass

    # 예상 총 길이(초): 엑셀/로그에 같이 남길 수 있게 추정
    # 표시/예상 시간은 1배속 기준(콘텐츠 기준)으로 유지한다.
    # (실제 재생 속도는 SequenceRunner.run(speed_scale=...)로 별도 적용)
    est_total = _estimate_sequence_total_duration_sec_for_log(parsed, speed_scale=1.0)
    est_text = f"{est_total:.2f}s" if isinstance(est_total, (float, int)) else "미확인"

    step_types: List[str] = []
    for step in parsed:
        if isinstance(step, dict):
            t = str(step.get("type", "")).strip().upper()
            if t:
                step_types.append(t)
    if step_types:
        preview = ",".join(step_types[:4]) + ("..." if len(step_types) > 4 else "")
    else:
        preview = "EMPTY"

    _append_anim_history_log(
        ext,
        f"[ANIM] 실행준비완료 | t={sim_time or '-'} | event={seq} | est={est_text} | action={action_text} | file={p.name} | steps={len(parsed)}({preview}) | runner={runner} | rule={rule_name or '-'}",
    )
    _push_story_json("PLAN (READY)")
    # 실제 실행 연결: 동시에 여러 run() 호출 시 기존 애니가 끊기는 문제를 막기 위해
    # "시뮬레이션 애니 실행 큐"로 직렬화한다.
    try:
        if not isinstance(getattr(ext, "_sim_anim_pending", None), list):
            ext._sim_anim_pending = []

        def _start_job_impl(job: Dict[str, Any]) -> None:
            # 화면별 runner/active/pending/pause로 분리한다(멀티에서 덮어쓰기/간섭 방지).
            try:
                scr_i = int(str(job.get("tbs_sim_screen", "1") or "1").strip() or "1")
            except Exception:
                scr_i = 1
            scr_i = max(1, scr_i)
            try:
                set_json_wall_busy(ext, scr_i, True)
            except Exception:
                pass
            try:
                notify_anim_started(ext, scr_i, str(job.get("file", "") or ""))
            except Exception:
                pass
            _playback = bool(getattr(ext, "_sim_playback_started", False))
            if not _playback:
                try:
                    _halt_screen_json_anim(ext, scr_i, join_sec=3.0)
                except Exception:
                    pass
            try:
                runners = getattr(ext, "_sim_runners_by_screen", None)
                if not isinstance(runners, dict):
                    runners = {}
                    ext._sim_runners_by_screen = runners
            except Exception:
                runners = {}
                ext._sim_runners_by_screen = runners
            try:
                runner_obj = runners.get(str(scr_i))
            except Exception:
                runner_obj = None
            try:
                if runner_obj is not None:
                    runner_obj._diag_ext = ext  # type: ignore[attr-defined]
                    runner_obj._diag_screen = scr_i  # type: ignore[attr-defined]
            except Exception:
                pass
            try:
                from .sequence_engine import SequenceRunner
                from .tbs_split_composed_loader import get_split_runtime_for_screen

                rt = get_split_runtime_for_screen(ext, scr_i)
                if rt is not None:
                    if runner_obj is None:
                        runner_obj = SequenceRunner(
                            registry=rt.registry,
                            scheduler=rt.scheduler,
                            evaluator=rt.evaluator,
                        )
                        runners[str(scr_i)] = runner_obj
                        if int(scr_i) == 1:
                            try:
                                ext._sim_runner = runner_obj
                            except Exception:
                                pass
                    else:
                        runner_obj._tbs_registry = rt.registry
                        runner_obj._tbs_scheduler = rt.scheduler
                        runner_obj._tbs_evaluator = rt.evaluator
                elif runner_obj is None:
                    runner_obj = SequenceRunner(
                        registry=getattr(ext, "_tbs_registry", None),
                        scheduler=getattr(ext, "_tbs_scheduler", None),
                        evaluator=getattr(ext, "_tbs_evaluator", None),
                    )
                    runners[str(scr_i)] = runner_obj
                    if int(scr_i) == 1:
                        try:
                            ext._sim_runner = runner_obj
                        except Exception:
                            pass
            except Exception:
                if runner_obj is None:
                    runner_obj = getattr(ext, "_sim_runner", None)

            try:
                pause_map = getattr(ext, "_sim_tick_pause_events_by_screen", None)
                if not isinstance(pause_map, dict):
                    pause_map = {}
                    ext._sim_tick_pause_events_by_screen = pause_map
            except Exception:
                pause_map = {}
                ext._sim_tick_pause_events_by_screen = pause_map
            pause_evt = pause_map.get(str(scr_i))
            if pause_evt is None:
                pause_evt = threading.Event()
                pause_map[str(scr_i)] = pause_evt

            sp = 1.0
            try:
                if bool(getattr(ext, "_sim_playback_started", False)):
                    sp = float(ensure_step_speed_locked(ext, scr_i))
                else:
                    m = getattr(ext, "_sim_speed_model", None)
                    if m is not None:
                        sp = max(0.1, float(m.get_value_as_float()))
            except Exception:
                sp = 1.0
            proc_sec_job = 0.0
            try:
                proc_sec_job = float(str(job.get("proc_sec", "") or "").strip() or "0")
            except Exception:
                proc_sec_job = 0.0
            est_total_job = job.get("est_total", None)
            est_total_f = 0.0
            try:
                if isinstance(est_total_job, (float, int)):
                    est_total_f = float(est_total_job)
            except Exception:
                est_total_f = 0.0
            parsed_steps = job.get("parsed", []) or []
            if est_total_f <= 1e-9 and isinstance(parsed_steps, list) and parsed_steps:
                try:
                    est_total_f = float(
                        _estimate_sequence_total_duration_sec_for_log(parsed_steps, speed_scale=1.0) or 0.0
                    )
                except Exception:
                    pass
            if proc_sec_job <= 1e-9:
                try:
                    from .json_playback_timing import timing_from_progress

                    _tm_ps = timing_from_progress(
                        {
                            "proc_sec": job.get("proc_sec"),
                            "anim_sec": est_total_f if est_total_f > 0 else job.get("anim_sec"),
                            "event_start_sim_time": job.get("t") or job.get("sim_time"),
                        },
                        steps=list(parsed_steps) if isinstance(parsed_steps, list) else None,
                    )
                    proc_sec_job = float(_tm_ps.get("proc") or proc_sec_job)
                except Exception:
                    pass
            eff_sp = compute_json_effective_speed(sp, proc_sec_job, est_total_f)
            json_wall_sec = json_wall_duration_sec(est_total_f, eff_sp)

            has_renewal = bool(job.get("has_renewal"))
            renewal_off = job.get("renewal_offset_sec")
            json_lead = 0.0
            try:
                from .json_playback_timing import json_lead_sec, renewal_info_from_steps

                if not has_renewal:
                    has_renewal, renewal_off = renewal_info_from_steps(list(parsed_steps))
                json_lead = json_lead_sec(proc_sec_job, est_total_f)
            except Exception:
                pass

            started_wall = time.monotonic()
            lead_wall = max(0.0, float(json_lead)) / max(0.1, float(sp))
            json_run_start_wall = float(started_wall) + float(lead_wall)
            total_wall_busy = float(lead_wall) + float(json_wall_sec)
            try:
                t0_sim = float(str(job.get("t") or job.get("sim_time") or "0").strip() or "0")
            except Exception:
                t0_sim = 0.0
            json_run_start_sim = float(t0_sim) + float(json_lead)
            active = dict(job)
            active["proc_sec"] = float(proc_sec_job)
            active["anim_sec"] = float(est_total_f)
            active["_event_start_sim"] = float(t0_sim)
            active["_json_run_start_sim"] = float(json_run_start_sim)
            active["_json_pending_sim_start"] = bool(_playback)
            active["_json_sequence_started"] = False
            active["_started_wall"] = started_wall
            active["_eff_sp"] = float(eff_sp)
            active["has_renewal"] = bool(has_renewal)
            active["renewal_offset_sec"] = renewal_off
            active["json_lead_sec"] = float(json_lead)
            active["_json_lead_wall_sec"] = float(lead_wall)
            active["_json_run_start_wall"] = float(json_run_start_wall)
            _set_renewal_port_defer(ext, scr_i, bool(has_renewal))
            _set_renewal_json_guard(ext, scr_i, bool(has_renewal))
            try:
                active_by = getattr(ext, "_sim_anim_active_by_screen", None)
                if not isinstance(active_by, dict):
                    active_by = {}
                    ext._sim_anim_active_by_screen = active_by
                active_by[str(scr_i)] = active
            except Exception:
                ext._sim_anim_active = active
            if not bool(has_renewal) and (not bool(getattr(ext, "_sim_playback_started", False))):
                try:
                    _flush_pending_post_anim_port_applies(ext, scr_i)
                except Exception:
                    pass
            _clear_post_anim_port_applied(ext, scr_i)
            try:
                if total_wall_busy > 0.0:
                    try:
                        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                        if not isinstance(until_by, dict):
                            until_by = {}
                            ext._sim_tick_pause_until_wall_by_screen = until_by
                        until_by[str(scr_i)] = float(started_wall) + float(total_wall_busy)
                    except Exception:
                        ext._sim_tick_pause_until_wall = float(started_wall) + float(total_wall_busy)
                else:
                    try:
                        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                        if isinstance(until_by, dict):
                            until_by[str(scr_i)] = None
                    except Exception:
                        ext._sim_tick_pause_until_wall = None
            except Exception:
                try:
                    until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                    if isinstance(until_by, dict):
                        until_by[str(scr_i)] = None
                except Exception:
                    ext._sim_tick_pause_until_wall = None

            def _on_done():
                # 재생(plan): 포트는 sim_now milestone replay 만. on_done·fail-safe 포트 적용 없음.
                src_done = dict(job or {})
                src_done["event"] = _normalize_anim_event_seq(str(src_done.get("event", "") or ""))
                src_done["event_start_sim_time"] = str(src_done.get("t") or src_done.get("sim_time") or "").strip()
                _playback = bool(getattr(ext, "_sim_playback_started", False))
                if (not bool(job.get("has_renewal"))) and (not _playback):
                    _queue_post_anim_port_apply(ext, int(scr_i), src_done)

                # 다음 JSON/ANIM_DONE 직전 — TIMESAMPLES·legacy translate 잔류 drain (BG thread).
                try:
                    from .sim_channel_scope import drain_channel_motion_complete, stop_channel_animations
                    from .tbs_split_composed_loader import get_split_runtime_for_screen

                    _ctx_done = _usd_context_name_for_sim_screen(ext, scr_i)
                    _rt_done = get_split_runtime_for_screen(ext, scr_i)
                    _reg_done = _rt_done.registry if _rt_done is not None else None
                    idle = drain_channel_motion_complete(
                        _ctx_done,
                        _reg_done,
                        max_sec=5.0,
                        stable_ticks=3,
                    )
                    if not idle:
                        try:
                            from . import sim_multi_diag as _mdiag

                            _mdiag.log_motion_drain_timeout(
                                ext, screen=int(scr_i), ctx=_ctx_done
                            )
                        except Exception:
                            pass
                        stop_channel_animations(
                            _ctx_done, diag_reason="on_done_drain_timeout"
                        )
                except Exception:
                    pass

                def _finish_on_main() -> None:
                    if not _playback:
                        try:
                            _flush_pending_post_anim_port_applies(ext, int(scr_i))
                        except Exception:
                            pass
                    # 화면별 pending 큐에서 다음 job만 이어서 실행
                    pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
                    pending = []
                    if isinstance(pending_by, dict):
                        pending = pending_by.get(str(scr_i), []) or []
                    if isinstance(pending, list) and pending:
                        try:
                            pending.sort(
                                key=lambda j: int((j or {}).get("_priority", 10)) if isinstance(j, dict) else 10
                            )
                        except Exception:
                            pass
                        nxt = pending.pop(0)
                        if isinstance(pending_by, dict):
                            pending_by[str(scr_i)] = pending
                        try:
                            from . import sim_multi_diag as _mdiag

                            _mdiag.log_anim_done(
                                ext,
                                screen=scr_i,
                                file_name=str((job or {}).get("file", "") or ""),
                                pending_left=len(pending),
                                next_file=str((nxt or {}).get("file", "") or ""),
                            )
                        except Exception:
                            pass
                        _dispatch_json_anim_job(ext, nxt)
                        return
                    try:
                        from . import sim_multi_diag as _mdiag

                        _mdiag.log_anim_done(
                            ext,
                            screen=scr_i,
                            file_name=str((job or {}).get("file", "") or ""),
                            pending_left=0,
                            next_file="",
                        )
                    except Exception:
                        pass
                    if pause_evt is not None:
                        try:
                            pause_evt.clear()
                        except Exception:
                            pass
                    try:
                        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                        if isinstance(until_by, dict):
                            until_by[str(scr_i)] = None
                    except Exception:
                        pass
                    try:
                        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
                        if isinstance(active_by, dict):
                            active_by[str(scr_i)] = {}
                    except Exception:
                        pass
                    if bool((job or {}).get("has_renewal")):
                        try:
                            src_by = getattr(ext, "_sim_post_anim_src_by_screen", None)
                            if isinstance(src_by, dict):
                                src_by.pop(str(scr_i), None)
                        except Exception:
                            pass
                    try:
                        try_release_json_wall_when_idle(ext, int(scr_i))
                    except Exception:
                        try:
                            set_json_wall_busy(ext, scr_i, False)
                        except Exception:
                            pass
                    try:
                        _clear_renewal_port_defer(ext, int(scr_i))
                        _set_renewal_json_guard(ext, int(scr_i), False)
                    except Exception:
                        pass
                    try:
                        notify_anim_finished(ext, int(scr_i))
                    except Exception:
                        pass
                    try:
                        _refresh_sim_progress_from_last(ext, int(scr_i))
                    except Exception:
                        pass

                try:
                    if threading.current_thread() is threading.main_thread():
                        _finish_on_main()
                    else:
                        from .tbs_main_dispatch import dispatch_main_wait

                        dispatch_main_wait(_finish_on_main, timeout=60.0)
                except Exception:
                    try:
                        _finish_on_main()
                    except Exception:
                        pass

            try:
                if runner_obj is not None:
                    runner_obj.on_sequence_completed = _on_done  # type: ignore[attr-defined]
            except Exception:
                pass

            def _on_renewal_step(_idx: int, _step: dict) -> None:
                active_r = _screen_active_json_job(ext, int(scr_i)) or {}
                src_r = dict(active_r if active_r else (job or {}))
                src_r["event"] = _normalize_anim_event_seq(
                    str(src_r.get("event") or src_r.get("event_seq") or "")
                )
                t0_src = (
                    active_r.get("_event_start_sim")
                    if isinstance(active_r, dict) and active_r.get("_event_start_sim") is not None
                    else src_r.get("t") or src_r.get("sim_time")
                )
                src_r["event_start_sim_time"] = str(t0_src or "").strip()
                if not str(src_r.get("path") or "").strip() and isinstance(job, dict):
                    src_r["path"] = str(job.get("path") or "").strip()
                if not str(src_r.get("file") or "").strip() and isinstance(job, dict):
                    src_r["file"] = str(job.get("file") or "").strip()
                if not src_r.get("parsed") and isinstance(job, dict) and job.get("parsed"):
                    src_r["parsed"] = job.get("parsed")
                if not src_r.get("has_renewal") and isinstance(job, dict):
                    src_r["has_renewal"] = bool(job.get("has_renewal"))
                if src_r.get("renewal_offset_sec") is None and isinstance(job, dict):
                    src_r["renewal_offset_sec"] = job.get("renewal_offset_sec")
                for _pk in ("lot_id", "from_port_id", "to_port_id", "port_id", "proc_sec", "anim_sec", "est_total"):
                    if not str(src_r.get(_pk) or "").strip() and isinstance(job, dict) and job.get(_pk):
                        src_r[_pk] = job.get(_pk)

                if bool(getattr(ext, "_sim_playback_started", False)):
                    def _apply_renewal_playback_ui() -> None:
                        try:
                            from .control_sim_playback_plan import apply_playback_renewal_from_wall

                            if apply_playback_renewal_from_wall(ext, int(scr_i), dict(src_r)):
                                src_apply = dict(src_r)
                                src_apply["_from_renewal_step"] = True
                                dedupe = _post_anim_port_dedupe_key(src_apply)
                                applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
                                if not isinstance(applied_by, dict):
                                    applied_by = {}
                                    ext._sim_post_anim_port_applied_by_screen = applied_by
                                applied_by[str(scr_i)] = dedupe
                        except Exception:
                            pass

                    try:
                        if threading.current_thread() is threading.main_thread():
                            _apply_renewal_playback_ui()
                        else:
                            from .tbs_main_dispatch import dispatch_main

                            dispatch_main(_apply_renewal_playback_ui)
                    except Exception:
                        try:
                            _apply_renewal_playback_ui()
                        except Exception:
                            pass
                    return

                def _apply_renewal_live_ui() -> None:
                    try:
                        src_apply = dict(src_r)
                        src_apply["_from_renewal_step"] = True
                        try:
                            from .control_sim_playback_plan import (
                                _register_removed_prim_hide_hold_for_renewal,
                                prim_occ_for_playback_visibility,
                            )

                            _register_removed_prim_hide_hold_for_renewal(
                                ext, int(scr_i), src_apply, None
                            )
                        except Exception:
                            pass
                        last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                        occ_now: Dict[str, Any] = {}
                        if isinstance(last_by, dict) and isinstance(last_by.get(str(scr_i)), dict):
                            occ_now = dict(last_by.get(str(scr_i)) or {})
                        occ_pred = _predict_ports_occupancy_after_anim(occ_now, src_apply)
                        sim_t = str(
                            src_r.get("t") or src_r.get("sim_time") or ""
                        ).strip()
                        if not sim_t:
                            lp_by = getattr(ext, "_sim_progress_last_payload_by_screen", None)
                            lp = lp_by.get(str(scr_i)) if isinstance(lp_by, dict) else None
                            sim_t = str((lp or {}).get("sim_time", "") or "") if isinstance(lp, dict) else ""
                        _apply_sim_event_state_only(
                            ext,
                            {
                                "ports_occupancy": dict(occ_pred),
                                "sim_time": sim_t,
                                "_from_renewal_step": True,
                            },
                            screen=int(scr_i),
                        )
                        dedupe = _post_anim_port_dedupe_key(src_apply)
                        applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
                        if not isinstance(applied_by, dict):
                            applied_by = {}
                            ext._sim_post_anim_port_applied_by_screen = applied_by
                        applied_by[str(scr_i)] = dedupe
                    except Exception:
                        pass
                    try:
                        _flush_renewal_bar_legacy(ext, int(scr_i), src_r, None)
                    except Exception:
                        pass

                try:
                    if threading.current_thread() is threading.main_thread():
                        _apply_renewal_live_ui()
                    else:
                        from .tbs_main_dispatch import dispatch_main

                        dispatch_main(_apply_renewal_live_ui)
                except Exception:
                    try:
                        _apply_renewal_live_ui()
                    except Exception:
                        pass

            try:
                if runner_obj is not None:
                    runner_obj.on_renewal_step = _on_renewal_step  # type: ignore[attr-defined]
            except Exception:
                pass

            # 요구사항: 공정시간 우선 기능은 사용하지 않음(항상 OFF 고정)
            proc_priority = False

            # (fail-safe) pause_until_wall — lead + eff_sp 반영 wall JSON 길이
            try:
                ub_by_screen = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                if isinstance(ub_by_screen, dict) and total_wall_busy > 0.0:
                    ub_by_screen[str(scr_i)] = float(started_wall) + float(total_wall_busy)
            except Exception:
                pass
            # JSON 종료·renewal 포트 갱신 fail-safe: wall-clock·pending 용 이벤트 스냅샷 보관
            try:
                by_src = getattr(ext, "_sim_post_anim_src_by_screen", None)
                if not isinstance(by_src, dict):
                    by_src = {}
                    ext._sim_post_anim_src_by_screen = by_src
                snap = _normalize_post_anim_port_src(dict(job))
                snap["has_renewal"] = bool(has_renewal)
                snap["_json_lead_wall_sec"] = float(lead_wall)
                snap["_json_run_start_wall"] = float(json_run_start_wall)
                if total_wall_busy > 0.0:
                    snap["_json_end_wall"] = float(json_run_start_wall) + float(json_wall_sec)
                if has_renewal and renewal_off is not None:
                    snap["_port_sync_wall"] = float(json_run_start_wall) + float(renewal_off) / max(
                        0.1, float(eff_sp)
                    )
                by_src[str(scr_i)] = snap
            except Exception:
                pass

            _ctx_run = _usd_context_name_for_sim_screen(ext, scr_i)

            def _run_json_sequence() -> None:
                if bool(active.get("_json_sequence_started")):
                    return
                active["_json_sequence_started"] = True
                try:
                    if isinstance(active_by, dict):
                        active_by[str(scr_i)] = active
                except Exception:
                    pass
                run_wall = time.monotonic()
                active["_json_run_start_wall"] = float(run_wall)
                try:
                    by_src = getattr(ext, "_sim_post_anim_src_by_screen", None)
                    if isinstance(by_src, dict) and isinstance(by_src.get(str(scr_i)), dict):
                        snap_live = dict(by_src[str(scr_i)])
                        snap_live["_json_run_start_wall"] = float(run_wall)
                        snap_live["_json_end_wall"] = float(run_wall) + float(json_wall_sec)
                        if has_renewal and renewal_off is not None:
                            snap_live["_port_sync_wall"] = float(run_wall) + float(renewal_off) / max(
                                0.1, float(eff_sp)
                            )
                        by_src[str(scr_i)] = snap_live
                except Exception:
                    pass
                if _playback:
                    try:
                        _halt_screen_json_anim(ext, scr_i, join_sec=0.25)
                    except Exception:
                        pass
                # 위치·TBS_OFFSET·TIMESAMPLES — JSON **시작** 시점에 초기화 (back-align lead 이후).
                try:
                    _reset_sim_motion_before_json_run(ext, job, runner_obj=runner_obj)
                except Exception as exc:
                    print(f"[TBS/SIM] pre-json motion reset failed: {exc}", flush=True)
                try:
                    from . import sim_multi_diag as _mdiag

                    _mdiag.log_anim_start(
                        ext,
                        screen=scr_i,
                        ctx=_ctx_run,
                        file_name=str((job or {}).get("file", "") or ""),
                        est_total=float(est_total_f),
                        eff_sp=float(eff_sp),
                        proc_sec=float(proc_sec_job),
                        runner=runner_obj,
                    )
                except Exception:
                    pass
                if runner_obj is not None:
                    try:
                        runner_obj._foup_proc_active_ep = _resolve_foup_proc_active_ep(
                            ext, scr_i, dict(job or {})
                        )
                    except Exception:
                        pass
                    runner_obj.run(
                        job.get("parsed", []),
                        usd_context_name=_ctx_run,
                        speed_scale=eff_sp,
                        wait_until_done=False,
                    )
                else:
                    try:
                        print(
                            f"[ANIM] 실행 스킵 — SequenceRunner 없음 screen={scr_i} ctx={_ctx_run!r}",
                            flush=True,
                        )
                    except Exception:
                        pass
                    try:
                        set_json_wall_busy(ext, scr_i, False)
                    except Exception:
                        pass

            active["_json_run_fn"] = _run_json_sequence

            # 배속>1일 때: 단일 화면에서만 애니·sim tick 동기를 위해 pause 사용.
            # 분할 N>1에서는 한 화면 애니가 다른 화면 엔진 틱까지 멈추어 공정시간·막대가 끊겨 보이므로 적용하지 않는다.
            if (
                (not proc_priority)
                and sp > 1.0
                and pause_evt is not None
                and (not _is_multi_viewport_sim(ext))
            ):
                try:
                    pause_evt.set()
                except Exception:
                    pass

            if _playback:
                try:
                    _pl = get_sim_playback_player(ext, scr_i)
                    if _pl is not None and float(_pl.sim_now(scr_i)) + 1e-9 >= float(json_run_start_sim):
                        _run_json_sequence()
                except Exception:
                    pass
            elif lead_wall > 1e-6:
                threading.Timer(float(lead_wall), _run_json_sequence).start()
            else:
                _run_json_sequence()

            try:
                _refresh_sim_progress_from_last(ext, scr_i)
            except Exception:
                pass

        ext._sim_json_start_fn = _start_job_impl

        def _start_job(job: Dict[str, Any]) -> None:
            _dispatch_json_anim_job(ext, job)

        try:
            _scr = int(str(payload.get("tbs_sim_screen", "1") or "1").strip() or "1")
        except Exception:
            _scr = 1
        _scr = max(1, _scr)
        job = {
            "t": sim_time,
            "event": seq,
            "file": p.name,
            "path": str(p),
            "action": action_text,
            "est": est_text,
            "est_total": float(est_total) if isinstance(est_total, (float, int)) else None,
            # simulation_engine 이벤트 payload의 공정시간(=proc_sec) — JSON이 더 길면 여기 기반으로 배속 처리한다.
            "proc_sec": str(payload.get("proc_sec", "") or "").strip(),
            "runner": runner,
            "rule": rule_name or "-",
            "lot_id": lot_id,
            "from_port_id": from_port,
            "to_port_id": to_port,
            "port_id": port,
            "parsed": parsed,
            "tbs_sim_screen": str(_scr),
            "foup_proc_active_ep": str(
                payload.get("foup_proc_active_ep", "") or ""
            ).strip().upper(),
        }
        try:
            from .json_playback_timing import renewal_info_from_steps

            _has_r, _r_off = renewal_info_from_steps(parsed)
            job["has_renewal"] = bool(_has_r)
            job["renewal_offset_sec"] = _r_off
        except Exception:
            job["has_renewal"] = False
            job["renewal_offset_sec"] = None
        # 우선순위: 생성(OHT->EP 직접투입 등) / 회수(REMOVED) 는 현재 애니가 끝나자마자 즉시 실행되어야 한다.
        # - 선점(interrupt)은 하지 않고, pending 큐의 "앞"에 삽입한다.
        try:
            is_pickup = str(seq).strip().upper() == str(xml_generator.SEQ_REMOVED).strip().upper()
        except Exception:
            is_pickup = False
        try:
            is_spawn = (
                str(seq).strip().upper() == str(xml_generator.SEQ_ARRIVED).strip().upper()
                and str(from_port).strip().upper() == "OHT"
                and str(to_port).strip().upper().startswith("EP")
            )
        except Exception:
            is_spawn = False
        job["_priority"] = 0 if (is_spawn or is_pickup) else 10
        try:
            bind_linked_anim_on_dispatch(
                ext,
                _scr,
                p.name,
                event_seq=str(seq or ""),
                sim_time=str(sim_time or ""),
            )
        except Exception:
            pass
        # 화면별 runner의 busy 여부를 본다.
        runner_busy = False
        try:
            runners = getattr(ext, "_sim_runners_by_screen", None)
            rr = runners.get(str(_scr)) if isinstance(runners, dict) else None
            if rr is None and int(_scr) == 1:
                rr = getattr(ext, "_sim_runner", None)
            runner_busy = bool(rr is not None and getattr(rr, "is_running", lambda: False)())
        except Exception:
            runner_busy = False
        if runner_busy:
            try:
                from . import sim_multi_diag as _mdiag

                _mdiag.log_anim_dispatch(
                    ext,
                    screen=_scr,
                    sim_time=str(sim_time or ""),
                    event=str(seq),
                    file_name=p.name,
                    est_total=float(est_total) if isinstance(est_total, (float, int)) else 0.0,
                    runner_busy=True,
                    decision="QUEUE",
                )
            except Exception:
                pass
            except Exception:
                pass
            pending_list: List[Any] = []
            try:
                pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
                if not isinstance(pending_by, dict):
                    pending_by = {}
                    ext._sim_anim_pending_by_screen = pending_by
                pending = pending_by.get(str(_scr), [])
                if not isinstance(pending, list):
                    pending = []
                if int(job.get("_priority", 10)) <= 0:
                    pending.insert(0, job)
                else:
                    pending.append(job)
                pending_by[str(_scr)] = pending
                pending_list = list(pending)
            except Exception:
                pending_list = []
            _append_anim_history_log(
                ext,
                f"[ANIM] 대기큐적재 | screen={_scr} | event={seq} | est={est_text} | action={action_text} | file={p.name}",
            )
            try:
                notify_anim_queued(
                    ext, _scr, p.name, len(pending_list), p.name
                )
            except Exception:
                pass
            try:
                _refresh_sim_progress_from_last(ext, _scr)
            except Exception:
                pass
            return
        try:
            from . import sim_multi_diag as _mdiag

            _mdiag.log_anim_dispatch(
                ext,
                screen=_scr,
                sim_time=str(sim_time or ""),
                event=str(seq),
                file_name=p.name,
                est_total=float(est_total) if isinstance(est_total, (float, int)) else 0.0,
                runner_busy=False,
                decision="START",
            )
        except Exception:
            pass
        _start_job(job)
    except Exception as e:
        _append_anim_history_log(ext, f"[ANIM] 실행실패 | event={seq} | action={action_text} | file={p.name} | err={e}")
        pause_evt = getattr(ext, "_sim_tick_pause_event", None)
        if pause_evt is not None:
            try:
                pause_evt.clear()
            except Exception:
                pass
        try:
            ext._sim_tick_pause_until_wall = None
        except Exception:
            pass

    if verbose:
        print(
            f"[ANIM MAP] 이벤트={seq} -> JSON 준비완료: {p} "
            f"(runner={runner}, rule={rule_name or '-'}, lot={lot_id_from_payload(payload)}, port={payload.get('port_id','')}, "
            f"from={payload.get('from_port_id','')}, to={payload.get('to_port_id','')})",
            flush=True,
        )


def _estimate_anim_duration_for_gate_payload(ext: Any, payload: Dict[str, str]) -> float:
    """
    simulation_engine의 on_gate에서 호출되는 "애니메이션 예상 길이" 계산기.
    - 게이트 시점에 XML 생성/역파싱 → rules/map 매핑 → JSON 파싱 → 총 duration 추정
    - 실패하면 0.0 반환(=애니 대기 없음)
    """
    try:
        seq_raw = str(payload.get("seq", "") or "").strip()
        if not seq_raw:
            return 0.0
        seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
        fr = str(payload.get("from_port_id", "") or "")
        to = str(payload.get("to_port_id", "") or "")
        port = str(payload.get("port_id", "") or "")

        # 1차: 원본 payload 기준(최소한 map fallback은 항상 시도)
        mapping_payload = dict(payload or {})
        mapping_payload["seq"] = seq

        # 2차: XML 표준화가 가능하면 덮어쓴다(우선 적용)
        try:
            if seq in xml_generator.FROM_TO_SEQS:
                xml_text = xml_generator.build_xml_string(
                    seq,
                    from_port_id=_parse_port_num(fr, 1),
                    to_port_id=_parse_port_num(to, 1),
                )
            elif seq in xml_generator.PORT_ID_ONLY_SEQS:
                xml_text = xml_generator.build_xml_string(seq, port_id=_parse_port_num(port, 1))
            else:
                xml_text = ""
            if xml_text:
                parsed = xml_generator.parse_xml_string(xml_text) or {}
                seq_for_mapping = str(parsed.get("sequence_name", "") or "").strip().upper() or seq
                mapping_payload["seq"] = seq_for_mapping
                mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
                seq = seq_for_mapping
        except Exception:
            # XML 표준화 실패해도 원본 payload로 rules/map 추정을 계속 시도한다.
            pass

        # 3) rules/map 매핑
        mapped_json, _meta, _rule, _src = _resolve_event_animation_entry(seq, mapping_payload)
        if not mapped_json:
            return 0.0

        # 4) JSON 파싱 + 총 길이 추정
        pth = _normalize_json_path(mapped_json)
        if not pth.is_file():
            return 0.0
        parsed_steps = json.loads(pth.read_text(encoding="utf-8"))
        if not isinstance(parsed_steps, list):
            return 0.0
        # 게이트 표시/예상 시간은 1배속 기준으로 유지한다(배속은 진행 속도만 변경).
        est = _estimate_sequence_total_duration_sec_for_log(parsed_steps, speed_scale=1.0)
        return max(0.0, float(est)) if isinstance(est, (float, int)) else 0.0
    except Exception:
        return 0.0


def _capture_per_screen_sim_settings(ext: Any, screen_1based: int = 1) -> Dict[str, Any]:
    """
    화면별 CASE 실시간 UI → 시뮬 엔진용 dict.

    화면1 = CASE A (HUD 와 공유), 화면2 = CASE B.
    """
    try:
        from .ebs_case_models import capture_case_sim_settings_for_screen

        return capture_case_sim_settings_for_screen(ext, int(screen_1based))
    except Exception:
        pass
    return {}


def _refresh_sim_per_screen_status_labels(ext: Any) -> None:
    """
    화면별 시뮼 설정 스냅샷(``_sim_per_screen_snapshots``) 존재 여부를 라벨에 반영한다.

    - 인덱스 i(0-based)는 화면 i+1 과 대응한다.
    - ``_on_save_sim_settings_to_screen`` / 자동 채움 후 호출되어 “(저장됨)/(미저장)” 만 갱신한다.
    """
    labels = getattr(ext, "_sim_per_screen_status_labels", None)
    snaps = getattr(ext, "_sim_per_screen_snapshots", None)
    if not isinstance(labels, list) or not isinstance(snaps, list):
        return
    for i, lbl in enumerate(labels):
        if lbl is None:
            continue
        try:
            ok = i < len(snaps) and snaps[i] is not None
            lbl.text = "(저장됨)" if ok else "(미저장)"
        except Exception:
            pass


def _refresh_sim_per_screen_rows(ext: Any) -> None:
    """분할 수(1~4, 3분할 포함)에 맞춰 화면별 설정 행 표시를 갱신한다."""
    block = getattr(ext, "_sim_per_screen_block", None)
    rows = getattr(ext, "_sim_per_screen_row_hstacks", None)
    if block is None or not isinstance(rows, list):
        return
    try:
        n = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        n = 1
    try:
        block.visible = n > 1
    except Exception:
        pass
    for i, h in enumerate(rows):
        if h is None:
            continue
        try:
            h.visible = n > 1 and (i + 1) <= n
        except Exception:
            pass
    _refresh_sim_per_screen_status_labels(ext)


def _on_save_sim_settings_to_screen(ext: Any, screen_1based: int) -> None:
    """제어창 현재 값을 해당 화면 인덱스(1~) 스냅샷에 저장한다."""
    if screen_1based < 1 or screen_1based > 4:
        return
    try:
        snaps = list(getattr(ext, "_sim_per_screen_snapshots", None) or [None, None, None, None])
    except Exception:
        snaps = [None, None, None, None]
    while len(snaps) < 4:
        snaps.append(None)
    snaps = snaps[:4]
    try:
        snaps[screen_1based - 1] = _capture_per_screen_sim_settings(ext, int(screen_1based))
        ext._sim_per_screen_snapshots = snaps
    except Exception:
        pass
    _refresh_sim_per_screen_status_labels(ext)
    try:
        _append_sim_log(ext, f"[SIM UI] 화면{screen_1based}에 현재 시뮼 설정(LOT·간격·적재/고장·시간) 저장")
    except Exception:
        pass
    try:
        _sync_ep3_port_cell_visibility(ext)
    except Exception:
        pass
    if int(screen_1based) >= 2:
        try:
            ctx_nm = _usd_context_name_for_sim_screen(ext, int(screen_1based))
            if ctx_nm:
                from .tbs_ep_port_visibility import apply_ep_port_layout_for_context

                apply_ep_port_layout_for_context(
                    ext,
                    str(ctx_nm),
                    int(screen_1based),
                    reason="screen_save",
                )
        except Exception:
            pass
    try:
        sim_multi_view.schedule_viewport_snapshot_hud_refresh(ext)
    except Exception:
        pass


def _auto_fill_per_screen_snapshots_on_start(ext: Any) -> None:
    """시작 시: 미저장 화면에는 제어창 현재값을 복사해 둔다(``prompt.md`` 자동 반영)."""
    try:
        n_ch = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        n_ch = 1
    # 요구사항: 기본화면(화면1) 외에는 "저장"을 눌렀을 때만 스냅샷에 반영되어야 한다.
    # 따라서 start 시점에 미저장(None) 화면을 자동으로 dict로 채워 (저장됨) 상태로 만드는 동작은 하지 않는다.
    # 엔진 생성 시에는 아래 on_sim_start_clicked()가 snaps[i]가 None이면 UI 캡처(cap)를 폴백으로 사용한다.
    _refresh_sim_per_screen_status_labels(ext)
    try:
        sim_multi_view.schedule_viewport_snapshot_hud_refresh(ext)
    except Exception:
        pass


def _sync_default_sim_snapshot_from_ui(ext: Any) -> None:
    """CASE A/B 실시간 모델 사용 — 스냅샷 자동 동기화는 사용하지 않는다."""
    return


def _fault_ports_from_snapshot(snap: Dict[str, Any], ep_count: int) -> Set[str]:
    """스냅샷의 고장 포트 체크박스를 집합으로 변환한다."""
    ebs_on = bool(snap.get("ebs_enabled", True)) if isinstance(snap, dict) else True
    out: Set[str] = set()
    if ebs_on:
        pairs = (
            ("INOUT", "fault_inout"),
            ("BP1", "fault_bp1"),
            ("BP2", "fault_bp2"),
            ("BP3", "fault_bp3"),
            ("BP4", "fault_bp4"),
            ("EP1", "fault_ep1"),
            ("EP2", "fault_ep2"),
            ("EP3", "fault_ep3"),
        )
    else:
        pairs = (
            ("EP1", "fault_ep1"),
            ("EP2", "fault_ep2"),
            ("EP3", "fault_ep3"),
        )
    for port, key in pairs:
        try:
            if bool(snap.get(key)):
                out.add(port)
        except Exception:
            pass
    if ep_count < 3:
        out.discard("EP3")
        out.discard("BP4")
    return out


def _timing_and_init_from_snapshot(ext: Any, snap: Dict[str, Any]) -> Tuple[SimulationTimingConfig, SimulationInitConfig]:
    """화면별 스냅샷으로 채널 전용 ``SimulationTimingConfig`` / ``SimulationInitConfig`` 를 만든다."""
    from .ebs_case_models import ep_count_from_snapshot

    ep_count = ep_count_from_snapshot(snap, default=int(_SIM_DEF.ep_count()))
    ep_count_idx = 1 if ep_count >= 3 else 0
    ebs_enabled = bool(snap.get("ebs_enabled", True))
    initial_full_ports: List[str] = []
    if ebs_enabled:
        if bool(snap.get("init_inout")):
            initial_full_ports.append("INOUT")
        if bool(snap.get("init_bp1")):
            initial_full_ports.append("BP1")
        if bool(snap.get("init_bp2")):
            initial_full_ports.append("BP2")
        if bool(snap.get("init_bp3")):
            initial_full_ports.append("BP3")
        if ep_count >= 3 and bool(snap.get("init_bp4")):
            initial_full_ports.append("BP4")
    if bool(snap.get("init_ep1")):
        initial_full_ports.append("EP1")
    if bool(snap.get("init_ep2")):
        initial_full_ports.append("EP2")
    if ep_count >= 3 and bool(snap.get("init_ep3")):
        initial_full_ports.append("EP3")
    try:
        spawn_imin = max(0.1, float(snap.get("spawn_min", _SIM_DEF.lot_spawn_min)))
        spawn_imax = max(0.1, float(snap.get("spawn_max", _SIM_DEF.lot_spawn_max)))
    except Exception:
        spawn_imin, spawn_imax = float(_SIM_DEF.lot_spawn_min), float(_SIM_DEF.lot_spawn_max)
    if spawn_imin > spawn_imax:
        spawn_imin, spawn_imax = spawn_imax, spawn_imin
    try:
        pue_min = max(0.1, float(snap.get("pue_min", _SIM_DEF.pickup_min)))
        pue_max = max(0.1, float(snap.get("pue_max", _SIM_DEF.pickup_max)))
    except Exception:
        pue_min, pue_max = float(_SIM_DEF.pickup_min), float(_SIM_DEF.pickup_max)
    if pue_min > pue_max:
        pue_min, pue_max = pue_max, pue_min

    def _f_snap(key: str, default: float = 5.0) -> float:
        try:
            return max(0.1, float(snap.get(key, default)))
        except Exception:
            return default

    timing = SimulationTimingConfig(
        oht_to_bp1_min=_f_snap("oht_bp1_min", float(_SIM_DEF.oht_to_bp1_min)),
        oht_to_bp1_max=_f_snap("oht_bp1_max", float(_SIM_DEF.oht_to_bp1_max)),
        bp1_to_bp_min=_f_snap("bp1_bp_min", float(_SIM_DEF.bp1_to_bp_min)),
        bp1_to_bp_max=_f_snap("bp1_bp_max", float(_SIM_DEF.bp1_to_bp_max)),
        bp_to_ep_min=_f_snap("bp_ep_min", float(_SIM_DEF.bp_to_ep_min)),
        bp_to_ep_max=_f_snap("bp_ep_max", float(_SIM_DEF.bp_to_ep_max)),
        ep_to_oht_min=_f_snap("ep_oht_min", float(_SIM_DEF.ep_to_oht_min)),
        ep_to_oht_max=_f_snap("ep_oht_max", float(_SIM_DEF.ep_to_oht_max)),
        lot_spawn_interval_min=spawn_imin,
        lot_spawn_interval_max=spawn_imax,
        pickup_event_interval_min=pue_min,
        pickup_event_interval_max=pue_max,
        foup_process_min=_f_snap("foup_proc_min", float(_SIM_DEF.foup_process_min)),
        foup_process_max=_f_snap("foup_proc_max", float(_SIM_DEF.foup_process_max)),
    )
    try:
        lot_count = max(1, int(snap.get("lot_count", _SIM_DEF.lot_count) or _SIM_DEF.lot_count))
    except Exception:
        lot_count = int(_SIM_DEF.lot_count)
    # 요구사항: 공정시간 우선(process_time_priority) 기능은 사용하지 않음
    proc_pri = False
    init = SimulationInitConfig(
        ep_count=ep_count,
        ebs_enabled=bool(ebs_enabled),
        initial_full_ports=initial_full_ports,
        max_oht_lots=lot_count,
        process_time_priority=proc_pri,
    )
    return timing, init


def _inject_lot_fix_proc_into_init(ext: Any, init_cfg: SimulationInitConfig) -> None:
    """fix 공정 입력 창이 비어 있으면 ``init_cfg`` 를 그대로 둔다 (기존 랜덤 동작)."""
    rows = read_lot_fix_proc_at_start(ext)
    if rows:
        init_cfg.lot_fix_proc_rows = rows


def notify_tbs_composed_usd_ready_for_split(ext: Any, usd_path: str = "") -> None:
    """합성 Master USD open 성공 후 분할화면 체크박스 행을 켠다 (구 ``load_window`` Load 완료와 동일)."""
    try:
        from . import sim_multi_view

        sim_multi_view.invalidate_split_layout_cache(ext)
    except Exception:
        pass
    p = str(usd_path or "").strip()
    if p:
        try:
            from .tbs_data_paths import resolve_local_data_path

            resolved = resolve_local_data_path(p) or p
            ext._tbs_last_loaded_usd_path = str(resolved).strip()
        except Exception:
            try:
                ext._tbs_last_loaded_usd_path = p
            except Exception:
                pass
    try:
        from .tbs_split_composed_loader import (
            register_main_composed_runtime,
            split_dual_usd_paths_enabled,
        )

        register_main_composed_runtime(ext)
    except Exception:
        pass
    try:
        from .tbs_split_composed_loader import schedule_split_composed_snapshot_prewarm

        if not split_dual_usd_paths_enabled(ext):
            schedule_split_composed_snapshot_prewarm(ext)
    except Exception:
        pass
    try:
        ext._tbs_multi_split_usd_ready = True
    except Exception:
        pass
    try:
        fn = getattr(ext, "_sync_sim_multi_split_row_visibility_fn", None)
        if callable(fn):
            fn(ext)
    except Exception:
        pass

    async def _sync_after_stage_settles() -> None:
        try:
            await app.get_app().next_update_async()
        except Exception:
            return
        f2 = getattr(ext, "_sync_sim_multi_split_row_visibility_fn", None)
        if callable(f2):
            try:
                f2(ext)
            except Exception:
                pass
        try:
            _refresh_sim_per_screen_rows(ext)
        except Exception:
            pass

    try:
        asyncio.ensure_future(_sync_after_stage_settles())
    except Exception:
        pass
    try:
        from . import sim_multi_view

        if bool(getattr(ext, "_tbs_split_deferred_aux_load_pending", False)):
            sim_multi_view.schedule_deferred_aux_usd_load_after_master(ext)
        else:
            sim_multi_view.schedule_split_rebuild_after_master_reload(ext)
    except Exception:
        pass

    async def _refresh_main_vp_after_open() -> None:
        try:
            from . import sim_multi_view

            n = int(getattr(ext, "_sim_viewport_split_count", 2) or 2)
            await sim_multi_view.wake_main_viewport_after_master_open(ext, n)
        except Exception:
            pass

    try:
        asyncio.ensure_future(_refresh_main_vp_after_open())
    except Exception:
        pass


def _preserve_split_layout_during_startup(ext: Any) -> bool:
    try:
        from . import sim_multi_view

        return bool(sim_multi_view.preserve_split_layout_during_startup(ext))
    except Exception:
        return False


def _sync_sim_multi_split_row_visibility(ext: Any) -> None:
    """
    제어창의 **시뮼 뷰포트 분할** 설정 행(``_sim_multi_split_row``) 표시 여부를 갱신한다.

    - USD 파일 스테이지가 로드되고(``is_usd_file_stage_loaded``)·내부 플래그 ``_tbs_multi_split_usd_ready`` 가 참일 때만 표시.
    - 조건 미충족 시 ``_force_sim_split_to_default`` 로 분할 UI·카운트를 1로 되돌린다.
    """
    row = getattr(ext, "_sim_multi_split_row", None)
    if row is None:
        return
    try:
        if get_stage() is None:
            if not _preserve_split_layout_during_startup(ext):
                try:
                    ext._tbs_multi_split_usd_ready = False
                except Exception:
                    pass
    except Exception:
        pass
    try:
        st_ok = bool(is_usd_file_stage_loaded())
        ready = bool(getattr(ext, "_tbs_multi_split_usd_ready", False))
        row.visible = bool(ready and st_ok)
    except Exception:
        try:
            row.visible = False
        except Exception:
            pass
    if not getattr(row, "visible", False):
        if _preserve_split_layout_during_startup(ext):
            return
        _force_sim_split_to_default(ext)


def _sync_sim_split_checkboxes_from_ext_count(ext: Any) -> None:
    """``ext._sim_viewport_split_count``(실제 적용 분할 수)에 맞춰 분할 체크박스를 맞춘다. ``apply`` 를 호출하지 않는다."""
    if getattr(ext, "_sim_split_mutate_guard", False):
        return
    try:
        from .sim_control_defaults import MAX_VIEWPORT_SPLIT_COUNT

        cap = max(1, int(MAX_VIEWPORT_SPLIT_COUNT))
    except Exception:
        cap = 2
    try:
        n = int(getattr(ext, "_sim_viewport_split_count", 1) or 1)
    except Exception:
        n = 1
    n = max(1, min(cap, n))
    models = getattr(ext, "_sim_split_cb_models", None)
    if not isinstance(models, list) or len(models) < 1:
        return
    ext._sim_split_mutate_guard = True
    try:
        for i, m in enumerate(models, start=1):
            if i > cap:
                break
            try:
                m.set_value(i == n)
            except Exception:
                pass
    finally:
        ext._sim_split_mutate_guard = False
    try:
        _refresh_sim_per_screen_rows(ext)
    except Exception:
        pass


def _force_sim_split_to_default(ext: Any) -> None:
    """분할 UI를 1개 시뮼만 선택된 상태로 되돌린다."""
    if _preserve_split_layout_during_startup(ext):
        return
    ext._sim_split_mutate_guard = True
    try:
        models = getattr(ext, "_sim_split_cb_models", None)
        if isinstance(models, list) and len(models) >= 1:
            for i, m in enumerate(models, start=1):
                try:
                    m.set_value(i == 1)
                except Exception:
                    pass
        try:
            ext._sim_viewport_split_count = 1
        except Exception:
            pass
    finally:
        ext._sim_split_mutate_guard = False
    try:
        ext._sim_per_screen_snapshots = [None, None, None, None]
    except Exception:
        pass
    try:
        sim_multi_view.apply_sim_viewport_split_layout(ext, 1)
    except Exception:
        pass
    try:
        _refresh_sim_per_screen_rows(ext)
    except Exception:
        pass


def _on_sim_split_choice_changed(ext: Any, idx: int, m: Any) -> None:
    """1~2 상호 배타 체크 + 분할 스텁 적용."""
    if getattr(ext, "_sim_split_mutate_guard", False):
        return
    try:
        from .sim_control_defaults import MAX_VIEWPORT_SPLIT_COUNT

        if int(idx) < 1 or int(idx) > int(MAX_VIEWPORT_SPLIT_COUNT):
            return
    except Exception:
        pass
    try:
        if not m.get_value_as_bool():
            cur = int(getattr(ext, "_sim_viewport_split_count", 1) or 1)
            if cur == idx:
                ext._sim_split_mutate_guard = True
                try:
                    m.set_value(True)
                finally:
                    ext._sim_split_mutate_guard = False
            return
        ext._sim_split_mutate_guard = True
        try:
            models = list(getattr(ext, "_sim_split_cb_models", []) or [])
            for j, md in enumerate(models, start=1):
                if j != idx:
                    try:
                        if md.get_value_as_bool():
                            md.set_value(False)
                    except Exception:
                        pass
        finally:
            ext._sim_split_mutate_guard = False
        sim_multi_view.apply_sim_viewport_split_layout(ext, idx)
    except Exception as err:
        try:
            print(f"[TBS multi-sim] split UI err: {err}", flush=True)
        except Exception:
            pass


def _usd_context_name_for_sim_screen(ext: Any, screen: int) -> Optional[str]:
    """
    시뮼 **화면 인덱스**에 대응하는 USD 컨텍스트 이름을 반환한다.

    - 화면 1: ``None`` → 기본 ``omni.usd`` 컨텍스트.
    - 화면 2 이상: ``sim_multi_view`` 가 실제 생성한 이름(``ext._sim_multi_context_names``) 우선,
      없으면 ``morph_tbs_split_aux_{screen-1}`` (절대 기본 ctx 로 폴백하지 않음 — 화면1 오염 방지).
    """
    try:
        s = int(screen)
    except Exception:
        s = 1
    if s <= 1:
        return None
    try:
        names = list(getattr(ext, "_sim_multi_context_names", []) or [])
    except Exception:
        names = []
    idx = s - 2
    if 0 <= idx < len(names):
        nm = str(names[idx] or "").strip()
        if nm:
            return nm
    return f"morph_tbs_split_aux_{s - 1}"


def _resolve_payload_sim_screen(ext: Any, payload: Optional[Dict[str, Any]]) -> Optional[int]:
    """payload ``tbs_sim_screen`` → 1..N. 멀티에서 태그 없으면 None(호출부가 drop)."""
    n = _sim_monitor_channel_count(ext)
    raw = ""
    try:
        if isinstance(payload, dict):
            raw = str(payload.get("tbs_sim_screen", "") or "").strip()
    except Exception:
        raw = ""
    if not raw:
        return 1 if n <= 1 else None
    try:
        scr = int(raw)
    except Exception:
        return 1 if n <= 1 else None
    if scr < 1:
        return 1 if n <= 1 else None
    return max(1, min(int(n), int(scr)))


def _sim_monitor_channel_count(ext: Any) -> int:
    """
    시뮼 모니터(포트상태·진행현황·SIM 로그) 열 개수를 반환한다.

    ``ext._sim_viewport_split_count``(1~4)와 동일하며, ``_rebuild_sim_monitor_split_ui`` 의 열 수와 맞춘다.
    """
    return max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))


def _sim_use_per_screen_windows(ext: Any) -> bool:
    """뷰포트 2분할 이상이면 모니터·타임테이블을 화면별 별도 ``ui.Window`` 로 분리한다."""
    return _sim_monitor_channel_count(ext) >= 2


def _sim_monitor_window_title(screen: int, n: int) -> str:
    if int(n) <= 1:
        return "TBS 시뮬 모니터"
    return f"TBS 시뮬 모니터 (화면{int(screen)})"


def _sim_timetable_window_title(screen: int, n: int) -> str:
    if int(n) <= 1:
        return "TBS 타임테이블"
    return f"TBS 타임테이블 (화면{int(screen)})"


def _destroy_workspace_window_by_title(title: str) -> None:
    try:
        ws = getattr(ui, "Workspace", None)
        if ws is not None and hasattr(ws, "get_window"):
            old = ws.get_window(str(title))
            if old is not None:
                try:
                    old.destroy()
                except Exception:
                    try:
                        old.visible = False
                    except Exception:
                        pass
    except Exception:
        pass


def _screen_dict(ext: Any, attr: str) -> Dict[str, Any]:
    by = getattr(ext, attr, None)
    if not isinstance(by, dict):
        by = {}
        setattr(ext, attr, by)
    return by


def _iter_sim_monitor_windows(ext: Any) -> List[Any]:
    n = _sim_monitor_channel_count(ext)
    if _sim_use_per_screen_windows(ext):
        by = _screen_dict(ext, "_sim_monitor_windows_by_screen")
        out: List[Any] = []
        for s in range(1, n + 1):
            w = by.get(str(s))
            if w is not None:
                out.append(w)
        return out
    w = getattr(ext, "_sim_monitor_window", None)
    return [w] if w is not None else []


def _iter_sim_timetable_windows(ext: Any) -> List[Any]:
    n = _sim_monitor_channel_count(ext)
    if _sim_use_per_screen_windows(ext):
        by = _screen_dict(ext, "_sim_timetable_windows_by_screen")
        out: List[Any] = []
        for s in range(1, n + 1):
            w = by.get(str(s))
            if w is not None:
                out.append(w)
        return out
    w = getattr(ext, "_sim_timetable_window", None)
    return [w] if w is not None else []


def _ensure_sim_monitor_window_shell(ext: Any, screen: int) -> None:
    """화면 ``screen`` 용 시뮬 모니터 ``ui.Window`` 셸(FOUP·split host)을 보장한다."""
    key = str(int(screen))
    wins = _screen_dict(ext, "_sim_monitor_windows_by_screen")
    if wins.get(key) is not None:
        return

    n = _sim_monitor_channel_count(ext)
    title = _sim_monitor_window_title(int(screen), n)
    _destroy_workspace_window_by_title(title)

    foup_hosts = _screen_dict(ext, "_sim_foup_outer_host_by_screen")
    foup_inners = _screen_dict(ext, "_sim_foup_inner_stack_by_screen")
    split_hosts = _screen_dict(ext, "_sim_monitor_split_host_by_screen")

    win = ui.Window(title, width=650, height=620)
    with win.frame:
        with ui.VStack(spacing=0, height=ui.Fraction(1.0)):
            with ui.Frame(style={"background_color": 0xFF1E2530}, height=ui.Fraction(1.0)):
                with ui.VStack(padding=8, spacing=4, height=ui.Fraction(1.0)):
                    if int(screen) == 1:
                        with ui.HStack(spacing=8, height=28):
                            ui.Label(
                                "시뮬 진행 모니터",
                                width=140,
                                height=24,
                                style={"color": 0xFFDDDDDD},
                            )
                            ui.Spacer()
                            ui.Button(
                                "진행현황+Sim로그 복사",
                                width=180,
                                clicked_fn=lambda: on_copy_sim_progress(ext),
                            )
                    foup_hdr = (
                        "FOUP 공정"
                        if n <= 1
                        else f"FOUP 공정 · 화면{int(screen)}"
                    )
                    ui.Label(foup_hdr, height=18, style={"color": 0xFFBFE7FF})
                    with ui.Frame(height=78, style={"background_color": 0xFF1A1E26, "border_width": 0}):
                        foup_host = ui.VStack(spacing=2, height=74)
                        with foup_host:
                            foup_inner = ui.VStack(spacing=2)
                    split_host = ui.Frame(
                        height=ui.Fraction(1.0),
                        style={"background_color": 0xFF1A1E26, "border_width": 1, "border_color": 0xFF3A3A3A},
                    )

    wins[key] = win
    foup_hosts[key] = foup_host
    foup_inners[key] = foup_inner
    split_hosts[key] = split_host

    if int(screen) == 1:
        ext._sim_monitor_window = win
        ext._sim_foup_outer_host = foup_host
        ext._sim_foup_inner_stack = foup_inner
        ext._sim_monitor_split_host = split_host
        try:
            ext._sim_foup_layout_n = 0
            ext._sim_foup_labels_by_screen = {}
        except Exception:
            pass
        if getattr(ext, "_rebuild_sim_monitor_split_ui_fn", None) is None:
            ext._rebuild_sim_monitor_split_ui_fn = lambda: _rebuild_all_sim_ui_panels(ext)
        if getattr(ext, "_rebuild_sim_timetable_split_ui_fn", None) is None:
            ext._rebuild_sim_timetable_split_ui_fn = lambda: _rebuild_all_sim_ui_panels(ext)
        if getattr(ext, "_sim_port_state_label", None) is None:
            ext._sim_port_state_label = ui.Label(
                "", word_wrap=False, width=0, height=0, visible=False
            )


def _destroy_sim_monitor_window_shell(ext: Any, screen: int) -> None:
    key = str(int(screen))
    wins = _screen_dict(ext, "_sim_monitor_windows_by_screen")
    win = wins.pop(key, None)
    if win is not None:
        try:
            win.destroy()
        except Exception:
            pass
    for attr in (
        "_sim_foup_outer_host_by_screen",
        "_sim_foup_inner_stack_by_screen",
        "_sim_monitor_split_host_by_screen",
        "_sim_monitor_split_inner_by_screen",
    ):
        d = getattr(ext, attr, None)
        if isinstance(d, dict):
            d.pop(key, None)


def _sync_sim_monitor_window_shells(ext: Any) -> None:
    """분할 수에 맞춰 모니터 창을 1개(단일) 또는 화면별 N개로 맞춘다."""
    n = _sim_monitor_channel_count(ext)
    if _sim_use_per_screen_windows(ext):
        for s in range(1, n + 1):
            _ensure_sim_monitor_window_shell(ext, s)
        for s in range(n + 1, 5):
            _destroy_sim_monitor_window_shell(ext, s)
    else:
        _ensure_sim_monitor_window_shell(ext, 1)
        for s in range(2, 5):
            _destroy_sim_monitor_window_shell(ext, s)


def _ensure_sim_timetable_window_shell(ext: Any, screen: int) -> None:
    """화면 ``screen`` 용 타임테이블 ``ui.Window`` 셸을 보장한다."""
    key = str(int(screen))
    wins = _screen_dict(ext, "_sim_timetable_windows_by_screen")
    if wins.get(key) is not None:
        return

    n = _sim_monitor_channel_count(ext)
    title = _sim_timetable_window_title(int(screen), n)
    _destroy_workspace_window_by_title(title)

    split_hosts = _screen_dict(ext, "_sim_timetable_split_host_by_screen")
    hdr = (
        "프리런 타임테이블 — 행 클릭으로 Seek"
        if n <= 1
        else f"프리런 타임테이블 · 화면{int(screen)} — 행 클릭으로 Seek"
    )

    win = ui.Window(title, width=450, height=560)
    with win.frame:
        with ui.VStack(spacing=0, height=ui.Fraction(1.0)):
            with ui.Frame(style={"background_color": 0xFF1E2530}, height=ui.Fraction(1.0)):
                with ui.VStack(padding=8, spacing=4, height=ui.Fraction(1.0)):
                    ui.Label(hdr, height=22, style={"color": 0xFFBFE7FF, "font_size": 13})
                    split_host = ui.Frame(
                        height=ui.Fraction(1.0),
                        style={"background_color": 0xFF1A1E26, "border_width": 1, "border_color": 0xFF3A3A3A},
                    )

    wins[key] = win
    split_hosts[key] = split_host
    if int(screen) == 1:
        ext._sim_timetable_window = win
        ext._sim_timetable_split_host = split_host
    _bind_sim_timetable_window_visibility(ext, win)


def _destroy_sim_timetable_window_shell(ext: Any, screen: int) -> None:
    key = str(int(screen))
    wins = _screen_dict(ext, "_sim_timetable_windows_by_screen")
    win = wins.pop(key, None)
    if win is not None:
        try:
            win.destroy()
        except Exception:
            pass
    for attr in ("_sim_timetable_split_host_by_screen", "_sim_timetable_split_inner_by_screen"):
        d = getattr(ext, attr, None)
        if isinstance(d, dict):
            d.pop(key, None)


def _sync_sim_timetable_window_shells(ext: Any) -> None:
    """분할 수에 맞춰 타임테이블 창을 1개(단일) 또는 화면별 N개로 맞춘다."""
    n = _sim_monitor_channel_count(ext)
    if _sim_use_per_screen_windows(ext):
        for s in range(1, n + 1):
            _ensure_sim_timetable_window_shell(ext, s)
        for s in range(n + 1, 5):
            _destroy_sim_timetable_window_shell(ext, s)
    else:
        _ensure_sim_timetable_window_shell(ext, 1)
        for s in range(2, 5):
            _destroy_sim_timetable_window_shell(ext, s)


def _monitor_split_host_for_screen(ext: Any, screen: int) -> Any:
    if _sim_use_per_screen_windows(ext):
        hosts = _screen_dict(ext, "_sim_monitor_split_host_by_screen")
        return hosts.get(str(int(screen)))
    return getattr(ext, "_sim_monitor_split_host", None)


def _timetable_split_host_for_screen(ext: Any, screen: int) -> Any:
    if _sim_use_per_screen_windows(ext):
        hosts = _screen_dict(ext, "_sim_timetable_split_host_by_screen")
        return hosts.get(str(int(screen)))
    return getattr(ext, "_sim_timetable_split_host", None)


def _clear_foup_inner_in_outer(outer: Any, inner: Any = None) -> Any:
    """지정 FOUP outer VStack 아래 inner 를 비우고 반환한다(추적 inner 우선).

    ``_clear_foup_inner_stack`` 과 동일 전략: 추적 중인 ``inner`` 를 그대로 ``clear()`` 한다.
    ``inner`` 가 없거나 outer 아래 VStack 이 여러 개로 누적된 경우에만 outer 를 통째로
    비우고 새 inner 를 만든다. (``outer.children[0]`` 재추정 금지 — 타이밍에 따라 새 inner 가
    추가되어 기존 라벨이 destroy 되지 않고 누적되던 버그 방지.)
    """
    if outer is None:
        return None
    try:
        kids = list(getattr(outer, "children", []) or [])
        if inner is None or len(kids) > 1:
            try:
                outer.clear()
            except Exception:
                pass
            inner = None
    except Exception:
        pass
    if inner is None:
        with outer:
            inner = ui.VStack(spacing=2)
    if inner is None:
        return None
    try:
        inner.clear()
    except Exception:
        try:
            for child in list(getattr(inner, "children", []) or []):
                try:
                    child.destroy()
                except Exception:
                    pass
        except Exception:
            pass
    return inner


def _clear_monitor_split_inner(ext: Any, host: Any, *, screen_key: Optional[str] = None) -> Any:
    if screen_key is not None:
        inners = _screen_dict(ext, "_sim_monitor_split_inner_by_screen")
        inn = inners.pop(screen_key, None)
    else:
        inn = getattr(ext, "_sim_monitor_split_inner", None)
        try:
            ext._sim_monitor_split_inner = None
        except Exception:
            pass
    if inn is not None:
        try:
            inn.destroy()
        except Exception:
            pass
    if host is None:
        return None
    with host:
        new_inner = ui.VStack(spacing=4, height=ui.Fraction(1.0))
    if screen_key is not None:
        _screen_dict(ext, "_sim_monitor_split_inner_by_screen")[screen_key] = new_inner
    else:
        ext._sim_monitor_split_inner = new_inner
    return new_inner


def _clear_timetable_split_inner(ext: Any, host: Any, *, screen_key: Optional[str] = None) -> Any:
    if screen_key is not None:
        inners = _screen_dict(ext, "_sim_timetable_split_inner_by_screen")
        inn = inners.pop(screen_key, None)
    else:
        inn = getattr(ext, "_sim_timetable_split_inner", None)
        try:
            ext._sim_timetable_split_inner = None
        except Exception:
            pass
    if inn is not None:
        try:
            inn.destroy()
        except Exception:
            pass
    if host is None:
        return None
    with host:
        new_inner = ui.VStack(spacing=4, height=ui.Fraction(1.0))
    if screen_key is not None:
        _screen_dict(ext, "_sim_timetable_split_inner_by_screen")[screen_key] = new_inner
    else:
        ext._sim_timetable_split_inner = new_inner
    return new_inner


def _snapshot_foup_label_state(ext: Any) -> Dict[int, Dict[str, Tuple[str, int]]]:
    """재빌드 전 화면별·EP별 FOUP 라벨 (text, color) 스냅샷."""
    out: Dict[int, Dict[str, Tuple[str, int]]] = {}
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list):
        return out
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si <= 0:
            continue
        labels = ch.get("foup_progress_labels")
        if not isinstance(labels, dict):
            continue
        row: Dict[str, Tuple[str, int]] = {}
        for ep_id, lbl in labels.items():
            if lbl is None:
                continue
            try:
                txt = str(getattr(lbl, "text", "") or "")
                col = int(getattr(lbl, "style", {}).get("color", 0xFF888888))  # type: ignore[union-attr]
            except Exception:
                txt, col = "", 0xFF888888
            row[str(ep_id)] = (txt, int(col))
        if row:
            out[si] = row
    return out


def _foup_label_idle_text(screen_num: int, ep_id: str) -> str:
    """EP 별 FOUP 라벨의 'idle(대기)' 텍스트를 한 곳에서 생성한다(생성/리셋 일관성)."""
    if int(screen_num or 1) <= 1:
        return f"{ep_id} FOUP 공정: 대기"
    return f"{ep_id} FOUP 공정(화면{int(screen_num)}): 대기"


def _make_foup_line_label(screen: int, ep_id: str) -> Any:
    """FOUP 한 줄 — 불투명 배경 ZStack 으로 형제 repaint 전파를 줄인다."""
    idle_text = _foup_label_idle_text(int(screen), str(ep_id))
    lbl: Any = None
    with ui.Frame(height=22):
        with ui.ZStack():
            ui.Rectangle(
                width=ui.Fraction(1.0),
                height=22,
                style={"background_color": 0xFF1A1E26},
            )
            lbl = ui.Label(
                idle_text,
                word_wrap=False,
                height=22,
                style={"color": 0xFF888888},
            )
    return lbl


def _create_foup_labels_in_vstack(ext: Any, ch: Dict[str, Any], screen: int) -> Dict[str, Any]:
    """채널 dict 에 EP1~3 FOUP Label 을 만들고 ext·채널에 등록한다(호출부가 VStack context 안)."""
    labels: Dict[str, Any] = {}
    if not isinstance(ch.get("_foup_label_cache"), dict):
        ch["_foup_label_cache"] = {}
    for ep_id in ("EP1", "EP2", "EP3"):
        labels[ep_id] = _make_foup_line_label(int(screen), str(ep_id))
    ch["foup_progress_labels"] = labels
    ch["foup_progress_label"] = labels.get("EP1")
    try:
        by = getattr(ext, "_sim_foup_labels_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_foup_labels_by_screen = by
        by[int(screen)] = dict(labels)
    except Exception:
        pass
    return labels


def _restore_foup_label_state(channels: List[Dict[str, Any]], saved: Dict[int, Dict[str, Tuple[str, int]]]) -> None:
    for ch in channels:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        row = saved.get(si)
        if not isinstance(row, dict):
            continue
        labels = ch.get("foup_progress_labels")
        if not isinstance(labels, dict):
            continue
        for ep_id, pair in row.items():
            lbl = labels.get(str(ep_id))
            if lbl is None or not isinstance(pair, tuple) or len(pair) < 2:
                continue
            txt, col = str(pair[0]), int(pair[1])
            _set_foup_progress_label(ch, str(ep_id), lbl, txt, {"color": col})


def _clear_foup_inner_stack(ext: Any) -> Any:
    """FOUP 전용 inner VStack 을 비운다(``destroy()`` 누적 방지 — ``clear()`` 우선)."""
    outer = getattr(ext, "_sim_foup_outer_host", None)
    if outer is None:
        return None
    inner = getattr(ext, "_sim_foup_inner_stack", None)
    # 과거 ``child.destroy()`` 실패로 outer 아래 VStack 이 여러 개 쌓인 경우 일괄 제거.
    try:
        kids = list(getattr(outer, "children", []) or [])
        if inner is None or len(kids) > 1:
            try:
                outer.clear()
            except Exception:
                pass
            inner = None
            ext._sim_foup_inner_stack = None
    except Exception:
        pass
    if inner is None:
        with outer:
            ext._sim_foup_inner_stack = ui.VStack(spacing=2)
        inner = getattr(ext, "_sim_foup_inner_stack", None)
    if inner is None:
        return None
    try:
        inner.clear()
    except Exception:
        try:
            for child in list(getattr(inner, "children", []) or []):
                try:
                    child.destroy()
                except Exception:
                    pass
        except Exception:
            pass
    return inner


def _foup_labels_mounted(labels: Any) -> bool:
    if not isinstance(labels, dict) or len(labels) < 3:
        return False
    for lbl in labels.values():
        if lbl is None:
            return False
    return True


def _foup_progress_labels_for_screen(ext: Any, ch: Optional[Dict[str, Any]], screen: int) -> Dict[str, Any]:
    """채널 dict 또는 ``_sim_foup_labels_by_screen`` 에서 EP별 FOUP 라벨 dict."""
    si = max(1, int(screen))
    if isinstance(ch, dict):
        labels = ch.get("foup_progress_labels")
        if isinstance(labels, dict) and _foup_labels_mounted(labels):
            return labels
    by = getattr(ext, "_sim_foup_labels_by_screen", None)
    if isinstance(by, dict):
        labels = by.get(si) or by.get(str(si))
        if isinstance(labels, dict) and _foup_labels_mounted(labels):
            if isinstance(ch, dict):
                ch["foup_progress_labels"] = dict(labels)
                ch["foup_progress_label"] = labels.get("EP1")
            return labels
    return {}


def _refresh_all_foup_playback_heartbeats(ext: Any) -> None:
    """N>1 재생 — 화면마다 FOUP 공정 라벨 heartbeat (한 화면만 갱신되는 누락 방지)."""
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    try:
        from .control_sim_screen_playback import iter_sim_playback_players

        for scr, player in iter_sim_playback_players(ext):
            if player is None:
                continue
            try:
                tnow = float(player.sim_now(int(scr)))
            except Exception:
                tnow = 0.0
            _refresh_foup_playback_heartbeat(ext, int(scr), tnow)
    except Exception:
        pass


def _rebind_foup_labels_to_channels(ext: Any, channels: List[Dict[str, Any]]) -> None:
    """기존 FOUP 위젯 참조만 채널 dict 에 연결(위젯 재생성·clear 없음)."""
    by = getattr(ext, "_sim_foup_labels_by_screen", None)
    if not isinstance(by, dict):
        by = {}
        ext._sim_foup_labels_by_screen = by
    for ch in channels:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si <= 0:
            continue
        labels = by.get(si) or by.get(str(si))
        if not _foup_labels_mounted(labels):
            continue
        ch["foup_progress_labels"] = labels
        ch["foup_progress_label"] = labels.get("EP1")
        if not isinstance(ch.get("_foup_label_cache"), dict):
            ch["_foup_label_cache"] = {}


def _foup_layout_ready(ext: Any, channels: List[Dict[str, Any]]) -> bool:
    by = getattr(ext, "_sim_foup_labels_by_screen", None)
    if not isinstance(by, dict):
        return False
    for ch in channels:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si <= 0:
            continue
        if not _foup_labels_mounted(by.get(si)):
            return False
    return bool(channels)


def _rebuild_sim_foup_outer_row(ext: Any, channels: List[Dict[str, Any]]) -> None:
    """
    FOUP 공정 3줄을 모니터 창 상단 고정 영역에 그린다.

    단일 화면: 기존처럼 한 창에 배치. 2분할 이상: 화면별 모니터 창마다 해당 화면 FOUP 만 그린다.
    """
    if not channels:
        return
    if _sim_use_per_screen_windows(ext):
        # 레이아웃·라벨이 이미 준비되어 있으면 재생성 없이 재연결만(누적·깜빡임 방지).
        layout_n = len(channels)
        prev_n = int(getattr(ext, "_sim_foup_layout_n", 0) or 0)
        if layout_n == prev_n and _foup_layout_ready(ext, channels):
            _rebind_foup_labels_to_channels(ext, channels)
            return
        saved = _snapshot_foup_label_state(ext)
        foup_hosts = _screen_dict(ext, "_sim_foup_outer_host_by_screen")
        foup_inners = _screen_dict(ext, "_sim_foup_inner_stack_by_screen")
        try:
            ext._sim_foup_labels_by_screen = {}
        except Exception:
            pass
        for ch in channels:
            if not isinstance(ch, dict):
                continue
            try:
                si = int(ch.get("screen", 1) or 1)
            except Exception:
                si = 1
            key = str(si)
            outer = foup_hosts.get(key)
            if outer is None:
                continue
            inner = _clear_foup_inner_in_outer(outer, foup_inners.get(key))
            if inner is None:
                continue
            foup_inners[key] = inner
            with inner:
                _create_foup_labels_in_vstack(ext, ch, si)
        _restore_foup_label_state(channels, saved)
        _rebind_foup_labels_to_channels(ext, channels)
        try:
            ext._sim_foup_layout_n = len(channels)
        except Exception:
            pass
        return

    layout_n = len(channels)
    prev_n = int(getattr(ext, "_sim_foup_layout_n", 0) or 0)
    if layout_n == prev_n and _foup_layout_ready(ext, channels):
        _rebind_foup_labels_to_channels(ext, channels)
        return
    saved = _snapshot_foup_label_state(ext)
    inner = _clear_foup_inner_stack(ext)
    if inner is None:
        return
    try:
        ext._sim_foup_layout_n = int(layout_n)
        ext._sim_foup_labels_by_screen = {}
    except Exception:
        pass
    with inner:
        if len(channels) == 1:
            _create_foup_labels_in_vstack(ext, channels[0], int(channels[0].get("screen", 1) or 1))
        else:
            with ui.HStack(spacing=8):
                for ch in channels:
                    if not isinstance(ch, dict):
                        continue
                    try:
                        si = int(ch.get("screen", 1) or 1)
                    except Exception:
                        si = 1
                    with ui.VStack(spacing=2, width=ui.Fraction(1.0)):
                        ui.Label(f"화면{si}", height=14, style={"color": 0xFF9AA4B2})
                        _create_foup_labels_in_vstack(ext, ch, si)
    _restore_foup_label_state(channels, saved)
    _rebind_foup_labels_to_channels(ext, channels)


def _sync_foup_labels_to_channels(ext: Any, channels: List[Dict[str, Any]]) -> None:
    """모니터 재빌드 후 FOUP 위젯을 유지·재연결만 한다(깜빡임 방지)."""
    if not channels:
        return
    _rebuild_sim_foup_outer_row(ext, channels)


def _snapshot_monitor_channel_texts(ext: Any) -> Tuple[Dict[int, str], Dict[int, str]]:
    """
    ``_sim_monitor_channels`` 에서 화면별 진행/이력 라벨 텍스트를 읽어 재빌드 시 복원한다.

    Returns:
        (history_by_screen, progress_by_screen) — 키는 1-based 화면 인덱스.
    """
    saved_h: Dict[int, str] = {}
    saved_p: Dict[int, str] = {}
    old = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(old, list):
        return saved_h, saved_p
    for ch in old:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si <= 0:
            continue
        pl = ch.get("progress_label")
        try:
            saved_h[si] = _get_channel_history_text(ch, ext)
        except Exception:
            pass
        if pl is not None:
            try:
                saved_p[si] = str(pl.text or "")
            except Exception:
                pass
    return saved_h, saved_p


def _ep_occ_timeline_layout_dims(ext: Any) -> Tuple[int, int, int, int, int]:
    """
    포트 아래 EP 타임라인(막대) Kit 가로 레이아웃.
    뷰포트 분할 시 열당 폭이 매우 좁아져, 막대·라벨·패딩을 함께 줄인다.

    Returns:
        (bar_w, name_w, val_w, frame_pad, row_sp) — 우측 누적 초 라벨 폭=val_w.
    """
    try:
        nsp = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        nsp = 1
    # BAR_W(막대 폭)는 2배. VAL_W(우측 값 라벨 최대 너비)는 1.5배(2줄 접힘 완화).
    if nsp <= 1:
        return (540, 64, 192, 2, 3)
    if nsp == 2:
        return (336, 48, 162, 3, 4)
    if nsp == 3:
        return (240, 44, 144, 2, 3)
    return (176, 40, 126, 2, 2)


_BAR_GRAPH_COPY_ROW_H = 26


def _sim_channel_upper_height(ext: Any) -> int:
    """포트·EP막대·진행현황 고정 높이(타임테이블 패널과 분리)."""
    return 84 + _ep_timeline_host_height(ext) + _BAR_GRAPH_COPY_ROW_H + 168 + 8


def _sim_snapshot_for_screen(ext: Any, screen_1based: int) -> Dict[str, Any]:
    snaps = getattr(ext, "_sim_per_screen_snapshots", None)
    if isinstance(snaps, list):
        try:
            si = max(1, int(screen_1based))
            if si <= len(snaps):
                s = snaps[si - 1]
                if isinstance(s, dict):
                    return dict(s)
        except Exception:
            pass
    return {}


def _effective_sim_settings_snapshot_for_screen(ext: Any, screen_1based: int) -> Dict[str, Any]:
    """
    프리런 export·웹 연동용 화면별 설정 dict.

    저장 슬롯이 비어 있어도 CASE A/B 실시간 UI(엔진 시작과 동일 SSOT)를 반환한다.
    """
    try:
        si = max(1, int(screen_1based))
    except Exception:
        si = 1
    try:
        from .ebs_case_models import capture_case_sim_settings_for_screen

        return dict(capture_case_sim_settings_for_screen(ext, si))
    except Exception:
        saved = _sim_snapshot_for_screen(ext, si)
        return dict(saved) if isinstance(saved, dict) else {}


def _ep_timeline_host_height(ext: Any) -> int:
    """막대 행 수(EP·ALL_EP·INOUT·BP)에 맞춘 ScrollingFrame 높이."""
    try:
        idx = int(get_sim_ep_count_idx(ext))
    except Exception:
        idx = 0
    try:
        from .ebs_control_panel_ui import get_sim_ebs_enabled

        ebs_on = bool(get_sim_ebs_enabled(ext))
    except Exception:
        ebs_on = True
    n_bars = len(bar_graph_row_order(idx, ebs_enabled=ebs_on))
    _, _, _, frame_pad, row_sp = _ep_occ_timeline_layout_dims(ext)
    bar_h = 10
    tick_h = 14
    inner_sp = 3
    # 우측 요약이 2줄로 접히는 경우가 많아 행당 최소 높이를 넉넉히 잡는다.
    row_h_est = 22
    h = (
        int(frame_pad) * 2
        + tick_h
        + inner_sp
        + n_bars * int(row_h_est)
        + max(0, n_bars - 1) * int(row_sp)
        + 2
    )
    return max(68, int(h))


def _ep_timeline_host_horizontal_scroll_policy(ext: Any) -> Any:
    """분할 시 막대 행이 열보다 넓으면 가로 스크롤로 잘림을 피한다."""
    try:
        nsp = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        nsp = 1
    if nsp <= 1:
        return ui.ScrollBarPolicy.SCROLLBAR_ALWAYS_OFF
    for name in ("SCROLLBAR_AS_NEEDED", "SCROLLBAR_AUTO", "SCROLLBAR_ALWAYS_ON"):
        pol = getattr(ui.ScrollBarPolicy, name, None)
        if pol is not None:
            return pol
    return ui.ScrollBarPolicy.SCROLLBAR_ALWAYS_OFF


def _create_sim_monitor_channel_column(ext: Any, screen: int) -> Dict[str, Any]:
    """
    단일 화면(채널)용 모니터 UI 블록을 만든다.

    - 구성: 포트상태 스크롤 → 진행현황 스크롤 → SIM 이력 스크롤.
    - 반환 dict 는 ``screen``, ``port_frame``/``port_cells``, ``progress_label``, ``history_label`` 등
      ``_rebuild_sim_monitor_split_ui``·``_update_sim_progress``·``_append_sim_log_channel`` 가 참조한다.
    """
    ch: Dict[str, Any] = {"screen": int(screen)}
    ch["port_cells"] = {}
    ch["port_cell_boxes"] = {}
    ch["column_root"] = ui.VStack(spacing=1, width=ui.Fraction(1.0), height=ui.Fraction(1.0))
    with ch["column_root"]:
        # 상단(포트·막대·진행)과 타임테이블을 분리 — 상단 갱신이 타임테이블 scroll_y 를 리셋하지 않게 한다.
        ch["monitor_upper_frame"] = ui.Frame(height=ui.Fraction(1.0))
        with ch["monitor_upper_frame"]:
            with ui.VStack(spacing=1):
                ch["port_frame"] = ui.ScrollingFrame(height=84, style={"background_color": 0xFF1A1E26, "border_width": 1, "border_color": 0xFF3A3A3A})
                with ch["port_frame"]:
                    with ui.VStack(spacing=2):
                        ch["port_header"] = ui.Label(
                            f"[포트상태·화면{screen}] 대기 중", height=18, style={"color": 0xFFBFE7FF}
                        )
                        with ui.VStack(spacing=2):
                            ch["port_buffer_row"] = ui.HStack(spacing=4, height=24)
                            with ch["port_buffer_row"]:
                                with ui.ZStack(width=90, height=24):
                                    ch["port_cell_boxes"]["BP1"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["BP1"] = ui.Label("BP1:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                with ui.ZStack(width=90, height=24):
                                    ch["port_cell_boxes"]["BP2"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["BP2"] = ui.Label("BP2:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                with ui.ZStack(width=90, height=24):
                                    ch["port_cell_boxes"]["BP3"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["BP3"] = ui.Label("BP3:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                ch["port_bp4_cell_container"] = ui.ZStack(width=90, height=24)
                                with ch["port_bp4_cell_container"]:
                                    ch["port_cell_boxes"]["BP4"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["BP4"] = ui.Label("BP4:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                            with ui.HStack(spacing=4, height=24):
                                ch["port_inout_cell_container"] = ui.ZStack(width=90, height=24)
                                with ch["port_inout_cell_container"]:
                                    ch["port_cell_boxes"]["INOUT"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["INOUT"] = ui.Label("IN/OUT:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                with ui.ZStack(width=90, height=24):
                                    ch["port_cell_boxes"]["EP1"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["EP1"] = ui.Label("EP1:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                with ui.ZStack(width=90, height=24):
                                    ch["port_cell_boxes"]["EP2"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_cells"]["EP2"] = ui.Label("EP2:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                ch["port_ep3_cell_container"] = ui.ZStack(width=90, height=24)
                                with ch["port_ep3_cell_container"]:
                                    ch["port_cell_boxes"]["EP3"] = ui.Rectangle(
                                        style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1}
                                    )
                                    ch["port_ep3_cell"] = ui.Label("EP3:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                ch["foup_progress_labels"] = {}
                ch["_foup_label_cache"] = {}
                ch["foup_progress_frame"] = None
                ch["foup_progress_label"] = None
                ch["ep_timeline_host"] = ui.ScrollingFrame(
                    height=_ep_timeline_host_height(ext),
                    horizontal_scrollbar_policy=_ep_timeline_host_horizontal_scroll_policy(ext),
                    vertical_scrollbar_policy=ui.ScrollBarPolicy.SCROLLBAR_ALWAYS_OFF,
                    style={"background_color": 0x221A1E26, "border_width": 1, "border_color": 0xFF3A3A3A},
                )
                with ch["ep_timeline_host"]:
                    ch["ep_timeline_busy_label"] = ui.Label("", height=1)
                ch["ep_timeline_widget"] = None
                bar_w, name_w, val_w, _, _ = _ep_occ_timeline_layout_dims(ext)
                with ui.HStack(height=_BAR_GRAPH_COPY_ROW_H, spacing=0):
                    ui.Spacer(width=int(name_w))
                    ui.Spacer(width=int(bar_w))
                    ch["ep_timeline_copy_btn"] = ui.Button(
                        "데이터 복사",
                        width=int(val_w),
                        height=22,
                        clicked_fn=lambda s=int(screen): _copy_bar_graph_prerun_json(ext, s),
                    )
                    try:
                        ch["ep_timeline_copy_btn"].enabled = False
                    except Exception:
                        pass
                ch["progress_frame"] = ui.ScrollingFrame(height=168, style={"background_color": 0xFF1A1E26, "border_width": 1, "border_color": 0xFF3A3A3A})
                with ch["progress_frame"]:
                    ch["progress_label"] = ui.Label("", word_wrap=True, height=158, style={"color": 0xFFFFFFFF})
                    ch["progress_ep_timeline_host"] = None
                    ch["progress_ep_timeline_widget"] = None
    try:
        by = getattr(ext, "_sim_timetable_display_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_timetable_display_by_screen = by
        by[str(int(screen))] = "[SIM] 대기 중" if screen == 1 else f"[SIM·화면{screen}] 대기 중"
    except Exception:
        pass
    ch["progress_label"].text = "[진행현황] 없음" if screen == 1 else f"[진행현황·화면{screen}] 없음"
    return ch


def _sim_ui_shell_rebuild_allowed(ext: Any) -> bool:
    """Start/Reset 직전에만 UI 셸(모니터·타임테이블) 전체 재조립을 허용한다."""
    if bool(getattr(ext, "_sim_timetable_allow_shell_rebuild", False)):
        return True
    return not timetable_rows_locked(ext)


def _resolve_timetable_channel_for_screen(ext: Any, screen: int) -> Optional[Dict[str, Any]]:
    """타임테이블 전용 창에 실제로 붙어 있는 채널 dict 를 찾는다."""
    try:
        si = int(screen)
    except Exception:
        return None
    by = getattr(ext, "_sim_timetable_channels", None)
    if isinstance(by, dict):
        ch = by.get(str(si))
        if isinstance(ch, dict) and ch.get("timetable_host") is not None:
            return ch
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and 0 < si <= len(chans):
        cand = chans[si - 1]
        if isinstance(cand, dict) and cand.get("timetable_host") is not None:
            return cand
    return None


def _rebuild_sim_monitor_split_ui(ext: Any) -> None:
    """
    뷰포트 분할 수(1~4)에 맞춰 시뮼 모니터 영역을 다시 그린다.

    - 1화면: 기존처럼 단일 ``TBS 시뮬 모니터`` 창.
    - 2화면 이상: 화면별 별도 ``ui.Window`` 에 채널 1개씩 전담 배치.
    """
    if not _sim_ui_shell_rebuild_allowed(ext):
        return
    _sync_sim_monitor_window_shells(ext)
    try:
        from .ebs_control_panel_ui import sync_aux_kit_window_visibility

        sync_aux_kit_window_visibility(ext)
    except Exception:
        pass

    saved_h, saved_p = _snapshot_monitor_channel_texts(ext)
    n = _sim_monitor_channel_count(ext)
    channels: List[Dict[str, Any]] = []
    per_screen = _sim_use_per_screen_windows(ext)

    if per_screen:
        for screen in range(1, n + 1):
            host = _monitor_split_host_for_screen(ext, screen)
            if host is None:
                continue
            inner = _clear_monitor_split_inner(ext, host, screen_key=str(screen))
            if inner is None:
                continue
            with inner:
                channels.append(_create_sim_monitor_channel_column(ext, screen))
    else:
        host = getattr(ext, "_sim_monitor_split_host", None)
        if host is None:
            return
        inner = _clear_monitor_split_inner(ext, host)
        if inner is None:
            return
        with inner:
            if n == 1:
                channels.append(_create_sim_monitor_channel_column(ext, 1))
            elif n == 2:
                with ui.HStack(spacing=6, height=ui.Fraction(1.0)):
                    with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                        channels.append(_create_sim_monitor_channel_column(ext, 1))
                    with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                        channels.append(_create_sim_monitor_channel_column(ext, 2))
            elif n == 3:
                with ui.VStack(spacing=4, height=ui.Fraction(1.0)):
                    with ui.VStack(spacing=4, height=ui.Fraction(0.5)):
                        channels.append(_create_sim_monitor_channel_column(ext, 1))
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 2))
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 3))
            else:
                with ui.VStack(spacing=4, height=ui.Fraction(1.0)):
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 1))
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 2))
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 3))
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            channels.append(_create_sim_monitor_channel_column(ext, 4))

    for ch in channels:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si in saved_h and str(saved_h.get(si) or "").strip():
            try:
                by_tb = getattr(ext, "_sim_timetable_display_by_screen", None)
                if not isinstance(by_tb, dict):
                    by_tb = {}
                    ext._sim_timetable_display_by_screen = by_tb
                by_tb[str(si)] = str(saved_h[si])
            except Exception:
                pass
        if si in saved_p and str(saved_p.get(si) or "").strip():
            try:
                ch["progress_label"].text = saved_p[si]
            except Exception:
                pass

    try:
        ext._sim_monitor_channels = channels
        ext._sim_monitor_layout_n = n
    except Exception:
        pass

    # 안정성(포트상태 깜빡임/초기화 방지):
    # 분할 레이아웃이 어떤 이유로 재조립되면(port cells가 기본값 "-"로 재생성),
    # 직후 포트상태가 "비었다가 채워졌다"처럼 보이는 깜빡임이 발생할 수 있다.
    # 마지막 점유 스냅샷이 있으면 즉시 복원해 화면을 안정화한다.
    try:
        by_occ = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if isinstance(by_occ, dict) and isinstance(channels, list):
            for ch in channels:
                if not isinstance(ch, dict):
                    continue
                try:
                    si = int(ch.get("screen", 1) or 1)
                except Exception:
                    si = 1
                occ = by_occ.get(str(si))
                if isinstance(occ, dict) and occ:
                    try:
                        _update_port_occupancy_panel(ext, occ, sim_time="", screen=si)
                    except Exception:
                        pass
    except Exception:
        pass

    try:
        _sync_foup_labels_to_channels(ext, channels)
    except Exception:
        pass

    if n == 1 and len(channels) == 1:
        c0 = channels[0]
        try:
            ext._sim_port_state_frame = c0["port_frame"]
            ext._sim_progress_frame = c0["progress_frame"]
            ext._sim_history_frame = c0["history_frame"]
            ext._sim_port_state_header_label = c0["port_header"]
            ext._sim_port_cells = c0["port_cells"]
            ext._sim_port_cell_boxes = c0["port_cell_boxes"]
            ext._sim_port_bp4_cell_container = c0.get("port_bp4_cell_container")
            ext._sim_port_buffer_row = c0.get("port_buffer_row")
            ext._sim_port_inout_cell_container = c0.get("port_inout_cell_container")
            ext._sim_port_ep3_cell_container = c0.get("port_ep3_cell_container")
            ext._sim_port_ep3_cell = c0.get("port_ep3_cell")
            ext._sim_progress_label = c0["progress_label"]
            ext._sim_history_label = c0["history_label"]
            if getattr(ext, "_sim_history_text", None) is not None:
                ext._sim_history_text.set_value(_get_channel_history_text(c0, ext))
            if getattr(ext, "_sim_progress_text", None) is not None:
                ext._sim_progress_text.set_value(c0["progress_label"].text)
        except Exception:
            pass
    else:
        try:
            ext._sim_port_state_frame = channels[0]["port_frame"]
            ext._sim_progress_frame = channels[0]["progress_frame"]
            ext._sim_history_frame = channels[0]["history_frame"]
            ext._sim_port_state_header_label = None
            ext._sim_port_cells = {}
            ext._sim_port_cell_boxes = {}
            ext._sim_port_bp4_cell_container = None
            ext._sim_port_buffer_row = None
            ext._sim_port_inout_cell_container = None
            ext._sim_port_ep3_cell_container = None
            ext._sim_port_ep3_cell = None
            ext._sim_progress_label = None
            ext._sim_history_label = None
        except Exception:
            pass

    try:
        _sync_ep3_port_cell_visibility(ext)
    except Exception:
        pass


def _clear_sim_timetable_storage(ext: Any) -> None:
    """이전 프리런 타임테이블 메타·백업을 비운다."""
    try:
        ext._sim_timetable_row_metas_by_screen = {}
        ext._sim_timetable_display_by_screen = {}
        ext._sim_timetable_channels = {}
        ext._sim_seek_snapshots_by_screen = {}
    except Exception:
        pass


def _reset_all_channel_timetables_to_idle(
    ext: Any,
    *,
    message: str = "타임테이블 대기 중 — Start 후 프리런 결과 표시",
) -> None:
    """모든 화면 타임테이블에서 기존 행을 제거하고 대기 문구만 남긴다."""
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list):
        return
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 1) or 1)
        except Exception:
            si = 1
        try:
            unlock_timetable_rows(ext)
            reset_timetable_channel_to_idle(ch, screen=si, message=message, ext=ext, force=True)
        except Exception:
            pass


def _clear_timetable_channel_widgets(ch: Dict[str, Any], ext: Any = None) -> None:
    """채널 dict 에 붙은 타임테이블 위젯만 제거(Start/Reset 시 스크롤 맨 위)."""
    if ext is not None and timetable_rows_locked(ext) and not bool(
        getattr(ext, "_sim_timetable_allow_shell_rebuild", False)
    ):
        return
    col = ch.get("timetable_column_root")
    if col is not None:
        try:
            col.destroy()
        except Exception:
            pass
    panel = ch.get("timetable_panel")
    if panel is not None:
        try:
            panel.destroy()
        except Exception:
            pass
    for key in (
        "timetable_panel",
        "timetable_viewport",
        "timetable_placer",
        "timetable_host",
        "timetable_scroll_track",
        "timetable_scroll_thumb_placer",
        "timetable_scroll_thumb",
        "timetable_inner",
        "timetable_busy_widget",
        "history_frame",
        "history_label",
        "timetable_column_root",
    ):
        ch[key] = None
    ch["timetable_row_buttons"] = []
    ch["timetable_row_labels"] = []
    ch["timetable_row_label_pairs"] = []
    ch["timetable_row_bgs"] = []
    ch["timetable_interactive"] = False
    ch["_timetable_row_style_cache"] = {}
    ch["_timetable_highlight_t"] = None
    ch["_tt_scroll_y"] = 0.0


def _mount_timetable_ui_on_channel(ch: Dict[str, Any], screen: int, ext: Any = None) -> None:
    """기존 채널 dict 에 타임테이블 전용 창용 UI를 장착한다."""
    _clear_timetable_channel_widgets(ch, ext)
    col = ui.VStack(spacing=2, width=ui.Fraction(1.0), height=ui.Fraction(1.0))
    ch["timetable_column_root"] = col
    with col:
        build_timetable_column_ui(ch, screen=int(screen), viewport_h=480)


def _rebuild_sim_timetable_split_ui(ext: Any) -> None:
    """뷰포트 분할 수에 맞춰 타임테이블 전용 창 영역을 다시 그린다."""
    if not _sim_ui_shell_rebuild_allowed(ext):
        return
    _sync_sim_timetable_window_shells(ext)
    try:
        from .ebs_control_panel_ui import sync_aux_kit_window_visibility

        sync_aux_kit_window_visibility(ext)
    except Exception:
        pass

    base_channels = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(base_channels, list) or not base_channels:
        return

    n = _sim_monitor_channel_count(ext)
    channels: List[Dict[str, Any]] = list(base_channels)
    per_screen = _sim_use_per_screen_windows(ext)
    ch_idx = 0

    def _next_ch(screen: int) -> Dict[str, Any]:
        nonlocal ch_idx
        if ch_idx < len(channels):
            ch = channels[ch_idx]
            ch_idx += 1
            if isinstance(ch, dict):
                return ch
        return {"screen": int(screen)}

    if per_screen:
        for screen in range(1, n + 1):
            host = _timetable_split_host_for_screen(ext, screen)
            if host is None:
                continue
            inner = _clear_timetable_split_inner(ext, host, screen_key=str(screen))
            if inner is None:
                continue
            with inner:
                _mount_timetable_ui_on_channel(_next_ch(screen), screen, ext)
    else:
        host = getattr(ext, "_sim_timetable_split_host", None)
        if host is None:
            return
        inner = _clear_timetable_split_inner(ext, host)
        if inner is None:
            return
        with inner:
            if n == 1:
                _mount_timetable_ui_on_channel(_next_ch(1), 1, ext)
            elif n == 2:
                with ui.HStack(spacing=6, height=ui.Fraction(1.0)):
                    with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                        _mount_timetable_ui_on_channel(_next_ch(1), 1, ext)
                    with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                        _mount_timetable_ui_on_channel(_next_ch(2), 2, ext)
            elif n == 3:
                with ui.VStack(spacing=4, height=ui.Fraction(1.0)):
                    with ui.VStack(spacing=4, height=ui.Fraction(0.5)):
                        _mount_timetable_ui_on_channel(_next_ch(1), 1, ext)
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(2), 2, ext)
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(3), 3, ext)
            else:
                with ui.VStack(spacing=4, height=ui.Fraction(1.0)):
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(1), 1, ext)
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(2), 2, ext)
                    with ui.HStack(spacing=6, height=ui.Fraction(0.5)):
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(3), 3, ext)
                        with ui.VStack(spacing=4, width=ui.Fraction(1.0), height=ui.Fraction(1.0)):
                            _mount_timetable_ui_on_channel(_next_ch(4), 4, ext)

    try:
        ext._sim_timetable_layout_n = n
    except Exception:
        pass


def _rebuild_all_sim_ui_panels(ext: Any) -> None:
    """모니터·타임테이블 창을 분할 수에 맞게 함께 재조립한다."""
    if not _sim_ui_shell_rebuild_allowed(ext):
        return
    _rebuild_sim_monitor_split_ui(ext)
    _rebuild_sim_timetable_split_ui(ext)
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans, list) and chans and isinstance(chans[0], dict):
            hf = chans[0].get("history_frame")
            if hf is not None:
                ext._sim_history_frame = hf
    except Exception:
        pass


def _sim_timetable_user_dismissed(ext: Any) -> bool:
    """사용자가 타임테이블 전용 창을 닫아 자동 표시를 원하지 않는 상태."""
    return bool(getattr(ext, "_sim_timetable_user_dismissed", False))


def _set_sim_timetable_user_dismissed(ext: Any, dismissed: bool) -> None:
    try:
        ext._sim_timetable_user_dismissed = bool(dismissed)
    except Exception:
        pass


def _on_sim_timetable_window_visibility_changed(ext: Any, visible: bool) -> None:
    """창 X/메뉴 토글 시 사용자 닫기 의도를 기억한다(프리런 후 자동 재오픈 방지)."""
    if bool(getattr(ext, "_sim_timetable_visibility_track_suppress", False)):
        return
    _set_sim_timetable_user_dismissed(ext, not bool(visible))


def _bind_sim_timetable_window_visibility(ext: Any, win: Any) -> None:
    if win is None:
        return
    try:
        win.set_visibility_changed_fn(lambda visible: _on_sim_timetable_window_visibility_changed(ext, bool(visible)))
    except Exception:
        pass


def build_sim_timetable_window(ext: Any) -> None:
    """프리런 타임테이블 전용 창 — 모니터 창과 분리해 스크롤·하이라이트 안정화."""
    _ensure_sim_timetable_window_shell(ext, 1)
    existing = getattr(ext, "_sim_timetable_window", None)
    if existing is not None:
        _bind_sim_timetable_window_visibility(ext, existing)


def build_sim_monitor_window(ext: Any) -> None:
    """
    시뮬 모니터 전용 창 — FOUP·포트·EP막대·진행현황.

    타임테이블은 ``build_sim_timetable_window`` 별도 창에서 표시한다.
    2분할 이상이면 ``_sync_sim_monitor_window_shells`` 가 화면별 창을 추가한다.
    """
    _ensure_sim_monitor_window_shell(ext, 1)


def build_control_window(ext: Any) -> None:
    """TBS 제어창을 만들고 ext에 위젯/모델 참조를 저장."""
    # destroy()가 실패하거나(Kit 이벤트/프레임 타이밍), 핫리로드로 ext 인스턴스가 바뀌면
    # 이전 창이 화면에 남은 채로 새 창이 생성되어 UI가 겹쳐 보일 수 있다.
    # 1) ext 참조 기준 중복 생성 방지
    if getattr(ext, "_control_window", None) is not None:
        return
    # 2) 워크스페이스에 남아있는 동명 창이 있으면 선제 제거(핫리로드/비정상 destroy 대비)
    try:
        ws = getattr(ui, "Workspace", None)
        if ws is not None and hasattr(ws, "get_window"):
            old = ws.get_window("EBS제어창(CASE A)")
            if old is None:
                old = ws.get_window("EBS 제어창")
            if old is not None:
                try:
                    old.destroy()
                except Exception:
                    try:
                        old.visible = False
                    except Exception:
                        pass
    except Exception:
        pass

    init_ebs_control_models(ext)

    ext._control_window = ui.Window("EBS제어창(CASE A)", width=520, height=500)
    with ext._control_window.frame:
        with ui.ScrollingFrame(
            height=ui.Fraction(1.0),
            style={"ScrollingFrame": {"padding": 4, "margin": 0}},
        ):
            with ui.VStack(spacing=0):
                build_ebs_control_panel_content(ext, compact=False, case_id=1)

    ext._control_window_b = ui.Window("EBS제어창(CASE B)", width=520, height=500)
    with ext._control_window_b.frame:
        with ui.ScrollingFrame(
            height=ui.Fraction(1.0),
            style={"ScrollingFrame": {"padding": 4, "margin": 0}},
        ):
            with ui.VStack(spacing=0):
                build_ebs_control_panel_content(ext, compact=False, case_id=2)
    build_sim_monitor_window(ext)
    build_sim_timetable_window(ext)
    build_fix_proc_window(ext)
    try:
        sync_aux_kit_window_visibility(ext)
    except Exception:
        pass
    try:
        _rebuild_all_sim_ui_panels(ext)
    except Exception:
        pass
    try:
        on_sim_ep_count_changed(ext)
    except Exception:
        pass
    try:
        _sync_default_sim_snapshot_from_ui(ext)
    except Exception:
        pass

    ext._sync_sim_multi_split_row_visibility_fn = _sync_sim_multi_split_row_visibility
    # 우선표시 이름 규칙 아래 UI를 숨겼으므로 목록 refresh도 비활성
    # refresh_object_list(ext)
    try:
        sim_multi_view.attach_stage_visibility_subscription(
            ext, lambda: _sync_sim_multi_split_row_visibility(ext)
        )
        _sync_sim_multi_split_row_visibility(ext)
        _refresh_sim_per_screen_rows(ext)
    except Exception:
        pass

    # -------------------------------------------------------------------
    # “기본값은 제어창 한 군데만 수정”을 위한 자동 동기화:
    # - UI 모델이 바뀌면 화면1 스냅샷(기본값) 갱신
    # - 저장하지 않은 화면(None)만 기본값을 따라가게 채움
    # -------------------------------------------------------------------
    try:
        for mdl in (
            getattr(ext, "_sim_lot_count_model", None),
            getattr(ext, "_sim_lot_spawn_min_model", None),
            getattr(ext, "_sim_lot_spawn_max_model", None),
            getattr(ext, "_sim_pickup_evt_min_model", None),
            getattr(ext, "_sim_pickup_evt_max_model", None),
            getattr(ext, "_sim_oht_bp1_min_model", None),
            getattr(ext, "_sim_oht_bp1_max_model", None),
            getattr(ext, "_sim_bp1_bp_min_model", None),
            getattr(ext, "_sim_bp1_bp_max_model", None),
            getattr(ext, "_sim_bp_ep_min_model", None),
            getattr(ext, "_sim_bp_ep_max_model", None),
            getattr(ext, "_sim_ep_oht_min_model", None),
            getattr(ext, "_sim_ep_oht_max_model", None),
            getattr(ext, "_sim_foup_proc_min_model", None),
            getattr(ext, "_sim_foup_proc_max_model", None),
            getattr(ext, "_sim_init_inout_model", None),
            getattr(ext, "_sim_init_bp1_model", None),
            getattr(ext, "_sim_init_bp2_model", None),
            getattr(ext, "_sim_init_bp3_model", None),
            getattr(ext, "_sim_init_bp4_model", None),
            getattr(ext, "_sim_init_ep1_model", None),
            getattr(ext, "_sim_init_ep2_model", None),
            getattr(ext, "_sim_init_ep3_model", None),
            getattr(ext, "_sim_fault_inout_model", None),
            getattr(ext, "_sim_fault_bp1_model", None),
            getattr(ext, "_sim_fault_bp2_model", None),
            getattr(ext, "_sim_fault_bp3_model", None),
            getattr(ext, "_sim_fault_bp4_model", None),
            getattr(ext, "_sim_fault_ep1_model", None),
            getattr(ext, "_sim_fault_ep2_model", None),
            getattr(ext, "_sim_fault_ep3_model", None),
        ):
            if mdl is None:
                continue
            try:
                mdl.add_value_changed_fn(lambda _m, e=ext: _sync_default_sim_snapshot_from_ui(e))
            except Exception:
                pass
    except Exception:
        pass


def on_xml_seq_changed(ext: Any) -> None:
    try:
        idx = ext._xml_seq_combo.model.get_item_value_model().as_int
    except Exception:
        idx = 0
    seqs = [
        xml_generator.SEQ_READYTOLOAD,
        xml_generator.SEQ_ARRIVED,
        xml_generator.SEQ_MOVE_TRANSFERING,
        xml_generator.SEQ_MOVE,
        xml_generator.SEQ_READYTOUNLOAD,
        xml_generator.SEQ_REMOVED,
    ]
    seq = seqs[idx] if 0 <= idx < len(seqs) else xml_generator.SEQ_READYTOLOAD
    ext._xml_ab_inputs_frame.visible = seq in xml_generator.FROM_TO_SEQS
    ext._xml_port_inputs_frame.visible = seq in xml_generator.PORT_ID_ONLY_SEQS


def on_xml_ok_clicked(ext: Any) -> None:
    try:
        idx = ext._xml_seq_combo.model.get_item_value_model().as_int
    except Exception:
        idx = 0
    seqs = [
        xml_generator.SEQ_READYTOLOAD,
        xml_generator.SEQ_ARRIVED,
        xml_generator.SEQ_MOVE_TRANSFERING,
        xml_generator.SEQ_MOVE,
        xml_generator.SEQ_READYTOUNLOAD,
        xml_generator.SEQ_REMOVED,
    ]
    seq = seqs[idx] if 0 <= idx < len(seqs) else xml_generator.SEQ_READYTOLOAD
    try:
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = ext._xml_from_port_model.get_value_as_int()
            to_port = ext._xml_to_port_model.get_value_as_int()
            xml = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port_id = ext._xml_port_id_model.get_value_as_int()
            xml = xml_generator.build_xml_string(seq, port_id=port_id)
        ext._last_generated_xml = xml
        print(xml, flush=True)
    except Exception as e:
        print(f"[morph.tbs_control_2][xml_generator] XML 생성 실패: {e}", flush=True)


def on_xml_run_clicked(ext: Any) -> None:
    xml_text = (ext._last_generated_xml or "").strip()
    if not xml_text:
        print("[morph.tbs_control_2][xml_generator] 저장된 XML이 없습니다. 먼저 OK로 XML을 생성하세요.", flush=True)
        return
    parsed = xml_generator.parse_xml_string(xml_text)
    if not parsed:
        print("[morph.tbs_control_2][xml_generator] XML 역파싱 실패.", flush=True)
        return
    lines = ["[XML PARSE RESULT]"]
    if parsed.get("action_desc"):
        lines.append("[ACTION]")
        lines.append(parsed.get("action_desc", ""))

    for k in (
        "sequence_name",
        "destination",
        "origination",
        "tid",
        "facility",
        "equipment_id",
        "port_id",
        "from_port_id",
        "to_port_id",
    ):
        lines.append(f"{k} = {parsed.get(k, '')}")
    print("\n".join(lines), flush=True)


def _append_sim_log_channel(ext: Any, screen: int, msg: str) -> None:
    """
    멀티 모니터에서 **지정 화면**의 ``history_label`` 에만 한 줄(또는 누적 블록)을 붙인다.

    - ``screen`` 은 1-based; ``_sim_monitor_channels[screen-1]`` 을 사용한다.
    - 최대 약 200줄을 넘기면 앞부분을 잘라 메모리·UI 부하를 줄인다.
    """
    if not msg:
        return
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list) or not chans:
        return
    try:
        idx = int(screen) - 1
    except Exception:
        idx = 0
    if idx < 0 or idx >= len(chans):
        idx = 0
    ch = chans[idx]
    if not isinstance(ch, dict):
        return
    try:
        by_tb = getattr(ext, "_sim_timetable_display_by_screen", None)
        if isinstance(by_tb, dict) and str(by_tb.get(str(int(screen)), "") or "").strip():
            return
    except Exception:
        pass
    if ch.get("timetable_interactive"):
        return
    prev = _get_channel_history_text(ch, ext).strip()
    merged = f"{prev}\n{msg}".strip() if prev else msg
    rows = merged.splitlines()
    if len(rows) > 200:
        merged = "\n".join(rows[-200:])
    try:
        _set_channel_history_text(ch, merged)
    except Exception:
        pass


def _append_sim_log(ext: Any, line: str) -> None:
    """
    UI 스레드 전용: 이력 로그 패널에 줄 추가. 시뮬 워커는 ``post_sim_history_line`` 로 큐에 넣는다.

    - 멀티(``_sim_monitor_channels`` 길이>1)이면 원문의 ``[화면N]`` 접두(엔진 on_log)로 채널을 고른 뒤,
      **접두 뒤 본문만** ``_format_history_line`` 한다. (포맷 단계에서 줄 앞에 이모지가 붙으면
      줄 처음의 ``[화면N]`` 패턴 매칭이 깨져 전부 화면1로 가는 문제를 피한다.)
    - 접두가 없으면 전역 메시지로 보고 화면1에만 붙인다.
    - 단일 모드는 ``_sim_history_text`` / ``_sim_history_label`` 레거시 경로.
    """
    raw = (line or "").strip()
    if not raw:
        return
    # 요구사항: SIM 현황(진행현황 아래 [SIM] 이력 영역)에는 타임테이블만 깔끔하게 남긴다.
    # - 일반 로그/안내(프리런 시작/완료 등)는 콘솔로만 보고, 이력 영역은 타임테이블 전용으로 유지.
    # - 타임테이블 블록은 _build_prerun_timetable_text에서 "[SIM] 타임테이블(프리런)" 헤더로 시작한다.
    try:
        only_tb = bool(getattr(ext, "_sim_history_timetable_only", True))
    except Exception:
        only_tb = True
    if only_tb:
        if not (raw.startswith("[SIM] 타임테이블(프리런)") or raw.startswith("[화면")):
            return
        try:
            by_tb = getattr(ext, "_sim_timetable_display_by_screen", None)
            if isinstance(by_tb, dict) and any(str(v or "").strip() for v in by_tb.values()):
                return
        except Exception:
            pass
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and len(chans) > 1:
        m0 = re.match(r"^\[화면(\d+)\]\s*", raw)
        if m0:
            try:
                si = int(m0.group(1))
            except Exception:
                si = 1
            try:
                si = max(1, min(len(chans), si))
            except Exception:
                si = 1
            rest = raw[m0.end() :].strip() if m0.end() <= len(raw) else raw
            msg = _format_history_line(rest)
            if not msg:
                return
            _append_sim_log_channel(ext, si, msg)
            return
        msg = _format_history_line(raw)
        if not msg:
            return
        _append_sim_log_channel(ext, 1, msg)
        return

    msg = _format_history_line(raw)
    if not msg:
        return
    prev = ext._sim_history_text.as_string if getattr(ext, "_sim_history_text", None) else ""
    merged = f"{prev}\n{msg}".strip() if prev else msg
    # 화면이 너무 길어지지 않게 최근 200줄만 유지
    rows = merged.splitlines()
    if len(rows) > 200:
        merged = "\n".join(rows[-200:])
    if getattr(ext, "_sim_history_text", None):
        ext._sim_history_text.set_value(merged)
    if getattr(ext, "_sim_history_label", None) is not None:
        ext._sim_history_label.text = merged


def _format_history_line(line: str) -> str:
    """
    이력로그 가독성을 위해 핵심 토큰을 앞에 배치하고 태그를 단순화한다.
    원문 정보는 유지하되 읽기만 쉽게 만든다.
    """
    s = (line or "").strip()
    if not s:
        return ""
    # 자주 보이는 태그를 직관적인 짧은 라벨로 치환
    tag_map = {
        "[STORY]": "[스토리]",
        "[MOVE]": "[이송]",
        "[ARRIVED]": "[도착]",
        "[PROCESS]": "[공정]",
        "[READYTOLOAD]": "[대기준비]",
        "[READYTOUNLOAD]": "[반출준비]",
        "[REMOVED]": "[반출완료]",
        "[INPUT]": "[투입]",
        "[WAIT]": "[대기]",
        "[SUMMARY LOT]": "[LOT 요약]",
        "[SUMMARY]": "[요약]",
    }
    for old, new in tag_map.items():
        if old in s:
            s = s.replace(old, new)
    # 시뮬 이벤트 원문은 너무 길어져서 핵심만 요약
    if "[SIM EVENT" in s and "seq=" in s:
        try:
            part = s.split("] ", 1)
            head = part[0] + "]" if len(part) == 2 else "[SIM EVENT]"
            body = part[1] if len(part) == 2 else s
            seq = ""
            lot = ""
            fr = ""
            to = ""
            for tok in body.split():
                if tok.startswith("seq="):
                    seq = tok[4:]
                elif tok.startswith("lot="):
                    lot = tok[4:]
                elif tok.startswith("from="):
                    fr = tok[5:]
                elif tok.startswith("to="):
                    to = tok[3:]
            route = f"{fr}->{to}" if fr and to else (to or fr or "-")
            return _with_history_color_icon(f"{head} [이벤트] seq={seq} lot={lot} route={route}")
        except Exception:
            return _with_history_color_icon(s)
    return _with_history_color_icon(s)


def _with_history_color_icon(s: str) -> str:
    """
    단일 Label 제약에서 줄 단위 강조를 위해 색상 아이콘을 앞에 붙인다.
    🟥 오류/실패, 🟨 대기/주의, 🟩 완료/성공, 🟦 이벤트/진행, ⬜ 일반
    """
    t = (s or "").upper()
    if any(k in t for k in ("실패", "ERROR", "EXCEPTION", "예외", "파싱실패")):
        return f"🟥 {s}"
    if any(k in t for k in ("대기", "[WAIT]", "파일없음", "매핑없음", "주의")):
        return f"🟨 {s}"
    if any(k in t for k in ("완료", "DONE", "저장 완료", "실행시작", "실행준비완료")):
        return f"🟩 {s}"
    if any(k in t for k in ("이벤트", "MOVE", "ARRIVED", "PROCESS", "STORY", "투입", "이송", "공정", "도착")):
        return f"🟦 {s}"
    return f"⬜ {s}"


def _append_anim_history_log(ext: Any, line: str) -> None:
    """애니메이션 실행이력 패널 제거됨. 호출은 호환을 위해 유지한다."""
    return


def _render_pending_dots(ext: Any) -> None:
    """(요청으로 제거) 점 표시 기능 비활성화."""
    return


def _port_cell_text(occ: Dict[str, Any], port: str) -> str:
    v = str(occ.get(port, "-")).strip()
    if not v or v.upper() in ("EMPTY", "-", "NONE"):
        return "-"
    if v.upper() == "FULL":
        return "FULL"
    return v


def _compact_cell_value(v: str, max_len: int = 10) -> str:
    s = (v or "-").strip()
    if len(s) <= max_len:
        return s
    return s[: max(1, max_len - 1)] + "..."


def _ep_count_idx_for_port_panel(ext: Any, screen_1based: int) -> int:
    from .tbs_ep_port_visibility import ep_count_idx_for_screen

    return ep_count_idx_for_screen(ext, int(screen_1based))


def _sync_ep3_port_cell_visibility_for_channel(ext: Any, ch: Dict[str, Any]) -> None:
    container = ch.get("port_ep3_cell_container")
    bp4_container = ch.get("port_bp4_cell_container")
    buffer_row = ch.get("port_buffer_row")
    inout_container = ch.get("port_inout_cell_container")
    try:
        si = int(ch.get("screen", 1) or 1)
    except Exception:
        si = 1
    ep_idx = _ep_count_idx_for_port_panel(ext, si)
    is_ep3 = bool(ep_idx == 1)
    try:
        from .tbs_ep_port_visibility import ebs_enabled_for_screen

        ebs_on = bool(ebs_enabled_for_screen(ext, si))
    except Exception:
        ebs_on = True
    if container is not None:
        container.visible = is_ep3
    if bp4_container is not None:
        bp4_container.visible = is_ep3 and ebs_on
    if buffer_row is not None:
        buffer_row.visible = ebs_on
    if inout_container is not None:
        inout_container.visible = ebs_on


def _sync_ep3_port_cell_visibility(ext: Any) -> None:
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and len(chans) > 0:
        for ch in chans:
            if isinstance(ch, dict):
                _sync_ep3_port_cell_visibility_for_channel(ext, ch)
        return
    ch = {
        "port_ep3_cell_container": getattr(ext, "_sim_port_ep3_cell_container", None),
        "port_bp4_cell_container": getattr(ext, "_sim_port_bp4_cell_container", None),
        "port_buffer_row": getattr(ext, "_sim_port_buffer_row", None),
        "port_inout_cell_container": getattr(ext, "_sim_port_inout_cell_container", None),
    }
    _sync_ep3_port_cell_visibility_for_channel(ext, ch)


def _set_port_box_style(ext: Any, port: str, value: str, cell_boxes: Any = None) -> None:
    boxes = cell_boxes if cell_boxes is not None else (getattr(ext, "_sim_port_cell_boxes", {}) or {})
    box = boxes.get(port) if isinstance(boxes, dict) else None
    if box is None:
        return
    v = (value or "").strip().upper()
    if not v or v in ("-", "EMPTY", "NONE"):
        fill = 0xFF2A2F38
    elif v == "FULL":
        fill = 0xFF6B5B2A
    else:
        fill = 0xFF1F4A36
    try:
        box.style = {"background_color": fill, "border_color": 0xFF7B8799, "border_width": 1}
    except Exception:
        pass


def _update_port_occupancy_panel(ext: Any, occ: Dict[str, Any], sim_time: str = "", screen: int = 1) -> None:
    if not isinstance(occ, dict):
        return
    try:
        scr_i = int(screen)
    except Exception:
        scr_i = 1
    scr_i = max(1, scr_i)
    # 재생 중: plan apply 경로(``_sim_playback_plan_panel_apply``)만 패널 SSOT.
    # 엔진·라이브 occ(공정 종료 proc_end 포함)는 plan 활성 여부와 무관하게 차단한다.
    # (plan 이 빌드되지 않은 경우에도 proc_end 라이브 occ 가 새지 않도록 무조건 막는다.)
    if bool(getattr(ext, "_sim_playback_started", False)):
        if not bool(getattr(ext, "_sim_playback_plan_panel_apply", False)):
            return
    # 중요(포트상태 "전부 비어버림" 방지):
    # occ가 빈 dict로 들어오면 _port_cell_text 기본값이 "-"가 되어 포트상태가 통째로 비어 보인다.
    # 타임라인 재생/프리런 플레이백에서 간헐적으로 occ가 누락/빈 값이 들어올 수 있으므로,
    # 이 경우에는 마지막 스냅샷으로 폴백하고, 폴백도 없으면 UI를 덮어쓰지 않는다.
    if not occ:
        try:
            by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
            if isinstance(by_prev, dict):
                occ_prev = by_prev.get(str(int(screen) if int(screen) > 0 else 1))
                if isinstance(occ_prev, dict) and occ_prev:
                    occ = dict(occ_prev)
        except Exception:
            pass
        if not occ:
            return
    ch: Optional[Dict[str, Any]] = None
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and len(chans) > 0:
        try:
            si = int(screen)
        except Exception:
            si = 1
        si = max(1, min(len(chans), si))
        ch = chans[si - 1] if isinstance(chans[si - 1], dict) else None
    inout = _port_cell_text(occ, "INOUT")
    bp1 = _port_cell_text(occ, "BP1")
    bp2 = _port_cell_text(occ, "BP2")
    bp3 = _port_cell_text(occ, "BP3")
    bp4 = _port_cell_text(occ, "BP4")
    ep1 = _port_cell_text(occ, "EP1")
    ep2 = _port_cell_text(occ, "EP2")
    ep3 = _port_cell_text(occ, "EP3")
    t = str(sim_time).strip()
    scr_lbl = int((ch or {}).get("screen", screen) or screen)
    hdr = (ch or {}).get("port_header") if ch else getattr(ext, "_sim_port_state_header_label", None)
    # 헤더의 t= 깜빡임 방지:
    # - 어떤 경로(예: 레이아웃 재조립 직후 복원)에서는 sim_time을 빈 문자열로 호출할 수 있다.
    # - 이때 헤더를 "[포트상태·화면N]"으로 덮어쓰면 t=가 사라졌다가, 다음 정상 업데이트에서 다시 나타나는
    #   깜빡임이 발생한다.
    # - 따라서 t가 비어있으면 헤더 텍스트는 유지(덮어쓰지 않음)한다.
    if hdr is not None and t:
        head = f"[포트상태·화면{scr_lbl} t={t}]"
        hdr.text = head
    _sync_ep3_port_cell_visibility(ext)
    cells = (ch or {}).get("port_cells") if ch else (getattr(ext, "_sim_port_cells", {}) or {})
    boxes = (ch or {}).get("port_cell_boxes") if ch else (getattr(ext, "_sim_port_cell_boxes", {}) or {})
    if not isinstance(cells, dict):
        cells = {}
    if not isinstance(boxes, dict):
        boxes = {}
    if "INOUT" in cells:
        cells["INOUT"].text = f"IN/OUT:{_compact_cell_value(inout)}"
        _set_port_box_style(ext, "INOUT", inout, boxes)
    if "BP1" in cells:
        cells["BP1"].text = f"BP1:{_compact_cell_value(bp1)}"
        _set_port_box_style(ext, "BP1", bp1, boxes)
    if "BP2" in cells:
        cells["BP2"].text = f"BP2:{_compact_cell_value(bp2)}"
        _set_port_box_style(ext, "BP2", bp2, boxes)
    if "BP3" in cells:
        cells["BP3"].text = f"BP3:{_compact_cell_value(bp3)}"
        _set_port_box_style(ext, "BP3", bp3, boxes)
    if "BP4" in cells:
        cells["BP4"].text = f"BP4:{_compact_cell_value(bp4)}"
        _set_port_box_style(ext, "BP4", bp4, boxes)
    if "EP1" in cells:
        cells["EP1"].text = f"EP1:{_compact_cell_value(ep1)}"
        _set_port_box_style(ext, "EP1", ep1, boxes)
    if "EP2" in cells:
        cells["EP2"].text = f"EP2:{_compact_cell_value(ep2)}"
        _set_port_box_style(ext, "EP2", ep2, boxes)
    ep3_cell = (ch or {}).get("port_ep3_cell") if ch else getattr(ext, "_sim_port_ep3_cell", None)
    if ep3_cell is not None:
        ep3_cell.text = f"EP3:{_compact_cell_value(ep3)}"
        _set_port_box_style(ext, "EP3", ep3, boxes)

    # 마지막 점유 스냅샷 저장(빈 dict는 저장하지 않음)
    try:
        by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_last_ports_occupancy_by_screen = by
        sk = str(int((ch or {}).get("screen", screen) or screen))
        if occ:
            by[sk] = dict(occ)
    except Exception:
        pass

    # 포트상태 아래 EP 타임라인(전용 영역) 갱신 — 멀티 채널 UI가 없어도(USD 미로드 등) 스냅샷/웹 막대용 상태는 누적
    try:
        ch_tl = ch if ch is not None else {"screen": int(screen), "ep_timeline_host": None}
        t_tl = t if t else f"{_resolve_ep_timeline_sim_time(ext, int(screen), ''):.2f}"
        _update_ep_timeline_under_port_state(ext, ch_tl, occ, t_tl)
    except Exception:
        pass


_PANEL_PORT_KEYS = ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")


def _ports_occ_is_complete_snapshot(occ: Dict[str, Any]) -> bool:
    """UI 패널 전체 키(INOUT~EP3)를 모두 포함한 스냅샷인지."""
    if not isinstance(occ, dict) or not occ:
        return False
    return all(k in occ for k in _PANEL_PORT_KEYS)


def _ports_occ_trust_all_empty(occ: Dict[str, Any], *, seq_u: str = "") -> bool:
    """전 포트가 비어 있어도 이전 스냅샷으로 되돌리면 안 되는 authoritative payload 인지."""
    if str(seq_u or "").strip().upper() == "PORT_OCC_REFRESH":
        return True
    if not isinstance(occ, dict) or not occ:
        return False
    if any(bool(str(v or "").strip()) for v in occ.values()):
        return False
    if _ports_occ_is_complete_snapshot(occ):
        return True
    keys = set(str(k).strip().upper() for k in occ.keys())
    if "INOUT" not in keys:
        return False
    ep_keys = {k for k in keys if k.startswith("EP")}
    # ep_count=2 엔진 스냅샷(INOUT+BP*+EP1+EP2)도 전 포트 비움을 신뢰
    return len(ep_keys) >= 2


def _merge_ports_occupancy_with_last(
    ext: Any, screen: int, occ: Dict[str, Any], *, seq_u: str = ""
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """부분/빈 ports_occupancy 를 화면별 마지막 스냅샷과 merge."""
    scr = max(1, int(screen))
    occ_in = dict(occ) if isinstance(occ, dict) else {}
    try:
        by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if not isinstance(by_prev, dict):
            by_prev = {}
            ext._sim_last_ports_occupancy_by_screen = by_prev
        occ_prev = by_prev.get(str(scr))
    except Exception:
        by_prev = {}
        occ_prev = None
    occ_out = dict(occ_in)
    if isinstance(occ_prev, dict) and occ_prev:
        if occ_out and any((k not in occ_out) for k in _PANEL_PORT_KEYS):
            merged = dict(occ_prev)
            merged.update(dict(occ_out))
            occ_out = merged
    if not occ_out:
        if isinstance(occ_prev, dict) and occ_prev:
            occ_out = dict(occ_prev)
    try:
        if (
            occ_out
            and (not any(bool(str(v or "").strip()) for v in occ_out.values()))
            and (not _ports_occ_trust_all_empty(occ_out, seq_u=seq_u))
        ):
            if isinstance(occ_prev, dict) and occ_prev and any(
                bool(str(v or "").strip()) for v in occ_prev.values()
            ):
                occ_out = dict(occ_prev)
    except Exception:
        pass
    return occ_out, by_prev


def _render_ep_bar_prerun_at_t(
    ext: Any,
    ch: Dict[str, Any],
    t_sim: float,
    occ: Optional[Dict[str, Any]] = None,
) -> bool:
    """``bar_pre`` truncate — plan/재생 resolve 실패 시에도 막대를 반드시 그린다."""
    if not isinstance(ch, dict) or ch.get("ep_timeline_host") is None:
        return False
    try:
        screen = max(1, int(ch.get("screen", 1) or 1))
    except Exception:
        screen = 1
    scr_key = str(screen)
    pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
    bar_pre = pre_by.get(scr_key) if isinstance(pre_by, dict) else None
    if not isinstance(bar_pre, EpBarPrecomputed):
        return False
    t_bar = max(0.0, float(t_sim))
    try:
        pm = getattr(ext, "_sim_bar_preview_model", None)
        preview_full = bool(pm.get_value_as_bool()) if pm is not None else False
    except Exception:
        preview_full = False
    rows = list(bar_pre.row_order) if bar_pre.row_order else []
    if not rows:
        try:
            ep_idx = int(_ep_count_idx_for_port_panel(ext, int(screen)))
            snap_cfg = _sim_snapshot_for_screen(ext, int(screen))
            ebs_on = bool(snap_cfg.get("ebs_enabled", True)) if snap_cfg else True
            rows = list(bar_graph_row_order(ep_idx, ebs_enabled=ebs_on))
        except Exception:
            rows = list(bar_graph_row_order(0, ebs_enabled=True))
    if preview_full and isinstance(bar_pre.rows, dict) and bar_pre.rows:
        rows_state = {str(k): list(v) for k, v in bar_pre.rows.items()}
    else:
        rows_state = truncate_bar_rows_at_t(bar_pre.rows or {}, t_bar)
    seed_dur = max(1e-6, float(t_bar))
    for r in rows:
        rk = str(r)
        segs = rows_state.get(rk)
        if not isinstance(segs, list) or not segs:
            rows_state[rk] = [{"state": BAR_STATE_EMPTY, "dur": float(seed_dur)}]
    total_est = float(getattr(bar_pre, "total_est", 0.0) or 0.0)
    if total_est <= 0.0:
        try:
            last_te = getattr(ext, "_sim_last_total_est_by_screen", None)
            if isinstance(last_te, dict):
                total_est = float(last_te.get(scr_key) or 0.0)
        except Exception:
            total_est = 0.0
    if total_est <= 0.0:
        total_est = max(30.0, float(t_bar) * 1.2)
    try:
        from .control_sim_playback_plan import PlaybackUIAxes, PlaybackUIState

        state = PlaybackUIState(
            screen=int(screen),
            axes=PlaybackUIAxes(t_display=float(t_bar), t_plan=float(t_bar)),
            ports=dict(occ or {}),
            bar_rows={str(k): list(v) for k, v in rows_state.items()},
            bar_total_est=float(total_est),
            row_order=tuple(str(r) for r in rows),
            preview_full=bool(preview_full),
        )
        _update_ep_timeline_under_port_state(
            ext,
            ch,
            dict(occ or {}),
            f"{float(t_bar):.2f}",
            honor_explicit_sim_time=True,
            playback_ui_state=state,
        )
        return ch.get("ep_timeline_widget") is not None
    except Exception:
        return False


def _apply_playback_bar_to_channel(
    ext: Any,
    screen: int,
    state: Any,
    *,
    t_fallback: float = 0.0,
) -> None:
    """재생 막대 — ``PlaybackUIState`` 또는 legacy fallback 으로 채널에 렌더."""
    try:
        scr = max(1, int(screen))
    except Exception:
        return
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list) or not (0 < scr <= len(chans)):
        return
    ch = chans[scr - 1]
    if not isinstance(ch, dict):
        return
    if state is not None:
        try:
            _update_ep_timeline_under_port_state(
                ext,
                ch,
                dict(getattr(state, "ports", {}) or {}),
                f"{float(state.axes.t_display):.2f}",
                honor_explicit_sim_time=False,
                playback_ui_state=state,
            )
            if ch.get("ep_timeline_widget") is not None:
                return
        except Exception:
            pass
    try:
        if _render_ep_bar_prerun_at_t(ext, ch, float(t_fallback), {}):
            return
    except Exception:
        pass
    try:
        _update_ep_timeline_under_port_state(
            ext,
            ch,
            {},
            f"{float(t_fallback):.2f}",
            honor_explicit_sim_time=False,
            playback_ui_state=None,
        )
    except Exception:
        pass


def apply_playback_ui_timeline(
    ext: Any,
    screen: int,
    state: Any,
) -> None:
    """재생 SSOT — ``refresh_playback_display_at_sim`` 로 포트·막대 공통 갱신."""
    _apply_playback_bar_to_channel(ext, int(screen), state)


def _refresh_playback_ep_bar_at_sim(
    ext: Any,
    screen: int,
    t_sim: float,
) -> None:
    """재생(plan) — ``refresh_playback_display_at_sim`` 단일 SSOT."""
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    scr = max(1, int(screen))
    try:
        from .control_sim_playback_plan import playback_plan_active, refresh_playback_display_at_sim

        if not playback_plan_active(ext, scr):
            return
        refresh_playback_display_at_sim(ext, scr, float(t_sim))
    except Exception:
        pass


def _flush_renewal_bar_legacy(
    ext: Any,
    screen: int,
    src_r: Dict[str, Any],
    sync_t: Optional[float],
) -> None:
    """라이브(비-재생) renewal — plan 없을 때 막대만 갱신."""
    scr = max(1, int(screen))
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list) or not (0 < scr <= len(chans)):
        return
    ch = chans[scr - 1]
    if not isinstance(ch, dict):
        return
    occ: Dict[str, Any] = {}
    by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
    if isinstance(by_prev, dict) and isinstance(by_prev.get(str(scr)), dict):
        occ = dict(by_prev.get(str(scr)) or {})
    sim_t = f"{float(sync_t):.2f}" if sync_t is not None else str(
        src_r.get("t") or src_r.get("sim_time") or ""
    ).strip()
    if not sim_t:
        lp_by = getattr(ext, "_sim_progress_last_payload_by_screen", None)
        lp = lp_by.get(str(scr)) if isinstance(lp_by, dict) else None
        sim_t = str((lp or {}).get("sim_time", "") or "") if isinstance(lp, dict) else ""
    honor = sync_t is not None
    _update_ep_timeline_under_port_state(
        ext, ch, occ, sim_t, honor_explicit_sim_time=honor
    )


def _sync_port_panel_from_engine_occ(
    ext: Any,
    screen: int,
    occ: Dict[str, Any],
    sim_time: str = "",
    *,
    allow_post_anim_block: bool = True,
    seq_u: str = "",
) -> None:
    """
    엔진 ports_occupancy → 포트 패널(텍스트) 동기화.

    visibility/prim 위치 스냅은 JSON 종료(post-anim) 또는 PORT_OCC_REFRESH 에서만 수행한다.
    """
    scr = max(1, int(screen))
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import playback_plan_active

            if playback_plan_active(ext, scr):
                return
        except Exception:
            pass
    if _should_defer_port_occ_sync_for_renewal(ext, scr):
        return
    if allow_post_anim_block:
        try:
            applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
            if isinstance(applied_by, dict) and applied_by.get(str(scr)):
                return
        except Exception:
            pass
    occ_m, by_prev = _merge_ports_occupancy_with_last(
        ext, scr, occ if isinstance(occ, dict) else {}, seq_u=str(seq_u or "")
    )
    if not occ_m or not any((k in occ_m) for k in _PANEL_PORT_KEYS):
        return
    try:
        if isinstance(by_prev, dict):
            by_prev[str(scr)] = dict(occ_m)
    except Exception:
        pass
    try:
        _update_port_occupancy_panel(ext, occ_m, str(sim_time or ""), screen=scr)
    except Exception:
        pass


def _bar_segment_rect_widths(
    segs: List[Dict[str, Any]],
    *,
    total_est: float,
    bar_w: int,
    t_cover: Optional[float] = None,
) -> List[Tuple[int, str]]:
    return allocate_bar_segment_pixels(
        segs,
        total_est=float(total_est),
        bar_w=int(bar_w),
        t_cover=t_cover,
    )


def _post_anim_affects_ep_port(src: Dict[str, Any]) -> bool:
    fr = _canonical_sim_port_key(str(src.get("from_port_id") or ""))
    to = _canonical_sim_port_key(str(src.get("to_port_id") or ""))
    port = _canonical_sim_port_key(str(src.get("port_id") or src.get("event_port_id") or ""))
    return any(str(p).startswith("EP") for p in (fr, to, port) if p)


def _occ_for_ep_timeline(
    ext: Any,
    screen: int,
    occ: Dict[str, Any],
    sim_time_text: str = "",
    *,
    progress_p: Optional[Dict[str, Any]] = None,
    t_sim: Optional[float] = None,
) -> Dict[str, Any]:
    """
    EP 막대용 ports_occupancy.

    재생(plan): ``t_sim`` / ``sim_now`` plan lookup. 라이브: 패널 스냅샷 + post-anim 예측.
    """
    scr = max(1, int(screen))
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import get_plan_ports_at_sim, playback_plan_active

            if playback_plan_active(ext, scr):
                if t_sim is not None:
                    t_bar = float(t_sim)
                else:
                    t_bar = _resolve_ep_timeline_sim_time(ext, scr, sim_time_text)
                occ_plan = get_plan_ports_at_sim(ext, scr, float(t_bar), honor_explicit=True)
                if isinstance(occ_plan, dict):
                    return dict(occ_plan)
        except Exception:
            pass
    occ_in = dict(occ) if isinstance(occ, dict) else {}
    last_snap: Dict[str, Any] = {}
    try:
        last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if isinstance(last_by, dict) and isinstance(last_by.get(str(scr)), dict):
            last_snap = dict(last_by.get(str(scr)) or {})
    except Exception:
        last_snap = {}
    occ_base = dict(occ_in)
    if last_snap:
        occ_base.update(last_snap)
    try:
        applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
        if isinstance(applied_by, dict) and applied_by.get(str(scr)) and last_snap:
            return dict(last_snap)
    except Exception:
        pass
    if progress_p is None:
        try:
            by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
            if isinstance(by_lp, dict) and isinstance(by_lp.get(str(scr)), dict):
                progress_p = by_lp.get(str(scr))
        except Exception:
            progress_p = None
    try:
        t_f = float(str(sim_time_text or "").strip() or "0.0")
    except Exception:
        t_f = 0.0
    if _should_defer_port_occ_sync_for_renewal(ext, scr):
        out = dict(occ_base)
        if last_snap and isinstance(progress_p, dict):
            if not _progress_event_affects_ep(progress_p):
                for ep in ("EP1", "EP2", "EP3"):
                    if ep in last_snap:
                        out[ep] = last_snap[ep]
        return out
    occ_pred = effective_ports_occupancy_at_t(
        occ_base,
        progress_p if isinstance(progress_p, dict) else None,
        t_f,
    )
    # 비-EP 이동 구간: EP 키는 패널 스냅샷 유지(빨간색으로 덮어쓰지 않음)
    out = dict(occ_pred)
    if last_snap and isinstance(progress_p, dict):
        if not _progress_event_affects_ep(progress_p):
            for ep in ("EP1", "EP2", "EP3"):
                if ep in last_snap:
                    out[ep] = last_snap[ep]
    return out


def _resolve_ep_timeline_sim_time(
    ext: Any,
    screen: int,
    sim_time_text: str,
    *,
    honor_explicit: bool = False,
) -> float:
    """
    EP 막대·진행현황이 같은 시계를 쓰도록 sim 시각을 통일한다.
    프리런 재생 중에는 ``SimTimelinePlayer.sim_now`` 를 최우선(진행현황 t(sim) 과 동일).

    ``honor_explicit=True`` — 호출자가 지정한 sim 시각을 그대로 사용(seek·명시 truncate).

    재생 중 **막대 시간축**은 ``sim_now`` 만 (replay floor 는 occ lookup 전용).
    """
    si = int(screen)
    t_parsed = 0.0
    try:
        t_parsed = float(str(sim_time_text or "").strip() or "0.0")
    except Exception:
        t_parsed = 0.0
    if honor_explicit and t_parsed > 1e-9:
        return float(t_parsed)
    player = get_sim_playback_player(ext, si)
    if player is not None:
        try:
            return float(player.sim_now(si))
        except Exception:
            pass
    if t_parsed > 1e-9:
        return float(t_parsed)
    try:
        by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
        if isinstance(by_lp, dict) and isinstance(by_lp.get(str(si)), dict):
            t2 = float(str(by_lp[str(si)].get("sim_time", "") or "0").strip() or "0")
            if t2 > 1e-9:
                return float(t2)
    except Exception:
        pass
    return float(t_parsed)


def _append_bar_live_interval(
    rows_state: Dict[str, List[Dict[str, Any]]],
    rows: List[str],
    *,
    dt: float,
    state_for_row: Callable[[str], str],
) -> None:
    """프리런 truncate 위 renewal lead 구간 live tail — 우측 누적 초·막대 전진."""
    if dt <= 1e-9:
        return
    for r in rows:
        st_seg = state_for_row(r)
        if r == "ALL_EP":
            ep_rows = [x for x in rows if x.startswith("EP")]
            ep_st = [state_for_row(x) for x in ep_rows]
            st_seg = _aggregate_all_ep_state(ep_st)
        segs = rows_state.get(r)
        if not isinstance(segs, list):
            segs = []
            rows_state[r] = segs
        if segs and isinstance(segs[-1], dict) and bar_state_from_seg(segs[-1]) == st_seg:
            segs[-1]["dur"] = float(segs[-1].get("dur", 0.0)) + float(dt)
        else:
            segs.append({"state": st_seg, "dur": float(dt)})
        if len(segs) > 4096:
            merged = merge_bar_row_segments(segs)
            if len(merged) > 220:
                merged = merged[-220:]
            rows_state[r] = merged


# 막대 마스크 배경색 — 막대 영역 bg(0xFF1A1E26)와 동일하게 덮어 "아직 안 지난 구간"으로 보이게 한다.
_BAR_MASK_BG = 0xFF1A1E26


def _estimate_wrapped_line_count(text: str, avail_px: float, font_size: int = 11) -> int:
    """``word_wrap`` 라벨이 차지할 대략적인 줄 수(공백 단위 그리디 줄바꿈 추정).

    우측 상태별 초 요약(예: ``empty:12s proc:5s ...``)이 좁은 폭에서 2줄 이상으로
    접힐 때를 미리 반영해, 막대 영역 높이를 내용에 맞춰 잡기 위한 추정값이다.
    """
    s = str(text or "").strip()
    if not s:
        return 1
    try:
        avail = max(1.0, float(avail_px))
    except Exception:
        avail = 1.0
    char_px = max(1.0, float(font_size) * 0.62)
    space_px = char_px
    lines = 1
    cur = 0.0
    for part in s.split(" "):
        if not part:
            continue
        w = len(part) * char_px
        if cur <= 0.0:
            cur = w
        elif cur + space_px + w <= avail:
            cur += space_px + w
        else:
            lines += 1
            cur = w
    return max(1, min(6, lines))


def _bar_row_value_label_height(
    seg_list: List[Any], *, val_w: int, bar_h: int, font_size: int = 11
) -> int:
    """행 우측 값 라벨이 필요로 하는 행 높이(막대 두께 ``bar_h`` 이상)를 반환한다."""
    try:
        txt = format_row_state_duration_summary(seg_list or [])
    except Exception:
        txt = ""
    lines = _estimate_wrapped_line_count(txt, val_w, font_size)
    if int(lines) <= 1:
        return int(bar_h)
    # 2줄 이상일 때만 행을 키우되, 줄 간격을 촘촘하게(폰트+2) 잡아 과도한 여백을 막는다.
    line_h = int(font_size) + 2
    return max(int(bar_h), int(lines) * int(line_h))


def _bar_playhead_arrow_offset_x(playhead_px: int, arrow_w: int = 12) -> int:
    """재생 위치 화살표(▼)가 playhead 중앙에 오도록 Placer offset_x."""
    return max(0, int(playhead_px) - int(arrow_w) // 2)


def _resolve_timetable_row_index_for_sim_time(
    ext: Any, screen: int, t_sim: float
) -> Optional[int]:
    """클릭 시각 t 에서 ``t <= t_sim`` 인 타임테이블 행 중 가장 늦은(큰 row_index) 행."""
    metas_by = getattr(ext, "_sim_timetable_row_metas_by_screen", None)
    metas: List[Any] = []
    if isinstance(metas_by, dict):
        metas = list(metas_by.get(str(int(screen)), []) or [])
    if not metas:
        results = getattr(ext, "_sim_prerun_results_by_screen", None)
        if isinstance(results, dict):
            res = results.get(int(screen))
            if res is not None:
                try:
                    metas = build_timetable_row_metas(res)
                except Exception:
                    metas = []
    if not metas:
        return None
    t_click = float(t_sim)
    best_idx: Optional[int] = None
    for m in metas:
        try:
            mt = float(getattr(m, "t", 0.0) or 0.0)
            ri = int(getattr(m, "row_index", 0))
        except Exception:
            continue
        if mt <= t_click + 1e-6:
            if best_idx is None or ri > int(best_idx):
                best_idx = ri
    if best_idx is None:
        return 0
    return int(best_idx)


def _on_bar_timeline_seek(ext: Any, screen: int, t_click: float) -> None:
    """막대 시간축 클릭 → 타임테이블 행 seek 과 동일 경로로 재생."""
    try:
        row_idx = _resolve_timetable_row_index_for_sim_time(ext, int(screen), float(t_click))
        if row_idx is None:
            return
        _on_timetable_row_seek(ext, int(screen), int(row_idx))
        print(
            f"[SIM] 막대 Seek 화면{int(screen)} t={float(t_click):.2f} → 행{int(row_idx)}",
            flush=True,
        )
    except Exception as e:
        print(f"[SIM] bar timeline seek 실패: {e}", flush=True)


def _attach_bar_time_axis_scrubber(
    ext: Any,
    st: Dict[str, Any],
    *,
    screen: int,
    bar_w: int,
    total_est: float,
    playhead_px: int,
    tick_step: float,
) -> None:
    """막대 상단 시간축 — 눈금 + 재생 위치 ▼ 화살표 + 클릭 시크(구간 투명 버튼).

    클릭 좌표(``screen_position_x``)는 omni.ui 버전에 따라 신뢰할 수 없어, 시간축 위에
    고정 시간을 가진 투명 버튼 구간들을 깔아 클릭 위치→시간을 좌표 계산 없이 결정한다.
    """
    bw = int(bar_w)
    te = float(total_est)
    st["_bar_seek_total_est"] = te
    st["_bar_seek_bar_w"] = bw
    st["_bar_seek_screen"] = int(screen)
    scr_i = int(screen)

    ph_x = _bar_playhead_arrow_offset_x(int(playhead_px))
    playhead_placer = None
    playhead_lbl = None

    tick_stack = ui.ZStack(width=bw, height=14)
    with tick_stack:
        ui.Rectangle(width=bw, height=14, style={"background_color": 0x441A1E26})
        try:
            ticks = max(1, int(te // float(tick_step)))
        except Exception:
            ticks = 1
        for i in range(ticks + 1):
            try:
                t_lbl = float(i) * float(tick_step)
            except Exception:
                t_lbl = 0.0
            x = int(round((float(t_lbl) / te) * float(bw))) if te > 1e-9 else 0
            x = max(0, min(bw - 1, x))
            with ui.Placer(offset_x=x, offset_y=0):
                ui.Label(
                    f"{int(round(t_lbl))}",
                    width=36,
                    height=14,
                    style={"color": 0xFFE0E6F0, "font_size": 10},
                )
        playhead_placer = ui.Placer(offset_x=int(ph_x), offset_y=0)
        with playhead_placer:
            playhead_lbl = ui.Label(
                "▼",
                width=12,
                height=14,
                alignment=ui.Alignment.CENTER,
                style={"color": 0xFFFFCC66, "font_size": 10},
            )
        # 최상위 레이어: 구간별 투명 클릭 영역(클릭 위치→고정 시간). 좌표 계산 불필요.
        # (타임테이블 행과 동일하게 ZStack + set_mouse_pressed_fn 사용 — 검증된 패턴)
        n_buckets = max(20, min(240, int(bw // 4) or 1))

        def _mk_seek(t_val: float) -> Callable[[float, float, int, Any], None]:
            def _press(_x: float, _y: float, button: int, _mods: Any) -> None:
                if int(button) != 0:
                    return
                _on_bar_timeline_seek(ext, scr_i, float(t_val))

            return _press

        with ui.HStack(width=bw, height=14, spacing=0):
            for bi in range(n_buckets):
                seg_w = (bw // n_buckets) if bi < n_buckets - 1 else (bw - (bw // n_buckets) * (n_buckets - 1))
                if seg_w <= 0:
                    continue
                t_center = ((float(bi) + 0.5) / float(n_buckets)) * te
                seg = ui.ZStack(width=int(seg_w), height=14)
                with seg:
                    ui.Spacer()
                try:
                    seg.set_mouse_pressed_fn(_mk_seek(float(t_center)))
                except Exception:
                    pass

    if playhead_placer is not None:
        st["_bar_playhead_placer"] = playhead_placer
    if playhead_lbl is not None:
        st["_bar_playhead_label"] = playhead_lbl


def _apply_bar_mask_widths(
    st: Dict[str, Any],
    rows: List[Any],
    *,
    bar_w: int,
    playhead_px: int,
    t_bar: float,
    preview_full: bool,
) -> None:
    """행별 마스크 사각형의 offset/width(또는 visible)만 갱신한다. UI 트리 재생성 없음.

    - 미리보기(preview_full) ON  → 마스크 숨김(막대 전체 노출)
    - OFF → 마스크가 playhead~끝 구간을 덮어 진행분만 노출
    값 라벨(누적 시간 요약)은 라벨 텍스트만 갱신(트리 재생성 아님).
    """
    masks = st.get("_mask_widgets") or {}
    full_rows = st.get("_mask_full_rows") or {}
    mask_w = max(0, int(bar_w) - int(playhead_px))

    trunc = None
    if not preview_full:
        try:
            trunc = truncate_bar_rows_at_t(full_rows, float(t_bar))
        except Exception:
            trunc = None

    for r in rows:
        rk = str(r)
        ent = masks.get(rk)
        if not ent:
            continue
        placer, rect, vlabel = ent
        try:
            if preview_full or mask_w <= 0:
                rect.visible = False
            else:
                rect.visible = True
                placer.offset_x = ui.Pixel(int(playhead_px))
                rect.width = ui.Pixel(int(mask_w))
        except Exception:
            pass
        if vlabel is not None:
            try:
                if preview_full:
                    seg_v = full_rows.get(rk, []) or []
                else:
                    seg_v = (trunc or {}).get(rk, []) or []
                vlabel.text = format_row_state_duration_summary(seg_v)
            except Exception:
                pass

    ph_placer = st.get("_bar_playhead_placer")
    if ph_placer is not None:
        try:
            ph_placer.offset_x = ui.Pixel(_bar_playhead_arrow_offset_x(int(playhead_px)))
        except Exception:
            pass
    ph_lbl = st.get("_bar_playhead_label")
    if ph_lbl is not None:
        try:
            ph_lbl.visible = True
        except Exception:
            pass


def _bar_graph_has_prerun_data(ext: Any, screen: int) -> bool:
    pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
    if not isinstance(pre_by, dict):
        return False
    bar_pre = pre_by.get(str(int(screen)))
    if not isinstance(bar_pre, EpBarPrecomputed):
        return False
    rows = getattr(bar_pre, "rows", None)
    return isinstance(rows, dict) and bool(rows)


def _copy_bar_graph_prerun_json(ext: Any, screen: int) -> None:
    """프리런 막대 전체를 시간 순 JSON 으로 클립보드·콘솔에 출력."""
    pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
    bar_pre = pre_by.get(str(int(screen))) if isinstance(pre_by, dict) else None
    if not isinstance(bar_pre, EpBarPrecomputed) or not _bar_graph_has_prerun_data(ext, int(screen)):
        msg = f"[SIM UI] 화면{int(screen)}: 프리런 막대 데이터가 없습니다."
        print(msg, flush=True)
        _append_sim_log(ext, msg)
        return
    try:
        snap = _effective_sim_settings_snapshot_for_screen(ext, int(screen))
    except Exception:
        snap = _sim_snapshot_for_screen(ext, int(screen))
    doc = build_bar_graph_copy_document(
        screen=int(screen),
        bar=bar_pre,
        sim_snapshot=snap,
    )
    text = json.dumps(doc, ensure_ascii=False, indent=2)
    header = f"[SIM UI] 막대그래프 JSON 복사 (화면{int(screen)})"
    print(f"{header}\n{text}", flush=True)
    try:
        import omni.kit.clipboard as cb  # type: ignore

        if hasattr(cb, "copy"):
            cb.copy(text)
        elif hasattr(cb, "set_text"):
            cb.set_text(text)
        else:
            raise RuntimeError("clipboard api not found")
        _append_sim_log(ext, f"{header} — 클립보드 복사 완료")
    except Exception as exc:
        print(f"[SIM UI] 클립보드 복사 실패({exc}); 위 콘솔 JSON 참고", flush=True)
        _append_sim_log(ext, f"{header} — 클립보드 미지원, 콘솔 출력")


def _sync_bar_graph_copy_button(ext: Any, ch: Dict[str, Any], screen: int) -> None:
    btn = ch.get("ep_timeline_copy_btn")
    if btn is None:
        return
    try:
        btn.enabled = _bar_graph_has_prerun_data(ext, int(screen))
    except Exception:
        pass


def _build_precomputed_bar_with_mask(
    ext: Any,
    ch: Dict[str, Any],
    st: Dict[str, Any],
    *,
    screen: int,
    rows: List[Any],
    bar_pre: Any,
    total_est: float,
    playhead_px: int,
    preview_full: bool,
    layout: Tuple[int, int, int, int, int],
) -> bool:
    """막대를 1회만 정적으로 그리고 행마다 마스크 사각형 1개를 얹는다.

    이후 heartbeat 에서는 ``_apply_bar_mask_widths`` 로 마스크 width/offset 만 바꾼다.
    """
    host = ch.get("ep_timeline_host")
    if host is None:
        return False

    BAR_W, NAME_W, VAL_W, frame_pad, row_sp = layout
    BAR_H = 10
    inner_sp = 3
    tick_h = 14
    tick_step = max(10.0, float(int((((float(total_est) / 8.0) + 9.999) // 10.0) * 10.0)))

    full_rows: Dict[str, List[Any]] = {
        str(k): list(v) for k, v in (getattr(bar_pre, "rows", {}) or {}).items()
    }
    # 행별 높이: 우측 값 라벨(최종 full 요약)이 2줄 이상이 되어도 잘리지 않도록 산정.
    # full 요약은 재생 중 표시되는 텍스트의 최댓값이라, 한 번만 계산하면 재생 내내 안전하다.
    row_h_map: Dict[str, int] = {
        str(r): _bar_row_value_label_height(
            full_rows.get(str(r), []) or [], val_w=int(VAL_W), bar_h=BAR_H
        )
        for r in rows
    }
    rows_total_h = sum(int(row_h_map.get(str(r), BAR_H)) for r in rows) if rows else BAR_H
    content_h = (
        int(frame_pad) * 2
        + tick_h
        + inner_sp
        + rows_total_h
        + max(0, len(rows) - 1) * int(row_sp)
        + 4
    )
    # 막대 영역(호스트) 높이를 내용에 맞춰 자동 조절 → 2줄 라벨에도 스크롤/잘림 없음.
    try:
        host.height = ui.Pixel(int(content_h))
    except Exception:
        pass

    mask_w0 = max(0, int(BAR_W) - int(playhead_px))
    masks: Dict[str, Any] = {}

    def _color(state: str) -> int:
        return bar_state_color(str(state or BAR_STATE_EMPTY))

    _clear_ep_timeline_host_content(ch)
    try:
        with host:
            root = ui.VStack(spacing=2, height=int(content_h))
            ch["ep_timeline_widget"] = root
            with root:
                with ui.Frame(style={"padding": int(frame_pad)}):
                    with ui.VStack(spacing=inner_sp):
                        with ui.HStack(height=14, spacing=0):
                            ui.Spacer(width=NAME_W)
                            _attach_bar_time_axis_scrubber(
                                ext,
                                st,
                                screen=int(screen),
                                bar_w=int(BAR_W),
                                total_est=float(total_est),
                                playhead_px=int(playhead_px),
                                tick_step=float(tick_step),
                            )
                            try:
                                t_end_lbl = float(total_est)
                                end_txt = (
                                    f"{int(round(t_end_lbl))}"
                                    if abs(float(t_end_lbl) - float(int(round(t_end_lbl)))) < 1e-6
                                    else f"{float(t_end_lbl):.1f}"
                                )
                            except Exception:
                                end_txt = f"{total_est:.1f}"
                            ui.Spacer(width=6)
                            ui.Label(
                                end_txt,
                                width=24,
                                height=14,
                                alignment=ui.Alignment.LEFT_CENTER,
                                style={"color": 0xFFE0E6F0, "font_size": 10},
                            )
                        for r in rows:
                            rk = str(r)
                            seg_list = full_rows.get(rk, []) or []
                            row_h = int(row_h_map.get(rk, BAR_H))
                            with ui.HStack(height=row_h, spacing=int(row_sp)):
                                ui.Label(rk, width=NAME_W, height=row_h, alignment=ui.Alignment.LEFT_CENTER, style={"color": 0xFFBFC7D5, "font_size": 11})
                                # 막대는 행 안에서 세로 중앙 정렬(2줄 라벨로 행이 커져도 위에 붙지 않게).
                                with ui.VStack(width=BAR_W, height=row_h):
                                    ui.Spacer()
                                    with ui.ZStack(width=BAR_W, height=BAR_H):
                                        ui.Rectangle(width=BAR_W, height=BAR_H, style={"background_color": 0xFF1A1E26})
                                        # 막대는 항상 전체(total_est) 기준으로 1회 그린다.
                                        rects = _bar_segment_rect_widths(
                                            seg_list,
                                            total_est=float(total_est),
                                            bar_w=int(BAR_W),
                                            t_cover=float(total_est),
                                        )
                                        with ui.HStack(height=BAR_H, spacing=0):
                                            used = 0
                                            for w, seg_st in rects:
                                                used += int(w)
                                                ui.Rectangle(
                                                    width=int(w),
                                                    height=BAR_H,
                                                    style={"background_color": _color(seg_st)},
                                                )
                                            if used < BAR_W:
                                                ui.Spacer(width=(BAR_W - used))
                                        # 진행 마스크(행당 1개) — 배경색으로 미래 구간을 덮는다.
                                        placer = ui.Placer(offset_x=int(playhead_px), offset_y=0)
                                        with placer:
                                            mrect = ui.Rectangle(
                                                width=ui.Pixel(max(1, int(mask_w0))),
                                                height=BAR_H,
                                                style={"background_color": _BAR_MASK_BG},
                                            )
                                        try:
                                            mrect.visible = bool((not preview_full) and mask_w0 > 0)
                                        except Exception:
                                            pass
                                    ui.Spacer()
                                # 값 라벨(누적 요약) — 텍스트만 추후 갱신
                                if preview_full:
                                    seg_v = seg_list
                                else:
                                    try:
                                        seg_v = truncate_bar_rows_at_t({rk: seg_list}, float(playhead_px) / float(BAR_W) * float(total_est) if BAR_W else 0.0).get(rk, [])
                                    except Exception:
                                        seg_v = seg_list
                                try:
                                    dur_txt = format_row_state_duration_summary(seg_v)
                                except Exception:
                                    dur_txt = ""
                                vlabel = ui.Label(
                                    dur_txt,
                                    width=int(VAL_W),
                                    height=row_h,
                                    word_wrap=True,
                                    alignment=ui.Alignment.LEFT_CENTER,
                                    style={"color": 0xFFDDDDDD, "font_size": 11},
                                )
                                masks[rk] = (placer, mrect, vlabel)
        st["_mask_widgets"] = masks
        st["_mask_full_rows"] = full_rows
        _sync_bar_graph_copy_button(ext, ch, int(screen))
        return True
    except Exception as ex:
        print(f"[SIM] EP 막대 마스크 렌더 실패(화면{screen}): {ex}", flush=True)
        _clear_ep_timeline_host_content(ch)
        st["_mask_widgets"] = {}
        st["_mask_full_rows"] = {}
        st["_mask_sig"] = None
        st.pop("_bar_playhead_placer", None)
        st.pop("_bar_playhead_label", None)
        st.pop("_bar_seek_tick_stack", None)
        return False


def _render_or_update_precomputed_bar_mask(
    ext: Any,
    ch: Dict[str, Any],
    st: Dict[str, Any],
    *,
    screen: int,
    t_bar: float,
    total_est: float,
    preview_full: bool,
    layout: Tuple[int, int, int, int, int],
) -> bool:
    """재생(프리런 사전계산) 막대를 정적 1회 렌더 + 행별 마스크로 갱신한다.

    구조가 동일하면(같은 행/스케일/레이아웃/프리런) 마스크 width/offset 만 갱신하고,
    구조가 바뀌면 1회 재빌드한다. 처리하면 True, 처리 못하면 False(레거시 경로로 폴백).
    """
    host = ch.get("ep_timeline_host")
    if host is None:
        return False

    pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
    bar_pre = pre_by.get(str(int(screen))) if isinstance(pre_by, dict) else None
    if not isinstance(bar_pre, EpBarPrecomputed):
        return False
    if not isinstance(getattr(bar_pre, "rows", None), dict) or not bar_pre.rows:
        return False
    try:
        total_est = float(total_est)
    except Exception:
        return False
    if total_est <= 1e-9:
        return False

    # 행 순서 — 프리런 row_order 우선
    rows: List[Any] = []
    if getattr(bar_pre, "row_order", None):
        try:
            rows = normalize_bar_graph_row_order(list(bar_pre.row_order))
        except Exception:
            rows = []
    if not rows:
        try:
            ep_idx = int(_ep_count_idx_for_port_panel(ext, int(screen)))
            snap = _sim_snapshot_for_screen(ext, int(screen))
            ebs_on = bool(snap.get("ebs_enabled", True)) if snap else True
            rows = list(bar_graph_row_order(ep_idx, ebs_enabled=ebs_on))
        except Exception:
            rows = []
    if not rows:
        return False

    BAR_W = int(layout[0])
    try:
        playhead_px = int(round((float(t_bar) / float(total_est)) * float(BAR_W)))
    except Exception:
        playhead_px = 0
    playhead_px = max(0, min(int(BAR_W), playhead_px))

    seg_ver = 0
    try:
        seg_ver = sum(len(bar_pre.rows.get(str(r), []) or []) for r in rows)
    except Exception:
        seg_ver = 0
    sig = (
        tuple(str(r) for r in rows),
        round(float(total_est), 3),
        tuple(int(x) for x in layout),
        id(bar_pre),
        int(seg_ver),
    )

    widget_alive = ch.get("ep_timeline_widget") is not None
    masks = st.get("_mask_widgets")
    cur_sig = st.get("_mask_sig")

    if widget_alive and cur_sig == sig and isinstance(masks, dict) and masks:
        # 변화 없음(dt=0 · preview 동일) → 아무 것도 안 함
        same_t = abs(float(st.get("_mask_last_t", -1.0)) - float(t_bar)) < 1e-6
        same_pv = bool(st.get("_mask_last_preview", None)) == bool(preview_full)
        if same_t and same_pv:
            return True
        _apply_bar_mask_widths(
            st,
            rows,
            bar_w=int(BAR_W),
            playhead_px=int(playhead_px),
            t_bar=float(t_bar),
            preview_full=bool(preview_full),
        )
        st["_mask_last_t"] = float(t_bar)
        st["_mask_last_preview"] = bool(preview_full)
        _sync_bar_graph_copy_button(ext, ch, int(screen))
        return True

    ok = _build_precomputed_bar_with_mask(
        ext,
        ch,
        st,
        screen=int(screen),
        rows=rows,
        bar_pre=bar_pre,
        total_est=float(total_est),
        playhead_px=int(playhead_px),
        preview_full=bool(preview_full),
        layout=layout,
    )
    if ok:
        st["_mask_sig"] = sig
        st["_mask_last_t"] = float(t_bar)
        st["_mask_last_preview"] = bool(preview_full)
        _sync_bar_graph_copy_button(ext, ch, int(screen))
        return True
    return False


def _update_ep_timeline_under_port_state(
    ext: Any,
    ch: Dict[str, Any],
    occ: Dict[str, Any],
    sim_time_text: str,
    *,
    honor_explicit_sim_time: bool = False,
    playback_ui_state: Any = None,
) -> None:
    """포트상태 아래 막대 — EP·ALL_EP·INOUT·BP (5상태, 프리런 사전계산 우선)."""
    host = ch.get("ep_timeline_host")
    try:
        screen = int(ch.get("screen", 1))
    except Exception:
        screen = 1
    scr_key = str(screen)

    # 재생 SSOT — ``resolve_playback_ui_at_sim`` 결과만 사용 (truncate/overlay/live occ 금지)
    if playback_ui_state is None:
        try:
            from .control_sim_playback_plan import (
                playback_plan_active,
                resolve_playback_ui_at_sim,
            )

            if bool(getattr(ext, "_sim_playback_started", False)) and playback_plan_active(
                ext, int(screen)
            ):
                t_probe = _resolve_ep_timeline_sim_time(
                    ext, screen, sim_time_text, honor_explicit=bool(honor_explicit_sim_time)
                )
                playback_ui_state = resolve_playback_ui_at_sim(
                    ext,
                    int(screen),
                    float(t_probe),
                    explicit=bool(honor_explicit_sim_time),
                )
        except Exception:
            playback_ui_state = None

    if playback_ui_state is not None:
        pui = playback_ui_state
        occ = dict(getattr(pui, "ports", {}) or {})
        t_display = float(pui.axes.t_display)
        t_bar = float(t_display)
        use_precomputed = True
        preview_full = bool(getattr(pui, "preview_full", False))
        rows = list(getattr(pui, "row_order", ()) or ())
        if not rows:
            try:
                from .control_sim_bar_graph import bar_graph_row_order

                ep_idx = int(_ep_count_idx_for_port_panel(ext, int(screen)))
                snap_cfg = _sim_snapshot_for_screen(ext, int(screen))
                ebs_on = bool(snap_cfg.get("ebs_enabled", True)) if snap_cfg else True
                rows = list(bar_graph_row_order(ep_idx, ebs_enabled=ebs_on))
            except Exception:
                rows = []
        rows_state = {
            str(k): list(v) for k, v in (getattr(pui, "bar_rows", {}) or {}).items()
        }
        total_est_fixed = float(getattr(pui, "bar_total_est", 0.0) or 0.0)
        if not any(isinstance(v, list) and v for v in rows_state.values()):
            pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
            bar_pre_fb = pre_by.get(scr_key) if isinstance(pre_by, dict) else None
            if isinstance(bar_pre_fb, EpBarPrecomputed) and isinstance(bar_pre_fb.rows, dict) and bar_pre_fb.rows:
                rows_state = truncate_bar_rows_at_t(bar_pre_fb.rows, float(t_bar))
                if float(total_est_fixed) <= 0.0 and float(getattr(bar_pre_fb, "total_est", 0.0) or 0.0) > 0.0:
                    total_est_fixed = float(bar_pre_fb.total_est)
        for r in rows:
            rk = str(r)
            if rk not in rows_state:
                rows_state[rk] = []
        _sync_ep_bar_virtual_time_to_sim(ext, screen, float(t_display))
        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if not isinstance(st_by, dict):
            st_by = {}
            ext._sim_ep_occ_timeline_state_by_screen = st_by
        st = st_by.get(scr_key)
        if not isinstance(st, dict):
            st = {"t_last": None, "rows": {}, "total_est_fixed": None}
            st_by[scr_key] = st
        st["t_last"] = float(t_display)
        st["rows"] = rows_state
        if total_est_fixed > 0.0:
            st["total_est_fixed"] = float(total_est_fixed)
        dt = 0.0
        bar_pre = None
        playback_lock = True
        # 아래 공통 렌더 경로로 fall-through (live/truncate/overlay 분기 생략)
        goto_render = True
    else:
        goto_render = False

    if not goto_render:
        t_playback = _resolve_ep_timeline_sim_time(
            ext, screen, sim_time_text, honor_explicit=bool(honor_explicit_sim_time)
        )
        playback_pure_bar = False
        try:
            from .control_sim_playback_plan import playback_plan_active

            playback_pure_bar = (
                bool(getattr(ext, "_sim_playback_started", False))
                and playback_plan_active(ext, int(screen))
            )
        except Exception:
            playback_pure_bar = False

        t_occ_lookup = float(t_playback)
        if not playback_pure_bar and not bool(honor_explicit_sim_time):
            try:
                from .control_sim_playback_plan import playback_plan_active, plan_lookup_sim_t

                if bool(getattr(ext, "_sim_playback_started", False)) and playback_plan_active(
                    ext, int(screen)
                ):
                    t_occ_lookup = float(
                        plan_lookup_sim_t(ext, int(screen), float(t_playback))
                    )
            except Exception:
                pass
        if not playback_pure_bar:
            occ = _occ_for_ep_timeline(
                ext,
                screen,
                occ if isinstance(occ, dict) else {},
                sim_time_text,
                t_sim=float(t_occ_lookup),
            )
        elif not isinstance(occ, dict):
            occ = {}

        pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
        bar_pre = pre_by.get(scr_key) if isinstance(pre_by, dict) else None
        use_precomputed = isinstance(bar_pre, EpBarPrecomputed)
        player = get_sim_playback_player(ext, screen)
        playback_lock = use_precomputed or (
            player is not None and hasattr(player, "is_playing") and player.is_playing()
        )

        if playback_lock:
            t_display = float(t_playback)
            t_bar = float(t_display)
            _sync_ep_bar_virtual_time_to_sim(ext, screen, float(t_display))
        else:
            try:
                vt_by = getattr(ext, "_sim_ep_timeline_virtual_time_by_screen", None)
                if not isinstance(vt_by, dict):
                    vt_by = {}
                    ext._sim_ep_timeline_virtual_time_by_screen = vt_by
            except Exception:
                vt_by = {}
                try:
                    ext._sim_ep_timeline_virtual_time_by_screen = vt_by
                except Exception:
                    pass
            try:
                vprev = float(vt_by.get(scr_key, 0.0) or 0.0)
            except Exception:
                vprev = 0.0
            if t_playback + 1e-9 < vprev:
                vprev = float(t_playback)
            try:
                now_wall = float(time.perf_counter())
            except Exception:
                now_wall = 0.0
            try:
                last_wall = float(vt_by.get(f"_wall_{scr_key}", 0.0) or 0.0)
            except Exception:
                last_wall = 0.0
            if last_wall <= 0.0 or now_wall <= 0.0:
                dt_wall = 0.0
            else:
                dt_wall = max(0.0, float(now_wall) - float(last_wall))
            try:
                vt_by[f"_wall_{scr_key}"] = float(now_wall)
            except Exception:
                pass
            try:
                sp_model = getattr(ext, "_sim_speed_model", None)
                sp = float(sp_model.get_value_as_float()) if sp_model is not None else 1.0
            except Exception:
                sp = 1.0
            if sp <= 0.0:
                sp = 1.0
            dt_adv_raw = float(dt_wall) * float(sp)
            dt_adv_cap = 0.20 * float(sp)
            dt_adv = min(float(dt_adv_cap), float(dt_adv_raw))
            vnow = float(vprev) + float(dt_adv)
            if vnow > float(t_playback):
                vnow = float(t_playback)
            try:
                vt_by[scr_key] = float(vnow)
            except Exception:
                pass
            t_bar = float(vnow)
            t_display = float(t_bar)

        preview_full = False
        if use_precomputed:
            try:
                pm = getattr(ext, "_sim_bar_preview_model", None)
                preview_full = bool(pm.get_value_as_bool()) if pm is not None else False
            except Exception:
                preview_full = False

        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if not isinstance(st_by, dict):
            st_by = {}
            ext._sim_ep_occ_timeline_state_by_screen = st_by
        st = st_by.get(scr_key)
        if not isinstance(st, dict):
            st = {"t_last": None, "rows": {}, "total_est_fixed": None}
            st_by[scr_key] = st
        t_last_prev = st.get("t_last", None)
        st["t_last"] = float(t_bar)
        if use_precomputed and bar_pre is not None:
            dt = 0.0
            if preview_full:
                rows_state = {k: list(v) for k, v in (bar_pre.rows or {}).items()}
            else:
                rows_state = truncate_bar_rows_at_t(bar_pre.rows, t_bar)
            st["rows"] = rows_state
            if float(bar_pre.total_est) > 0.0:
                st["total_est_fixed"] = float(bar_pre.total_est)
        else:
            if t_last_prev is None:
                dt = 0.0
            else:
                dt = max(0.0, float(t_bar) - float(t_last_prev))

        ep_idx = int(_ep_count_idx_for_port_panel(ext, int(screen)))
        snap = _sim_snapshot_for_screen(ext, int(screen))
        ebs_on = bool(snap.get("ebs_enabled", True)) if snap else True
        rows = list(bar_graph_row_order(ep_idx, ebs_enabled=ebs_on))
        if use_precomputed and bar_pre is not None and bar_pre.row_order:
            rows = normalize_bar_graph_row_order(list(bar_pre.row_order))

        rows_state = st.get("rows", {})
        if not isinstance(rows_state, dict):
            rows_state = {}
            st["rows"] = rows_state
        for r in rows:
            if r not in rows_state or not isinstance(rows_state.get(r), list):
                rows_state[r] = []

        ep_count = 3 if ep_idx else 2
        fault_ports = _fault_ports_from_snapshot(snap, ep_count) if snap else set()

        if not use_precomputed and dt > 1e-9:
            def _live_port_bar_state(port: str) -> str:
                p = str(port or "").strip().upper()
                if p in fault_ports:
                    return "down"
                has_lot = bool(str(occ.get(p, "") or "").strip())
                if p.startswith("EP"):
                    if not has_lot:
                        return BAR_STATE_EMPTY
                    try:
                        by_f = getattr(ext, "_sim_foup_proc_active_ep_by_screen", None)
                        ap = str((by_f or {}).get(str(screen), "") or "").strip().upper() if isinstance(by_f, dict) else ""
                        if ap == p:
                            return "proc"
                    except Exception:
                        pass
                    return BAR_STATE_LOAD
                return BAR_STATE_EMPTY if not has_lot else BAR_STATE_LOAD

            _append_bar_live_interval(
                rows_state, rows, dt=float(dt), state_for_row=_live_port_bar_state
            )

        empty_acc: Dict[str, float] = {}
        for r in rows:
            try:
                empty_acc[r] = sum(
                    float(s.get("dur", 0.0))
                    for s in rows_state.get(r, [])
                    if isinstance(s, dict) and bar_state_from_seg(s) == BAR_STATE_EMPTY
                )
            except Exception:
                empty_acc[r] = 0.0
    else:
        ep_idx = int(_ep_count_idx_for_port_panel(ext, int(screen)))
        snap = _sim_snapshot_for_screen(ext, int(screen))
        ep_count = 3 if ep_idx else 2
        fault_ports = _fault_ports_from_snapshot(snap, ep_count) if snap else set()
        empty_acc: Dict[str, float] = {}
        for r in rows:
            if r not in rows_state or not isinstance(rows_state.get(r), list):
                rows_state[r] = []
        for r in rows:
            try:
                empty_acc[r] = sum(
                    float(s.get("dur", 0.0))
                    for s in rows_state.get(r, [])
                    if isinstance(s, dict) and bar_state_from_seg(s) == BAR_STATE_EMPTY
                )
            except Exception:
                empty_acc[r] = 0.0

    # total_est(막대 스케일): 폴백 max(30,t*1.2)로 먼저 고정된 뒤 엔진 sim_total_est 가 늦게 들어오면
    # 이전에는 상향이 안 되어 30칸이 전부 빨간 EMPTY 세그로만 채워진 것처럼 보였다 → 확정값이 더 크면 상향만 허용.
    total_est = st.get("total_est_fixed", None)
    try:
        total_est = float(total_est) if total_est is not None else None
    except Exception:
        total_est = None
    cand: Optional[float] = None
    try:
        last_te = getattr(ext, "_sim_last_total_est_by_screen", None)
        if isinstance(last_te, dict):
            cand = float(last_te.get(scr_key) or 0.0)
    except Exception:
        cand = None
    if cand is not None and cand <= 0.0:
        cand = None
    if total_est is None or total_est <= 0.0:
        if cand is not None:
            total_est = cand
        else:
            total_est = max(30.0, float(t_bar) * 1.2)
        st["total_est_fixed"] = float(total_est)
    elif cand is not None and cand > float(total_est) + 1e-3:
        st["total_est_fixed"] = float(cand)

    # ep_timeline_host 없음(레거시 단일 패널·웹 스냅샷만): 상태만 갱신하고 omni.ui 는 건너뜀
    if host is None:
        return

    BAR_W, NAME_W, VAL_W, frame_pad, row_sp = _ep_occ_timeline_layout_dims(ext)
    cur_layout = (int(BAR_W), int(NAME_W), int(VAL_W), int(frame_pad), int(row_sp))

    # ── 재생(프리런 사전계산) 막대: 정적 1회 렌더 + 행별 마스크 ──
    # 기존에는 heartbeat 마다 막대 VStack 전체를 destroy/rebuild 하여 3D 뷰가 끊겼다.
    # 막대는 한 번만 그리고, 행마다 배경색 마스크 1개의 offset/width 만 갱신한다.
    # (미리보기 ON 이면 마스크 숨김 → 막대 전체 노출) 처리하면 즉시 반환.
    if use_precomputed:
        if _render_or_update_precomputed_bar_mask(
            ext,
            ch,
            st,
            screen=int(screen),
            t_bar=float(t_bar),
            total_est=float(total_est),
            preview_full=bool(preview_full),
            layout=cur_layout,
        ):
            _sync_bar_graph_copy_button(ext, ch, int(screen))
            return

    # 동일 시뮼 시각(dt=0)·막대 스케일·EP 점유가 같으면 VStack 전체 destroy/rebuild 생략.
    # (매 tick마다 트리를 갈아엎으면 단일 모니터에서 막대 영역 전체가 깜빡인다.)
    try:
        te_snap = float(total_est)
    except Exception:
        te_snap = 0.0
    eps_fp = [r for r in rows if str(r).startswith("EP")]

    def _port_occ_empty(port_key: str) -> bool:
        return not bool(str(occ.get(str(port_key), "") or "").strip())

    try:
        all_empty_fp = all(_port_occ_empty(ep) for ep in eps_fp) if eps_fp else True
        occ_fp = tuple((str(ep), _port_occ_empty(ep)) for ep in eps_fp) + (bool(all_empty_fp),)
    except Exception:
        occ_fp = ()
    old = ch.get("ep_timeline_widget", None)
    last_te = st.get("_ep_tl_last_ui_te")
    last_fp = st.get("_ep_tl_last_ui_occ_fp")
    last_layout = st.get("_ep_tl_last_ui_layout")
    last_t_bar = st.get("_ep_tl_last_ui_t_bar")
    last_preview = st.get("_ep_tl_last_preview")
    skip_render = False
    time_stable = False
    if ch.get("ep_timeline_widget") is None:
        time_stable = False
    elif use_precomputed:
        time_stable = (
            isinstance(last_t_bar, (int, float))
            and abs(float(last_t_bar) - float(t_bar)) <= 1e-4
        )
    else:
        time_stable = dt <= 1e-9
    if old is not None and time_stable:
        layout_ok = (
            isinstance(last_layout, tuple)
            and len(last_layout) == 5
            and tuple(int(x) for x in last_layout) == cur_layout
        )
        base_ok = (
            isinstance(last_te, (int, float))
            and abs(float(last_te) - te_snap) <= 1e-2
            and last_fp == occ_fp
            and layout_ok
        )
        if use_precomputed:
            # 프리컴pute 슬라이스: virtual time(t_bar)·미리보기 토글이 바뀌면 반드시 다시 그린다.
            if (
                base_ok
                and isinstance(last_t_bar, (int, float))
                and abs(float(last_t_bar) - float(t_bar)) <= 1e-4
                and bool(last_preview) == bool(preview_full)
            ):
                skip_render = True
            elif (
                base_ok
                and bool(last_preview) == bool(preview_full)
                and last_fp == occ_fp
            ):
                # t_bar만 전진: UI rebuild 를 ~12Hz 로 제한(형제 영역 합성 깜빡임 완화).
                try:
                    lw = float(st.get("_ep_tl_last_render_wall", 0.0) or 0.0)
                    if (time.perf_counter() - lw) < 0.08:
                        skip_render = True
                except Exception:
                    pass
        elif base_ok:
            skip_render = True
    if skip_render and ch.get("ep_timeline_widget") is not None:
        return

    _clear_ep_timeline_host_content(ch)

    if not rows:
        try:
            ep_idx_fb = int(_ep_count_idx_for_port_panel(ext, int(screen)))
            snap_fb = _sim_snapshot_for_screen(ext, int(screen))
            ebs_fb = bool(snap_fb.get("ebs_enabled", True)) if snap_fb else True
            rows = list(bar_graph_row_order(ep_idx_fb, ebs_enabled=ebs_fb))
        except Exception:
            rows = list(bar_graph_row_order(0, ebs_enabled=True))

    BAR_H = 10
    inner_sp = 3
    tick_h = 14
    tick_step = max(10.0, float(int((((float(total_est) / 8.0) + 9.999) // 10.0) * 10.0)))
    # 행별 높이: 우측 값 라벨이 2줄 이상 접혀도 잘리지 않도록 현재 세그먼트 기준으로 산정.
    row_h_map: Dict[str, int] = {
        str(r): _bar_row_value_label_height(
            rows_state.get(str(r), []) or [], val_w=int(VAL_W), bar_h=BAR_H
        )
        for r in rows
    }
    rows_total_h = sum(int(row_h_map.get(str(r), BAR_H)) for r in rows) if rows else BAR_H
    content_h = (
        int(frame_pad) * 2
        + tick_h
        + inner_sp
        + rows_total_h
        + max(0, len(rows) - 1) * int(row_sp)
        + 4
    )
    # 막대 영역(호스트) 높이를 내용에 맞춰 자동 조절 → 2줄 라벨에도 스크롤/잘림 없음.
    try:
        host.height = ui.Pixel(int(content_h))
    except Exception:
        pass

    try:
        legacy_playhead_px = int(round((float(t_bar) / float(total_est)) * float(BAR_W)))
    except Exception:
        legacy_playhead_px = 0
    legacy_playhead_px = max(0, min(int(BAR_W), legacy_playhead_px))

    def _color(state: str) -> int:
        return bar_state_color(str(state or BAR_STATE_EMPTY))

    try:
        with host:
            root = ui.VStack(spacing=2, height=int(content_h))
            ch["ep_timeline_widget"] = root
            with root:
                with ui.Frame(style={"padding": int(frame_pad)}):
                    with ui.VStack(spacing=inner_sp):
                        # 시간 라벨(너무 촘촘하면 안 보이므로 최대 8개 정도만)
                        with ui.HStack(height=14, spacing=0):
                            ui.Spacer(width=NAME_W)
                            _attach_bar_time_axis_scrubber(
                                ext,
                                st,
                                screen=int(screen),
                                bar_w=int(BAR_W),
                                total_est=float(total_est),
                                playhead_px=int(legacy_playhead_px),
                                tick_step=float(tick_step),
                            )
                            try:
                                t_end_lbl = float(total_est)
                                end_txt = (
                                    f"{int(round(t_end_lbl))}"
                                    if abs(float(t_end_lbl) - float(int(round(t_end_lbl)))) < 1e-6
                                    else f"{float(t_end_lbl):.1f}"
                                )
                            except Exception:
                                end_txt = f"{total_est:.1f}"
                            ui.Spacer(width=6)
                            ui.Label(
                                end_txt,
                                width=24,
                                height=14,
                                alignment=ui.Alignment.LEFT_CENTER,
                                style={"color": 0xFFE0E6F0, "font_size": 10},
                            )
                        for r in rows:
                            row_h = int(row_h_map.get(str(r), BAR_H))
                            with ui.HStack(height=row_h, spacing=int(row_sp)):
                                ui.Label(r, width=NAME_W, height=row_h, alignment=ui.Alignment.LEFT_CENTER, style={"color": 0xFFBFC7D5, "font_size": 11})
                                # 막대는 행 안에서 세로 중앙 정렬(2줄 라벨로 행이 커져도 위에 붙지 않게).
                                with ui.VStack(width=BAR_W, height=row_h):
                                    ui.Spacer()
                                    with ui.ZStack(width=BAR_W, height=BAR_H):
                                        ui.Rectangle(width=BAR_W, height=BAR_H, style={"background_color": 0xFF1A1E26})
                                        segs = rows_state.get(r, []) or []
                                        seg_list = segs if isinstance(segs, list) else []
                                        try:
                                            t_cover = sum(
                                                float(s.get("dur", 0.0))
                                                for s in seg_list
                                                if isinstance(s, dict)
                                            )
                                        except Exception:
                                            t_cover = float(t_bar)
                                        if use_precomputed and bool(preview_full):
                                            t_cover = float(total_est)
                                        elif use_precomputed or playback_lock:
                                            t_cover = min(float(t_bar), float(t_cover))
                                        rects = _bar_segment_rect_widths(
                                            seg_list,
                                            total_est=float(total_est),
                                            bar_w=int(BAR_W),
                                            t_cover=float(t_cover),
                                        )
                                        with ui.HStack(height=BAR_H, spacing=0):
                                            used = 0
                                            for w, seg_st in rects:
                                                used += int(w)
                                                ui.Rectangle(
                                                    width=int(w),
                                                    height=BAR_H,
                                                    style={"background_color": _color(seg_st)},
                                                )
                                            if used < BAR_W:
                                                ui.Spacer(width=(BAR_W - used))
                                    ui.Spacer()
                                try:
                                    dur_txt = format_row_state_duration_summary(seg_list)
                                except Exception:
                                    dur_txt = f"{float(empty_acc.get(r, 0.0) or 0.0):.1f}s"
                                ui.Label(
                                    dur_txt,
                                    width=int(VAL_W),
                                    height=row_h,
                                    word_wrap=True,
                                    alignment=ui.Alignment.LEFT_CENTER,
                                    style={"color": 0xFFDDDDDD, "font_size": 11},
                                )
        _sync_bar_graph_copy_button(ext, ch, int(screen))
    except Exception as ex:
        print(f"[SIM] EP 막대 UI 렌더 실패(화면{screen}): {ex}", flush=True)
        _clear_ep_timeline_host_content(ch)

    try:
        st["_ep_tl_last_ui_te"] = float(te_snap)
        st["_ep_tl_last_ui_occ_fp"] = occ_fp
        st["_ep_tl_last_ui_layout"] = cur_layout
        st["_ep_tl_last_ui_t_bar"] = float(t_bar)
        st["_ep_tl_last_preview"] = bool(preview_full) if use_precomputed else False
        st["_ep_tl_last_render_wall"] = float(time.perf_counter())
    except Exception:
        pass


def _sync_all_ep_occ_timelines_from_engines(ext: Any) -> None:
    """
    멀티 시뮬에서 한 화면의 ANIM/큐 폭주로 다른 화면의 ``timeline_only`` 가 밀리면
    포트 아래 EP 막대가 멈춘 것처럼 보인다. 재생 중에는 ``SimTimelinePlayer.sim_now`` 를 쓴다.
    """
    chans = getattr(ext, "_sim_monitor_channels", None)
    engs = getattr(ext, "_sim_engines", None)
    if not isinstance(chans, list) or len(chans) < 2:
        return
    if not isinstance(engs, list) or len(engs) < 2:
        return
    last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
    if not isinstance(last_by, dict):
        last_by = {}
    empty_occ = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
    for i, ch in enumerate(chans):
        if not isinstance(ch, dict):
            continue
        si = i + 1
        sk = str(si)
        eng = engs[i] if i < len(engs) else None
        if eng is None or bool(getattr(eng, "is_done", False)):
            continue
        t_now = _resolve_ep_timeline_sim_time(ext, si, "")
        if get_sim_playback_player(ext, si) is None:
            try:
                t_now = float(getattr(getattr(eng, "env", None), "now", 0.0) or 0.0)
            except Exception:
                t_now = 0.0
        occ = last_by.get(sk) if isinstance(last_by.get(sk), dict) else None
        if occ is None:
            occ = dict(empty_occ)
        try:
            lp_by = getattr(ext, "_sim_progress_last_payload_by_screen", None)
            lp = lp_by.get(sk) if isinstance(lp_by, dict) else None
            occ = _occ_for_ep_timeline(ext, si, occ, f"{t_now:.2f}", progress_p=lp if isinstance(lp, dict) else None)
        except Exception:
            pass
        try:
            _update_ep_timeline_under_port_state(ext, ch, occ, f"{t_now:.2f}")
        except Exception:
            pass


def _enqueue_sim_log(ext: Any, line: str) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.HISTORY_LINE, (line or "").strip()))
    except Exception:
        pass


def post_sim_history_line(ext: Any, line: str) -> None:
    """시뮬 워커 스레드에서 호출: 스토리/상태 텍스트를 '이력 로그' 패널로 보낸다."""
    _enqueue_sim_log(ext, line)


def _enqueue_anim_event(ext: Any, payload: Dict[str, str]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        p2 = dict(payload or {})
        try:
            p2["_run_gen"] = str(int(getattr(ext, "_sim_run_gen", 0) or 0))
        except Exception:
            p2["_run_gen"] = "0"
        q.put_nowait((SimUiQueueKind.ANIM_EVENT, p2))
    except Exception:
        pass


def post_sim_anim_event(ext: Any, payload: Dict[str, str]) -> None:
    """시뮬 워커 스레드에서 호출: 애니메이션 파이프라인(포트 패널 + 애니 이력)으로 이벤트를 넘긴다."""
    _enqueue_anim_event(ext, payload)


def _enqueue_control_action(ext: Any, action: str) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.ACTION, action))
    except Exception:
        pass


def _enqueue_gate_request(ext: Any, payload: Dict[str, Any]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.GATE, dict(payload or {})))
    except Exception:
        pass


def _show_sim_gate_dialog(ext: Any, payload: Dict[str, Any]) -> None:
    # 공정확인 모드에서는 "확인 전까지 완전 정지"가 목표이므로,
    # 확인창이 이미 떠 있으면 새 창으로 교체하지 않고(=pause가 풀리는 부작용 방지) 그냥 대기시킨다.
    if getattr(ext, "_sim_gate_dialog", None) is not None:
        return
    title = str(payload.get("title", "공정 확인"))
    msg = str(payload.get("message", "다음 공정을 진행할까요?"))
    done = payload.get("_done_event", None)
    # 웹(HTTP 브리지)에서 공정 확인을 대신 처리할 수 있도록 직렬화 가능한 요약 + done 이벤트 참조를 남긴다.
    try:
        ext._sim_web_gate_pending = {
            "title": title,
            "message": msg,
            "gate_seq_raw": str(payload.get("gate_seq_raw", "") or "").strip(),
            "gate_seq_canonical": str(payload.get("gate_seq_canonical", "") or "").strip(),
            "gate_xml_sequence_name": str(payload.get("gate_xml_sequence_name", "") or "").strip(),
        }
        ext._sim_web_gate_done_event = done
    except Exception:
        pass
    g_raw = str(payload.get("gate_seq_raw", "")).strip()
    g_can = str(payload.get("gate_seq_canonical", "")).strip()
    g_xml = str(payload.get("gate_xml_sequence_name", "")).strip()
    win_suffix = f" [{g_raw}]" if g_raw else ""
    ext._sim_gate_dialog = ui.Window(f"[SIM 확인] {title}{win_suffix}", width=580, height=400)
    with ext._sim_gate_dialog.frame:
        with ui.VStack(spacing=8, padding=10):
            with ui.Frame(style={"background_color": 0xFF2A3140, "border_width": 1, "border_color": 0xFF5A6A80}):
                with ui.VStack(spacing=4, padding=8):
                    ui.Label("이벤트 (sequence_name)", height=22, style={"color": 0xFF8EC8FF})
                    if g_raw and g_can and g_raw == g_can:
                        seq_line = f"sequence_name: {g_raw}"
                    elif g_raw or g_can:
                        seq_line = f"시뮬 seq: {g_raw or '-'}  → 규격/별칭: {g_can or '-'}"
                    else:
                        seq_line = "sequence_name: -"
                    ui.Label(seq_line, word_wrap=True, height=36)
                    if g_xml:
                        ui.Label(f"XML SEQUENCE_NAME: {g_xml}", height=22, style={"color": 0xFFC8E0FF})
            with ui.ScrollingFrame(height=240):
                with ui.VStack(spacing=4):
                    ui.Label(msg, word_wrap=True, height=200)
            with ui.HStack(spacing=8, height=30):
                ui.Button("확인", width=80, clicked_fn=lambda: _close_sim_gate_dialog(ext, done))


def _close_sim_gate_dialog(ext: Any, done_event: Any) -> None:
    w = getattr(ext, "_sim_gate_dialog", None)
    if w is not None:
        try:
            w.visible = False
            # 이벤트/드로우 중 destroy 호출 금지: 다음 프레임으로 지연
            def _defer_destroy(_e=None):
                try:
                    w.destroy()
                except Exception:
                    pass
            try:
                app.get_app().get_post_update_event_stream().create_subscription_to_pop(
                    _defer_destroy,
                    name="morph.tbs_control_2:sim_gate_destroy",
                )
            except Exception:
                pass
        except Exception:
            pass
    ext._sim_gate_dialog = None
    # 이벤트 확인창이 닫히면 gate pause 해제 (애니 pause는 별도 이벤트로 유지)
    try:
        gp = getattr(ext, "_sim_gate_pause_event", None)
        if gp is not None:
            gp.clear()
    except Exception:
        pass
    try:
        if done_event is not None:
            done_event.set()
    except Exception:
        pass
    try:
        ext._sim_web_gate_pending = None
        ext._sim_web_gate_done_event = None
    except Exception:
        pass


def _enqueue_sim_progress(ext: Any, payload: Dict[str, str]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.PROGRESS, dict(payload or {})))
    except Exception:
        pass


def post_sim_progress_update(ext: Any, payload: Dict[str, str]) -> None:
    """시뮬 워커 스레드에서 호출: 공정 진행률/상태를 '진행현황' 패널로 보낸다."""
    _enqueue_sim_progress(ext, payload)


def _timeline_event_needs_json_gate(seq_u: str) -> bool:
    if not seq_u:
        return False
    if seq_u == "PORT_OCC_REFRESH":
        return False
    if seq_u in ("FOUP_PROCESS_START", "FOUP_PROCESS_END"):
        return False
    return True


def _screen_anim_worker_has_pending(ext: Any, screen: int) -> bool:
    try:
        workers = getattr(ext, "_sim_anim_workers_by_screen", None)
        if not isinstance(workers, dict):
            return False
        ent = workers.get(str(max(1, int(screen))))
        if not isinstance(ent, dict):
            return False
        q = ent.get("queue")
        return isinstance(q, list) and len(q) > 0
    except Exception:
        return False


def _prepare_playback_emit_environment(
    ext: Any,
    results: Dict[int, Any],
    *,
    scope_screens_only: bool = False,
) -> None:
    """
    프리런 재생 직전: 잔류 JSON wall·러너 busy 를 비워 첫 타임라인 event 가 막히지 않게 한다.

    ``emit_due_items`` 는 event 게이트에 걸리면 커서를 전진시키지 못해 시계만 흐르는 증상이 난다.

    ``scope_screens_only=True`` — 다른 화면이 이미 재생 중일 때 ``results`` 에 해당하는 화면만 정리한다.
    """
    scoped: set = set()
    if scope_screens_only:
        for scr_k in (results or {}).keys():
            try:
                scoped.add(int(scr_k))
            except Exception:
                continue
    try:
        if scope_screens_only and scoped:
            for scr_i in scoped:
                try:
                    set_json_wall_busy(ext, scr_i, False)
                except Exception:
                    pass
            try:
                from .control_sim_playback_gate import clear_playback_step_speed_locks

                clear_playback_step_speed_locks(ext)
            except Exception:
                pass
        else:
            clear_playback_gate_state(ext)
            clear_playback_step_speed_locks(ext)
    except Exception:
        pass
    try:
        if scope_screens_only and scoped:
            by = getattr(ext, "_sim_playback_json_jobs_by_screen", None)
            if isinstance(by, dict):
                for scr_i in scoped:
                    by.pop(str(int(scr_i)), None)
        else:
            _clear_playback_json_job_queues(ext)
    except Exception:
        pass
    try:
        if scope_screens_only and scoped:
            foup_by = getattr(ext, "_sim_foup_playback_last_by_screen", None)
            if isinstance(foup_by, dict):
                for scr_i in scoped:
                    foup_by.pop(str(int(scr_i)), None)
                    foup_by.pop(int(scr_i), None)
        else:
            ext._sim_foup_playback_last_by_screen = {}
    except Exception:
        pass
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans, list) and chans:
            _sync_foup_labels_to_channels(ext, chans)
    except Exception:
        pass
    for scr_k in (results or {}).keys():
        try:
            scr_i = max(1, int(scr_k))
        except Exception:
            continue
        try:
            set_json_wall_busy(ext, scr_i, False)
        except Exception:
            pass
        try:
            _halt_screen_json_anim(
                ext,
                scr_i,
                join_sec=0.15 if is_multi_playback_instances(ext) else 0.5,
            )
        except Exception:
            pass
        try:
            set_json_wall_busy(ext, scr_i, False)
        except Exception:
            pass


def _deliver_playback_timeline_emit(ext: Any, kind: str, payload: Any, screen: int) -> None:
    """
    프리런 재생 타임라인 emit — 메인(UI) 스레드에서 **동기** 처리.

    재생 중 큐를 쓰면 2화면 로그 폭주 시 진행률·애니가 지연된다.
    event 는 ``json_wall_busy`` 레이스 방지를 위해 반드시 동기 처리한다.
    """
    scr = max(1, int(screen))
    if kind == "log":
        line = payload if isinstance(payload, str) else str(payload)
        if scr > 1:
            line = f"[화면{scr}] {line}"
        try:
            _sim_ui_sink_history_line(ext, line, SimLogPanelMode.ALL)
        except Exception:
            post_sim_history_line(ext, line)
        return
    if kind == "event" and isinstance(payload, dict):
        pl = dict(payload)
        pl["tbs_sim_screen"] = str(scr)
        # 프리런에 저장된 구 세대 토큰이 있으면 sink 가 이벤트를 버린다.
        pl.pop("_run_gen", None)
        seq_u = _normalize_anim_event_seq(str(pl.get("seq") or ""))
        needs_json_gate = _timeline_event_needs_json_gate(seq_u)
        if needs_json_gate:
            try:
                set_json_wall_busy(ext, scr, True)
            except Exception:
                pass
            # proc_wait: JSON wall 이 먼저 풀려도 공정 종료 전 다음 gated emit 금지
            try:
                t0 = float(
                    str(
                        pl.get("event_start_sim_time")
                        or pl.get("t")
                        or pl.get("sim_time")
                        or "0"
                    ).strip()
                    or "0"
                )
                proc = float(str(pl.get("proc_sec") or "0").strip() or "0")
                if proc > 1e-9:
                    set_proc_gate_end(ext, scr, float(t0) + float(proc))
            except Exception:
                pass
        try:
            _sim_ui_sink_anim_event(ext, pl, SimLogPanelMode.ALL)
        except Exception:
            post_sim_anim_event(ext, pl)
        else:
            # 매핑 실패 등으로 JSON 이 안 떴는데 wall 만 True 인 경우 다음 event 영구 차단 방지
            if needs_json_gate:
                try:
                    from .control_sim_playback_gate import is_screen_runner_busy

                    active = _screen_active_json_job(ext, scr)
                    in_lead = (
                        isinstance(active, dict)
                        and bool(active.get("_json_pending_sim_start"))
                        and not bool(active.get("_json_sequence_started"))
                    )
                    if (not in_lead) and (not is_screen_runner_busy(ext, scr)) and (
                        not _screen_anim_worker_has_pending(ext, scr)
                    ):
                        try_release_json_wall_when_idle(ext, scr)
                except Exception:
                    pass
        return
    if kind == "progress" and isinstance(payload, dict):
        p = dict(payload)
        p["tbs_sim_screen"] = str(scr)
        try:
            st = str(p.get("status", "") or "").upper()
            el = float(str(p.get("elapsed", "0") or "0").strip() or "0")
        except Exception:
            st, el = "", 0.0
        if st == "RUNNING" and el <= 1e-9:
            lock_playback_step_speed(ext, scr)
        elif st == "DONE":
            unlock_playback_step_speed(ext, scr)
        try:
            _sim_ui_sink_progress(ext, p)
        except Exception:
            post_sim_progress_update(ext, p)
        return


def _deliver_playback_heartbeat_progress(ext: Any, payload: Dict[str, Any]) -> None:
    scr_opt = _resolve_payload_sim_screen(ext, payload if isinstance(payload, dict) else {})
    if scr_opt is None:
        try:
            print("[TBS/port-screen] drop heartbeat (no tbs_sim_screen)", flush=True)
        except Exception:
            pass
        return
    scr = int(scr_opt)
    try:
        _sim_ui_sink_progress(ext, dict(payload or {}))
    except Exception:
        try:
            post_sim_progress_update(ext, dict(payload or {}))
        except Exception:
            pass
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    try:
        tnow = float(str((payload or {}).get("sim_time", "0") or "0").strip() or "0")
        player = get_sim_playback_player(ext, scr)
        if player is not None:
            try:
                tnow = float(player.sim_now(scr))
            except Exception:
                pass
        from .control_sim_playback_plan import refresh_playback_display_at_sim

        refresh_playback_display_at_sim(ext, scr, tnow)
        _refresh_foup_playback_heartbeat(ext, scr, tnow)
    except Exception:
        pass


def _build_playback_prog_payload_for_session(
    scr: int,
    tnow: float,
    lp: Optional[Dict[str, Any]],
    ext: Any,
) -> Dict[str, Any]:
    try:
        results = getattr(ext, "_sim_prerun_results_by_screen", None)
        te_val = None
        if isinstance(results, dict) and results.get(int(scr)) is not None:
            te_val = float(results[int(scr)].final_sim_time)
    except Exception:
        te_val = None
    p3 = _build_playback_time_tick_payload(
        int(scr),
        float(tnow),
        lp if isinstance(lp, dict) else None,
        final_sim_time=te_val,
        ext=ext,
    )
    return p3


def _sync_playback_engine_now(ext: Any, screen: int, tnow: float) -> None:
    """UI 막대 동기화용 playback 엔진 now."""
    try:
        engs = getattr(ext, "_sim_engines", None)
        if not isinstance(engs, list):
            return
        idx = int(screen) - 1
        eng = engs[idx] if 0 <= idx < len(engs) else None
        if eng is None:
            return
        if hasattr(eng, "_set_now"):
            eng._set_now(float(tnow))  # type: ignore[attr-defined]
        elif hasattr(eng, "env") and eng.env is not None:
            eng.env.now = float(tnow)  # type: ignore[attr-defined]
    except Exception:
        pass


def _make_playback_emit_fn(ext: Any, results: Dict[int, Any]) -> Callable[[str, Any, int], None]:
    def _emit(kind: str, payload: Any, screen: int) -> None:
        if isinstance(payload, dict):
            try:
                rr2 = results.get(int(screen))
                if rr2 is not None:
                    payload = dict(payload)
                    payload["sim_total_est_sec"] = f"{float(rr2.final_sim_time):.2f}"
            except Exception:
                pass
        _deliver_playback_timeline_emit(ext, kind, payload, int(screen))

    return _emit


def _is_playback_time_tick_payload(payload: Dict[str, Any]) -> bool:
    """프리런 재생 heartbeat — sim_time·elapsed·percent 를 보간 갱신(포트/애니 상태는 재전송 안 함)."""
    return str(payload.get("playback_time_tick", "")).strip() in ("1", "true", "True", "ON", "on")


_PLAYBACK_TIME_TICK_STRIP_KEYS = (
    "ports_occupancy",
    "ep_occ",
    "all_ep_empty",
    "foup_proc_active_ep",
)


def _apply_playback_step_progress_from_sim(
    p3: Dict[str, Any],
    tnow: float,
    *,
    ext: Any = None,
    screen: int = 1,
) -> None:
    """
    타임라인 progress 항목 사이(애니·공정 대기) heartbeat 에서 elapsed/percent 를 sim_now 로 보간.

    **MOVE/ARRIVED 등 현재 이벤트** 진행현황 전용. FOUP 공정은 ``_apply_foup_playback_progress_from_sim`` 사용.
    """
    try:
        t0 = float(str(p3.get("event_start_sim_time") or "").strip() or "0")
    except Exception:
        t0 = 0.0
    proc = 0.0
    try:
        proc = float(str(p3.get("proc_sec") or "").strip() or "0")
    except Exception:
        proc = 0.0
    if proc <= 1e-9:
        try:
            proc = float(str(p3.get("total") or "").strip() or "0")
        except Exception:
            proc = 0.0
    if ext is not None:
        try:
            active_by = getattr(ext, "_sim_anim_active_by_screen", None)
            act = active_by.get(str(int(screen))) if isinstance(active_by, dict) else None
            if isinstance(act, dict):
                try:
                    at = float(str(act.get("t") or "").strip() or "0")
                except Exception:
                    at = 0.0
                try:
                    ap = float(str(act.get("proc_sec") or "").strip() or "0")
                except Exception:
                    ap = 0.0
                if at > 1e-9:
                    t0 = at
                if ap > 1e-9:
                    proc = ap
        except Exception:
            pass
    if proc <= 1e-9:
        return
    el = max(0.0, min(float(proc), float(tnow) - float(t0)))
    pct = min(100.0, 100.0 * el / float(proc))
    p3["elapsed"] = f"{el:.1f}"
    p3["total"] = f"{float(proc):.1f}"
    p3["percent"] = str(int(pct))


def _apply_foup_playback_progress_from_sim(p3: Dict[str, Any], tnow: float) -> None:
    """
    FOUP 공정 라벨 heartbeat 보간 — **FOUP 전용** ``event_start_sim_time``·``proc_sec`` 만 사용.

    MOVE/ARRIVED 등 현재 JSON 애니(``_sim_anim_active_by_screen``)와 무관하다.
    +Y / 공정(설정 시간) / -Y 단계는 타임라인에 기록된 FOUP_PROCESS payload 기준.
    """
    if str(p3.get("event_seq") or p3.get("sequence_name") or "").strip().upper() != "FOUP_PROCESS":
        return
    try:
        t0 = float(str(p3.get("event_start_sim_time") or "").strip() or "0")
    except Exception:
        t0 = 0.0
    proc = 0.0
    try:
        proc = float(str(p3.get("proc_sec") or "").strip() or "0")
    except Exception:
        proc = 0.0
    if proc <= 1e-9:
        try:
            proc = float(str(p3.get("total") or "").strip() or "0")
        except Exception:
            proc = 0.0
    if proc <= 1e-9:
        return
    el = max(0.0, min(float(proc), float(tnow) - float(t0)))
    pct = min(100.0, 100.0 * el / float(proc))
    p3["elapsed"] = f"{el:.1f}"
    p3["total"] = f"{float(proc):.1f}"
    p3["percent"] = str(int(pct))


def _build_playback_time_tick_payload(
    scr: int,
    tnow: float,
    lp: Optional[Dict[str, Any]],
    *,
    final_sim_time: Optional[float] = None,
    ext: Any = None,
) -> Dict[str, Any]:
    """프리런 재생 heartbeat — ProgressStepState 기준 보간(lp 레거시 인자는 무시)."""
    _ = lp
    return build_playback_tick_payload(
        ext,
        int(scr),
        float(tnow),
        final_sim_time=final_sim_time,
        apply_step_progress=_apply_playback_step_progress_from_sim,
    )


def _sim_ui_sink_progress(ext: Any, payload: Dict[str, Any]) -> None:
    p = payload if isinstance(payload, dict) else {}
    scr_opt = _resolve_payload_sim_screen(ext, p)
    if scr_opt is None:
        try:
            print(
                f"[TBS/port-screen] drop PROGRESS (no tbs_sim_screen)",
                flush=True,
            )
        except Exception:
            pass
        return
    scr = max(1, int(scr_opt))
    playback_tick = _is_playback_time_tick_payload(p)
    if not playback_tick:
        try:
            apply_engine_progress_payload(ext, scr, p)
        except Exception:
            pass
    if not playback_tick:
        if (
            (not _should_defer_port_occ_sync_for_renewal(ext, scr))
            and _json_playback_wall_started(ext, scr)
            and (not bool(getattr(ext, "_sim_playback_started", False)))
        ):
            try:
                _flush_pending_post_anim_port_applies(ext, scr)
            except Exception:
                pass
        # fail-safe(비-renewal): JSON wall 시작 후 + sim json_end 경과 시 1회 (라이브 전용).
        # 재생 중 포트는 sim plan milestone replay 만 (heartbeat).
        if not bool(getattr(ext, "_sim_playback_started", False)):
            try:
                if str(p.get("status", "")).strip().upper() == "RUNNING":
                    ev = _normalize_anim_event_seq(str(p.get("event_seq") or p.get("sequence_name") or ""))
                    if ev in _ANIM_PORT_UPDATE_SEQS:
                        ctx = _active_json_timing_context(ext, scr)
                        if bool(ctx.get("has_renewal")):
                            pass
                        elif not _json_playback_wall_started(ext, scr):
                            pass
                        else:
                            elapsed = float(str(p.get("elapsed") or "0").strip() or "0")
                            try:
                                from .json_playback_timing import timing_from_progress

                                info = timing_from_progress(
                                    dict(p),
                                    json_path=ctx.get("json_path"),
                                    steps=ctx.get("steps") if isinstance(ctx.get("steps"), list) else None,
                                )
                                t0 = float(info.get("t0", 0.0))
                                t_sync = info.get("t_port_sync")
                                sync_elapsed = (
                                    float(t_sync) - float(t0)
                                    if t_sync is not None
                                    else float(str(p.get("anim_sec") or "0").strip() or "0")
                                )
                            except Exception:
                                sync_elapsed = float(str(p.get("anim_sec") or "0").strip() or "0")
                            if sync_elapsed > 1e-6 and elapsed + 1e-6 >= sync_elapsed:
                                src = dict(p)
                                src["event"] = ev
                                src["event_start_sim_time"] = str(
                                    p.get("event_start_sim_time") or p.get("t") or ""
                                ).strip()
                                _queue_post_anim_port_apply(ext, scr, src)
                                _flush_pending_post_anim_port_applies(ext, scr)
            except Exception:
                pass
    _update_sim_progress(ext, p)
    if playback_tick:
        return
    # progress payload의 ports_occupancy — 재생(plan) 중에는 sim milestone replay 만.
    occ = p.get("ports_occupancy", {})
    _skip_prog_occ = bool(_renewal_json_guard_active(ext, scr))
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import playback_plan_active

            if playback_plan_active(ext, scr):
                _skip_prog_occ = True
        except Exception:
            pass
        if not _skip_prog_occ:
            _ev_po = _normalize_anim_event_seq(str(p.get("event_seq") or p.get("sequence_name") or ""))
            if _ev_po in _ANIM_PORT_UPDATE_SEQS:
                _skip_prog_occ = True
    elif _screen_active_json_has_renewal(ext, scr):
        _ev_po = _normalize_anim_event_seq(str(p.get("event_seq") or p.get("sequence_name") or ""))
        if _ev_po in _ANIM_PORT_UPDATE_SEQS:
            _skip_prog_occ = True
    if (
        (not _skip_prog_occ)
        and isinstance(occ, dict)
        and occ
        and any((k in occ) for k in _PANEL_PORT_KEYS)
        and (not _should_defer_port_occ_sync_for_renewal(ext, scr))
    ):
        try:
            _sync_port_panel_from_engine_occ(
                ext,
                scr,
                occ,
                str(p.get("sim_time", "") or ""),
                allow_post_anim_block=True,
            )
        except Exception:
            pass
    # FOUP 공정중 EP도 progress에서 최신값을 기억(위치 초기화에서 plateau 유지에 사용)
    try:
        _remember_foup_active_ep(ext, scr, p)
    except Exception:
        pass


def _build_sim_gate_request_payload(ext: Any, p: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    공정확인(게이트) 창에 표시할 title/message를 구성해 큐 payload(dict)로 반환.

    목적:
    - 게이트 메시지 구성 로직을 한 함수로 모아, 유지보수 시 흐름 추적 비용을 줄인다.
    - 기능 동작(표시 순서/매핑/타이머/XML)은 기존과 동일하게 유지한다.
    """
    try:
        seq_raw = str(p.get("seq", "") or "")
        seq_can = SIM_SEQ_ALIAS.get(seq_raw.strip(), seq_raw.strip()) if seq_raw else ""
        lot = str(p.get("lot_id", "") or "")
        lot_seq = str(p.get("lot_seq", "") or "")
        foup_id = str(p.get("foup_id", "") or "")
        fr = str(p.get("from_port_id", "") or "")
        to = str(p.get("to_port_id", "") or "")
        port = str(p.get("port_id", "") or "")
        t = str(p.get("sim_time", "") or "")
        title = f"EVENT t={t}" if t else "EVENT"

        # 공정확인 창 메시지 구성(요구사항 순서):
        # 1) 발생 이벤트명  2) 이벤트 동작 설명  3) 연계 애니 파일/존재/빈파일/불필요
        # 4) TIMER(생성/회수)  5) XML 표시
        lines: List[str] = []
        lines.append(f"[EVENT] sequence_name={seq_can or '-'} (raw={seq_raw or '-'})")
        lines.append(
            f"[EVENT] lot={lot or '-'}"
            + (f" (seq={lot_seq})" if lot_seq else "")
            + (f" foup={foup_id}" if foup_id else "")
            + f" | from={fr or '-'} to={to or '-'} port={port or '-'}"
        )

        xml_text = ""
        seq_for_mapping = seq_can
        parsed: Dict[str, Any] = {}
        try:
            if seq_can in xml_generator.FROM_TO_SEQS:
                xml_text = xml_generator.build_xml_string(
                    seq_can,
                    from_port_id=_parse_port_num(fr, 1),
                    to_port_id=_parse_port_num(to, 1),
                )
            else:
                xml_text = xml_generator.build_xml_string(seq_can, port_id=_parse_port_num(port, 1))
            parsed = xml_generator.parse_xml_string(xml_text) or {}
            parsed_seq = str(parsed.get("sequence_name", "") or "").strip().upper()
            if parsed_seq:
                seq_for_mapping = parsed_seq
        except Exception:
            xml_text = ""
            parsed = {}

        # 2) 이벤트 동작 설명
        action_desc = str(parsed.get("action_desc", "") or "").strip() if isinstance(parsed, dict) else ""
        if action_desc:
            lines.append(f"[ACTION] {action_desc}")
        elif seq_for_mapping:
            lines.append(f"[ACTION] (설명 없음) seq={seq_for_mapping}")
        else:
            lines.append("[ACTION] (설명 없음)")

        # 3) 연계된 애니메이션 파일/존재여부/비어있는 파일 여부
        try:
            seq_u = str(seq_can or "").strip().upper()
            is_anim_event = seq_u in (
                str(xml_generator.SEQ_ARRIVED).strip().upper(),
                str(xml_generator.SEQ_MOVE_TRANSFERING).strip().upper(),
                str(xml_generator.SEQ_MOVE_REQ).strip().upper(),
                str(xml_generator.SEQ_REMOVED).strip().upper(),
            )
        except Exception:
            is_anim_event = False

        if not is_anim_event:
            lines.append("[ANIM] 이 이벤트는 애니메이션이 필요없는 이벤트입니다.")
        else:
            try:
                mapping_payload = dict(p or {})
                mapping_payload["seq"] = seq_for_mapping
                if parsed:
                    mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                    mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                    mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
                mapped_json, _meta, rule_name, source_name = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
                if not mapped_json:
                    lines.append(f"[ANIM] 매핑 없음 (event={seq_for_mapping})")
                else:
                    jp = _normalize_json_path(mapped_json)
                    exists_txt = "존재" if jp.is_file() else "없음"
                    empty_txt = ""
                    if jp.is_file():
                        try:
                            raw = json.loads(jp.read_text(encoding="utf-8"))
                            if isinstance(raw, list) and len(raw) == 0:
                                empty_txt = " / EMPTY(빈 파일)"
                        except Exception:
                            empty_txt = ""
                    lines.append(
                        f"[ANIM] file={jp.name} ({exists_txt}{empty_txt}) | source={source_name or '-'} rule={rule_name or '-'}"
                    )
            except Exception as e:
                lines.append(f"[ANIM] 매핑 확인 실패: {e}")

        # 4) TIMER
        try:
            sim = None
            me = getattr(ext, "_sim_engines", None)
            si_s = str(p.get("tbs_sim_screen", "") or "").strip()
            if isinstance(me, list) and len(me) > 0 and si_s.isdigit():
                idx = int(si_s) - 1
                if 0 <= idx < len(me):
                    sim = me[idx]
            if sim is None and isinstance(me, list) and len(me) > 0:
                sim = me[0]
            if sim is None:
                sim = getattr(ext, "_sim_engine", None)
            now_t = float(p.get("sim_time", "0.0") or 0.0)
            spawn_at = getattr(sim, "_next_spawn_at", None) if sim is not None else None
            pickup_at = getattr(sim, "_next_pickup_at", None) if sim is not None else None
            lines_t: List[str] = []
            if isinstance(spawn_at, (int, float)):
                lines_t.append(f"다음 생성까지: {max(0.0, float(spawn_at) - now_t):.2f}s (sim)")
            if isinstance(pickup_at, (int, float)):
                lines_t.append(f"다음 회수티켓까지: {max(0.0, float(pickup_at) - now_t):.2f}s (sim)")
            if lines_t:
                lines.append("")
                lines.append("TIMER:")
                lines.extend(lines_t)
        except Exception:
            pass

        # 5) XML
        if xml_text:
            lines.append("")
            lines.append("XML:")
            lines.append(xml_text)

        message = "\n".join([ln for ln in lines if ln is not None])
        return {
            "title": title,
            "message": message,
            "_done_event": threading.Event(),
            "gate_seq_raw": seq_raw,
            "gate_seq_canonical": seq_can,
            "gate_xml_sequence_name": "",
        }
    except Exception:
        return None


def _sim_ui_sink_anim_event(ext: Any, payload: Dict[str, Any], panel_mode: SimLogPanelMode) -> None:
    """
    시뮼 엔진에서 올라온 이벤트(큐 ``ANIM_EVENT``)를 메인 스레드에서 처리한다.

    - ``tbs_sim_screen`` 으로 보조 USD 컨텍스트를 골라 **포트 LOT prim 가시성**을 맞춘다.
    - ``handle_sim_event_for_animation`` → rules/map → JSON ``SequenceRunner``(화면별 USD 컨텍스트).
    """
    p = payload if isinstance(payload, dict) else {}
    # stop/reset 이후 들어온 잔여 이벤트는 무시한다.
    try:
        gen_now = int(getattr(ext, "_sim_run_gen", 0) or 0)
        gen_evt = int(str(p.get("_run_gen", gen_now) or gen_now).strip() or gen_now)
        if gen_evt != gen_now:
            if bool(getattr(ext, "_sim_playback_started", False)):
                try:
                    scr_b = _resolve_payload_sim_screen(ext, p)
                    if scr_b is not None:
                        set_json_wall_busy(ext, int(scr_b), False)
                except Exception:
                    pass
            return
    except Exception:
        pass
    scr_opt = _resolve_payload_sim_screen(ext, p)
    if scr_opt is None:
        try:
            print(
                f"[TBS/port-screen] drop ANIM_EVENT (no tbs_sim_screen) "
                f"seq={p.get('seq') or p.get('event') or '-'}",
                flush=True,
            )
        except Exception:
            pass
        return
    scr = int(scr_opt)
    occ = p.get("ports_occupancy", {})
    if not isinstance(occ, dict):
        occ = {}
    # 안정성:
    # - ports_occupancy가 빈 dict이거나(누락), 부분 dict(키 누락) 또는 전부 빈 값이면
    #   포트상태 패널이 통째로 비어 보이거나 "전 포트 EMPTY"로 덮이는 현상이 생길 수 있다.
    # - 따라서 마지막 스냅샷으로 폴백/merge하고, 빈 값으로는 last snapshot을 덮어쓰지 않는다.
    _REQ_PORT_KEYS = ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")
    try:
        by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if not isinstance(by_prev, dict):
            by_prev = {}
            ext._sim_last_ports_occupancy_by_screen = by_prev
        occ_prev = by_prev.get(str(scr))
    except Exception:
        by_prev = {}
        occ_prev = None
    seq_u = str(p.get("seq", "") or "").strip().upper()
    # 1) 부분 dict는 merge (누락 키가 '-'로 보이는 문제 방지)
    try:
        if isinstance(occ_prev, dict) and occ_prev:
            if occ and any((k not in occ) for k in _REQ_PORT_KEYS):
                merged = dict(occ_prev)
                merged.update(dict(occ))
                occ = merged
    except Exception:
        pass
    # 2) 빈 dict는 마지막 스냅샷으로 폴백
    if not occ:
        try:
            if isinstance(occ_prev, dict) and occ_prev:
                occ = dict(occ_prev)
        except Exception:
            pass
    # 3) "전부 빈 값"도 마지막 스냅샷으로 폴백 — 단, 엔진 authoritative payload(회수 완료 등)는 신뢰
    try:
        if (
            occ
            and (not any(bool(str(v or "").strip()) for v in occ.values()))
            and (not _ports_occ_trust_all_empty(occ, seq_u=seq_u))
        ):
            if isinstance(occ_prev, dict) and occ_prev and any(bool(str(v or "").strip()) for v in occ_prev.values()):
                occ = dict(occ_prev)
    except Exception:
        pass
    # PORT_OCC_REFRESH: 재생 중에는 last-occ 도 쓰지 않음(잘못된 screen 오염 방지). live 만 반영.
    if seq_u == "PORT_OCC_REFRESH" and bool(getattr(ext, "_sim_playback_started", False)):
        return
    try:
        if (
            seq_u == "PORT_OCC_REFRESH"
            and isinstance(by_prev, dict)
            and occ
            and any((k in occ) for k in _REQ_PORT_KEYS)
        ):
            by_prev[str(scr)] = dict(occ)
    except Exception:
        pass

    if seq_u == "PORT_OCC_REFRESH":
        # 프리런 재생: 포트·visibility 는 plan milestone replay 만.
        if bool(getattr(ext, "_sim_playback_started", False)):
            return
        # renewal JSON: renewal 시점에만 갱신 — 공정 종료 PORT_OCC_REFRESH 는 생략.
        if _renewal_json_guard_active(ext, scr):
            return
        if _screen_active_json_has_renewal(ext, scr):
            return
        try:
            if _anim_mapped_json_has_renewal(ext, p, seq_u):
                return
        except Exception:
            pass
        try:
            src_by = getattr(ext, "_sim_post_anim_src_by_screen", None)
            stored = src_by.get(str(scr)) if isinstance(src_by, dict) else None
            if isinstance(stored, dict) and bool(stored.get("has_renewal")):
                return
        except Exception:
            pass
        _clear_post_anim_port_applied(ext, scr)
        _clear_renewal_port_defer(ext, scr)
        ctx_nm = _usd_context_name_for_sim_screen(ext, scr)
        active_ep = _remember_foup_active_ep(ext, scr, p)
        try:
            apply_port_lot_prim_visibility_for_context(ctx_nm, occ)
        except Exception as exc:
            try:
                print(
                    f"[TBS/port-screen] visibility skip scr={scr} ctx={ctx_nm!r}: {exc}",
                    flush=True,
                )
            except Exception:
                pass
        try:
            sync_port_lot_positions_after_visibility(ctx_nm, foup_proc_active_ep=active_ep)
        except Exception:
            pass
        _update_port_occupancy_panel(ext, occ, str(p.get("sim_time", "")), screen=scr)
        return

    # 일반 애니: JSON 매핑·실행 후 renewal 이 아니면 패널만 동기화(visibility 는 renewal/종료 시)
    verbose = panel_mode != SimLogPanelMode.PROGRESS_ONLY
    handle_sim_event_for_animation(ext, p, verbose=verbose)
    if (
        _is_anim_port_event(seq_u)
        and (not _should_defer_port_occ_sync_for_renewal(ext, scr))
        and (not _anim_mapped_json_has_renewal(ext, p, seq_u))
        and (not bool(getattr(ext, "_sim_playback_started", False)))
    ):
        try:
            _sync_port_panel_from_engine_occ(
                ext,
                scr,
                occ,
                str(p.get("sim_time", "") or ""),
                allow_post_anim_block=False,
                seq_u=seq_u,
            )
        except Exception:
            pass
    try:
        _sync_all_ep_occ_timelines_from_engines(ext)
    except Exception:
        pass


def _sim_ui_sink_history_line(ext: Any, line: str, panel_mode: SimLogPanelMode) -> None:
    if not line:
        return
    if panel_mode == SimLogPanelMode.PROGRESS_ONLY:
        return
    _append_sim_log(ext, line)


def _finalize_sim_timeline_on_done(ext: Any) -> None:
    """
    시뮬레이션이 자연 종료(완료)되었을 때, EP 타임라인 막대그래프 진행을 즉시 종료 상태로 고정한다.

    목표(요구사항):
    - 종료 후에도 UI 가상 시간 ticker가 계속 돌아 막대가 진행되는 현상 방지
    - 요약 로그의 총 시간(=env.now)과 막대그래프의 총 길이(스케일)가 일치하도록 확정
    """
    # 1) UI 가상 시간 ticker 중단(종료 후 막대가 계속 전진하는 원인)
    try:
        sub = getattr(ext, "_sim_ep_timeline_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_ep_timeline_ui_sub = None
    except Exception:
        pass

    # 2) 화면별 최종 sim_time(env.now)을 얻어 스케일/표시를 확정한다.
    final_by_screen: Dict[str, float] = {}
    try:
        engs = list(getattr(ext, "_sim_engines", None) or [])
    except Exception:
        engs = []
    if engs:
        for i, eng in enumerate(engs):
            if eng is None:
                continue
            scr_key = str(i + 1)
            try:
                t = float(getattr(getattr(eng, "env", None), "now", 0.0) or 0.0)
            except Exception:
                t = 0.0
            final_by_screen[scr_key] = max(0.0, t)
    else:
        eng = getattr(ext, "_sim_engine", None)
        if eng is not None:
            try:
                t = float(getattr(getattr(eng, "env", None), "now", 0.0) or 0.0)
            except Exception:
                t = 0.0
            final_by_screen["1"] = max(0.0, t)

    # 3) UI 그래프들이 사용하는 total_est / virtual_time / state를 최종값으로 덮어쓴다.
    try:
        by_te = getattr(ext, "_sim_last_total_est_by_screen", None)
        if not isinstance(by_te, dict):
            by_te = {}
            ext._sim_last_total_est_by_screen = by_te
        for scr_key, t in final_by_screen.items():
            if t > 0.0:
                by_te[str(scr_key)] = float(t)
    except Exception:
        pass
    try:
        vt_by = getattr(ext, "_sim_ep_timeline_virtual_time_by_screen", None)
        if not isinstance(vt_by, dict):
            vt_by = {}
            ext._sim_ep_timeline_virtual_time_by_screen = vt_by
        for scr_key, t in final_by_screen.items():
            vt_by[str(scr_key)] = float(t)
    except Exception:
        pass
    try:
        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if not isinstance(st_by, dict):
            st_by = {}
            ext._sim_ep_occ_timeline_state_by_screen = st_by
        for scr_key, t in final_by_screen.items():
            st = st_by.get(str(scr_key))
            if not isinstance(st, dict):
                continue
            # total_est_fixed는 포트상태 아래 타임라인의 "전체 길이(스케일)"로 사용된다.
            st["total_est_fixed"] = float(max(0.0, t))
            # 종료 시점 이후 추가 누적이 발생하지 않도록 t_last도 최종값으로 맞춘다.
            st["t_last"] = float(max(0.0, t))
    except Exception:
        pass

    # 4) 즉시 1회 재렌더(최종 시각/스케일 반영)
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and chans:
        last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        for ch in chans:
            if not isinstance(ch, dict):
                continue
            try:
                scr_key = str(int(ch.get("screen", 1) or 1))
            except Exception:
                scr_key = "1"
            t_final = float(final_by_screen.get(scr_key, final_by_screen.get("1", 0.0)) or 0.0)
            occ = None
            try:
                occ = last_by.get(str(scr_key)) if isinstance(last_by, dict) else None
            except Exception:
                occ = None
            if isinstance(occ, dict):
                try:
                    _update_ep_timeline_under_port_state(ext, ch, occ, f"{t_final:.2f}")
                except Exception:
                    pass
                try:
                    # 진행현황 패널 하단 EP 타임라인도 최종 total로 맞춘다.
                    payload = {
                        "sim_time": f"{t_final:.2f}",
                        "sim_total_est_sec": f"{t_final:.2f}",
                        "ep_ports": [k for k in ("EP1", "EP2", "EP3") if k in occ],
                        "ep_occ": {k: ("EMPTY" if not str(occ.get(k, "") or "").strip() else "FULL") for k in ("EP1", "EP2", "EP3")},
                        "all_ep_empty": "1" if all(not str(occ.get(k, "") or "").strip() for k in ("EP1", "EP2", "EP3") if k in occ) else "0",
                    }
                    _update_progress_ep_timeline_widget(ext, ch, payload)
                except Exception:
                    pass


def _sim_ui_sink_action(ext: Any, payload: Any) -> None:
    if str(payload) == SimUiControlAction.EXPORT_XLSX.value:
        _finalize_sim_timeline_on_done(ext)
        _export_sim_logs_to_xlsx(ext)


def _sim_ui_sink_gate(ext: Any, payload: Dict[str, Any]) -> None:
    _show_sim_gate_dialog(ext, payload if isinstance(payload, dict) else {})


def _coerce_sim_ui_queue_kind(kind: Any) -> str:
    """
    큐에서 꺼낸 kind가 SimUiQueueKind 멤버일 때 str(kind)는 'SimUiQueueKind.GATE'처럼
    값이 아니라 멤버 이름이 되어 라우팅이 깨진다. 항상 실제 큐 문자열 값으로 맞춘다.
    """
    if isinstance(kind, SimUiQueueKind):
        return kind.value
    return str(kind)


def _dispatch_sim_ui_queue_item(ext: Any, kind: str, payload: Any, panel_mode: SimLogPanelMode) -> None:
    if kind == SimUiQueueKind.PROGRESS.value:
        _sim_ui_sink_progress(ext, payload if isinstance(payload, dict) else {})
    elif kind == SimUiQueueKind.ANIM_EVENT.value:
        _sim_ui_sink_anim_event(ext, payload if isinstance(payload, dict) else {}, panel_mode)
    elif kind == SimUiQueueKind.ACTION.value:
        _sim_ui_sink_action(ext, payload)
    elif kind == SimUiQueueKind.GATE.value:
        _sim_ui_sink_gate(ext, payload if isinstance(payload, dict) else {})
    elif kind == SimUiQueueKind.HISTORY_LINE.value:
        line = payload if isinstance(payload, str) else str(payload)
        _sim_ui_sink_history_line(ext, line, panel_mode)
    else:
        line = payload if isinstance(payload, str) else str(payload)
        _sim_ui_sink_history_line(ext, line, panel_mode)


def _sync_ep_bar_virtual_time_to_sim(ext: Any, screen: int, t_sim: float) -> None:
    """EP 막대 virtual time 을 현재 sim 시각에 맞춘다(미리보기 해제·Seek 직후)."""
    sk = str(int(screen))
    try:
        vt_by = getattr(ext, "_sim_ep_timeline_virtual_time_by_screen", None)
        if not isinstance(vt_by, dict):
            vt_by = {}
            ext._sim_ep_timeline_virtual_time_by_screen = vt_by
        vt_by[sk] = float(t_sim)
        vt_by[f"_wall_{sk}"] = float(time.perf_counter())
    except Exception:
        pass
    try:
        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if isinstance(st_by, dict) and isinstance(st_by.get(sk), dict):
            st_by[sk]["t_last"] = float(t_sim)
    except Exception:
        pass


def _on_sim_bar_preview_toggled(ext: Any) -> None:
    """결과 미리보기 토글 시 막대 캐시 무효화 → 즉시 한 프레임 재렌더."""
    try:
        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if isinstance(st_by, dict):
            for st in st_by.values():
                if isinstance(st, dict):
                    for k in (
                        "_ep_tl_last_ui_te",
                        "_ep_tl_last_ui_occ_fp",
                        "_ep_tl_last_ui_layout",
                        "_ep_tl_last_ui_t_bar",
                        "_ep_tl_last_preview",
                    ):
                        st.pop(k, None)
    except Exception:
        pass
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if not isinstance(chans, list) or not chans:
            return
        empty_occ = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
        for i, ch in enumerate(chans):
            if not isinstance(ch, dict):
                continue
            si = i + 1
            player = get_sim_playback_player(ext, si)
            if player is None:
                continue
            try:
                tnow = float(player.sim_now(si))
            except Exception:
                tnow = 0.0
            _sync_ep_bar_virtual_time_to_sim(ext, si, tnow)
            occ = empty_occ
            if isinstance(last_by, dict) and isinstance(last_by.get(str(si)), dict):
                occ = dict(last_by.get(str(si)) or empty_occ)
            _update_ep_timeline_under_port_state(ext, ch, occ, f"{tnow:.2f}")
    except Exception:
        pass


def _set_sim_start_enabled(ext: Any, enabled: bool) -> None:
    btns = getattr(ext, "_sim_start_buttons", None)
    if isinstance(btns, list) and btns:
        for btn in btns:
            if btn is None:
                continue
            try:
                btn.enabled = bool(enabled)
            except Exception:
                pass
        return
    btn = getattr(ext, "_sim_start_button", None)
    if btn is None:
        return
    try:
        btn.enabled = bool(enabled)
    except Exception:
        pass


def _clear_ep_timeline_host_content(ch: Dict[str, Any]) -> None:
    """``ep_timeline_host`` 자식 전부 제거 — orphan Frame/VStack 누적 방지."""
    host = ch.get("ep_timeline_host")
    if host is None:
        return
    try:
        host.clear()
    except Exception:
        try:
            for child in list(getattr(host, "children", []) or []):
                try:
                    child.destroy()
                except Exception:
                    pass
        except Exception:
            pass
    ch["ep_timeline_widget"] = None
    ch["ep_timeline_busy_label"] = None


def _clear_ep_prerun_busy_labels(ext: Any) -> None:
    """EP 막대 '프리런 계산 중…' 라벨만 끈다."""
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list):
        return
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        try:
            bl = ch.get("ep_timeline_busy_label")
            if bl is not None:
                bl.text = ""
                bl.visible = False
        except Exception:
            pass


def _set_sim_prerun_ui_busy(ext: Any, busy: bool) -> None:
    """프리런 시작 시에만 막대·타임테이블 busy 표시. 해제 시 타임테이블 행은 건드리지 않는다."""
    _set_sim_start_enabled(ext, not bool(busy))
    try:
        ext._sim_prerun_ui_busy = bool(busy)
    except Exception:
        pass
    if not bool(busy):
        _clear_ep_prerun_busy_labels(ext)
        return
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list):
        return
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        si = int(ch.get("screen", 1) or 1)
        try:
            bl = ch.get("ep_timeline_busy_label")
            if bl is not None:
                bl.text = "프리런 계산 중…"
                bl.visible = True
        except Exception:
            pass
        try:
            set_timetable_busy_label(ch, True, screen=si, ext=ext)
        except Exception:
            pass


_SIM_REQ_PORT_KEYS = ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")

_ANIM_PORT_UPDATE_SEQS = frozenset({
    "ARRIVED",
    "MOVE_TRANSFERING",
    "MOVE_REQ",
    "MOVE",
    "REMOVED",
})

_CANONICAL_TO_SHORT_ANIM_EVENT: Dict[str, str] = {
    str(xml_generator.SEQ_ARRIVED).strip().upper(): "ARRIVED",
    str(xml_generator.SEQ_MOVE_TRANSFERING).strip().upper(): "MOVE_TRANSFERING",
    str(xml_generator.SEQ_MOVE_REQ).strip().upper(): "MOVE_REQ",
    str(xml_generator.SEQ_MOVE).strip().upper(): "MOVE",
    str(xml_generator.SEQ_REMOVED).strip().upper(): "REMOVED",
}


def _normalize_anim_event_seq(ev: str) -> str:
    """짧은 이름(ARRIVED) 또는 정식명(EAPEIS_PORT_ARRIVED) → 짧은 이름으로 통일."""
    e = str(ev or "").strip().upper()
    if not e:
        return ""
    if e in _ANIM_PORT_UPDATE_SEQS:
        return e
    mapped = _CANONICAL_TO_SHORT_ANIM_EVENT.get(e)
    if mapped:
        return mapped
    if e in SIM_SEQ_ALIAS:
        return str(e).strip().upper()
    return e


def _is_anim_port_event(ev: str) -> bool:
    return _normalize_anim_event_seq(ev) in _ANIM_PORT_UPDATE_SEQS


def _anim_mapped_json_has_renewal(ext: Any, payload: Dict[str, Any], seq_u: str) -> bool:
    """이벤트에 매핑된 JSON 에 renewal 마커가 있으면 True."""
    try:
        mapping = dict(payload or {})
        mapping["seq"] = str(seq_u or "").strip().upper()
        mapped_json, _, _, _ = _resolve_event_animation_entry(
            str(seq_u or "").strip().upper(), mapping
        )
        if not mapped_json:
            return False
        from .sim_sequence_json import renewal_info_from_basename_or_path

        has_r, _, _ = renewal_info_from_basename_or_path(str(mapped_json))
        return bool(has_r)
    except Exception:
        return False


def _set_renewal_json_guard(ext: Any, screen: int, active: bool) -> None:
    """renewal JSON 공정 중 — proc_end PORT_OCC·progress occ 가 패널을 덮지 않게."""
    try:
        by = getattr(ext, "_sim_renewal_json_guard_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_renewal_json_guard_by_screen = by
        sk = str(max(1, int(screen)))
        if active:
            by[sk] = True
        else:
            by.pop(sk, None)
    except Exception:
        pass


def _renewal_json_guard_active(ext: Any, screen: int) -> bool:
    try:
        by = getattr(ext, "_sim_renewal_json_guard_by_screen", None)
        if isinstance(by, dict) and by.get(str(max(1, int(screen)))):
            return True
    except Exception:
        pass
    return False


def _set_renewal_port_defer(ext: Any, screen: int, defer: bool) -> None:
    try:
        by = getattr(ext, "_sim_renewal_port_defer_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_renewal_port_defer_by_screen = by
        sk = str(max(1, int(screen)))
        if defer:
            by[sk] = True
        else:
            by.pop(sk, None)
    except Exception:
        pass


def _clear_renewal_port_defer(ext: Any, screen: int) -> None:
    _set_renewal_port_defer(ext, int(screen), False)


def _screen_active_json_job(ext: Any, screen: int) -> Optional[Dict[str, Any]]:
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if isinstance(active_by, dict):
            active = active_by.get(str(max(1, int(screen))))
            if isinstance(active, dict) and active:
                return active
    except Exception:
        pass
    return None


def _json_playback_wall_started(ext: Any, screen: int) -> bool:
    """
    back-align lead 구간이 끝나 JSON 러너가 실제 시작된 뒤에만 True.

    sim 축 ``_json_run_start_sim`` 과 ``_json_sequence_started`` 로 판정한다.
    """
    active = _screen_active_json_job(ext, int(screen))
    if not isinstance(active, dict) or not active:
        return True
    try:
        if float(active.get("json_lead_sec") or 0.0) <= 1e-6:
            return True
    except Exception:
        pass
    return bool(active.get("_json_sequence_started"))


def _poll_playback_sim_aligned_json_starts(ext: Any) -> None:
    """프리런 재생 — ``sim_now >= t0+json_lead`` 일 때 JSON 러너를 시작한다 (wall Timer 대신)."""
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    active_by = getattr(ext, "_sim_anim_active_by_screen", None)
    if not isinstance(active_by, dict) or not active_by:
        return
    for scr_s, active in list(active_by.items()):
        if not isinstance(active, dict) or not active:
            continue
        if bool(active.get("_json_sequence_started")):
            continue
        if not bool(active.get("_json_pending_sim_start")):
            continue
        try:
            scr_i = int(scr_s)
        except Exception:
            continue
        try:
            t_start = float(active.get("_json_run_start_sim", 0.0))
        except Exception:
            t_start = 0.0
        player = get_sim_playback_player(ext, scr_i)
        if player is None:
            continue
        try:
            t_now = float(player.sim_now(scr_i))
        except Exception:
            t_now = 0.0
        if t_now + 1e-9 < t_start:
            continue
        run_fn = active.get("_json_run_fn")
        if callable(run_fn):
            try:
                run_fn()
            except Exception:
                pass


def _should_defer_port_occ_sync_for_renewal(ext: Any, screen: int) -> bool:
    """
    renewal JSON — 포트는 renewal wall(또는 재생 plan)만. 엔진 progress·DONE·occ 로 덮지 않음.

    재생(plan active): defer·guard 무시 — ``plan.lookup(sim_now)`` 가 패널 SSOT.
    """
    scr = max(1, int(screen))
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import playback_plan_active

            if playback_plan_active(ext, scr):
                return False
        except Exception:
            pass
    if _renewal_json_guard_active(ext, scr):
        return True
    return _screen_active_json_has_renewal(ext, scr)


def _screen_active_json_has_renewal(ext: Any, screen: int) -> bool:
    """화면의 활성·보관 JSON job 이 renewal 이면 True (공정 종료 occ 차단용)."""
    scr = max(1, int(screen))
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if isinstance(active_by, dict):
            active = active_by.get(str(scr))
            if isinstance(active, dict) and active.get("has_renewal"):
                return True
    except Exception:
        pass
    try:
        src_by = getattr(ext, "_sim_post_anim_src_by_screen", None)
        if isinstance(src_by, dict):
            stored = src_by.get(str(scr))
            if isinstance(stored, dict) and bool(stored.get("has_renewal")):
                return True
    except Exception:
        pass
    return False


def _active_json_timing_context(ext: Any, screen: int) -> Dict[str, Any]:
    """progress fail-safe 용 — active job 의 path·parsed·has_renewal."""
    out: Dict[str, Any] = {"json_path": None, "steps": None, "has_renewal": False}
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if not isinstance(active_by, dict):
            return out
        active = active_by.get(str(max(1, int(screen))))
        if not isinstance(active, dict):
            return out
        out["json_path"] = str(active.get("path") or "").strip() or None
        out["steps"] = active.get("parsed")
        out["has_renewal"] = bool(active.get("has_renewal"))
    except Exception:
        pass
    return out


def _canonical_sim_port_key(port: str) -> str:
    """엔진·UI 공통 포트 키(INOUT, BP1, EP1 …)로 정규화."""
    o = str(port or "").strip().upper()
    if not o:
        return ""
    if o in ("IN/OUT", "INOUT"):
        return "INOUT"
    if o.startswith("BP"):
        try:
            n = int(o.replace("BP", ""))
            if 1 <= n <= 4:
                return f"BP{n}"
        except Exception:
            pass
    if o.startswith("EP"):
        try:
            n = int(o.replace("EP", ""))
            if 1 <= n <= 3:
                return f"EP{n}"
        except Exception:
            pass
    return o


def _post_anim_port_dedupe_key(src: Dict[str, Any]) -> str:
    ev = _normalize_anim_event_seq(
        str(src.get("event") or src.get("event_seq") or src.get("seq") or "")
    )
    lot = str(src.get("lot_id") or "").strip()
    fr = _canonical_sim_port_key(str(src.get("from_port_id") or ""))
    to = _canonical_sim_port_key(str(src.get("to_port_id") or ""))
    port = _canonical_sim_port_key(str(src.get("port_id") or src.get("event_port_id") or ""))
    # sim_time 은 progress 중 계속 변하므로 dedupe 키에 쓰지 않는다.
    t = str(src.get("event_start_sim_time") or src.get("t") or "").strip()
    if not t:
        t = str(src.get("label") or "").strip()
    return f"{ev}|{fr}|{to}|{port}|{lot}|{t}"


def _predict_ports_occupancy_after_anim(occ_base: Dict[str, Any], src: Dict[str, Any]) -> Dict[str, Any]:
    """JSON(이동·안착·회수) 종료 직후 기대되는 ports_occupancy 를 예측한다."""
    occ_pred = dict(occ_base or {})
    ev = _normalize_anim_event_seq(
        str(src.get("event") or src.get("event_seq") or src.get("seq") or "")
    )
    lot_id = str(src.get("lot_id") or "").strip()
    fr = _canonical_sim_port_key(str(src.get("from_port_id") or ""))
    to = _canonical_sim_port_key(str(src.get("to_port_id") or ""))
    # progress payload에서는 event_port_id로 들어올 수 있다(진행현황 라우팅용 port_id와 구분).
    port = _canonical_sim_port_key(str(src.get("port_id") or src.get("event_port_id") or ""))
    if ev in ("MOVE_TRANSFERING", "MOVE_REQ", "MOVE"):
        if fr:
            occ_pred[fr] = ""
        if to and lot_id:
            occ_pred[to] = lot_id
    elif ev == "ARRIVED":
        dest = port or to
        if dest and lot_id:
            occ_pred[dest] = lot_id
    elif ev == "REMOVED":
        if port:
            occ_pred[port] = ""
    return occ_pred


def _clear_post_anim_port_applied(ext: Any, screen: int) -> None:
    try:
        by_ap = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
        if isinstance(by_ap, dict):
            by_ap.pop(str(int(screen)), None)
    except Exception:
        pass


def _normalize_post_anim_port_src(src: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(src or {})
    out["event"] = _normalize_anim_event_seq(
        str(out.get("event") or out.get("event_seq") or out.get("seq") or "")
    )
    if not str(out.get("event_start_sim_time") or "").strip():
        out["event_start_sim_time"] = str(out.get("t") or "").strip()
    return out


def _queue_post_anim_port_apply(ext: Any, screen: int, src: Dict[str, Any]) -> None:
    """JSON 종료 직후 포트 갱신 요청을 화면별 pending 에 적재(bg thread 안전)."""
    try:
        pending_by = getattr(ext, "_sim_pending_post_anim_port_by_screen", None)
        if not isinstance(pending_by, dict):
            pending_by = {}
            ext._sim_pending_post_anim_port_by_screen = pending_by
        pending_by[str(int(screen))] = _normalize_post_anim_port_src(src)
    except Exception:
        pass


def _flush_pending_post_anim_port_applies(ext: Any, screen: Optional[int] = None) -> None:
    """main thread 에서 pending 포트 갱신을 실제로 반영한다."""
    try:
        pending_by = getattr(ext, "_sim_pending_post_anim_port_by_screen", None)
        if not isinstance(pending_by, dict) or not pending_by:
            return
        keys = [str(int(screen))] if screen is not None else list(pending_by.keys())
        for k in keys:
            src = pending_by.get(str(k))
            if not isinstance(src, dict):
                pending_by.pop(str(k), None)
                continue
            if _try_apply_port_state_after_json_anim(ext, int(k), src, force=True):
                pending_by.pop(str(k), None)
                try:
                    src_by = getattr(ext, "_sim_post_anim_src_by_screen", None)
                    if isinstance(src_by, dict):
                        src_by.pop(str(k), None)
                except Exception:
                    pass
    except Exception:
        pass


def _maybe_queue_post_anim_port_by_wall_clock(ext: Any) -> None:
    """JSON wall-clock 종료 시점 fail-safe (라이브 전용). 재생 중 plan replay 사용."""
    if bool(getattr(ext, "_sim_playback_started", False)):
        return
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
        src_by = getattr(ext, "_sim_post_anim_src_by_screen", None)
        now = time.monotonic()
        scr_keys: Set[str] = set()
        if isinstance(active_by, dict):
            scr_keys.update(str(k) for k in active_by.keys())
        if isinstance(src_by, dict):
            scr_keys.update(str(k) for k in src_by.keys())
        for scr_s in scr_keys:
            active = active_by.get(scr_s) if isinstance(active_by, dict) else None
            stored = src_by.get(scr_s) if isinstance(src_by, dict) else None
            src: Optional[Dict[str, Any]] = None
            end_wall: Optional[float] = None
            if isinstance(active, dict) and active:
                src = _normalize_post_anim_port_src(active)
                if not _is_anim_port_event(str(src.get("event") or "")):
                    continue
                if isinstance(until_by, dict) and until_by.get(str(scr_s)) is not None:
                    end_wall = float(until_by.get(str(scr_s)))
                elif isinstance(stored, dict):
                    try:
                        end_wall = float(stored.get("_json_end_wall") or 0.0)
                    except Exception:
                        end_wall = None
            elif isinstance(stored, dict) and stored:
                src = _normalize_post_anim_port_src(dict(stored))
                try:
                    end_wall = float(stored.get("_json_end_wall") or 0.0)
                except Exception:
                    end_wall = None
            if not isinstance(src, dict) or not src:
                continue
            if not _is_anim_port_event(str(src.get("event") or "")):
                continue
            has_renewal = bool((active or stored or {}).get("has_renewal"))
            port_sync_wall: Optional[float] = None
            try:
                pw = (stored or active or {}).get("_port_sync_wall")
                if pw is not None:
                    port_sync_wall = float(pw)
            except Exception:
                port_sync_wall = None
            if has_renewal:
                if port_sync_wall is None or now + 0.02 < float(port_sync_wall):
                    continue
                if not _json_playback_wall_started(ext, int(scr_s)):
                    continue
                dedupe = _post_anim_port_dedupe_key(src)
                applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
                if isinstance(applied_by, dict) and applied_by.get(str(scr_s)) == dedupe:
                    continue
                src_apply = dict(src)
                src_apply["_from_renewal_step"] = True
                _clear_renewal_port_defer(ext, int(scr_s))
                _queue_post_anim_port_apply(ext, int(scr_s), src_apply)
                continue
            if end_wall is None or end_wall <= 0.0:
                started = float((active or stored or {}).get("_started_wall") or 0.0)
                est = (active or stored or {}).get("est_total")
                try:
                    est_f = float(est) if est is not None else 0.0
                except Exception:
                    est_f = 0.0
                if started > 0.0 and est_f > 0.0:
                    end_wall = float(started) + float(est_f)
            if end_wall is None or now + 0.02 < float(end_wall):
                continue
            dedupe = _post_anim_port_dedupe_key(src)
            applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
            if isinstance(applied_by, dict) and applied_by.get(str(scr_s)) == dedupe:
                continue
            _queue_post_anim_port_apply(ext, int(scr_s), src)
    except Exception:
        pass


def _try_apply_port_state_after_json_anim(
    ext: Any,
    screen: int,
    src: Dict[str, Any],
    *,
    force: bool = False,
) -> bool:
    """
    JSON 시퀀스 종료 직후 포트상태/visibility/위치를 1회 반영한다.
    renewal JSON 은 renewal 스텝(``_from_renewal_step``) 또는 fail-safe 에서만 허용.
    """
    if not isinstance(src, dict):
        return False
    scr = max(1, int(screen))
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import playback_plan_active

            if playback_plan_active(ext, scr):
                return False
        except Exception:
            pass
    if _should_defer_port_occ_sync_for_renewal(ext, scr) and not bool(
        src.get("_from_renewal_step")
    ):
        return False
    ev = _normalize_anim_event_seq(
        str(src.get("event") or src.get("event_seq") or src.get("seq") or "")
    )
    if ev not in _ANIM_PORT_UPDATE_SEQS:
        return False
    dedupe = _post_anim_port_dedupe_key(src)
    applied_by = getattr(ext, "_sim_post_anim_port_applied_by_screen", None)
    if not isinstance(applied_by, dict):
        applied_by = {}
        ext._sim_post_anim_port_applied_by_screen = applied_by
    if (not force) and applied_by.get(str(scr)) == dedupe:
        return False

    last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
    occ_now: Dict[str, Any] = {}
    if isinstance(last_by, dict) and isinstance(last_by.get(str(scr)), dict):
        occ_now = dict(last_by.get(str(scr)) or {})

    by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
    lp = by_lp.get(str(scr)) if isinstance(by_lp, dict) else None
    sim_t = str((lp or {}).get("sim_time", "") or "") if isinstance(lp, dict) else str(src.get("sim_time", "") or "")
    active_ep = _remember_foup_active_ep(ext, scr, lp if isinstance(lp, dict) else src)
    occ_pred = _predict_ports_occupancy_after_anim(occ_now, src)

    try:
        if isinstance(last_by, dict):
            last_by[str(scr)] = dict(occ_pred)
    except Exception:
        pass

    try:
        _apply_sim_event_state_only(
            ext,
            {
                "ports_occupancy": dict(occ_pred),
                "sim_time": sim_t,
                "foup_proc_active_ep": active_ep,
            },
            screen=scr,
        )
    except Exception:
        return False

    applied_by[str(scr)] = dedupe
    return True


def _remember_foup_active_ep(ext: Any, screen: int, payload: Dict[str, Any]) -> str:
    """엔진 payload 의 ``foup_proc_active_ep`` 를 화면별로 기억·조회."""
    scr = int(screen)
    ep = str(payload.get("foup_proc_active_ep", "") or "").strip().upper()
    try:
        by = getattr(ext, "_sim_foup_proc_active_ep_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_foup_proc_active_ep_by_screen = by
        if ep:
            by[str(scr)] = ep
        elif "foup_proc_active_ep" in payload:
            by[str(scr)] = ""
    except Exception:
        pass
    if ep:
        return ep
    try:
        by = getattr(ext, "_sim_foup_proc_active_ep_by_screen", None)
        if isinstance(by, dict):
            return str(by.get(str(scr), "") or "").strip().upper()
    except Exception:
        pass
    return ""


def _resolve_foup_proc_active_ep(
    ext: Any,
    screen: int,
    payload: Optional[Dict[str, Any]] = None,
) -> str:
    """화면 캐시 → payload → 시뮬 엔진 순으로 FOUP active EP 를 조회."""
    ep = _remember_foup_active_ep(ext, int(screen), payload or {})
    if ep:
        return ep
    scr = int(screen)
    try:
        engines = getattr(ext, "_sim_engines", None)
        eng = None
        if isinstance(engines, list) and 0 <= (scr - 1) < len(engines):
            eng = engines[scr - 1]
        elif isinstance(engines, dict):
            eng = engines.get(str(scr)) or engines.get(scr)
        if eng is not None:
            ep = str(getattr(eng, "_foup_proc_active_ep", "") or "").strip().upper()
            if ep:
                return ep
    except Exception:
        pass
    try:
        if scr == 1:
            eng = getattr(ext, "_sim_engine", None)
            if eng is not None:
                ep = str(getattr(eng, "_foup_proc_active_ep", "") or "").strip().upper()
                if ep:
                    return ep
    except Exception:
        pass
    return ""


def _apply_sim_event_state_only(ext: Any, payload: Dict[str, Any], *, screen: int) -> None:
    """Seek Fast-apply: 포트·FOUP 가시성만 반영(JSON 애니 생략)."""
    if not isinstance(payload, dict):
        return
    scr = int(screen)
    if _should_defer_port_occ_sync_for_renewal(ext, scr) and not bool(
        payload.get("_from_renewal_step")
    ) and not bool(payload.get("_from_playback_plan")):
        return
    occ = payload.get("ports_occupancy", {})
    if not isinstance(occ, dict):
        occ = {}
    occ_panel = dict(occ)
    occ_prims = dict(occ)
    # REMOVED renewal / hide-hold: 패널 EMPTY 여도 prim 은 hold 구간 유지
    try:
        from .control_sim_playback_plan import prim_occ_for_playback_visibility

        occ_prims = prim_occ_for_playback_visibility(ext, scr, dict(occ_panel))
    except Exception:
        if bool(payload.get("_from_renewal_step")):
            try:
                by_hold = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                prev = by_hold.get(str(scr)) if isinstance(by_hold, dict) else None
                if isinstance(prev, dict):
                    for pk, pv in prev.items():
                        pu = str(pk or "").strip().upper()
                        if not pu:
                            continue
                        if str(pv or "").strip() and not str(
                            occ_panel.get(pu) or occ_panel.get(pk) or ""
                        ).strip():
                            occ_prims[pu] = str(pv)
            except Exception:
                pass
    ctx_nm = _usd_context_name_for_sim_screen(ext, scr)
    active_ep = _remember_foup_active_ep(ext, scr, payload)
    try:
        apply_port_lot_prim_visibility_for_context(ctx_nm, occ_prims)
    except Exception as exc:
        try:
            print(
                f"[TBS/port-screen] state_only visibility skip scr={scr} ctx={ctx_nm!r}: {exc}",
                flush=True,
            )
        except Exception:
            pass
    try:
        sync_port_lot_positions_after_visibility(ctx_nm, foup_proc_active_ep=active_ep)
    except Exception:
        pass
    try:
        by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if not isinstance(by_prev, dict):
            by_prev = {}
            ext._sim_last_ports_occupancy_by_screen = by_prev
        if occ_panel:
            by_prev[str(scr)] = dict(occ_panel)
    except Exception:
        pass
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            by_prim = getattr(ext, "_sim_last_prim_ports_occupancy_by_screen", None)
            if not isinstance(by_prim, dict):
                by_prim = {}
                ext._sim_last_prim_ports_occupancy_by_screen = by_prim
            by_prim[str(scr)] = dict(occ_prims)
        except Exception:
            pass
    try:
        _update_port_occupancy_panel(ext, occ_panel, str(payload.get("sim_time", "")), screen=scr)
    except Exception:
        pass


def _extract_ep_id_from_foup_payload(payload: Dict[str, Any]) -> str:
    ep_id = str(payload.get("port_id", "") or "").strip().upper()
    if ep_id:
        return ep_id
    try:
        import re as _re

        src_txt = (str(payload.get("label", "") or "") + " " + str(payload.get("detail", "") or "")).upper()
        m = _re.search(r"\bEP(\d+)\b", src_txt)
        if m:
            n = int(m.group(1))
            if 1 <= n <= 3:
                return f"EP{n}"
    except Exception:
        pass
    return ""


def _remember_foup_playback_progress(ext: Any, screen: int, ep_id: str, payload: Dict[str, Any]) -> None:
    """프리런 재생 — EP별 마지막 FOUP_PROCESS payload (heartbeat 보간용)."""
    ep = str(ep_id or "").strip().upper()
    if not ep:
        return
    sk = str(max(1, int(screen)))
    by = getattr(ext, "_sim_foup_playback_last_by_screen", None)
    if not isinstance(by, dict):
        by = {}
        ext._sim_foup_playback_last_by_screen = by
    slot = by.get(sk)
    if not isinstance(slot, dict):
        slot = {}
        by[sk] = slot
    slot[ep] = dict(payload or {})


def _forget_foup_playback_progress(ext: Any, screen: int, ep_id: str) -> None:
    ep = str(ep_id or "").strip().upper()
    if not ep:
        return
    by = getattr(ext, "_sim_foup_playback_last_by_screen", None)
    if not isinstance(by, dict):
        return
    slot = by.get(str(max(1, int(screen))))
    if isinstance(slot, dict):
        slot.pop(ep, None)


def _refresh_foup_playback_heartbeat(ext: Any, screen: int, tnow: float) -> None:
    """
    프리런 재생 heartbeat — 타임라인 FOUP progress 항목 사이에서 EP별 FOUP 라벨을 sim_now 로 보간.

    FOUP 공정 시간(설정 proc_sec·단계별 +Y/-Y)은 **현재 MOVE/ARRIVED 이벤트와 별개**이다.
    """
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    by = getattr(ext, "_sim_foup_playback_last_by_screen", None)
    if not isinstance(by, dict):
        return
    slot = by.get(str(max(1, int(screen))))
    if not isinstance(slot, dict) or not slot:
        return
    scr = max(1, int(screen))
    for _ep_id, lp in list(slot.items()):
        if not isinstance(lp, dict):
            continue
        st = str(lp.get("status", "")).strip().upper()
        lab_u = str(lp.get("label", "") or "").upper()
        if st == "DONE" and "-Y" in lab_u:
            continue
        p = dict(lp)
        p["tbs_sim_screen"] = str(scr)
        p["sim_time"] = f"{float(tnow):.2f}"
        p["event_seq"] = "FOUP_PROCESS"
        p["_foup_heartbeat_tick"] = "1"
        if st == "DONE":
            _update_sim_progress(ext, p)
            continue
        p["status"] = "RUNNING"
        try:
            _apply_foup_playback_progress_from_sim(p, float(tnow))
        except Exception:
            pass
        _update_sim_progress(ext, p)


def _cancel_foup_label_reset_subs(ext: Any) -> None:
    """Seek/리셋 시 보류 중인 FOUP 라벨 자동 복귀 타이머를 취소."""
    try:
        holders = getattr(ext, "_foup_label_reset_subs", None)
        if not isinstance(holders, list):
            return
        for h in holders:
            try:
                s = h.get("sub")
                if s is not None:
                    s.unsubscribe()
            except Exception:
                pass
        ext._foup_label_reset_subs = []
    except Exception:
        pass


def _sync_foup_labels_at_seek(
    ext: Any,
    *,
    screen: int,
    items: List[Any],
    play_cursor: int,
    foup_by_ep: Optional[Dict[str, Dict[str, Any]]] = None,
) -> None:
    """Seek 시점까지 누적된 FOUP_PROCESS progress 로 FOUP 공정 라벨을 동기화."""
    _cancel_foup_label_reset_subs(ext)
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if not isinstance(chans, list) or not (0 < int(screen) <= len(chans)):
            return
        ch = chans[int(screen) - 1]
        if not isinstance(ch, dict):
            return
        labels = ch.get("foup_progress_labels") or {}
        if not _foup_labels_mounted(labels if isinstance(labels, dict) else {}):
            labels = _foup_progress_labels_for_screen(ext, ch, int(screen))
        if not isinstance(labels, dict):
            labels = {}
        screen_num = int(ch.get("screen", screen) or screen)
        try:
            ch["_foup_label_cache"] = {}
        except Exception:
            pass
        for ep_id, lbl in labels.items():
            if lbl is None:
                continue
            _set_foup_progress_label(
                ch,
                str(ep_id),
                lbl,
                _foup_label_idle_text(screen_num, str(ep_id)),
                {"color": 0xFF888888},
            )
        last_foup_by_ep: Dict[str, Dict[str, Any]] = {}
        if isinstance(foup_by_ep, dict) and foup_by_ep:
            last_foup_by_ep = {str(k): dict(v) for k, v in foup_by_ep.items() if isinstance(v, dict)}
        else:
            for i in range(max(0, int(play_cursor))):
                if i >= len(items):
                    break
                it = items[i]
                kind = str(getattr(it, "kind", "") or "").strip().lower()
                p = getattr(it, "payload", None)
                if kind != "progress" or not isinstance(p, dict):
                    continue
                ev_seq = str(p.get("event_seq") or p.get("sequence_name") or "").strip().upper()
                if ev_seq != "FOUP_PROCESS":
                    continue
                ep_id = _extract_ep_id_from_foup_payload(p)
                if ep_id:
                    last_foup_by_ep[ep_id] = dict(p)
        for _ep_id, p in last_foup_by_ep.items():
            st = str(p.get("status", "")).strip().upper()
            lab_u = str(p.get("label", "") or "").upper()
            if st == "DONE" and "-Y" in lab_u:
                continue
            if not str(p.get("label", "") or "").strip():
                continue
            try:
                last_key = getattr(ext, "_sim_progress_last_key", None)
                if isinstance(last_key, dict):
                    slot = str(p.get("tbs_sim_screen", screen) or screen).strip() or "1"
                    last_key.pop(f"_panel_{slot}", None)
            except Exception:
                pass
            _update_sim_progress(ext, p)
    except Exception:
        pass


def _apply_seek_progress_panel(ext: Any, *, screen: int, t_target: float) -> None:
    """Seek 후 진행현황 텍스트를 마지막 progress payload 기준으로 갱신."""
    sk = str(int(screen))
    try:
        by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
        lp = by_lp.get(sk) if isinstance(by_lp, dict) else None
        if not isinstance(lp, dict) or not str(lp.get("label", "") or "").strip():
            return
        last_key = getattr(ext, "_sim_progress_last_key", None)
        if isinstance(last_key, dict):
            last_key.pop(f"_panel_{sk}", None)
        p_prog = dict(lp)
        p_prog["sim_time"] = f"{float(t_target):.2f}"
        _update_sim_progress(ext, p_prog)
    except Exception:
        pass


def _seek_extra_steps_for_restore(ext: Any, *, screen: int) -> List[Dict[str, Any]]:
    """Seek 직전 재생 중이던 JSON step — prim 경로 수집용."""
    out: List[Dict[str, Any]] = []
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if isinstance(active_by, dict):
            job = active_by.get(str(int(screen)))
            if isinstance(job, dict):
                parsed = job.get("parsed")
                if isinstance(parsed, list):
                    out.extend(s for s in parsed if isinstance(s, dict))
    except Exception:
        pass
    return out


def _purge_anim_events_from_sim_queue(ext: Any, *, screen: int) -> None:
    """Seek 직전 해당 화면 ANIM_EVENT 큐 항목을 폐기(잔여 이벤트로 애니 재시작 방지)."""
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    scr_s = str(int(screen))
    pending_items: List[Any] = []
    while True:
        try:
            pending_items.append(q.get_nowait())
        except Exception:
            break
    for item in pending_items:
        try:
            if isinstance(item, tuple) and len(item) == 2:
                kind, payload = item
            else:
                kind, payload = SimUiQueueKind.HISTORY_LINE.value, item
            if _coerce_sim_ui_queue_kind(kind) == SimUiQueueKind.ANIM_EVENT and isinstance(payload, dict):
                try:
                    ps = str(payload.get("tbs_sim_screen", "1") or "1").strip() or "1"
                except Exception:
                    ps = "1"
                if ps == scr_s:
                    continue
            q.put_nowait(item)
        except Exception:
            pass


def _halt_anim_for_prerun_seek(ext: Any, *, screen: int) -> None:
    """타임테이블 Seek — 재생 중 JSON·pending·러너를 화면별로 정리."""
    scr_s = str(int(screen))
    try:
        pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
        if isinstance(pending_by, dict):
            pending_by[scr_s] = []
    except Exception:
        pass
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if isinstance(active_by, dict):
            active_by[scr_s] = {}
    except Exception:
        pass
    if int(screen) == 1 and not _is_multi_viewport_sim(ext):
        try:
            ext._sim_anim_pending = []
            ext._sim_anim_active = {}
        except Exception:
            pass
    try:
        pause_map = getattr(ext, "_sim_tick_pause_events_by_screen", None)
        if isinstance(pause_map, dict):
            pe = pause_map.get(scr_s)
            if pe is not None:
                pe.clear()
    except Exception:
        pass
    try:
        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
        if isinstance(until_by, dict):
            until_by[scr_s] = None
    except Exception:
        pass
    try:
        runners = getattr(ext, "_sim_runners_by_screen", None)
        runner_paused = False
        if isinstance(runners, dict):
            r = runners.get(scr_s)
            if r is not None:
                try:
                    if getattr(r, "is_running", lambda: False)():
                        r.pause()
                    runner_paused = True
                except Exception:
                    pass
        if not runner_paused and int(screen) == 1:
            r0 = getattr(ext, "_sim_runner", None)
            if r0 is not None:
                try:
                    if getattr(r0, "is_running", lambda: False)():
                        r0.pause()
                except Exception:
                    pass
    except Exception:
        pass
    try:
        from .sim_channel_scope import stop_channel_animations

        ctx = _usd_context_name_for_sim_screen(ext, int(screen))
        stop_channel_animations(ctx)
    except Exception:
        pass


def _fast_apply_prerun_seek(ext: Any, *, screen: int, row_index: int) -> Tuple[float, int]:
    """
    클릭한 타임테이블 행까지 items 를 state-only 로 적용.
    반환: (target_t, play_cursor) — Visible 재생은 ``play_cursor`` 부터.
    """
    results = getattr(ext, "_sim_prerun_results_by_screen", None)
    metas_by = getattr(ext, "_sim_timetable_row_metas_by_screen", None)
    if not isinstance(results, dict):
        return 0.0, 0
    res = results.get(int(screen))
    if res is None:
        return 0.0, 0
    metas = []
    if isinstance(metas_by, dict):
        metas = list(metas_by.get(str(int(screen)), []) or [])
    if not metas:
        metas = build_timetable_row_metas(res)
    t_target, through = resolve_seek_through_index(metas, int(row_index))
    play_cursor = max(0, int(through))

    scr_i = int(screen)
    ctx = _usd_context_name_for_sim_screen(ext, scr_i)

    player = get_sim_playback_player(ext, scr_i)
    if player is not None:
        try:
            if getattr(player, "is_playing", lambda: False)():
                player.stop()
        except Exception:
            pass

    try:
        _purge_anim_events_from_sim_queue(ext, screen=scr_i)
    except Exception:
        pass
    try:
        _halt_anim_for_prerun_seek(ext, screen=scr_i)
    except Exception:
        pass
    try:
        extra = _seek_extra_steps_for_restore(ext, screen=scr_i)
        _restore_sim_prim_motion_to_initial(
            ext,
            extra_steps=extra if extra else None,
            usd_context_name=ctx,
            preserve_foup_offsets=True,
            foup_proc_active_ep=_remember_foup_active_ep(ext, scr_i, {}),
        )
    except Exception:
        pass

    items = res.items
    sk = str(int(screen))
    snap_list = None
    snap_by = getattr(ext, "_sim_seek_snapshots_by_screen", None)
    if isinstance(snap_by, dict):
        snap_list = snap_by.get(sk)
    used_snapshot = (
        isinstance(snap_list, list) and len(snap_list) > int(play_cursor) and snap_list[int(play_cursor)] is not None
    )
    used_plan_seek = False
    if bool(getattr(ext, "_sim_playback_started", False)):
        try:
            from .control_sim_playback_plan import seek_playback_ui_at_sim

            if seek_playback_ui_at_sim(ext, scr_i, float(t_target)):
                used_plan_seek = True
        except Exception:
            used_plan_seek = False
    if used_snapshot and not used_plan_seek:
        snap = snap_list[int(play_cursor)]
        if bool(getattr(snap, "needs_state_apply", False)):
            try:
                apply_payload = dict(snap.apply_payload)
                lp_snap = getattr(snap, "progress_last_payload", None)
                if isinstance(lp_snap, dict) and float(t_target) > 1e-9:
                    try:
                        from .control_sim_prerun_playback import effective_ports_occupancy_at_t

                        occ0 = dict(apply_payload.get("ports_occupancy") or {})
                        occ_eff = effective_ports_occupancy_at_t(
                            occ0, dict(lp_snap), float(t_target)
                        )
                        apply_payload["ports_occupancy"] = dict(occ_eff)
                        apply_payload["sim_time"] = f"{float(t_target):.2f}"
                    except Exception:
                        pass
                _apply_sim_event_state_only(ext, apply_payload, screen=int(screen))
            except Exception:
                pass
        if isinstance(getattr(snap, "progress_last_payload", None), dict):
            try:
                by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
                if not isinstance(by_lp, dict):
                    by_lp = {}
                    ext._sim_progress_last_payload_by_screen = by_lp
                by_lp[sk] = dict(snap.progress_last_payload)
            except Exception:
                pass
    elif not used_plan_seek:
        for i in range(play_cursor):
            if i >= len(items):
                break
            it = items[i]
            kind = str(it.kind or "").strip().lower()
            if kind == "event" and isinstance(it.payload, dict):
                _apply_sim_event_state_only(ext, dict(it.payload), screen=int(screen))
            elif kind == "progress" and isinstance(it.payload, dict):
                p = dict(it.payload)
                occ = p.get("ports_occupancy", {})
                if isinstance(occ, dict) and occ:
                    _apply_sim_event_state_only(ext, p, screen=int(screen))
                try:
                    by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
                    if not isinstance(by_lp, dict):
                        by_lp = {}
                        ext._sim_progress_last_payload_by_screen = by_lp
                    by_lp[sk] = dict(p)
                except Exception:
                    pass

    # 막대 virtual time / 슬라이스 기준 시각
    _sync_ep_bar_virtual_time_to_sim(ext, int(screen), float(t_target))
    try:
        st_by = getattr(ext, "_sim_ep_occ_timeline_state_by_screen", None)
        if not isinstance(st_by, dict):
            st_by = {}
            ext._sim_ep_occ_timeline_state_by_screen = st_by
        st_by[sk] = {"t_last": float(t_target), "rows": {}, "total_est_fixed": None}
    except Exception:
        pass

    try:
        ext._sim_playback_done = False
    except Exception:
        pass

    player = get_sim_playback_player(ext, int(screen))
    if player is not None and hasattr(player, "seek"):
        try:
            player.seek(int(screen), target_t=float(t_target), item_cursor=int(play_cursor))
        except Exception:
            pass

    engs = getattr(ext, "_sim_engines", None)
    if isinstance(engs, list) and int(screen) - 1 < len(engs):
        eng = engs[int(screen) - 1]
        if eng is not None:
            try:
                if hasattr(eng, "_set_now"):
                    eng._set_now(float(t_target))  # type: ignore[attr-defined]
                elif hasattr(eng, "env") and eng.env is not None:
                    eng.env.now = float(t_target)  # type: ignore[attr-defined]
            except Exception:
                pass

    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans, list) and 0 < int(screen) <= len(chans):
            ch = chans[int(screen) - 1]
            occ = {}
            by_prev = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
            if isinstance(by_prev, dict) and isinstance(by_prev.get(sk), dict):
                occ = dict(by_prev.get(sk) or {})
            _update_ep_timeline_under_port_state(ext, ch, occ, f"{float(t_target):.2f}")
    except Exception:
        pass

    try:
        foup_prefill = None
        if used_snapshot and isinstance(snap_list, list) and len(snap_list) > int(play_cursor):
            cand = snap_list[int(play_cursor)]
            if isinstance(getattr(cand, "foup_by_ep", None), dict):
                foup_prefill = dict(cand.foup_by_ep)
        _sync_foup_labels_at_seek(
            ext,
            screen=int(screen),
            items=list(items),
            play_cursor=int(play_cursor),
            foup_by_ep=foup_prefill,
        )
    except Exception:
        pass
    try:
        _apply_seek_progress_panel(ext, screen=int(screen), t_target=float(t_target))
    except Exception:
        pass
    return float(t_target), int(play_cursor)


def _on_timetable_row_seek(ext: Any, screen: int, row_index: int) -> None:
    try:
        t_target, _pc = _fast_apply_prerun_seek(ext, screen=int(screen), row_index=int(row_index))
        print(f"[SIM] 타임테이블 Seek 화면{int(screen)} 행{int(row_index)} → t={float(t_target):.2f}", flush=True)
        try:
            refresh_timetable_row_highlight(ext, screen=int(screen), sim_now=float(t_target))
        except Exception:
            pass
    except Exception as e:
        print(f"[SIM] timetable seek 실패: {e}", flush=True)


def _get_channel_history_text(ch: Dict[str, Any], ext: Any = None) -> str:
    try:
        si = int(ch.get("screen", 0))
    except Exception:
        si = 0
    if ext is not None and si > 0:
        by = getattr(ext, "_sim_timetable_display_by_screen", None)
        if isinstance(by, dict):
            txt = str(by.get(str(si), "") or "").strip()
            if txt:
                return txt
    lbl = ch.get("history_label")
    if lbl is not None:
        try:
            return str(getattr(lbl, "text", "") or "")
        except Exception:
            pass
    model = ch.get("history_model")
    if model is not None:
        try:
            return str(model.as_string or "")
        except Exception:
            pass
    return ""


def _set_channel_history_text(ch: Dict[str, Any], text: str, *, ext: Any = None) -> None:
    txt = str(text or "")
    if ch.get("timetable_interactive"):
        return
    try:
        si = int(ch.get("screen", 0))
    except Exception:
        si = 0
    if ext is not None and si > 0:
        try:
            by = getattr(ext, "_sim_timetable_display_by_screen", None)
            if not isinstance(by, dict):
                by = {}
                ext._sim_timetable_display_by_screen = by
            by[str(si)] = txt
        except Exception:
            pass
    lbl = ch.get("history_label")
    if lbl is not None:
        try:
            lbl.text = txt
            try:
                lbl.visible = True
            except Exception:
                pass
        except Exception:
            pass
    model = ch.get("history_model")
    if model is not None:
        try:
            model.set_value(txt)
        except Exception:
            pass


def _scroll_sim_monitor_to_timetable(ext: Any) -> None:
    """프리런 직후 타임테이블 전용 창을 앞으로 가져온다(사용자가 닫아 둔 경우 생략)."""
    if _sim_timetable_user_dismissed(ext):
        return
    for tw in _iter_sim_timetable_windows(ext):
        if tw is None:
            continue
        try:
            tw.visible = True
            if hasattr(tw, "focus"):
                tw.focus()
        except Exception:
            pass


def _build_timetable_label_text(header: str, metas: List[Any]) -> str:
    lines = [str(header or "").strip()]
    for m in metas or []:
        lines.append(str(getattr(m, "display_line", m)))
    txt = "\n".join([ln for ln in lines if ln]).strip()
    return txt if txt else "[SIM] 타임테이블(프리런) — 항목 없음"


def _schedule_deferred_timetable_refresh(ext: Any, *, screen: int, text: str) -> None:
    """Kit 레이아웃 1프레임 뒤 타임테이블 Label 을 다시 채운다(StringField 잔여·클리핑 대비)."""
    pending = {"screen": int(screen), "text": str(text or "")}

    def _once(_e: Any) -> None:
        try:
            sub = getattr(ext, "_sim_timetable_deferred_sub", None)
            if sub is not None:
                try:
                    sub.unsubscribe()
                except Exception:
                    pass
                ext._sim_timetable_deferred_sub = None
        except Exception:
            pass
        try:
            si = int(pending.get("screen", 1))
            txt = str(pending.get("text", "") or "").strip()
            if not txt:
                return
            chans = getattr(ext, "_sim_monitor_channels", None)
            ch = None
            if isinstance(chans, list) and 0 < si <= len(chans):
                ch = chans[si - 1]
            if isinstance(ch, dict):
                _apply_timetable_text_to_channel(ext, ch, screen=si, text=txt, _deferred=True)
        except Exception:
            pass

    try:
        sub_old = getattr(ext, "_sim_timetable_deferred_sub", None)
        if sub_old is not None:
            try:
                sub_old.unsubscribe()
            except Exception:
                pass
        ext._sim_timetable_deferred_sub = app.get_app().get_update_event_stream().create_subscription_to_pop(
            _once,
            name="morph.tbs_control_2:timetable_deferred_refresh",
        )
    except Exception:
        pass


def _apply_timetable_text_to_channel(
    ext: Any,
    ch: Dict[str, Any],
    *,
    screen: int,
    text: str,
    _deferred: bool = False,
    show_label: bool = True,
) -> None:
    """타임테이블 본문을 ext 에 백업하고, 필요 시 Label 폴백으로 표시한다."""
    try:
        txt = str(text or "").strip()
        if not txt:
            return
        try:
            by = getattr(ext, "_sim_timetable_display_by_screen", None)
            if not isinstance(by, dict):
                by = {}
                ext._sim_timetable_display_by_screen = by
            by[str(int(screen))] = txt
        except Exception:
            pass
        if int(screen) == 1 and getattr(ext, "_sim_history_text", None) is not None:
            try:
                ext._sim_history_text.set_value(txt)
            except Exception:
                pass
        if not show_label or ch.get("timetable_interactive"):
            return
        _set_channel_history_text(ch, txt, ext=ext)
        try:
            hf = ch.get("history_frame")
            if hf is not None:
                hf.visible = True
        except Exception:
            pass
        try:
            lbl = ch.get("history_label")
            if lbl is not None and hasattr(lbl, "height"):
                n_lines = max(1, len(txt.splitlines()))
                lbl.height = max(230, min(8000, 14 * n_lines))
        except Exception:
            pass
        shown = _get_channel_history_text(ch, ext)
        print(
            f"[SIM] 타임테이블 UI 적용(화면{int(screen)}): {len(txt)}자, {len(txt.splitlines())}줄"
            f", label표시={len(shown)}자"
            f"{', deferred' if _deferred else ''}",
            flush=True,
        )
        if not _deferred and len(shown) < max(32, len(txt) // 4):
            _schedule_deferred_timetable_refresh(ext, screen=int(screen), text=txt)
    except Exception:
        pass


def _restore_timetable_display(ext: Any, *, screen: Optional[int] = None) -> None:
    """다른 UI 갱신으로 타임테이블 Label 이 덮였을 때 백업본을 복원."""
    by = getattr(ext, "_sim_timetable_display_by_screen", None)
    if not isinstance(by, dict) or not by:
        return
    chans = getattr(ext, "_sim_monitor_channels", None)
    if not isinstance(chans, list):
        return
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        try:
            si = int(ch.get("screen", 0))
        except Exception:
            continue
        if si <= 0:
            continue
        if screen is not None and int(screen) != si:
            continue
        if ch.get("timetable_interactive"):
            continue
        txt = str(by.get(str(si), "") or "").strip()
        if not txt:
            continue
        cur = _get_channel_history_text(ch, ext).strip()
        if cur == txt:
            continue
        if (not cur.startswith("[SIM] 타임테이블(프리런)")) or (len(cur) < max(32, len(txt) // 3)):
            try:
                _apply_timetable_text_to_channel(ext, ch, screen=si, text=txt, show_label=True)
            except Exception:
                pass


def _merge_prerun_ui_screen_dict(
    ext: Any,
    attr: str,
    new_by: Dict[Any, Any],
    *,
    merge: bool,
) -> None:
    """화면별 프리런 UI 자산 dict — ``merge=True`` 이면 기존 화면 키는 유지한다."""
    if not isinstance(new_by, dict) or not new_by:
        return
    if merge:
        prev = getattr(ext, attr, None)
        merged = dict(prev) if isinstance(prev, dict) else {}
        merged.update(new_by)
        setattr(ext, attr, merged)
    else:
        setattr(ext, attr, new_by)


def _finalize_prerun_ui_assets(
    ext: Any,
    results: Dict[int, SimPreRunResult],
    *,
    merge: bool = False,
) -> None:
    """프리런 완료 후 막대 사전 계산·인터랙티브 타임테이블 장착."""
    if not isinstance(results, dict) or not results:
        return
    try:
        ext._sim_prerun_ui_busy = False
    except Exception:
        pass
    _set_sim_start_enabled(ext, True)
    _clear_ep_prerun_busy_labels(ext)
    sp = 1.0
    try:
        m = getattr(ext, "_sim_speed_model", None)
        if m is not None:
            sp = max(0.1, float(m.get_value_as_float()))
    except Exception:
        sp = 1.0
    sched_by: Dict[int, Any] = {}
    try:
        from .playback_schedule import build_schedules_by_screen

        sched_by = build_schedules_by_screen(results, user_sp=float(sp))
    except Exception as e:
        print(f"[SIM] 재생 스케줄 사전계산 실패: {e}", flush=True)
        sched_by = {}

    bar_by: Dict[str, EpBarPrecomputed] = {}
    meta_by: Dict[str, List[Any]] = {}
    seek_by: Dict[str, List[Any]] = {}
    export_by: Dict[str, Dict[str, Any]] = {}
    init_occ_by: Dict[str, Dict[str, str]] = {}
    plan_by: Dict[str, Any] = {}
    for scr, res in results.items():
        try:
            si = int(scr)
        except Exception:
            continue
        snap = _effective_sim_settings_snapshot_for_screen(ext, si)
        try:
            ep_idx = int(snap.get("ep_count_idx", _ep_count_idx_for_port_panel(ext, si)) or 0)
        except Exception:
            ep_idx = int(_ep_count_idx_for_port_panel(ext, si))
        ep_count = 3 if ep_idx else 2
        ebs_on = bool(snap.get("ebs_enabled", True))
        faults = _fault_ports_from_snapshot(snap, ep_count) if snap else set()
        sched = sched_by.get(int(si)) if isinstance(sched_by, dict) else None
        bar = None
        ports: List[str] = []
        try:
            from .control_sim_bar_graph import _initial_bar_occ_at_t0
            from .control_sim_bar_graph import bar_graph_row_order
            from .playback_plan import build_playback_plan_snapshot, replay_ports_occ_at_t

            row_o = bar_graph_row_order(int(ep_idx), ebs_enabled=bool(ebs_on))
            ports = [r for r in row_o if r != "ALL_EP"]
            init0 = dict(_initial_bar_occ_at_t0(res.items, ports))
            if sched is not None:
                plan_by[str(si)] = build_playback_plan_snapshot(
                    int(si),
                    sched,
                    res.items,
                    initial_occ=dict(init0),
                    port_keys=ports,
                    final_sim_time=float(res.final_sim_time),
                )
                init_occ_by[str(si)] = dict(
                    replay_ports_occ_at_t(
                        sched.ui_milestones,
                        t_sim=0.0,
                        all_ports=ports,
                        initial_occ=init0,
                    )
                )
            else:
                init_occ_by[str(si)] = dict(init0)
            if sched is not None:
                bar = build_ep_bar_from_playback_schedule(
                    sched,
                    res.items,
                    final_sim_time=float(res.final_sim_time),
                    ep_count_idx=int(ep_idx),
                    ebs_enabled=bool(ebs_on),
                    fault_ports=faults,
                )
            else:
                bar = build_ep_bar_from_progress_items(
                    res.items,
                    final_sim_time=float(res.final_sim_time),
                    ep_count_idx=int(ep_idx),
                    ebs_enabled=bool(ebs_on),
                    fault_ports=faults,
                )
        except Exception:
            init_occ_by[str(si)] = {}
        if bar is None:
            try:
                bar = build_ep_bar_from_progress_items(
                    res.items,
                    final_sim_time=float(res.final_sim_time),
                    ep_count_idx=int(ep_idx),
                    ebs_enabled=bool(ebs_on),
                    fault_ports=faults,
                )
            except Exception:
                continue
        bar_by[str(si)] = bar
        metas = build_timetable_row_metas(res)
        meta_by[str(si)] = metas
        try:
            seek_by[str(si)] = build_seek_snapshots_by_item_index(res.items)
        except Exception:
            seek_by[str(si)] = []
        try:
            try:
                _sp_m = getattr(ext, "_sim_speed_model", None)
                _sim_speed = max(0.1, float(_sp_m.get_value_as_float())) if _sp_m is not None else 1.0
            except Exception:
                _sim_speed = 1.0
            export_doc = build_prerun_export_document(
                screen=si,
                result=res,
                bar=bar,
                timetable_metas=metas,
                seek_snapshots_count=len(seek_by.get(str(si)) or []),
                sim_snapshot=snap,
                sim_speed=_sim_speed,
            )
            export_by[str(si)] = export_doc
            _prerun_export_json = True
            try:
                from .sim_control_defaults import SIM_PRERUN_EXPORT_JSON

                _prerun_export_json = bool(SIM_PRERUN_EXPORT_JSON)
            except Exception:
                _prerun_export_json = True
            if _prerun_export_json:
                try:
                    from pathlib import Path
                    from datetime import datetime

                    out_dir = Path(__file__).resolve().parents[2] / "data" / "sim_prerun"
                    out_dir.mkdir(parents=True, exist_ok=True)
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    out_path = out_dir / f"prerun_screen{si}_{stamp}.json"
                    write_prerun_export_json(str(out_path), export_doc)
                    print(f"[SIM] 프리런 export JSON (화면{si}): {out_path}", flush=True)
                    try:
                        slim_doc = build_prerun_export_document_web_slim(export_doc)
                        out_path_slim = out_dir / f"prerun_screen{si}_{stamp}_temp.json"
                        write_prerun_export_json(str(out_path_slim), slim_doc)
                        print(
                            f"[SIM] 프리런 export JSON (웹 슬림·화면{si}): {out_path_slim}",
                            flush=True,
                        )
                    except Exception as ex2:
                        print(
                            f"[SIM] 프리런 export JSON 슬림 저장 실패(화면{si}): {ex2}",
                            flush=True,
                        )
                except Exception as ex:
                    print(f"[SIM] 프리런 export JSON 저장 실패(화면{si}): {ex}", flush=True)
        except Exception as ex:
            print(f"[SIM] 프리런 export 문서 구성 실패(화면{si}): {ex}", flush=True)
        header = f"[SIM] 타임테이블(프리런) — 화면{si}"
        ch = _resolve_timetable_channel_for_screen(ext, si)
        if ch is None:
            chans = getattr(ext, "_sim_monitor_channels", None)
            if isinstance(chans, list) and 0 < si <= len(chans):
                ch = chans[si - 1]
        if isinstance(ch, dict):
            tb_map = _build_prerun_timetable_text({si: res}) or {}
            txt = str(tb_map.get(si) or "").strip()
            if not txt:
                txt = _build_timetable_label_text(header, metas)
            _apply_timetable_text_to_channel(ext, ch, screen=si, text=txt, show_label=not metas)
            if metas:
                def _mk_seek_cb(screen_i: int = si) -> Callable[[int], None]:
                    return lambda row_i: _on_timetable_row_seek(ext, int(screen_i), int(row_i))

                try:
                    mount_interactive_timetable(
                        ext,
                        ch,
                        screen=si,
                        header=header,
                        row_metas=metas,
                        on_row_clicked=_mk_seek_cb(),
                    )
                    ch["timetable_interactive"] = True
                    print(
                        f"[SIM] 타임테이블 인터랙티브(화면{si}): {len(metas)}행, 클릭·하이라이트 활성",
                        flush=True,
                    )
                    try:
                        refresh_timetable_row_highlight(ext, screen=si, sim_now=0.0)
                    except Exception:
                        pass
                except Exception as e:
                    ch["timetable_interactive"] = False
                    print(f"[SIM] 타임테이블 인터랙티브 장착 실패(화면{si}): {e}", flush=True)
                    _apply_timetable_text_to_channel(ext, ch, screen=si, text=txt, show_label=True)
            else:
                ch["timetable_interactive"] = False
                if "[SIM] 타임테이블(프리런)" not in txt:
                    print(f"[SIM] 타임테이블(화면{si}): 표시할 event/step 항목 없음", flush=True)
        # 프리런 타임테이블 JSON 콘솔 덤프 — SIM_PRERUN_CONSOLE_LOG 플래그로 함께 on/off.
        # (UI 타임테이블 패널 적용은 위에서 이미 끝났고, 여기서는 콘솔 출력만 제어한다)
        _prerun_console = True
        try:
            from .sim_control_defaults import SIM_PRERUN_CONSOLE_LOG

            _prerun_console = bool(SIM_PRERUN_CONSOLE_LOG)
        except Exception:
            _prerun_console = True
        if _prerun_console:
            try:
                print(header, flush=True)
                for m in metas:
                    print(m.display_line, flush=True)
            except Exception:
                pass
    try:
        _merge_prerun_ui_screen_dict(ext, "_sim_ep_bar_prerun_by_screen", bar_by, merge=merge)
        _merge_prerun_ui_screen_dict(ext, "_sim_timetable_row_metas_by_screen", meta_by, merge=merge)
        _merge_prerun_ui_screen_dict(ext, "_sim_seek_snapshots_by_screen", seek_by, merge=merge)
        _merge_prerun_ui_screen_dict(ext, "_sim_prerun_export_json_by_screen", export_by, merge=merge)
        _merge_prerun_ui_screen_dict(
            ext,
            "_sim_playback_schedule_by_screen",
            sched_by if isinstance(sched_by, dict) else {},
            merge=merge,
        )
        _merge_prerun_ui_screen_dict(
            ext,
            "_sim_playback_plan_by_screen",
            plan_by if isinstance(plan_by, dict) else {},
            merge=merge,
        )
        _merge_prerun_ui_screen_dict(
            ext,
            "_sim_playback_plan_initial_occ_by_screen",
            init_occ_by,
            merge=merge,
        )
    except Exception:
        pass
    try:
        from .control_sim_playback_plan import refresh_playback_display_at_sim

        screens_to_paint: set = set()
        for sk in (bar_by or {}).keys():
            try:
                screens_to_paint.add(int(sk))
            except Exception:
                continue
        for sk in (init_occ_by or {}).keys():
            try:
                screens_to_paint.add(int(sk))
            except Exception:
                continue
        chans = getattr(ext, "_sim_monitor_channels", None)
        for si in sorted(screens_to_paint):
            init = {}
            if isinstance(init_occ_by, dict):
                raw = init_occ_by.get(str(si))
                if not isinstance(raw, dict):
                    raw = init_occ_by.get(int(si))
                if isinstance(raw, dict):
                    init = dict(raw)
            try:
                if init:
                    _update_port_occupancy_panel(ext, dict(init), sim_time="0.00", screen=int(si))
            except Exception:
                pass
            rendered = False
            if bool(getattr(ext, "_sim_playback_started", False)):
                try:
                    refresh_playback_display_at_sim(ext, int(si), 0.0, force=True)
                    rendered = True
                except Exception:
                    pass
            if isinstance(chans, list) and 0 < int(si) <= len(chans):
                ch0 = chans[int(si) - 1]
                if isinstance(ch0, dict) and ch0.get("ep_timeline_widget") is None:
                    try:
                        if _render_ep_bar_prerun_at_t(ext, ch0, 0.0, init):
                            rendered = True
                    except Exception:
                        pass
                    if not rendered:
                        try:
                            _update_ep_timeline_under_port_state(
                                ext,
                                ch0,
                                dict(init),
                                "0.00",
                                honor_explicit_sim_time=False,
                                playback_ui_state=None,
                            )
                        except Exception:
                            pass
    except Exception:
        pass
    try:
        refresh_all_timetable_highlights(ext)
    except Exception:
        pass
    _scroll_sim_monitor_to_timetable(ext)
    # 재시작용 프리런 번들 보관 (정지해도 유지, 리셋/신규 시작 finalize 시 갱신)
    try:
        full_res = getattr(ext, "_sim_prerun_results_by_screen", None)
        if isinstance(full_res, dict) and full_res:
            _stash_prerun_restart_bundle(ext, full_res)
        else:
            _stash_prerun_restart_bundle(ext, results)
    except Exception as exc:
        print(f"[SIM] prerun restart stash failed: {exc}", flush=True)


_RESTART_BUNDLE_ATTRS = (
    "_sim_ep_bar_prerun_by_screen",
    "_sim_timetable_row_metas_by_screen",
    "_sim_seek_snapshots_by_screen",
    "_sim_prerun_export_json_by_screen",
    "_sim_playback_schedule_by_screen",
    "_sim_playback_plan_by_screen",
    "_sim_playback_plan_initial_occ_by_screen",
)


def _stash_prerun_restart_bundle(ext: Any, results: Dict[int, Any]) -> None:
    """직전 성공 프리런을 재시작용으로 보관."""
    if not isinstance(results, dict) or not results:
        return
    res_copy: Dict[int, Any] = {}
    for k, v in results.items():
        try:
            res_copy[int(k)] = v
        except Exception:
            continue
    if not res_copy:
        return
    extras: Dict[str, Any] = {}
    for attr in _RESTART_BUNDLE_ATTRS:
        raw = getattr(ext, attr, None)
        if isinstance(raw, dict):
            extras[attr] = dict(raw)
        else:
            extras[attr] = {}
    ext._sim_restart_prerun_bundle = {
        "results": res_copy,
        "extras": extras,
    }


def _clear_prerun_restart_bundle(ext: Any) -> None:
    try:
        ext._sim_restart_prerun_bundle = None
    except Exception:
        pass


def _restore_prerun_restart_bundle(ext: Any) -> Optional[Dict[int, Any]]:
    """번들에서 프리런/재생 자산을 복원하고 results 를 반환."""
    bundle = getattr(ext, "_sim_restart_prerun_bundle", None)
    if not isinstance(bundle, dict):
        return None
    results = bundle.get("results")
    if not isinstance(results, dict) or not results:
        return None
    out: Dict[int, Any] = {}
    for k, v in results.items():
        try:
            out[int(k)] = v
        except Exception:
            continue
    if not out:
        return None
    extras = bundle.get("extras")
    if isinstance(extras, dict):
        for attr in _RESTART_BUNDLE_ATTRS:
            raw = extras.get(attr)
            if isinstance(raw, dict):
                setattr(ext, attr, dict(raw))
    ext._sim_prerun_results_by_screen = dict(out)
    return out


def _build_prerun_timetable_text(results_by_screen: Any) -> Dict[int, str]:
    """
    프리런 결과(SimPreRunResult.items)를 **JSON 라인 형식의 타임테이블**로 만든다.

    출력 정책(요구사항):
    - 한 줄에 한 JSON 객체. 각 줄은 두 종류 중 하나.
        ┌─ kind="event": 시뮬 이벤트 발생 시점(ARRIVED/MOVE_*/REMOVED/FOUP_PROCESS_START/END 등)
        └─ kind="step":  공정/애니 동작 시작 시점(progress 의 RUNNING 첫 emit, elapsed=0.0)
    - 같은 시각이면 event → step 순서로 정렬.
    - port_id 등은 문자열("EP1", "BP3" 등)로 그대로 유지(시뮬 내부 표기와 일치).
    - 동작 라인(step)에는 anim 파일명/공정시간/애니시간/동작 설명/공정시간우선 등을 함께 동봉.

    출력 헤더는 ``"[SIM] 타임테이블(프리런) — 화면N"`` 으로 두어, ``_append_sim_log`` 의
    "타임테이블만 표시" 필터(timetable_only)를 그대로 통과한다.
    """
    out: Dict[int, str] = {}
    if not isinstance(results_by_screen, dict):
        return out

    def _f(x: Any, d: float = 0.0) -> float:
        try:
            return float(str(x).strip() or d)
        except Exception:
            return float(d)

    def _s(v: Any) -> str:
        try:
            return str(v).strip() if v is not None else ""
        except Exception:
            return ""

    for scr, res in results_by_screen.items():
        try:
            si = int(scr)
        except Exception:
            continue
        items = getattr(res, "items", None)
        if not isinstance(items, (list, tuple)):
            continue

        rows: List[Dict[str, Any]] = []
        for it in items:
            try:
                kind = str(getattr(it, "kind", "") or "").strip().lower()
                p = getattr(it, "payload", None)
                t_val = round(_f(getattr(it, "t", 0.0), 0.0), 2)

                # 1) 시뮬 이벤트 라인 (kind="event")
                if kind == "event" and isinstance(p, dict):
                    seq = _s(p.get("seq")).upper()
                    if not seq:
                        continue
                    row: Dict[str, Any] = {
                        "t": t_val,
                        "screen": si,
                        "kind": "event",
                        "event": seq,
                    }
                    # 있으면 동봉(없으면 키 자체 생략 → JSON 한 줄을 깔끔하게)
                    for k in ("port_id", "from_port_id", "to_port_id", "lot_id", "foup_id", "lot_seq"):
                        v = _s(p.get(k))
                        if v:
                            row[k] = v
                    rows.append(row)

                # 2) 동작 시작 라인 (kind="step") = progress.RUNNING 첫 emit (elapsed=0.0)
                elif kind == "progress" and isinstance(p, dict):
                    st = _s(p.get("status")).upper()
                    el = _f(p.get("elapsed", 0.0), 0.0)
                    if st != "RUNNING" or abs(el) > 1e-9:
                        continue
                    ev = _s(p.get("event_seq") or p.get("sequence_name")).upper()
                    if not ev:
                        continue
                    row = {
                        "t": t_val,
                        "screen": si,
                        "kind": "step",
                        "event": ev,
                    }
                    pid = _s(p.get("port_id"))
                    if pid:
                        row["port_id"] = pid
                    label = _s(p.get("label"))
                    if label:
                        row["label"] = label
                    # anim 파일명: 비어 있어도 명시적으로 빈 문자열로 둔다(필드 존재 자체가 의미)
                    row["anim"] = _s(p.get("linked_anim_json"))
                    row["proc_sec"] = round(_f(p.get("proc_sec", 0.0), 0.0), 2)
                    row["anim_sec"] = round(_f(p.get("anim_sec", 0.0), 0.0), 2)
                    detail = _s(p.get("detail"))
                    if detail:
                        row["detail"] = detail
                    ptp = _s(p.get("process_time_priority"))
                    if ptp:
                        row["process_time_priority"] = ptp
                    rows.append(row)
            except Exception:
                continue

        # 같은 시각이면 event 를 먼저, step 을 다음에
        kind_prio = {"event": 0, "step": 1}
        try:
            rows.sort(key=lambda r: (
                float(r.get("t", 0.0)),
                int(kind_prio.get(str(r.get("kind", "")), 9)),
            ))
        except Exception:
            pass

        lines: List[str] = []
        lines.append(f"[SIM] 타임테이블(프리런) — 화면{si}")
        if not rows:
            lines.append('{"kind":"info","message":"표시할 event/step 항목 없음"}')
        for r in rows:
            try:
                lines.append(json.dumps(r, ensure_ascii=False))
            except Exception:
                continue
        out[si] = "\n".join(lines).strip()
    return out


def _bootstrap_partial_prerun_playback(
    ext: Any, new_results: Dict[int, SimPreRunResult], partial_targets: List[int]
) -> None:
    """다른 화면 재생 중 — 신규 화면만 프리런 결과를 붙여 재생한다."""
    if not new_results:
        return
    try:
        set_sim_playback_active(ext, True)
    except Exception:
        pass
    try:
        by = getattr(ext, "_sim_last_total_est_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_last_total_est_by_screen = by
        for scr, res in new_results.items():
            try:
                by[str(int(scr))] = float(res.final_sim_time)
            except Exception:
                continue
    except Exception:
        pass
    try:
        _finalize_prerun_ui_assets(ext, new_results, merge=True)
    except Exception as e:
        print(f"[SIM] 부분 프리런 UI 자산 구성 실패: {e}", flush=True)

    full_results: Dict[int, SimPreRunResult] = {}
    try:
        merged_src = getattr(ext, "_sim_prerun_results_by_screen", None)
        if isinstance(merged_src, dict):
            for k, v in merged_src.items():
                try:
                    full_results[int(k)] = v
                except Exception:
                    continue
    except Exception:
        pass
    if not full_results:
        full_results = dict(new_results)

    def _speed() -> float:
        try:
            m = getattr(ext, "_sim_speed_model", None)
            return float(m.get_value_as_float()) if m is not None else 1.0
        except Exception:
            return 1.0

    def _timeline_event_gate(scr: int) -> bool:
        return can_emit_timeline_event(ext, int(scr))

    try:
        _prepare_playback_emit_environment(ext, new_results, scope_screens_only=True)
    except Exception:
        pass
    try:
        add_playback_sessions_after_prerun(
            ext,
            new_results,
            _make_playback_emit_fn(ext, full_results),
            _speed,
            gate_fn=_timeline_event_gate,
        )
    except Exception as e:
        print(f"[SIM] 부분 재생 세션 추가 실패: {e}", flush=True)
        return
    try:
        from .control_sim_playback_plan import refresh_playback_display_at_sim

        for scr_k in new_results:
            try:
                refresh_playback_display_at_sim(ext, int(scr_k), 0.0, force=True)
            except Exception:
                pass
    except Exception:
        pass
    try:
        for scr, rr in new_results.items():
            try:
                scr_i = int(scr)
            except Exception:
                scr_i = 1
            p0 = {
                "tbs_sim_screen": str(scr_i),
                "sim_time": "0.00",
                "sim_total_est_sec": f"{float(rr.final_sim_time):.2f}",
                "label": "대기",
                "detail": "",
                "status": "RUNNING",
                "elapsed": "0.0",
                "total": "0.0",
                "percent": "0",
            }
            _update_sim_progress(ext, p0)
    except Exception:
        pass
    try:
        sub = getattr(ext, "_sim_playback_ui_sub", None)
        if sub is None:
            ext._sim_playback_ui_sub = app.get_app().get_update_event_stream().create_subscription_to_pop(
                lambda _e: _tick_playback_timeline(ext),
                name="morph.tbs_control_2:sim_playback_tick",
            )
    except Exception:
        pass
    try:
        _append_sim_log(
            ext,
            f"[SIM] 화면 {','.join(str(s) for s in partial_targets)} 프리런 완료 — 재생 시작",
        )
    except Exception:
        pass
    try:
        ext._sim_partial_prerun_screens = None
    except Exception:
        pass


def _drain_sim_log_queue(ext: Any) -> None:
    try:
        # 프리런 완료 시점에 타임라인 플레이어를 시작한다(메인 스레드에서만).
        try:
            ev = getattr(ext, "_sim_prerun_done_evt", None)
            started = bool(getattr(ext, "_sim_playback_started", False))
            playback_done = bool(getattr(ext, "_sim_playback_done", False))
            partial_targets = getattr(ext, "_sim_partial_prerun_screens", None)
            if (
                partial_targets
                and ev is not None
                and hasattr(ev, "is_set")
                and ev.is_set()
            ):
                results_all = getattr(ext, "_sim_prerun_results_by_screen", None)
                partial_set = {int(x) for x in partial_targets}
                new_results: Dict[int, SimPreRunResult] = {}
                if isinstance(results_all, dict):
                    for k, v in results_all.items():
                        try:
                            ki = int(k)
                        except Exception:
                            continue
                        if ki in partial_set:
                            new_results[ki] = v
                if new_results:
                    _bootstrap_partial_prerun_playback(ext, new_results, list(partial_targets))
                else:
                    try:
                        ext._sim_partial_prerun_screens = None
                    except Exception:
                        pass
            elif (
                (not started)
                and (not playback_done)
                and ev is not None
                and hasattr(ev, "is_set")
                and ev.is_set()
            ):
                results = getattr(ext, "_sim_prerun_results_by_screen", None)
                if isinstance(results, dict) and results:
                    try:
                        stopped_scr = {
                            int(x)
                            for x in (getattr(ext, "_sim_stopped_screens", None) or set())
                        }
                        if stopped_scr:
                            results = {
                                int(k): v
                                for k, v in results.items()
                                if int(k) not in stopped_scr
                            }
                    except Exception:
                        pass
                if isinstance(results, dict) and results:
                    try:
                        set_sim_playback_active(ext, True)
                        ext._sim_anim_pending = []
                        ext._sim_anim_pending_by_screen = {}
                    except Exception:
                        pass

                    # 화면별 총 시뮬 시간 확정(막대 스케일/표시)
                    try:
                        by = getattr(ext, "_sim_last_total_est_by_screen", None)
                        if not isinstance(by, dict):
                            by = {}
                            ext._sim_last_total_est_by_screen = by
                        for scr, res in results.items():
                            try:
                                by[str(int(scr))] = float(res.final_sim_time)
                            except Exception:
                                continue
                    except Exception:
                        pass

                    # 프리런 완료 시 막대 사전계산 + 인터랙티브 타임테이블(매 Start 마다 갱신)
                    try:
                        _finalize_prerun_ui_assets(ext, results)
                    except Exception as e:
                        print(f"[SIM] 프리런 UI 자산 구성 실패: {e}", flush=True)
                        try:
                            tb_by = _build_prerun_timetable_text(results) or {}
                            for si, txt in tb_by.items():
                                if str(txt or "").strip():
                                    print(str(txt), flush=True)
                        except Exception:
                            pass
                    try:
                        ext._sim_prerun_timetable_printed = True
                    except Exception:
                        pass

                    # UI 그래프 동기화용 playback 엔진 생성
                    playback_engs: List[Any] = []
                    try:
                        max_scr = max(int(s) for s in results.keys())
                    except Exception:
                        max_scr = 1
                    for i in range(1, max_scr + 1):
                        rr = results.get(int(i))
                        if rr is None:
                            continue
                        playback_engs.append(PlaybackEngine(final_sim_time=float(rr.final_sim_time)))
                    try:
                        ext._sim_engines = playback_engs
                        ext._sim_engine = playback_engs[0] if playback_engs else None
                    except Exception:
                        pass

                    def _speed() -> float:
                        try:
                            m = getattr(ext, "_sim_speed_model", None)
                            return float(m.get_value_as_float()) if m is not None else 1.0
                        except Exception:
                            return 1.0

                    def _timeline_event_gate(scr: int) -> bool:
                        return can_emit_timeline_event(ext, int(scr))

                    try:
                        _prepare_playback_emit_environment(ext, results)
                    except Exception:
                        pass

                    bootstrap_playback_after_prerun(
                        ext,
                        results,
                        _make_playback_emit_fn(ext, results),
                        _speed,
                        gate_fn=_timeline_event_gate,
                    )
                    # 첫 공정(첫 이벤트) 전에도 진행현황/시간이 계속 증가하도록 초기 payload를 즉시 세팅한다.
                    try:
                        for scr, rr in results.items():
                            try:
                                scr_i = int(scr)
                            except Exception:
                                scr_i = 1
                            p0 = {
                                "tbs_sim_screen": str(scr_i),
                                "sim_time": "0.00",
                                "sim_total_est_sec": f"{float(rr.final_sim_time):.2f}",
                                "label": "대기",
                                "detail": "",
                                "status": "RUNNING",
                                "elapsed": "0.0",
                                "total": "0.0",
                                "percent": "0",
                            }
                            _update_sim_progress(ext, p0)
                    except Exception:
                        pass

                    # UI update stream에서 재생 tick
                    try:
                        sub = getattr(ext, "_sim_playback_ui_sub", None)
                        if sub is not None:
                            try:
                                sub.unsubscribe()
                            except Exception:
                                pass
                        ext._sim_playback_ui_sub = app.get_app().get_update_event_stream().create_subscription_to_pop(
                            lambda _e: _tick_playback_timeline(ext),
                            name="morph.tbs_control_2:sim_playback_tick",
                        )
                    except Exception:
                        pass
                    try:
                        _append_sim_log(ext, "[SIM] 프리런 완료: 결과 타임라인을 재생합니다.")
                    except Exception:
                        pass
        except Exception:
            pass

        q = getattr(ext, "_sim_log_queue", None)
        if q is None:
            return

        # 공정설정 시간 우선 모드에서의 애니 중단은 tick 스레드가 아니라 UI(메인) 스레드에서만 수행한다.
        # tick 스레드에서 stop_all_* / runner.stop() 등을 호출하면 Kit 내부가 스레드-unsafe로 크래시할 수 있다.
        try:
            # 화면별 interrupt 우선 처리
            by = getattr(ext, "_sim_interrupt_anim_event_by_screen", None)
            fn_by = getattr(ext, "_sim_interrupt_anim_apply_fn_by_screen", None)
            if isinstance(by, dict) and isinstance(fn_by, dict):
                for scr, ev in list(by.items()):
                    try:
                        if ev is None or not hasattr(ev, "is_set") or not ev.is_set():
                            continue
                        try:
                            ev.clear()
                        except Exception:
                            pass
                        fn = fn_by.get(str(scr))
                        if callable(fn):
                            try:
                                fn()
                            except Exception:
                                pass
                    except Exception:
                        continue
        except Exception:
            pass
        try:
            ie = getattr(ext, "_sim_interrupt_anim_event", None)
            if ie is not None and hasattr(ie, "is_set") and ie.is_set():
                try:
                    ie.clear()
                except Exception:
                    pass
                fn = getattr(ext, "_sim_interrupt_anim_apply_fn", None)
                if callable(fn):
                    try:
                        fn()
                    except Exception:
                        pass
        except Exception:
            pass

        # 표시모드 제거: 항상 둘다(진행현황+이력로그)
        panel_mode = SimLogPanelMode.ALL
        try:
            _maybe_queue_post_anim_port_by_wall_clock(ext)
            _flush_pending_post_anim_port_applies(ext)
        except Exception:
            pass
        drain_limit = sim_log_ui_drain_limit(ext)
        hist_limit = sim_log_ui_history_drain_limit(ext)
        hist_count = 0
        batch: List[Any] = []
        while len(batch) < drain_limit:
            try:
                batch.append(q.get_nowait())
            except Exception:
                break
        requeue: List[Any] = []
        gate_break = False
        for item in batch:
            if gate_break:
                requeue.append(item)
                continue
            kind, payload = (
                item if isinstance(item, tuple) and len(item) == 2 else (SimUiQueueKind.HISTORY_LINE.value, item)
            )
            kind_s = _coerce_sim_ui_queue_kind(kind)
            if (
                bool(getattr(ext, "_sim_playback_started", False))
                and kind_s == SimUiQueueKind.HISTORY_LINE.value
                and hist_count >= hist_limit
            ):
                requeue.append(item)
                continue
            if kind_s == SimUiQueueKind.HISTORY_LINE.value:
                hist_count += 1
            _dispatch_sim_ui_queue_item(ext, kind_s, payload, panel_mode)

            # 공정확인 체크 + gate pause 상태면, "확인창 1개를 띄울 때까지만" 처리하고 멈춘다.
            # (gate pause를 너무 이르게 걸어도 UI가 1개 이벤트를 처리해 창을 띄울 수 있어야 한다)
            try:
                confirm_each = bool(
                    getattr(ext, "_sim_confirm_each_step_model", None) is not None
                    and ext._sim_confirm_each_step_model.get_value_as_bool()
                )
            except Exception:
                confirm_each = False
            if confirm_each:
                try:
                    gp = getattr(ext, "_sim_gate_pause_event", None)
                    if gp is not None and gp.is_set() and getattr(ext, "_sim_gate_dialog", None) is not None:
                        gate_break = True
                except Exception:
                    gate_break = True
        for item in requeue:
            try:
                q.put_nowait(item)
            except Exception:
                pass
    except Exception as e:
        # UI 드레인 예외가 발생해도 구독이 끊기지 않도록 보호
        print(f"[SIM UI] 로그 드레인 예외: {e}", flush=True)


def _finalize_playback_if_done(ext: Any) -> None:
    rt = get_playback_runtime(ext)
    results = getattr(ext, "_sim_prerun_results_by_screen", None)
    if rt is None or not isinstance(results, dict) or not results:
        return
    if not rt.all_reached_end() or bool(getattr(ext, "_sim_playback_done", False)):
        return
    ext._sim_playback_done = True
    try:
        _finalize_sim_timeline_on_done(ext)
    except Exception:
        pass
    for scr_k, res in results.items():
        try:
            scr_i = int(scr_k)
        except Exception:
            scr_i = 1
        try:
            p_done = {
                "tbs_sim_screen": str(scr_i),
                "sim_time": f"{float(res.final_sim_time):.2f}",
                "sim_total_est_sec": f"{float(res.final_sim_time):.2f}",
                "label": "완료",
                "detail": "",
                "status": "DONE",
                "elapsed": f"{float(res.final_sim_time):.1f}",
                "total": f"{float(res.final_sim_time):.1f}",
                "percent": "100",
            }
            _deliver_playback_heartbeat_progress(ext, p_done)
        except Exception:
            pass
    try:
        _enqueue_control_action(ext, SimUiControlAction.EXPORT_XLSX.value)
    except Exception:
        pass
    try:
        stop_playback_runtime(ext)
    except Exception:
        pass
    try:
        set_sim_playback_active(ext, False)
    except Exception:
        pass
    try:
        sub = getattr(ext, "_sim_playback_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_playback_ui_sub = None
    except Exception:
        pass
    # 재생 1회 완료 후 drain 이 프리런→재생 bootstrap 을 다시 타지 않게 정리(시뮬 로직·결과 표시는 유지).
    try:
        ev = getattr(ext, "_sim_prerun_done_evt", None)
        if ev is not None and hasattr(ev, "clear"):
            ev.clear()
        ext._sim_prerun_results_by_screen = None
        ext._sim_playback_schedule_by_screen = None
        try:
            from .control_sim_playback_plan import clear_playback_plan_runtime_state

            clear_playback_plan_runtime_state(ext)
        except Exception:
            pass
    except Exception:
        pass


def _tick_playback_timeline(ext: Any) -> None:
    """프리런 재생 — 1·N 화면 동일 ``SimPlaybackRuntime.tick_all``."""
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    rt = get_playback_runtime(ext)
    if rt is None:
        legacy = getattr(ext, "_sim_playback_player", None)
        if legacy is not None and callable(getattr(legacy, "is_playing", None)) and legacy.is_playing():
            try:
                legacy.tick()
            except Exception:
                pass
            try:
                refresh_all_timetable_highlights(ext)
            except Exception:
                pass
            _finalize_playback_if_done(ext)
        return

    def _after_tick(e: Any) -> None:
        try:
            refresh_all_timetable_highlights(e)
        except Exception:
            pass
        try:
            _refresh_all_foup_playback_heartbeats(e)
        except Exception:
            pass
        _finalize_playback_if_done(e)

    rt.tick_all(
        ext,
        max_emits_per_screen=20,
        progress_sink=_deliver_playback_heartbeat_progress,
        timeline_only_sink=_sim_ui_sink_progress,
        build_prog_payload=_build_playback_prog_payload_for_session,
        sync_engine_now=_sync_playback_engine_now,
        on_after_tick=_after_tick,
    )


def _tick_playback(ext: Any) -> None:
    _tick_playback_timeline(ext)


def _sim_active_anim_owner_screen(ext: Any) -> int:
    """
    현재(또는 직전) JSON 애니 job 이 어느 시뮼 화면에서 시작됐는지 1-based 인덱스로 반환한다.

    ``_execute_mapped_sequence_stub`` 가 job 에 넣은 ``tbs_sim_screen`` 과 ``_sim_anim_active`` 를 읽는다.
    값이 없거나 파싱 실패 시 1(메인).
    """
    active = getattr(ext, "_sim_anim_active", None) or {}
    if isinstance(active, dict):
        try:
            v = int(str(active.get("tbs_sim_screen", "1") or "1").strip() or "1")
            return max(1, v)
        except Exception:
            pass
    return 1


def _ensure_tick_pause_map_for_multi(ext: Any, n_ch: int) -> None:
    """분할 N>1 시 화면마다 독립 Event — 없으면 전역 pause 로 떨어져 전 엔진 틱이 같이 멈출 수 있다."""
    if n_ch <= 1:
        return
    try:
        m: Dict[str, threading.Event] = {}
        for i in range(1, n_ch + 1):
            m[str(i)] = threading.Event()
        ext._sim_tick_pause_events_by_screen = m
        ub: Dict[str, Any] = {str(i): None for i in range(1, n_ch + 1)}
        ext._sim_tick_pause_until_wall_by_screen = ub
    except Exception:
        pass


def _is_multi_viewport_sim(ext: Any) -> bool:
    try:
        nsp = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
        return nsp > 1
    except Exception:
        return False


_ANIM_SCREEN_WORKER_STOP = object()


def _playback_json_job_queue(ext: Any, screen: int):
    from collections import deque

    by = getattr(ext, "_sim_playback_json_jobs_by_screen", None)
    if not isinstance(by, dict):
        by = {}
        ext._sim_playback_json_jobs_by_screen = by
    key = str(max(1, int(screen)))
    q = by.get(key)
    if not isinstance(q, deque):
        q = deque()
        by[key] = q
    return q


def _clear_playback_json_job_queues(ext: Any) -> None:
    try:
        ext._sim_playback_json_jobs_by_screen = {}
    except Exception:
        pass


def _enqueue_playback_json_job(ext: Any, job: Dict[str, Any]) -> None:
    """N>1 프리런 — emit sink 는 즉시 반환, ``tick_all`` 에서 ``_start_job_impl`` 실행."""
    try:
        scr = int(str(job.get("tbs_sim_screen", "1") or "1").strip() or "1")
    except Exception:
        scr = 1
    _playback_json_job_queue(ext, scr).append(dict(job))


def _dispatch_json_anim_job(ext: Any, job: Dict[str, Any]) -> None:
    if bool(getattr(ext, "_sim_playback_started", False)) and is_multi_playback_instances(ext):
        _enqueue_playback_json_job(ext, job)
        return
    fn = getattr(ext, "_sim_json_start_fn", None)
    if callable(fn):
        try:
            fn(dict(job))
        except Exception:
            pass


def _drain_playback_json_job_queues(ext: Any) -> None:
    """N>1 — 화면별 대기 job 을 runner idle 일 때 1건씩 시작."""
    if not bool(getattr(ext, "_sim_playback_started", False)):
        return
    if not is_multi_playback_instances(ext):
        return
    fn = getattr(ext, "_sim_json_start_fn", None)
    if not callable(fn):
        return
    by = getattr(ext, "_sim_playback_json_jobs_by_screen", None)
    if not isinstance(by, dict) or not by:
        return
    for key in sorted(by.keys(), key=lambda x: int(x)):
        q = by.get(key)
        if q is None:
            continue
        try:
            scr = int(key)
        except Exception:
            continue
        if is_screen_runner_busy(ext, scr):
            continue
        try:
            job = q.popleft()
        except Exception:
            continue
        if isinstance(job, dict):
            try:
                fn(job)
            except Exception:
                pass


def _try_release_all_playback_json_walls(ext: Any) -> None:
    rt = get_playback_runtime(ext)
    if rt is None:
        return
    for scr in list(rt.sessions.keys()):
        try:
            try_release_json_wall_when_idle(ext, int(scr))
        except Exception:
            pass


def _anim_screen_worker_name(screen_idx: int) -> str:
    return f"morph.tbs_anim_scr_{max(1, int(screen_idx))}"


def _ensure_anim_screen_worker(ext: Any, screen_idx: int) -> None:
    """멀티 뷰: 화면별 JSON 애니 전용 워커(1화면=1스레드 직렬 실행)."""
    scr = max(1, int(screen_idx))
    key = str(scr)
    workers = getattr(ext, "_sim_anim_workers_by_screen", None)
    if not isinstance(workers, dict):
        workers = {}
        ext._sim_anim_workers_by_screen = workers
    ent = workers.get(key)
    if isinstance(ent, dict):
        th = ent.get("thread")
        if th is not None and getattr(th, "is_alive", lambda: False)():
            return
    lock = threading.Lock()
    cond = threading.Condition(lock)
    queue: list = []

    def _loop() -> None:
        while True:
            with cond:
                while not queue:
                    cond.wait()
                item = queue.pop(0)
            if item is _ANIM_SCREEN_WORKER_STOP:
                break
            try:
                fn = item.get("run_fn") if isinstance(item, dict) else None
                if callable(fn):
                    fn()
            except Exception:
                pass

    th = threading.Thread(target=_loop, name=_anim_screen_worker_name(scr), daemon=True)
    workers[key] = {"thread": th, "lock": lock, "cond": cond, "queue": queue}
    th.start()


def _halt_screen_json_anim(ext: Any, screen_idx: int, *, join_sec: float = 5.0) -> None:
    """해당 화면의 진행 중 JSON·main dispatch·MOVE 를 즉시 중단(다음 JSON 선행)."""
    scr = max(1, int(screen_idx))
    ctx = _usd_context_name_for_sim_screen(ext, scr)
    reg = None
    try:
        from .tbs_split_composed_loader import get_split_runtime_for_screen

        rt = get_split_runtime_for_screen(ext, scr)
        if rt is not None:
            reg = rt.registry
    except Exception:
        reg = None
    try:
        runners = getattr(ext, "_sim_runners_by_screen", None)
        rr = runners.get(str(scr)) if isinstance(runners, dict) else None
        if rr is not None:
            try:
                if bool(getattr(rr, "is_running", lambda: False)()):
                    rr.pause(cancel_all_move_rotate=True)
            except Exception:
                pass
            th = getattr(rr, "_lam_thread", None)
            if th is not None and getattr(th, "is_alive", lambda: False)():
                try:
                    th.join(timeout=max(0.5, float(join_sec)))
                except Exception:
                    pass
    except Exception:
        pass
    try:
        from .sim_channel_scope import drain_channel_motion_complete, stop_channel_animations

        stop_channel_animations(ctx, diag_reason="halt_screen_json")
        drain_channel_motion_complete(ctx, reg, max_sec=max(1.0, float(join_sec)), stable_ticks=3)
    except Exception:
        pass


def _enqueue_anim_screen_job(
    ext: Any,
    screen_idx: int,
    run_fn: Any,
    *,
    priority: int = 10,
) -> None:
    """화면 워커 큐에 job 추가. priority<=0 이면 맨 앞. 새 이벤트 시 이전 JSON 선행 중단."""
    scr = max(1, int(screen_idx))
    try:
        _halt_screen_json_anim(ext, scr, join_sec=2.0)
    except Exception:
        pass
    _ensure_anim_screen_worker(ext, scr)
    ent = ext._sim_anim_workers_by_screen[str(scr)]
    lock = ent["lock"]
    cond = ent["cond"]
    queue = ent["queue"]
    payload = {"run_fn": run_fn}
    with cond:
        # 이전 대기 job 은 버림 — 최신 이벤트 JSON 만 처음부터 실행.
        queue.clear()
        if int(priority) <= 0:
            queue.insert(0, payload)
        else:
            queue.append(payload)
        cond.notify()


def _stop_anim_screen_workers(ext: Any) -> None:
    workers = getattr(ext, "_sim_anim_workers_by_screen", None)
    if not isinstance(workers, dict):
        return
    for key, ent in list(workers.items()):
        if not isinstance(ent, dict):
            continue
        lock = ent.get("lock")
        cond = ent.get("cond")
        queue = ent.get("queue")
        th = ent.get("thread")
        try:
            if lock is not None and cond is not None and queue is not None:
                with cond:
                    queue.append(_ANIM_SCREEN_WORKER_STOP)
                    cond.notify()
        except Exception:
            pass
        try:
            if th is not None:
                th.join(timeout=2.0)
        except Exception:
            pass
        try:
            workers.pop(key, None)
        except Exception:
            pass


def _pause_event_for_screen(ext: Any, screen_idx: int) -> Optional[threading.Event]:
    try:
        m = getattr(ext, "_sim_tick_pause_events_by_screen", None)
        if not isinstance(m, dict):
            return getattr(ext, "_sim_tick_pause_event", None)
        key = str(max(1, int(screen_idx)))
        ev = m.get(key)
        if ev is None:
            ev = threading.Event()
            m[key] = ev
        return ev
    except Exception:
        return getattr(ext, "_sim_tick_pause_event", None)


def _multi_tick_should_skip_for_screen(ext: Any, screen_idx: int, anim_running: bool) -> bool:
    """
    멀티 뷰 전용: ``_sim_tick_pause_event`` 가 켜져 있을 때 이 화면만 tick 을 건너뛸지.

    - SequenceRunner 가 돌면: 해당 애니가 시작된 화면(owner)만 sim 을 잠시 멈춘다(배속>1 동기).
    - Kit translate/rotate/curve 만 돌면: 메인 스테이지(화면 1)만 멈춘다.
    - 그 외 pause 구간(보수적): 모든 화면 멈춤(기존 단일 루프와 동일).
    """
    pause_evt = getattr(ext, "_sim_tick_pause_event", None)
    if pause_evt is None or not pause_evt.is_set():
        return False
    owner = _sim_active_anim_owner_screen(ext)
    try:
        ru = bool(
            getattr(ext, "_sim_runner", None) is not None
            and getattr(ext._sim_runner, "is_running", lambda: False)()
        )
    except Exception:
        ru = False
    if ru:
        return screen_idx == owner
    kit_only = bool(anim_running) and not ru
    if kit_only:
        return screen_idx == 1
    if anim_running:
        return True
    uw = getattr(ext, "_sim_tick_pause_until_wall", None)
    if isinstance(uw, (float, int)) and time.monotonic() < float(uw):
        return screen_idx == owner
    return True


def _sim_multi_engine_tick_worker(
    ext: Any,
    sim: Any,
    screen_idx: int,
    stop_evt: threading.Event,
    export_lock: threading.Lock,
) -> None:
    """
    한 개 ``TBSSimulationEngine`` 전용 tick 루프(별도 스레드에서 실행).

    - ``screen_idx`` 에 해당하는 엔진만 ``sim.tick(scaled)``; 배속은 매 루프 ``_sim_speed_model`` 에서 읽는다.
    - ``_sim_tick_pause_event`` 가 켜져 있으면 ``_multi_tick_should_skip_for_screen`` 으로 **이 화면만**
      tick 을 건너뛸지 결정해, 다른 화면은 계속 진행한다.
    - ``ext._sim_engines`` 전원 ``is_done`` 이면 export 액션을 한 번만 enqueue 하고 ``_sim_multi_tick_shutdown`` 을 세운다.
    """
    last = time.perf_counter()
    printed = False
    while not stop_evt.is_set():
        if getattr(ext, "_sim_multi_tick_shutdown", False):
            break
        pause_evt = _pause_event_for_screen(ext, screen_idx)
        gate_pause_evt = getattr(ext, "_sim_gate_pause_event", None)
        try:
            confirm_each = bool(
                getattr(ext, "_sim_confirm_each_step_model", None) is not None
                and ext._sim_confirm_each_step_model.get_value_as_bool()
            )
        except Exception:
            confirm_each = False
        if not confirm_each and gate_pause_evt is not None and gate_pause_evt.is_set():
            try:
                gate_pause_evt.clear()
            except Exception:
                pass
        # 화면별 pause: 분할 N>1 에서는 애니 동기용 tick 정지를 쓰지 않는다(각 엔진 sim 시간이 끊기지 않게).
        multi_vp = _is_multi_viewport_sim(ext)
        if (not multi_vp) and pause_evt is not None and pause_evt.is_set():
            # fail-safe 1) 예상 애니 길이(벽시계) 경과 시 자동 해제
            try:
                until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
                until_t = until_by.get(str(screen_idx)) if isinstance(until_by, dict) else None
                if isinstance(until_t, (float, int)) and time.monotonic() >= float(until_t):
                    try:
                        pause_evt.clear()
                    except Exception:
                        pass
            except Exception:
                pass
            # fail-safe 2) runner/active가 없는데 pause만 남아있으면 해제
            try:
                runner_alive = False
                runners_by = getattr(ext, "_sim_runners_by_screen", None)
                rr = runners_by.get(str(screen_idx)) if isinstance(runners_by, dict) else None
                runner_alive = bool(rr is not None and getattr(rr, "is_running", lambda: False)())
            except Exception:
                runner_alive = False
            try:
                active_by = getattr(ext, "_sim_anim_active_by_screen", None)
                act = active_by.get(str(screen_idx)) if isinstance(active_by, dict) else None
                active_has = bool(isinstance(act, dict) and act)
            except Exception:
                active_has = False
            if (not runner_alive) and (not active_has):
                try:
                    pause_evt.clear()
                except Exception:
                    pass
            time.sleep(0.02)
            if pause_evt.is_set():
                continue

        now = time.perf_counter()
        dt = now - last
        last = now
        dt = max(0.001, min(dt, 0.1))
        try:
            sp = max(0.1, float(ext._sim_speed_model.get_value_as_float()))
        except Exception:
            sp = 1.0
        scaled = dt * sp
        try:
            if sim is not None and not getattr(sim, "is_done", False):
                sim.tick(scaled)
        except Exception:
            pass
        try:
            from . import sim_multi_diag as _mdiag

            _mdiag.log_tick_heartbeat(ext, screen=screen_idx, sim=sim)
        except Exception:
            pass

        if not printed:
            try:
                print(f"[SIM] tick 동작 확인 (screen{screen_idx} worker)", flush=True)
            except Exception:
                pass
            printed = True

        eng_list = getattr(ext, "_sim_engines", None)
        if isinstance(eng_list, list) and len(eng_list) > 0:
            if all(getattr(s, "is_done", False) for s in eng_list if s is not None):
                try:
                    print("[SIM] 멀티 종료 감지", flush=True)
                except Exception:
                    pass
                with export_lock:
                    if not getattr(ext, "_sim_multi_export_done", False):
                        try:
                            ext._sim_multi_export_done = True
                        except Exception:
                            pass
                        try:
                            ext._sim_multi_tick_shutdown = True
                        except Exception:
                            pass
                        _enqueue_control_action(ext, SimUiControlAction.EXPORT_XLSX.value)
                break
        time.sleep(0.02)


def _sim_anim_status_key(ext: Any) -> Tuple[bool, str, int, str]:
    """진행 패널 중복 스킵용: 재생 여부·현재 파일·대기 큐·다음 파일."""
    runner = getattr(ext, "_sim_runner", None)
    try:
        running = bool(runner is not None and runner.is_running())
    except Exception:
        running = False
    active = getattr(ext, "_sim_anim_active", None) or {}
    cur_file = str(active.get("file", "") or "").strip() if isinstance(active, dict) else ""
    pend = getattr(ext, "_sim_anim_pending", None)
    plist = pend if isinstance(pend, list) else []
    q = len(plist)
    next_f = ""
    if plist and isinstance(plist[0], dict):
        next_f = str(plist[0].get("file", "") or "").strip()
    return (running, cur_file, q, next_f)


def _format_anim_status_footer(ext: Any) -> str:
    """진행현황 패널 하단: 현재 재생 JSON 파일·대기열.

    SequenceRunner.run()은 _begin_sequence()를 다음 프레임으로 미루므로,
    잠깐 동안은 _sim_anim_active.file 은 있는데 is_running() 이 False 인 구간이 있다.
    이 경우에도 파일명을 보여준다(진행현황에 '재생 없음'으로 보이는 문제 방지).
    시퀀스가 끝나면 _on_done 경로에서 _sim_anim_active 를 비운다.
    """
    running, cur_file, q, next_f = _sim_anim_status_key(ext)
    if cur_file:
        if running:
            lines = [f"애니메이션 파일(재생 중): {cur_file}"]
        else:
            lines = [f"애니메이션 파일: {cur_file}"]
        if q > 0 and next_f:
            lines.append(f"대기열: {q}건 (다음 {next_f})")
        return "\n".join(lines)
    if q > 0 and next_f:
        return "애니메이션: 대기 — 다음 " + next_f + (f" (큐 {q}건)" if q > 1 else "")
    return "애니메이션 파일: 재생 없음"


def _format_progress_anim_footer(ext: Any, payload: Dict[str, str]) -> str:
    """진행현황 하단 — ProgressStepState 단일 출처(러너는 보조 줄만)."""
    try:
        scr = int(str(payload.get("tbs_sim_screen", "1") or "1").strip() or "1")
    except Exception:
        scr = 1
    return format_progress_anim_footer(ext, scr)


def _refresh_sim_progress_from_last(ext: Any, screen: Optional[int] = None) -> None:
    """애니 런타임 변화 후 ProgressStepState 기준으로 패널만 다시 그린다."""
    screens: List[int] = []
    if screen is not None:
        screens = [max(1, int(screen))]
    else:
        by = getattr(ext, "_sim_progress_step_by_screen", None)
        if isinstance(by, dict) and by:
            for k in by.keys():
                try:
                    screens.append(max(1, int(str(k).strip() or "1")))
                except Exception:
                    continue
        if not screens:
            screens = [1]
    for scr in screens:
        try:
            sync_anim_runtime_from_ext(ext, scr)
        except Exception:
            pass
        p = build_payload_from_step(ext, scr)
        if isinstance(p, dict):
            p["_force_progress_ui"] = "1"
            _update_sim_progress(ext, p)


def _update_sim_progress(ext: Any, payload: Dict[str, str]) -> None:
    """
    진행현황 텍스트를 갱신한다.

    - ``payload["tbs_sim_screen"]``(엔진 ``event_tags`` 병합)으로 **멀티 모니터** 중 해당 열의
      ``progress_label`` 에만 쓴다. 단일 모드는 첫 채널 + ``_sim_progress_text`` 레거시 모델.
    - RUNNING 일 때 동일 내용 반복 갱신을 줄이기 위해 ``_sim_progress_last_key`` 로 디듀프한다.
    - 멀티에서 ``tbs_sim_screen`` 누락 시 drop (화면1 오염 방지).
    """
    scr_opt = _resolve_payload_sim_screen(ext, payload if isinstance(payload, dict) else {})
    if scr_opt is None:
        try:
            print("[TBS/port-screen] drop progress UI (no tbs_sim_screen)", flush=True)
        except Exception:
            pass
        return
    panel_slot = str(int(scr_opt))
    label = str(payload.get("label", "")).strip()
    # EP 타임라인 전용 업데이트는 텍스트를 덮어쓰지 않고 그래프만 갱신한다.
    try:
        if str(payload.get("timeline_only", "")).strip() in ("1", "true", "True", "ON", "on"):
            chans2 = getattr(ext, "_sim_monitor_channels", None)
            try:
                si_tl = int(panel_slot)
            except Exception:
                si_tl = 1
            # 재생(plan): 막대·포트는 renewal/heartbeat plan replay 전용 — timeline_only 가 sim_now 로 되돌리지 않게
            if bool(getattr(ext, "_sim_playback_started", False)):
                try:
                    from .control_sim_playback_plan import playback_plan_active, refresh_playback_display_at_sim

                    if playback_plan_active(ext, si_tl):
                        t_bar = float(str(payload.get("sim_time", "") or "0").strip() or "0")
                        try:
                            pl = get_sim_playback_player(ext, si_tl)
                            if pl is not None:
                                t_bar = float(pl.sim_now(si_tl))
                        except Exception:
                            pass
                        try:
                            refresh_playback_display_at_sim(ext, si_tl, float(t_bar))
                        except Exception:
                            pass
                        try:
                            chans_tl = getattr(ext, "_sim_monitor_channels", None)
                            if isinstance(chans_tl, list) and 0 < int(si_tl) <= len(chans_tl):
                                ch_tl = chans_tl[int(si_tl) - 1]
                                if isinstance(ch_tl, dict) and ch_tl.get("ep_timeline_widget") is None:
                                    occ_fb: Dict[str, Any] = {}
                                    try:
                                        last_by_fb = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                                        if isinstance(last_by_fb, dict) and isinstance(
                                            last_by_fb.get(str(si_tl)), dict
                                        ):
                                            occ_fb = dict(last_by_fb.get(str(si_tl)) or {})
                                    except Exception:
                                        pass
                                    _render_ep_bar_prerun_at_t(ext, ch_tl, float(t_bar), occ_fb)
                        except Exception:
                            pass
                        try:
                            t_foup = float(t_bar)
                            _refresh_foup_playback_heartbeat(ext, si_tl, t_foup)
                        except Exception:
                            pass
                        return
                except Exception:
                    pass
            # 포트상태 아래 전용 EP 타임라인을 대기 구간에도 전진시키기 위해
            # 마지막 ports_occupancy 스냅샷을 사용한다.
            try:
                last_by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                if not isinstance(last_by, dict):
                    last_by = {}
                    ext._sim_last_ports_occupancy_by_screen = last_by
                sk_occ = str(panel_slot)
                last_occ = last_by.get(sk_occ)
                if not isinstance(last_occ, dict):
                    # 시작 직후·첫 이벤트 전: 이벤트로 점유 스냅샷이 아직 없어도 EP 타임라인은 진행되어야 한다.
                    last_occ = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
                    last_by[sk_occ] = dict(last_occ)
            except Exception:
                last_occ = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
            try:
                sim_t = str(payload.get("sim_time", ""))
                last_occ = _occ_for_ep_timeline(
                    ext,
                    int(str(panel_slot or "1").strip() or "1"),
                    last_occ if isinstance(last_occ, dict) else {},
                    sim_t,
                    progress_p=payload if isinstance(payload, dict) else None,
                )
            except Exception:
                pass
            if isinstance(chans2, list) and len(chans2) > 1:
                try:
                    pslot_i = int(str(panel_slot or "1").strip() or "1")
                except Exception:
                    pslot_i = 1
                pslot_i = max(1, min(len(chans2), pslot_i))
                chp = chans2[pslot_i - 1]
                if isinstance(chp, dict):
                    sim_t = str(payload.get("sim_time", ""))
                    _update_ep_timeline_under_port_state(ext, chp, last_occ, sim_t)
            elif isinstance(chans2, list) and len(chans2) == 1:
                chp0 = chans2[0]
                if isinstance(chp0, dict):
                    sim_t = str(payload.get("sim_time", ""))
                    _update_ep_timeline_under_port_state(ext, chp0, last_occ, sim_t)
            if isinstance(chans2, list) and len(chans2) > 1:
                try:
                    _sync_all_ep_occ_timelines_from_engines(ext)
                except Exception:
                    pass
            if bool(getattr(ext, "_sim_playback_started", False)):
                try:
                    t_foup = float(str(payload.get("sim_time", "") or "0").strip() or "0")
                except Exception:
                    t_foup = 0.0
                try:
                    si_tl = int(str(panel_slot or "1").strip() or "1")
                except Exception:
                    si_tl = 1
                try:
                    _refresh_foup_playback_heartbeat(ext, si_tl, t_foup)
                except Exception:
                    pass
            return
    except Exception:
        pass
    if not label:
        return
    status = str(payload.get("status", "RUNNING"))
    percent = str(payload.get("percent", "0"))
    elapsed = str(payload.get("elapsed", "0.0"))
    total = str(payload.get("total", "0.0"))
    sim_time = str(payload.get("sim_time", "0.00"))
    detail = str(payload.get("detail", ""))
    # panel_slot 은 함수 상단 _resolve_payload_sim_screen 에서 확정
    event_seq = str(payload.get("event_seq") or payload.get("sequence_name") or "").strip()
    linked_anim = str(payload.get("linked_anim_json") or "").strip()
    try:
        st_p = build_payload_from_step(ext, int(panel_slot))
        if isinstance(st_p, dict):
            la = str(st_p.get("linked_anim_json") or "").strip()
            if la:
                linked_anim = la
    except Exception:
        pass
    proc_sec = str(payload.get("proc_sec", "")).strip()
    anim_sec = str(payload.get("anim_sec", "")).strip()
    proc_pri = str(payload.get("process_time_priority", "")).strip()
    ep_occ = payload.get("ep_occ", {})
    all_ep_empty = str(payload.get("all_ep_empty", "")).strip()

    anim_key = _sim_anim_status_key(ext)

    # FOUP 공정 진행은 EP 포트별로 "줄을 다르게 고정"한 라벨에만 갱신한다.
    # - 단일 라벨에 다른 EP/단계 텍스트가 덮어써져 깜빡이는 문제를 근본적으로 차단.
    # - payload["port_id"] (또는 detail/label 의 EPn 패턴) 으로 라우팅한다.
    # - -Y 단계 DONE 수신 시 1.05초 뒤 해당 EP 라벨만 "대기" 로 되돌린다(다른 EP 라벨은 무관).
    try:
        if str(event_seq or "").strip().upper() == "FOUP_PROCESS":
            chansf = getattr(ext, "_sim_monitor_channels", None)
            if isinstance(chansf, list) and len(chansf) > 0:
                try:
                    si = int(str(panel_slot or "1").strip() or "1")
                except Exception:
                    si = 1
                si = max(1, min(len(chansf), si))
                chf = chansf[si - 1] if isinstance(chansf[si - 1], dict) else None
            else:
                chf = None
            # EP 식별자 결정: payload.port_id → label/detail 안의 EPn 패턴 폴백
            ep_id = ""
            try:
                ep_id = str(payload.get("port_id", "") or "").strip().upper()
            except Exception:
                ep_id = ""
            if not ep_id:
                try:
                    import re as _re
                    src_txt = (str(payload.get("label", "") or "") + " " + str(payload.get("detail", "") or "")).upper()
                    m = _re.search(r"\bEP(\d+)\b", src_txt)
                    if m:
                        n = int(m.group(1))
                        if 1 <= n <= 3:
                            ep_id = f"EP{n}"
                except Exception:
                    pass
            try:
                lab_probe = str(payload.get("label", "") or "")
                det_probe = str(payload.get("detail", "") or "")
                if ep_id and ("공정" in lab_probe or "공정" in det_probe) and ("-Y" not in lab_probe):
                    by_f = getattr(ext, "_sim_foup_proc_active_ep_by_screen", None)
                    if not isinstance(by_f, dict):
                        by_f = {}
                        ext._sim_foup_proc_active_ep_by_screen = by_f
                    by_f[str(si)] = str(ep_id)
            except Exception:
                pass
            labels = (chf or {}).get("foup_progress_labels") if chf else None
            lbl = None
            if isinstance(labels, dict) and ep_id:
                lbl = labels.get(ep_id)
            if lbl is None and chf is not None:
                labels_fb = _foup_progress_labels_for_screen(ext, chf, si)
                if isinstance(labels_fb, dict) and ep_id:
                    lbl = labels_fb.get(ep_id)
            if lbl is None:
                # 폴백: 구조가 갱신되기 전이거나 EP 식별 실패 시 단일 라벨에라도 표시(원래 동작 유지)
                lbl = (chf or {}).get("foup_progress_label") if chf else None
            if lbl is not None:
                t_sim = str(payload.get("sim_time", "0.00"))
                st = str(payload.get("status", "RUNNING"))
                pct = str(payload.get("percent", "0"))
                el = str(payload.get("elapsed", "0.0"))
                tot = str(payload.get("total", "0.0"))
                lab = str(payload.get("label", "") or "").strip()
                det = str(payload.get("detail", "") or "").strip()
                screen_num = int((chf or {}).get("screen", si) or si)
                if bool(getattr(ext, "_sim_playback_started", False)) and ep_id:
                    try:
                        if str(payload.get("_foup_heartbeat_tick", "")).strip() not in (
                            "1",
                            "true",
                            "True",
                            "ON",
                            "on",
                        ):
                            _remember_foup_playback_progress(ext, si, ep_id, dict(payload))
                            st_r = str(st or "").strip().upper()
                            lab_r = str(lab or "").upper()
                            if st_r == "DONE" and "-Y" in lab_r:
                                _forget_foup_playback_progress(ext, si, ep_id)
                    except Exception:
                        pass
                head = (
                    f"{ep_id} FOUP 공정"
                    if (ep_id and screen_num == 1)
                    else (
                        f"{ep_id} FOUP 공정(화면{screen_num})"
                        if ep_id
                        else ("FOUP 공정" if screen_num == 1 else f"FOUP 공정(화면{screen_num})")
                    )
                )
                # 단계 표기(+Y/-Y/공정)는 detail 보다 label 이 더 짧고 깔끔
                stage = ""
                try:
                    if "+Y" in lab:
                        stage = "+Y"
                    elif "-Y" in lab:
                        stage = "-Y"
                    elif "공정" in lab or "공정" in det:
                        stage = "공정"
                except Exception:
                    stage = ""
                body = f"{head}: 진행 [{stage}]" if stage else f"{head}: 진행"
                try:
                    t_disp = f"{float(t_sim):.1f}"
                except Exception:
                    t_disp = str(t_sim)
                new_txt = f"{body} | {st} {pct}% ({el}/{tot}) | t={t_disp}"
                color = 0xFFFFE08A if str(st).upper() == "RUNNING" else 0xFF9FBFA0
                if chf is not None and ep_id:
                    _set_foup_progress_label(chf, ep_id, lbl, new_txt, {"color": color})
                else:
                    try:
                        lbl.text = new_txt
                        lbl.style = {"color": color}
                    except Exception:
                        pass
            # -Y(공정 종료 복귀) DONE 수신 시 해당 EP 만 1.05초 뒤 "대기" 로 되돌린다.
            try:
                lab_u = str(payload.get("label", "") or "").upper()
                if (
                    str(payload.get("status", "")).strip().upper() == "DONE"
                    and "-Y" in lab_u
                    and ep_id
                    and isinstance(labels, dict)
                    and labels.get(ep_id) is not None
                ):
                    _schedule_foup_label_reset(ext, si, ep_id, delay_sec=1.05)
            except Exception:
                pass
            return
    except Exception:
        pass
    # 화면별 마지막 진행현황 payload 저장(플레이백에서 DONE 상태에도 t(sim)을 부드럽게 갱신하기 위함)
    # playback_time_tick 은 표시용 sim_time 만 바꾸므로 lp 스냅샷을 덮어쓰지 않는다.
    if not _is_playback_time_tick_payload(payload if isinstance(payload, dict) else {}):
        try:
            by_lp = getattr(ext, "_sim_progress_last_payload_by_screen", None)
            if not isinstance(by_lp, dict):
                by_lp = {}
                ext._sim_progress_last_payload_by_screen = by_lp
            by_lp[str(panel_slot)] = dict(payload)
        except Exception:
            pass
    dedupe_key = f"_panel_{panel_slot}"
    # total_est는 포트상태 아래 전용 그래프에서도 사용하므로 화면별로 저장
    try:
        pslot_g = str(panel_slot)
        try:
            te = float(str(payload.get("sim_total_est_sec", "")).strip() or "0.0")
        except Exception:
            te = 0.0
        if te > 0.0:
            by = getattr(ext, "_sim_last_total_est_by_screen", None)
            if not isinstance(by, dict):
                by = {}
                ext._sim_last_total_est_by_screen = by
            by[str(pslot_g)] = float(te)
    except Exception:
        pass

    if status == "RUNNING":
        try:
            last_key = getattr(ext, "_sim_progress_last_key", None)
            step_id, disp_rev = progress_dedupe_extra(payload if isinstance(payload, dict) else {})
            force_ui = str(payload.get("_force_progress_ui", "")).strip().lower() in (
                "1",
                "true",
                "on",
            )
            # 진행현황 디듀프 키에는 실제 표시 문자열에 영향을 주는 값들을 포함해야 한다.
            # (proc_sec/anim_sec/priority가 바뀌었는데도 elapsed/total이 같으면 UI가 갱신되지 않는 문제 방지)
            key = (
                panel_slot,
                # heartbeat(첫 공정 전/대기/DONE 포함)에서 sim_time만 바뀌는 업데이트도 표시되어야 한다.
                str(sim_time),
                str(percent),
                str(elapsed),
                str(total),
                str(status),
                label,
                event_seq,
                linked_anim,
                anim_key,
                proc_sec,
                anim_sec,
                proc_pri,
                int(step_id),
                int(disp_rev),
                # 총 시간(총=XXXs)은 header에 직접 반영되므로 키에 포함
                str(payload.get("sim_total_est_sec", "") or "").strip(),
            )
            if (
                not force_ui
                and isinstance(last_key, dict)
                and last_key.get(dedupe_key) == key
            ):
                return
            if isinstance(last_key, dict):
                last_key[dedupe_key] = key
        except Exception:
            pass
    else:
        try:
            last_key = getattr(ext, "_sim_progress_last_key", None)
            if isinstance(last_key, dict):
                last_key.pop(dedupe_key, None)
        except Exception:
            pass

    head = "[진행현황] 단계 완료" if status == "DONE" else "[진행현황] 진행 중"
    ev_line = f"이벤트: {event_seq}\n" if event_seq else ""
    anim_footer = _format_progress_anim_footer(ext, payload if isinstance(payload, dict) else {})
    sec_line = ""
    if proc_sec or anim_sec:
        pri_txt = "ON" if proc_pri in ("1", "true", "True", "ON", "on") else "OFF"
        sec_line = f"시간: 공정={proc_sec or '-'}s | 애니={anim_sec or '-'}s | 공정시간우선={pri_txt}\n"
    te_txt = ""
    try:
        te_txt = str(payload.get("sim_total_est_sec", "") or "").strip()
    except Exception:
        te_txt = ""
    total_head = f" | 총={te_txt}s" if te_txt else ""
    text = (
        f"{head}{total_head} | t(sim)={sim_time}s\n"
        f"{ev_line}"
        f"{label}\n"
        f"{sec_line}"
        f"진행률: {percent}% ({elapsed} / {total}s)\n"
        f"{detail}\n"
        f"---\n"
        f"{anim_footer}"
    )
    try:
        ext._sim_progress_last_payload = dict(payload)
    except Exception:
        ext._sim_progress_last_payload = payload
    try:
        chans2 = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans2, list) and len(chans2) > 1:
            try:
                pslot_i = int(str(panel_slot or "1").strip() or "1")
            except Exception:
                pslot_i = 1
            pslot_i = max(1, min(len(chans2), pslot_i))
            chp = chans2[pslot_i - 1]
            if isinstance(chp, dict) and chp.get("progress_label") is not None:
                chp["progress_label"].text = text
                try:
                    _update_progress_ep_timeline_widget(ext, chp, payload if isinstance(payload, dict) else {})
                except Exception:
                    pass
        elif isinstance(chans2, list) and len(chans2) == 1:
            chp0 = chans2[0]
            if isinstance(chp0, dict) and chp0.get("progress_label") is not None:
                chp0["progress_label"].text = text
                try:
                    _update_progress_ep_timeline_widget(ext, chp0, payload if isinstance(payload, dict) else {})
                except Exception:
                    pass
    except Exception:
        pass
    try:
        chans3 = getattr(ext, "_sim_monitor_channels", None)
        if not (isinstance(chans3, list) and len(chans3) > 1) or str(panel_slot or "1").strip() in ("", "1"):
            ext._sim_progress_text.set_value(text)
    except Exception:
        try:
            ext._sim_progress_text.set_value(text)
        except Exception:
            pass
    if getattr(ext, "_sim_progress_label", None) is not None:
        ext._sim_progress_label.text = text


def _update_progress_ep_timeline_widget(ext: Any, ch: Dict[str, Any], payload: Dict[str, Any]) -> None:
    """
    진행현황 패널 하단: EP 점유 상태를 시뮬 시간 기준으로 누적 막대그래프로 표현한다.
    프리런 막대가 있으면 EP·ALL_EP 행만 동일 5상태 데이터로 표시한다.
    """
    host = ch.get("progress_ep_timeline_host")
    if host is None:
        return
    try:
        screen = int(ch.get("screen", 1))
    except Exception:
        screen = 1
    scr_key = str(screen)

    sim_time = None
    try:
        sim_time = float(str(payload.get("sim_time", "")).strip() or "0.0")
    except Exception:
        sim_time = None
    if sim_time is None:
        return

    pre_by = getattr(ext, "_sim_ep_bar_prerun_by_screen", None)
    bar_pre = pre_by.get(scr_key) if isinstance(pre_by, dict) else None
    use_pre = isinstance(bar_pre, EpBarPrecomputed)

    rows_state: Dict[str, List[Dict[str, Any]]] = {}
    rows: List[str] = []
    if use_pre and bar_pre is not None:
        rows_state = truncate_bar_rows_at_t(bar_pre.rows, float(sim_time))
        eps_pre = [r for r in (bar_pre.row_order or ()) if str(r).startswith("EP")]
        rows = (["ALL_EP"] if "ALL_EP" in (bar_pre.row_order or ()) else []) + list(eps_pre)
    else:
        st_by = getattr(ext, "_sim_ep_timeline_state_by_screen", None)
        if not isinstance(st_by, dict):
            st_by = {}
            ext._sim_ep_timeline_state_by_screen = st_by
        st = st_by.get(scr_key)
        if not isinstance(st, dict):
            st = {"t_last": None, "rows": {}}
            st_by[scr_key] = st
        t_last = st.get("t_last", None)
        st["t_last"] = sim_time
        if t_last is None:
            return
        dt = max(0.0, float(sim_time) - float(t_last))
        if dt <= 1e-9:
            return

        occ_raw = payload.get("ports_occupancy", {})
        if not isinstance(occ_raw, dict):
            occ_raw = {}
        occ_eff = _occ_for_ep_timeline(
            ext,
            screen,
            occ_raw,
            f"{sim_time:.2f}",
            progress_p=payload if isinstance(payload, dict) else None,
        )

        eps: List[str] = []
        ep_ports = payload.get("ep_ports", [])
        if isinstance(ep_ports, list) and ep_ports:
            eps = [str(x).strip().upper() for x in ep_ports if str(x).strip().upper().startswith("EP")]
        if not eps:
            eps = [str(k).strip().upper() for k in occ_eff.keys() if str(k).strip().upper().startswith("EP")]
        eps = sorted(eps, key=lambda x: int(str(x).upper().replace("EP", "") or "0"))
        if not eps:
            eps = ["EP1", "EP2"]
        rows = ["ALL_EP"] + list(eps)

        rows_state = st.get("rows", {})
        if not isinstance(rows_state, dict):
            rows_state = {}
            st["rows"] = rows_state

        snap = _sim_snapshot_for_screen(ext, screen)
        ep_count = 3 if int(_ep_count_idx_for_port_panel(ext, screen)) else 2
        fault_ports = _fault_ports_from_snapshot(snap, ep_count) if snap else set()

        def _live_ep_state(ep: str, occ_disp: Dict[str, Any]) -> str:
            p = str(ep or "").strip().upper()
            if p in fault_ports:
                return BAR_STATE_DOWN
            if not bool(str(occ_disp.get(p, "") or "").strip()):
                return BAR_STATE_EMPTY
            try:
                by_f = getattr(ext, "_sim_foup_proc_active_ep_by_screen", None)
                ap = str((by_f or {}).get(scr_key, "") or "").strip().upper() if isinstance(by_f, dict) else ""
                if ap == p:
                    return BAR_STATE_PROC
            except Exception:
                pass
            return BAR_STATE_LOAD

        def _push(row: str, state: str, dur: float) -> None:
            segs = rows_state.get(row)
            if not isinstance(segs, list):
                segs = []
                rows_state[row] = segs
            st_seg = str(state or BAR_STATE_EMPTY)
            if segs and isinstance(segs[-1], dict) and bar_state_from_seg(segs[-1]) == st_seg:
                segs[-1]["dur"] = float(segs[-1].get("dur", 0.0)) + float(dur)
            else:
                segs.append({"state": st_seg, "dur": float(dur)})
            if len(segs) > 220:
                del segs[:-200]

        for r in rows:
            if r not in rows_state or not isinstance(rows_state.get(r), list):
                rows_state[r] = []

        t_cursor = float(t_last)
        for dt_part, occ_part in interval_occ_parts(occ_raw, payload, float(t_last), float(sim_time)):
            if dt_part <= 1e-9:
                continue
            t_cursor += float(dt_part)
            occ_disp = _occ_for_ep_timeline(
                ext,
                screen,
                occ_part if isinstance(occ_part, dict) else {},
                f"{t_cursor:.6f}",
                progress_p=payload if isinstance(payload, dict) else None,
            )
            ep_states = [_live_ep_state(ep, occ_disp) for ep in eps]
            all_ep_st = _aggregate_all_ep_state(ep_states)
            for ep, ep_st in zip(eps, ep_states):
                _push(ep, ep_st, dt_part)
            _push("ALL_EP", all_ep_st, dt_part)

    old = ch.get("progress_ep_timeline_widget", None)
    if old is not None:
        try:
            old.destroy()
        except Exception:
            pass
        ch["progress_ep_timeline_widget"] = None

    # 요구사항: 막대 전체 길이를 "총 시뮬레이션 시간(예상)"으로 고정하고,
    # 그 안에서 비율만큼 채워지게 한다(슬라이딩 윈도우 금지).
    BAR_W = 320         # 진행현황 컬럼 기본 폭(창 너비에 맞게 짧게)
    BAR_H = 14
    NAME_W = 56

    t_end = float(sim_time)
    # total_est가 없거나 0이면(리셋/초기화 상태) "총시간 라벨"은 표시하지 않는다.
    # 단, 막대 스케일 계산을 위해 내부 total_est는 최소값(10s)을 사용한다.
    _total_raw = 0.0
    if use_pre and bar_pre is not None and float(bar_pre.total_est) > 0.0:
        _total_raw = float(bar_pre.total_est)
    else:
        try:
            _total_raw = float(str(payload.get("sim_total_est_sec", "")).strip() or "0.0")
        except Exception:
            _total_raw = 0.0
    show_total_label = bool(isinstance(_total_raw, (float, int)) and float(_total_raw) > 0.0)
    total_est = max(10.0, float(_total_raw) if show_total_label else 0.0)
    if total_est <= 0.0:
        total_est = 10.0
    t_start = 0.0

    # 라벨이 너무 많으면(폭이 1px) 안 보이므로, 화면에 보일 만큼만 샘플링한다.
    # 목표: 최대 7~9개 정도만 표시.
    try:
        max_labels = 8
        raw_step = max(10.0, float(total_est) / float(max_labels))
        # 10초 단위로 올림
        tick_step = float(int(((raw_step + 9.999) // 10.0) * 10.0))
    except Exception:
        tick_step = 10.0
    tick_step = max(10.0, tick_step)

    def _seg_color(state: str) -> int:
        return bar_state_color(str(state or BAR_STATE_EMPTY))

    with host:
        ch["progress_ep_timeline_widget"] = ui.VStack(spacing=4)

        # 상단 눈금(가독성 위해 최대 8개 내)
        with ui.HStack(height=12, spacing=0):
            ui.Spacer(width=NAME_W)
            with ui.ZStack(width=BAR_W, height=12):
                ui.Rectangle(width=BAR_W, height=12, style={"background_color": 0x441A1E26})
                # 막대 끝(우측)에 총 시뮬 시간을 표시(리셋/초기 상태에서는 숨김)
                if show_total_label:
                    try:
                        with ui.Placer(offset_x=max(0, BAR_W - 72), offset_y=0):
                            try:
                                _end_txt = (
                                    f"{int(round(total_est))}s"
                                    if abs(float(total_est) - float(int(round(total_est)))) < 1e-6
                                    else f"{float(total_est):.1f}s"
                                )
                            except Exception:
                                _end_txt = f"{total_est:.1f}s"
                            ui.Label(
                                _end_txt,
                                width=72,
                                height=12,
                                alignment=ui.Alignment.RIGHT_CENTER,
                                style={"color": 0xFFBFC7D5, "font_size": 10},
                            )
                    except Exception:
                        pass
                # 눈금 라벨을 절대좌표로 배치(겹침/폭 축소로 안 보이는 문제 방지)
                try:
                    ticks = int(total_est // tick_step)
                    ticks = max(1, ticks)
                except Exception:
                    ticks = 1
                for i in range(ticks + 1):
                    t_lbl = int(round(t_start + i * tick_step))
                    x = int(round((float(t_lbl) / float(total_est)) * float(BAR_W)))
                    x = max(0, min(BAR_W - 1, x))
                    with ui.Placer(offset_x=x, offset_y=0):
                        ui.Label(
                            f"{t_lbl}",
                            width=36,
                            height=12,
                            style={"color": 0xFFE0E6F0, "font_size": 10},
                        )
                # 마지막 총시간 라벨(정확값)을 반드시 끝에 추가(50단위만 보이는 문제 방지)
                try:
                    t_end_lbl = float(total_est)
                    try:
                        _end_tick_txt = (
                            f"{int(round(t_end_lbl))}"
                            if abs(float(t_end_lbl) - float(int(round(t_end_lbl)))) < 1e-6
                            else f"{float(t_end_lbl):.1f}"
                        )
                    except Exception:
                        _end_tick_txt = f"{t_end_lbl:.1f}"
                    with ui.Placer(offset_x=max(0, BAR_W - 36), offset_y=0):
                        ui.Label(
                            _end_tick_txt,
                            width=36,
                            height=12,
                            alignment=ui.Alignment.RIGHT_CENTER,
                            style={"color": 0xFFE0E6F0, "font_size": 10},
                        )
                except Exception:
                    pass

        for row in rows:
            with ui.HStack(height=BAR_H, spacing=6):
                ui.Label(str(row), width=NAME_W, height=BAR_H, style={"color": 0xFFBFC7D5})
                with ui.ZStack(width=BAR_W, height=BAR_H):
                    ui.Rectangle(width=BAR_W, height=BAR_H, style={"background_color": 0xFF1A1E26})
                    # 막대 세그먼트
                    segs = rows_state.get(row, [])
                    if not isinstance(segs, list):
                        segs = []
                    rects = _bar_segment_rect_widths(
                        segs,
                        total_est=float(total_est),
                        bar_w=int(BAR_W),
                        t_cover=float(sim_time),
                    )
                    with ui.HStack(height=BAR_H, spacing=0):
                        used = 0
                        for w, seg_st in rects:
                            used += int(w)
                            ui.Rectangle(
                                width=int(w),
                                height=BAR_H,
                                style={"background_color": _seg_color(seg_st)},
                            )
                        if used < BAR_W:
                            ui.Spacer(width=(BAR_W - used))


def _on_sim_event(ext: Any, payload: Dict[str, str]) -> None:
    seq_raw = (payload.get("seq") or "").strip()
    if not seq_raw:
        return
    seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
    lot_id = lot_id_from_payload(payload)
    sim_time = payload.get("sim_time", "")

    try:
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = _parse_port_num(str(payload.get("from_port_id", "1")), 1)
            to_port = _parse_port_num(str(payload.get("to_port_id", "1")), 1)
            xml_text = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port = _parse_port_num(str(payload.get("port_id", "1")), 1)
            xml_text = xml_generator.build_xml_string(seq, port_id=port)
        ext._last_generated_xml = xml_text
        parsed = xml_generator.parse_xml_string(xml_text) or {}
        story = f"[SIM EVENT t={sim_time}] seq={seq_raw}->{seq} lot={lot_id} port={payload.get('port_id','')} from={payload.get('from_port_id','')} to={payload.get('to_port_id','')}"
        if parsed.get("action_desc"):
            story += f" | action={parsed.get('action_desc')}"
        _append_sim_log(ext, story)
    except Exception as e:
        _append_sim_log(ext, f"[SIM EVENT] XML 생성/역파싱 실패: seq={seq}, err={e}")


def _parse_port_num(port_text: str, default_value: int = 1) -> int:
    """
    내부 포트 텍스트를 EAPEIS 포트 ID로 변환한다.
    매핑 규칙:
    - OHT -> 10 (MOVE FROM 가상 포트; EP/BP/INOUT과 충돌 없음)
    - EP1/2/3 -> 1/2/3
    - INOUT(=IN/OUT) -> 5
    - BP1/2/3/4 -> 6/7/8/9
    """
    txt = (port_text or "").strip().upper()
    if not txt:
        return default_value
    if txt in ("INOUT", "IN/OUT"):
        return 5
    if txt.startswith("OHT"):
        return 10
    if txt.startswith("EP"):
        try:
            n = int(txt.replace("EP", ""))
            if 1 <= n <= 3:
                return n
        except Exception:
            return default_value
    if txt.startswith("BP"):
        try:
            n = int(txt.replace("BP", ""))
            if 1 <= n <= 4:
                return 5 + n
        except Exception:
            return default_value
    if txt.startswith("PORT_"):
        txt = txt.replace("PORT_", "")
    try:
        return int(txt)
    except Exception:
        return default_value


def _port_kind(port_text: str) -> str:
    t = (port_text or "").strip().upper()
    if t.startswith("BP"):
        return "버퍼포트(BP)"
    if t.startswith("EP"):
        return "공정포트(EP)"
    if t.startswith("OHT"):
        return "이송장치(OHT)"
    return "미확인"


def _normalize_port_text_from_xml(parsed_val: str, original_text: str) -> str:
    """
    XML 역파싱 값(parsed_val)을 기준으로 포트 문자열을 표준화한다.
    - 원본이 BP/EP 접두사를 가지고 있으면 같은 접두사를 유지
    - 없으면 XML 숫자값 그대로 사용
    """
    p = (parsed_val or "").strip()
    if not p:
        return ""
    o = (original_text or "").strip().upper()
    try:
        n = int(p)
    except Exception:
        n = None
    if o.startswith("BP"):
        # XML ID 6~9는 BP1~4에 대응
        if n is not None and 6 <= n <= 9:
            return f"BP{n - 5}"
        return f"BP{p}"
    if o.startswith("EP"):
        # XML ID 1~3은 EP1~3에 대응
        if n is not None and 1 <= n <= 3:
            return f"EP{n}"
        return f"EP{p}"
    if o.startswith("INOUT") or o.startswith("IN/OUT"):
        return "INOUT"
    if o.startswith("OHT"):
        return "OHT"
    if n is not None:
        if 1 <= n <= 3:
            return f"EP{n}"
        if n == 5:
            return "INOUT"
        if 6 <= n <= 9:
            return f"BP{n - 5}"
    return p


def _set_foup_progress_label(ch: Dict[str, Any], ep_id: str, lbl: Any, text: str, style: Dict[str, Any]) -> None:
    """동일 텍스트·스타일이면 FOUP 라벨을 다시 쓰지 않아 깜빡임을 줄인다."""
    if lbl is None:
        return
    try:
        cache = ch.get("_foup_label_cache")
        if not isinstance(cache, dict):
            cache = {}
            ch["_foup_label_cache"] = cache
        key = str(ep_id or "").strip().upper()
        sig = (str(text or ""), int(style.get("color", 0)) if isinstance(style, dict) else 0)
        if cache.get(key) == sig:
            return
        cache[key] = sig
        lbl.text = str(text or "")
        try:
            cur_col = int(getattr(lbl, "style", {}).get("color", 0))  # type: ignore[union-attr]
        except Exception:
            cur_col = None
        new_col = int(style.get("color", 0)) if isinstance(style, dict) else 0
        if cur_col != new_col:
            lbl.style = dict(style or {})
    except Exception:
        try:
            lbl.text = str(text or "")
            lbl.style = dict(style or {})
        except Exception:
            pass


def _reset_foup_label_now(ext: Any, screen_idx: int, ep_id: str) -> None:
    """지정된 화면(screen_idx, 1-based) / EP 라벨 한 줄을 즉시 'idle(대기)' 표시로 되돌린다."""
    try:
        chans = getattr(ext, "_sim_monitor_channels", None) or []
        if not isinstance(chans, list) or not chans:
            return
        si = max(1, min(len(chans), int(screen_idx or 1)))
        chf = chans[si - 1] if isinstance(chans[si - 1], dict) else None
        if not chf:
            return
        labels = chf.get("foup_progress_labels") or {}
        lbl = labels.get(str(ep_id or "").strip().upper()) if isinstance(labels, dict) else None
        if lbl is None:
            return
        screen_num = int(chf.get("screen", si) or si)
        _set_foup_progress_label(
            chf,
            str(ep_id),
            lbl,
            _foup_label_idle_text(screen_num, str(ep_id)),
            {"color": 0xFF888888},
        )
    except Exception:
        pass


def _reset_all_foup_labels(ext: Any) -> None:
    """모든 화면/EP 의 FOUP 라벨을 'idle(대기)' 표시로 일괄 리셋(Reset/Stop 안전망용)."""
    try:
        chans = getattr(ext, "_sim_monitor_channels", None) or []
        if not isinstance(chans, list):
            return
        for ch in chans:
            if not isinstance(ch, dict):
                continue
            try:
                screen_num = int(ch.get("screen", 1) or 1)
            except Exception:
                screen_num = 1
            labels = ch.get("foup_progress_labels") or {}
            if not isinstance(labels, dict):
                continue
            for ep_id, lbl in labels.items():
                if lbl is None:
                    continue
                _set_foup_progress_label(
                    ch,
                    str(ep_id),
                    lbl,
                    _foup_label_idle_text(screen_num, str(ep_id)),
                    {"color": 0xFF888888},
                )
            try:
                ch["_foup_label_cache"] = {}
            except Exception:
                pass
    except Exception:
        pass


def _schedule_foup_label_reset(ext: Any, screen_idx: int, ep_id: str, delay_sec: float = 1.05) -> None:
    """
    FOUP_PROCESS 의 -Y(공정 종료 복귀) DONE 시점에서 ``delay_sec`` 후
    해당 화면/EP 라벨만 'idle(대기)' 로 되돌리는 1회성 update event subscription 을 등록한다.

    동작:
    - opts/A 의 ``_schedule_foup_inprogress_unmark`` 와 동일한 패턴(1회성 sub, deadline 기반).
    - subscription 객체는 GC 방지를 위해 ``ext._foup_label_reset_subs`` 에 보관.
    - Stop/Reset 시 해당 리스트도 일괄 정리(아래 on_sim_stop_clicked 안전망 참조).
    """
    try:
        import time as _t
        deadline = _t.monotonic() + max(0.0, float(delay_sec))
        sub_holder: Dict[str, Any] = {"sub": None, "done": False}

        def _on_update(_ev):
            try:
                if sub_holder.get("done"):
                    return
                if _t.monotonic() < deadline:
                    return
                sub_holder["done"] = True
                try:
                    _reset_foup_label_now(ext, int(screen_idx), str(ep_id))
                except Exception:
                    pass
                try:
                    s = sub_holder.get("sub")
                    if s is not None:
                        s.unsubscribe()
                except Exception:
                    pass
                sub_holder["sub"] = None
            except Exception:
                pass

        sub_holder["sub"] = app.get_app().get_update_event_stream().create_subscription_to_pop(
            _on_update, name="morph.tbs_control_2.foup_label_reset"
        )
        try:
            holders = getattr(ext, "_foup_label_reset_subs", None)
            if not isinstance(holders, list):
                holders = []
                ext._foup_label_reset_subs = holders
            holders.append(sub_holder)
            try:
                ext._foup_label_reset_subs = [h for h in holders if not h.get("done")]
            except Exception:
                pass
        except Exception:
            pass
    except Exception:
        try:
            _reset_foup_label_now(ext, int(screen_idx), str(ep_id))
        except Exception:
            pass


def _schedule_foup_inprogress_unmark(
    ext: Any,
    prim_path: str,
    delay_sec: float = 1.05,
    *,
    usd_context_name: Optional[str] = None,
) -> None:
    """
    FOUP_PROCESS_END 가 발생하면 -Y 복귀 애니가 끝나는 시점(약 1초 후)에
    port_lot_visibility 의 FOUP 진행중 표시를 (해당 화면 컨텍스트에서) 해제한다.

    동작:
    - 1회성 update event subscription 으로 deadline 도달 시 unmark + self-unsubscribe.
    - subscription 객체가 GC 로 사라지지 않게 ext._foup_unmark_subs 에 보관한다.
    - 어떤 단계에서 실패하더라도, 마지막 안전망으로 즉시 unmark 호출(보호가 너무 길게 남는 것 방지).
    """
    try:
        from . import port_lot_visibility  # type: ignore
    except Exception:
        return
    p = str(prim_path or "").strip()
    if not p:
        return
    try:
        import time as _t
        deadline = _t.monotonic() + max(0.0, float(delay_sec))
        sub_holder: Dict[str, Any] = {"sub": None, "done": False}

        def _on_update(_ev):
            try:
                if sub_holder.get("done"):
                    return
                if _t.monotonic() < deadline:
                    return
                sub_holder["done"] = True
                try:
                    port_lot_visibility.mark_foup_in_progress(
                        p, False, usd_context_name=usd_context_name
                    )
                except Exception:
                    pass
                try:
                    s = sub_holder.get("sub")
                    if s is not None:
                        s.unsubscribe()
                except Exception:
                    pass
                sub_holder["sub"] = None
            except Exception:
                pass

        sub_holder["sub"] = app.get_app().get_update_event_stream().create_subscription_to_pop(
            _on_update, name="morph.tbs_control_2.foup_unmark"
        )
        try:
            holders = getattr(ext, "_foup_unmark_subs", None)
            if not isinstance(holders, list):
                holders = []
                ext._foup_unmark_subs = holders
            holders.append(sub_holder)
            # 다 끝난 항목 정리(메모리 누수 방지)
            try:
                ext._foup_unmark_subs = [h for h in holders if not h.get("done")]
            except Exception:
                pass
        except Exception:
            pass
    except Exception:
        try:
            port_lot_visibility.mark_foup_in_progress(
                p, False, usd_context_name=usd_context_name
            )
        except Exception:
            pass


def handle_sim_event_for_animation(ext: Any, payload: Dict[str, str], verbose: bool = True) -> None:
    """
    시뮬레이션 이벤트 → 애니메이션 실행 훅.

    이 함수가 하는 일(전체 흐름 한눈에):
    1) FOUP 공정 전용 이벤트(FOUP_PROCESS_START/END) → 별도 분기에서 prim Y축 ±이동만 실행 후 종료
    2) 그 외 일반 이벤트 → "이벤트 → XML 빌드 → 역파싱 → rules/map 매칭 → JSON 시퀀스 실행" 파이프라인
    3) 디버그 로그는 verbose 모드에서만 출력

    호출 경로(요약):
    - simulation_engine._emit_event(...) → ANIM_EVENT 큐 → _sim_ui_sink_anim_event(...) → 본 함수
    """

    # 0) 시퀀스 식별자(seq) 가져오기
    #    - 비어 있으면 처리할 게 없으므로 즉시 종료
    seq_raw = (payload.get("seq") or "").strip()
    if not seq_raw:
        return

    # ─────────────────────────────────────────────────────────────────────
    # 1) FOUP 공정 전용 분기 (FOUP_PROCESS_START / FOUP_PROCESS_END)
    #    - XML/매핑/JSON 시퀀스 파이프라인을 "타지 않는다".
    #    - port_lot_prim_paths.json 매핑으로 EP 포트의 FOUP prim 경로를 찾고,
    #      START 시 +Y320, END 시 -Y320 만큼 1초 동안 이동시킨다.
    #    - 분할화면(보조 USD 컨텍스트)도 고려한다(ctx_nm).
    # ─────────────────────────────────────────────────────────────────────
    try:
        seq_u0 = str(seq_raw or "").strip().upper()
    except Exception:
        seq_u0 = ""
    if seq_u0 in ("FOUP_PROCESS_START", "FOUP_PROCESS_END"):
        # 1-A) 어떤 EP 포트의 FOUP 인지 확인
        #     - port_id 가 없으면 매핑할 prim 을 못 찾으므로 종료
        try:
            port_id = str(payload.get("port_id", "") or "").strip().upper()
        except Exception:
            port_id = ""
        if not port_id:
            return
        # 1-B) port_lot_prim_paths.json → 해당 EP 포트의 FOUP prim 경로 조회
        try:
            from . import port_lot_visibility  # type: ignore
            mp = port_lot_visibility.load_port_lot_prim_paths() or {}
        except Exception:
            mp = {}
        prim_path = str((mp or {}).get(port_id, "") or "").strip()
        if not prim_path:
            if verbose:
                print(f"[FOUP] prim path missing: port={port_id}", flush=True)
            return
        # 1-C) 분할 화면별 USD 컨텍스트 결정(어떤 화면의 stage 위에서 움직일지)
        scr_opt_f = _resolve_payload_sim_screen(ext, payload if isinstance(payload, dict) else {})
        if scr_opt_f is None:
            try:
                print(
                    f"[TBS/port-screen] drop FOUP (no tbs_sim_screen) seq={seq_u0}",
                    flush=True,
                )
            except Exception:
                pass
            return
        scr = int(scr_opt_f)
        ctx_nm = _usd_context_name_for_sim_screen(ext, scr)
        # 1-D) FOUP 진행중 보호 마킹(옵션 A):
        #     - START: 즉시 mark(True) → 다른 시퀀스가 중간에 시작해도 baseline 복원에서 제외되어
        #              +Y320 오프셋이 유지된다.
        #     - END: -Y320 복귀 애니(약 1초)가 끝난 시점에 mark(False) 로 해제(지연 unmark).
        #            (END 직후 즉시 해제하면 -Y 복귀 도중 다른 시퀀스의 restore 가 prim 을
        #             baseline 으로 스냅시켜 시각적 점프가 발생할 수 있어 1초 지연한다.)
        try:
            from . import port_lot_visibility as _plv  # type: ignore
            if seq_u0 == "FOUP_PROCESS_START":
                _plv.mark_foup_in_progress(prim_path, True, usd_context_name=ctx_nm)
                _remember_foup_active_ep(ext, scr, {"foup_proc_active_ep": port_id})
                try:
                    runners = getattr(ext, "_sim_runners_by_screen", None)
                    if isinstance(runners, dict):
                        r = runners.get(str(scr)) or runners.get(scr)
                        if r is not None:
                            r._foup_proc_active_ep = port_id
                    r0 = getattr(ext, "_sim_runner", None)
                    if r0 is not None and scr == 1:
                        r0._foup_proc_active_ep = port_id
                except Exception:
                    pass
            else:  # FOUP_PROCESS_END
                _plv.mark_foup_lifted(prim_path, False, usd_context_name=ctx_nm)
                _remember_foup_active_ep(ext, scr, {"foup_proc_active_ep": ""})
                _schedule_foup_inprogress_unmark(
                    ext, prim_path, delay_sec=1.05, usd_context_name=ctx_nm
                )
        except Exception:
            pass
        # 1-D-2) Material 바인딩(요청 사양):
        #     - START: 공정 진행 중 material(MATERIAL_PATH_FOUP_PROCESSING, 기본 CASE_02)
        #     - END  : 공정 종료(회수 대기) material(MATERIAL_PATH_FOUP_DONE, 기본 CASE_03)
        #     - 회수(REMOVED) 후 “포트상태 초기화 시점”에 phong1 로 복귀하는 처리는
        #       port_lot_visibility.apply_port_lot_prim_visibility_for_context 의
        #       has_lot=False 분기에서 일괄 처리한다.
        #     - 분할화면(보조 USD 컨텍스트)을 고려해 ctx_nm 을 그대로 전달.
        try:
            from . import port_lot_visibility as _plv  # type: ignore
            mat_path = (
                _plv.MATERIAL_PATH_FOUP_PROCESSING
                if seq_u0 == "FOUP_PROCESS_START"
                else _plv.MATERIAL_PATH_FOUP_DONE
            )
            _plv.apply_port_lot_prim_material_for_context(ctx_nm, prim_path, mat_path)
        except Exception:
            pass
        # 1-E) START/END: 현재 Y → 목표 Y 까지 1초 부드럽게 이동(이미 도달 시 생략).
        try:
            from . import port_lot_visibility as _plv_anim  # type: ignore

            if seq_u0 == "FOUP_PROCESS_START":
                _plv_anim.run_foup_smooth_y_anim(
                    prim_path,
                    usd_context_name=ctx_nm,
                    toward_lifted=True,
                    foup_proc_active_ep=port_id,
                )
            else:
                _plv_anim.run_foup_smooth_y_anim(
                    prim_path,
                    usd_context_name=ctx_nm,
                    toward_lifted=False,
                )
        except Exception:
            pass
        # 1-F) FOUP 분기는 여기서 종료(매핑/JSON 파이프라인을 타지 않는다)
        return

    # ─────────────────────────────────────────────────────────────────────
    # 2) 일반 이벤트(ARRIVED/MOVE_TRANSFERING/MOVE_REQ/REMOVED 등) 처리
    # ─────────────────────────────────────────────────────────────────────

    # 2-A) seq 별칭 정규화(내부 alias → 표준 시퀀스명)
    seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
    # 2-B) payload 에서 자주 쓰는 필드 미리 추출(가독성/로그용)
    sim_time = payload.get("sim_time", "")
    lot_id = lot_id_from_payload(payload)
    from_port_txt = str(payload.get("from_port_id", ""))
    to_port_txt = str(payload.get("to_port_id", ""))
    port_txt = str(payload.get("port_id", ""))
    from_kind = _port_kind(from_port_txt)
    to_kind = _port_kind(to_port_txt)
    port_kind = _port_kind(port_txt)

    # 2-C) 디버그 한 줄 로그(verbose 모드)
    if verbose:
        print(
            f"[ANIM HOOK t={sim_time}] 이벤트={seq_raw}->{seq} lot={lot_id} "
            f"port={port_txt}({port_kind}) from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind})",
            flush=True,
        )

    # 2-D) 정책상 애니메이션이 없는 이벤트는 즉시 종료
    #     - READYTOLOAD/READYTOUNLOAD 는 "준비됨" 알림 성격이라 연계 애니가 없다.
    #     - rules/map 에 잘못 남아 있어도 여기서 차단(과실행 방지)
    try:
        if str(seq).strip().upper() in (
            str(xml_generator.SEQ_READYTOLOAD).strip().upper(),
            str(xml_generator.SEQ_READYTOUNLOAD).strip().upper(),
        ):
            if verbose:
                print(f"[ANIM HOOK] no-anim event skip: {seq}", flush=True)
            return
    except Exception:
        pass

    # 2-E) 주 실행 파이프라인:
    #      이벤트 → XML 생성 → 역파싱 → 표준화된 mapping_payload → rules/map 매칭 → JSON 시퀀스 실행
    #
    #      유지보수 규칙(중요):
    #      - rules/map 매칭 입력은 "반드시" XML 역파싱 결과를 기준으로 표준화한다.
    #      - 시뮬 payload 원본을 바로 매칭에 쓰지 않는다(포맷 드리프트 방지).
    parsed: Dict[str, Any] = {}
    xml_text = ""
    seq_for_mapping = seq
    try:
        # 2-E-1) 이벤트 종류에 따라 from/to 또는 port 만으로 XML 텍스트 생성
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = _parse_port_num(from_port_txt, 1)
            to_port = _parse_port_num(to_port_txt, 1)
            xml_text = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port = _parse_port_num(port_txt, 1)
            xml_text = xml_generator.build_xml_string(seq, port_id=port)
        # 2-E-2) XML 을 다시 dict 로 역파싱하여 표준 sequence_name 추출
        parsed = xml_generator.parse_xml_string(xml_text) or {}
        parsed_seq = str(parsed.get("sequence_name", "")).strip().upper()
        if parsed_seq:
            # 2-E-3) XML 이 알려주는 정식 sequence 를 매칭 키로 최우선 채택
            seq_for_mapping = parsed_seq
    except Exception as e:
        if verbose:
            print(f"[ANIM HOOK] XML 생성/역파싱 실패: seq={seq}, err={e}", flush=True)
        return

    # 2-F) rules/map 매칭 입력 표준화
    #     - 원본 payload 복사본에 XML 역파싱 결과로 from/to/port/seq 를 정규화해서 덮어쓴다.
    #     - 원본 텍스트의 접두사(BP/EP/OHT) 힌트는 _normalize_port_text_from_xml 가 보존한다.
    mapping_payload = dict(payload or {})
    parsed_from = str(parsed.get("from_port_id", "") or "")
    parsed_to = str(parsed.get("to_port_id", "") or "")
    parsed_port = str(parsed.get("port_id", "") or "")
    mapping_payload["seq"] = seq_for_mapping
    mapping_payload["from_port_id"] = _normalize_port_text_from_xml(parsed_from, from_port_txt)
    mapping_payload["to_port_id"] = _normalize_port_text_from_xml(parsed_to, to_port_txt)
    mapping_payload["port_id"] = _normalize_port_text_from_xml(parsed_port, port_txt)
    mapping_payload["_xml_sequence_name"] = seq_for_mapping
    mapping_payload["_xml_text"] = xml_text
    mapping_payload["_xml_parsed"] = parsed

    # 2-G) 매칭(이벤트 → JSON 경로)
    #     - rules.json 우선 → fallback map.json 순으로 _resolve_event_animation_entry 가 결정
    #     - 매칭 성공 시 _execute_mapped_sequence_stub 가 SequenceRunner.run(...) 까지 트리거한다.
    mapped_json, mapped_meta, matched_rule, matched_source = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
    if mapped_json:
        _append_anim_history_log(
            ext,
            f"[ANIM MAP] source={matched_source or '-'} rule={matched_rule or '-'} event={seq_for_mapping} file={Path(str(mapped_json)).name}",
        )
        _execute_mapped_sequence_stub(ext, seq_for_mapping, mapping_payload, mapped_json, mapped_meta, matched_rule, verbose)
    elif verbose:
        # 매칭 없음: 콘솔에 어느 파일을 확인해야 하는지 힌트 출력
        print(
            f"[ANIM MAP] 이벤트={seq_for_mapping} 매핑 없음 "
            f"(config/event_animation_rules.json / event_animation_map.json 확인)",
            flush=True,
        )
    # 2-H) 매칭이 없으면 이력 로그(애니 히스토리) 한 줄 추가(필요 예시 파일명 추정)
    if not mapped_json:
        hint_name = f"{seq_for_mapping.lower()}.json".replace("eapeis_port_", "")
        _append_anim_history_log(
            ext,
            f"[ANIM] 매핑없음 | event={seq_for_mapping} | 필요한 예시파일={hint_name}",
        )
    # 2-I) 이후 출력에서 쓰는 텍스트도 XML 표준화 결과로 갱신(일관성 유지)
    from_port_txt = str(mapping_payload.get("from_port_id", ""))
    to_port_txt = str(mapping_payload.get("to_port_id", ""))
    port_txt = str(mapping_payload.get("port_id", ""))
    from_kind = _port_kind(from_port_txt)
    to_kind = _port_kind(to_port_txt)
    port_kind = _port_kind(port_txt)

    # 2-J) action_desc 가 있으면 디버그용 액션 한 줄 출력
    action_desc = parsed.get("action_desc", "")
    if action_desc and verbose:
        if port_txt:
            action_desc += f" | 대상포트={port_txt}({port_kind})"
        if from_port_txt or to_port_txt:
            action_desc += f" | 이동경로={from_port_txt}({from_kind})->{to_port_txt}({to_kind})"
        print(f"[ANIM HOOK ACTION] {action_desc}", flush=True)

    # 2-K) 시퀀스 종류별 디버그 분기
    #     - 실제 애니 실행은 위 _execute_mapped_sequence_stub 에서 이미 끝났고,
    #       여기서는 “어떤 종류의 이벤트가 통과했는지” 한 줄 로그만 남긴다.
    if seq_for_mapping == xml_generator.SEQ_READYTOLOAD:
        if verbose:
            print(f"[ANIM PLAN] READY_TO_LOAD 대기 상태 애니메이션 | port={port_txt}({port_kind})", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_ARRIVED:
        if verbose:
            print(f"[ANIM PLAN] ARRIVED 안착 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_MOVE_TRANSFERING:
        if verbose:
            print(f"[ANIM PLAN] MOVE_TRANSFERING 이송 애니메이션 | from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_MOVE:
        if verbose:
            print(f"[ANIM PLAN] MOVE 이동 애니메이션 | from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_READYTOUNLOAD:
        if verbose:
            print(f"[ANIM PLAN] READY_TO_UNLOAD 회수 준비 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_REMOVED:
        if verbose:
            print(f"[ANIM PLAN] REMOVED 회수 완료 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    else:
        if verbose:
            print(f"[ANIM PLAN] 미분류 이벤트 | seq={seq_for_mapping} payload={payload}", flush=True)


def _is_progress_only_mode(ext: Any) -> bool:
    # 표시모드 제거: 진행현황 전용 모드는 더 이상 사용하지 않는다.
    return False


def _export_sim_logs_to_xlsx(ext: Any) -> None:
    """
    시뮬 종료 시 자동 호출되어 ``data/sim_logs/sim_logs_YYYYmmdd_HHMMSS.xlsx`` 를 생성한다.

    출력 정책(요구사항):
    - 엑셀에는 **프리런 JSON 타임테이블만** 남긴다(진행현황/이력로그/리포트 시트는 만들지 않는다).
    - 시트는 **단 한 장(`타임테이블`)** 으로 통합한다. 화면 구분은 각 JSON 라인의 ``"screen"`` 키로 한다.
    - 1열에 JSON 라인을 한 줄씩 기록한다(헤더 라인 ``[SIM] 타임테이블(프리런) — 화면N`` 은 제외).
    - 정렬 우선순위: ``t`` (시뮬 시간) 오름차순 → ``screen`` 오름차순 → kind(event 우선, step 다음).

    데이터 출처(우선순위):
    1) ``ext._sim_prerun_results_by_screen`` (프리런 결과) → ``_build_prerun_timetable_text`` 로 재빌드
       (가장 신뢰할 수 있는 원본; UI 라벨에 잘려 표시되었어도 전체 라인을 복원할 수 있음)
    2) 화면별 ``history_label.text`` 에 이미 들어 있는 타임테이블 블록 (Fallback)
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        _append_sim_log(ext, f"[SIM EXPORT] openpyxl import 실패: {e}")
        return

    out_dir = _extension_root_dir() / "data" / "sim_logs"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"sim_logs_{ts}.xlsx"

    # 1순위: 프리런 결과로 JSON 타임테이블을 다시 빌드(원본 신뢰)
    timetable_by: Dict[int, str] = {}
    try:
        results = getattr(ext, "_sim_prerun_results_by_screen", None)
        if isinstance(results, dict) and results:
            built = _build_prerun_timetable_text(results) or {}
            if isinstance(built, dict):
                timetable_by = {int(k): str(v) for k, v in built.items() if str(v or "").strip()}
    except Exception:
        timetable_by = {}

    # 2순위(Fallback): UI 의 history_label.text 에서 타임테이블 블록을 그대로 가져온다.
    if not timetable_by:
        try:
            chans_x = getattr(ext, "_sim_monitor_channels", None) or []
            if isinstance(chans_x, list):
                for ch in chans_x:
                    if not isinstance(ch, dict):
                        continue
                    try:
                        si = int(ch.get("screen", 0))
                    except Exception:
                        continue
                    if si <= 0:
                        continue
                    hl = ch.get("history_label")
                    if hl is None:
                        continue
                    txt = str(getattr(hl, "text", "") or "").strip()
                    if txt:
                        timetable_by[si] = txt
        except Exception:
            pass

    # 모든 화면의 JSON 라인을 하나로 모아 시간순 정렬
    # all_rows 의 항목: (t_float, screen_int, kind_str, raw_json_line_str)
    all_rows: List[Tuple[float, int, str, str]] = []
    for si, txt in (timetable_by.items() if isinstance(timetable_by, dict) else []):
        for line in (txt or "").splitlines():
            s = (line or "").strip()
            if not s:
                continue
            if s.startswith("[SIM] 타임테이블(프리런)"):
                continue
            t_val = 0.0
            scr_val = int(si)
            kind_val = ""
            try:
                obj = json.loads(s)
                if isinstance(obj, dict):
                    try:
                        t_val = float(obj.get("t", 0.0))
                    except Exception:
                        t_val = 0.0
                    try:
                        scr_val = int(obj.get("screen", si))
                    except Exception:
                        scr_val = int(si)
                    kind_val = str(obj.get("kind", "") or "")
            except Exception:
                # JSON 파싱 실패해도 원문은 보존(정렬 키만 기본값)
                pass
            all_rows.append((float(t_val), int(scr_val), str(kind_val), s))

    kind_prio = {"event": 0, "step": 1}
    try:
        all_rows.sort(
            key=lambda r: (
                float(r[0]),
                int(r[1]),
                int(kind_prio.get(str(r[2]), 9)),
            )
        )
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "타임테이블"
    if not all_rows:
        ws.cell(row=1, column=1, value="(타임테이블 없음)")
    else:
        for i, (_t, _s, _k, raw) in enumerate(all_rows, start=1):
            ws.cell(row=i, column=1, value=raw)

    wb.save(str(out_path))
    _append_sim_log(ext, f"[SIM EXPORT] 저장 완료: {out_path}")


def _detach_sim_update(ext: Any) -> None:
    """
    시뮼 tick·UI 구독을 정리한다.

    - ``_sim_thread_stop`` 을 set 한 뒤, **멀티 시** ``_sim_tick_threads`` 의 worker 를 순서대로 join 하고,
      단일 tick 스레드 ``_sim_thread`` 가 남아 있으면 join 한다.
    - 시뮼 로그 큐 드레인용 ``_sim_log_ui_sub`` 구독을 해제한다.
    """
    sub = getattr(ext, "_sim_update_sub", None)
    if sub is not None:
        try:
            sub.unsubscribe()
        except Exception:
            pass
        ext._sim_update_sub = None

    stop_evt = getattr(ext, "_sim_thread_stop", None)
    th = getattr(ext, "_sim_thread", None)
    tick_threads = list(getattr(ext, "_sim_tick_threads", None) or [])
    if stop_evt is not None:
        try:
            stop_evt.set()
        except Exception:
            pass
    for tth in tick_threads:
        if tth is not None:
            try:
                tth.join(timeout=1.0)
            except Exception:
                pass
    try:
        ext._sim_tick_threads = []
    except Exception:
        pass
    try:
        _stop_anim_screen_workers(ext)
    except Exception:
        pass
    if th is not None:
        try:
            th.join(timeout=1.0)
        except Exception:
            pass
    ext._sim_thread = None
    ext._sim_thread_stop = None
    ui_sub = getattr(ext, "_sim_log_ui_sub", None)
    if ui_sub is not None:
        try:
            ui_sub.unsubscribe()
        except Exception:
            pass
        ext._sim_log_ui_sub = None

    # 공정설정 시간 우선 모드: 애니 중단 요청 플래그 정리
    try:
        ie = getattr(ext, "_sim_interrupt_anim_event", None)
        if ie is not None:
            try:
                ie.clear()
            except Exception:
                pass
    except Exception:
        pass
    try:
        ext._sim_interrupt_anim_event = None
        ext._sim_interrupt_anim_apply_fn = None
    except Exception:
        pass


def on_sim_start_clicked(ext: Any) -> None:
    """
    시뮬레이션 시작: 엔진(들) 생성, 콜백 연결, tick 스레드(들) 기동.

    - ``ext._sim_viewport_split_count`` 로 채널 수 N 을 정한다. N<=1 이면 단일 ``TBSSimulationEngine`` +
      기존 ``_tick_loop`` 스레드 한 개. N>1 이면 화면별 스냅샷으로 N 개 엔진을 만들고
      ``_sim_multi_engine_tick_worker`` 를 **화면마다 한 스레드씩** 띄운다(게이트·pause 독립).
    - 시작 전 ``on_sim_stop_clicked`` 로 이전 스레드·엔진을 정리한다.

    데이터/상태의 큰 흐름(추적용):
    - (UI) 이 함수에서 run 세대 토큰 ``ext._sim_run_gen`` 을 증가시킨다.
      - 목적: stop/reset 이후 큐에 남은 UI 업데이트/애니 이벤트가 "이전 실행" 것이라면 무시하도록 하기 위함.
    - (UI→엔진) `simulation_engine.py`의 ``TBSSimulationEngine``(또는 멀티 엔진 리스트)을 생성하고,
      엔진이 emit 하는 on_log/on_progress/on_anim_event 콜백을 `control_window.py`의 UI sink로 연결한다.
    - (엔진→UI) 엔진이 emit 한 progress payload에는 `ep_occ`, `all_ep_empty`, `sim_total_est_sec` 등이 포함되고,
      이는 아래 EP 타임라인 그래프/우측 누적 시간 표시로 연결된다.
      - 종료 요약(`[SUMMARY] EPn_EMPTY / ALL_EP_EMPTY`)은 엔진(`simulation_engine.py`)에서 계산/출력한다.
      - UI 막대 우측의 초 표시는 UI가 가진 그래프 세그먼트(rows_state) 합산값이다(개념/정의는 동일).

    스레드/구독(중요):
    - tick 스레드: 시뮬 시간을 실제로 전진시키는 워커(단일 1개 또는 화면별 N개).
    - EP 타임라인(포트상태 아래)은 **엔진의 SimPy ``env.now``를 ``timeline_only`` progress의 ``sim_time``으로** 갱신한다.
      (내부 ``virtual_now`` 예산 보간은 막대 축에 쓰지 않아, 진행현황 t(sim)과 동일한 시계를 유지한다.)
    """
    try:
        n_ch = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        n_ch = 1

    target_screens_raw = getattr(ext, "_sim_startup_target_screens", None)
    partial_startup = isinstance(target_screens_raw, (list, tuple)) and len(target_screens_raw) > 0
    if not partial_startup:
        try:
            from .control_sim_screen_playback import is_simulation_in_progress

            if is_simulation_in_progress(ext):
                try:
                    _append_sim_log(ext, "[SIM] 시뮬레이션 진행 중 — 시작(재시작)을 건너뜁니다.")
                except Exception:
                    pass
                return
        except Exception:
            pass
    target_screens: List[int] = []
    if partial_startup:
        for x in target_screens_raw:
            try:
                target_screens.append(max(1, min(4, int(x))))
            except Exception:
                pass

    try:
        unlock_timetable_rows(ext)
        if not partial_startup:
            _clear_sim_timetable_storage(ext)
    except Exception:
        pass
    if partial_startup:
        for sc in target_screens:
            _stop_sim_screen_only(ext, int(sc))
    else:
        on_sim_stop_clicked(ext, freeze_ep_timeline=False)
    try:
        stopped = _sim_stopped_screens_set(ext)
        if partial_startup:
            for sc in target_screens:
                stopped.discard(int(sc))
        else:
            stopped.clear()
    except Exception:
        pass
    if partial_startup:
        for sc in target_screens:
            try:
                ctx = _usd_context_name_for_sim_screen(ext, int(sc))
                _restore_sim_prim_motion_to_initial(ext, usd_context_name=ctx)
            except Exception:
                pass
    else:
        try:
            _restore_all_sim_channels_prim_motion(ext)
        except Exception:
            pass
    # FOUP 진행중 보호 표시 추가 안전망(이전 세션 잔여 차단):
    # - on_sim_stop_clicked 에서 이미 정리하지만, 어떤 경로로든 잔여가 남는 것을 막기 위해 한 번 더 비운다.
    try:
        from . import port_lot_visibility as _plv  # type: ignore
        _plv.clear_foup_in_progress()
    except Exception:
        pass
    # 포트 LOT prim 의 baseline transform 을 "시뮬 시작 직전(원위치)" 에 미리 캡처한다.
    # 이유:
    #   - baseline 캐시는 평소 SequenceRunner.run 진입 시 처음 잡힌다.
    #   - 그런데 첫 시퀀스가 FOUP_PROCESS_START 같은 비-SequenceRunner 이벤트면 baseline 이
    #     캡처되지 않은 채 +Y 320 애니가 먼저 prim 을 옮긴다. 이후 다른 시퀀스가 처음 SequenceRunner
    #     를 돌릴 때 이미 +320 위치가 baseline 으로 잘못 캡처되어, 후속 보정에서 +320+320=640 으로
    #     점프하는 누적 오프셋 버그가 발생한다.
    #   - 시작 시점에 prim 이 진짜 author 위치인 상태에서 캐시를 채워두면 이 경합이 원천 차단된다.
    try:
        from . import port_lot_visibility as _plv  # type: ignore
        _plv.clear_port_lot_authoring_cache()
        _plv.ensure_port_lot_authoring_captured()
    except Exception:
        pass
    # 시작을 반복할 때 모니터 UI(포트/그래프/진행/이력) 영역에 위젯이 누적되어
    # 버튼 아래의 빈 공간이 점점 커지는 현상이 발생할 수 있어,
    # 시작 시점에 모니터 영역을 한 번 깨끗하게 재빌드한다.
    if not partial_startup:
        try:
            ext._sim_timetable_allow_shell_rebuild = True
            _rebuild_all_sim_ui_panels(ext)
            ext._sim_timetable_allow_shell_rebuild = False
        except Exception:
            try:
                ext._sim_timetable_allow_shell_rebuild = False
            except Exception:
                pass
    # 실행 세대 토큰: stop/reset 후 남은 이벤트/애니 job을 무시하기 위해 사용
    # (화면별 시작 시 다른 화면 재생 이벤트는 유지 — 전역 gen 은 올리지 않음)
    if not partial_startup:
        try:
            ext._sim_run_gen = int(getattr(ext, "_sim_run_gen", 0) or 0) + 1
        except Exception:
            ext._sim_run_gen = 1
    _auto_fill_per_screen_snapshots_on_start(ext)
    if partial_startup and not target_screens:
        return
    if n_ch > 1:
        try:
            _ensure_tick_pause_map_for_multi(ext, n_ch)
        except Exception:
            pass

    # 공정 시간/간격/초기포트/고장포트 등 “시뮬 입력값”은 스냅샷(dict) 하나로 통일한다.
    # - 분할(N>1): 화면별로 저장된 스냅샷을 사용
    # - 단일(N==1): 화면1 스냅샷(없으면 현재 UI값을 캡처한 dict)을 사용
    ep_count = _SIM_DEF.ep_count()
    timing = SimulationTimingConfig()
    init_cfg = SimulationInitConfig(
        ep_count=ep_count,
        initial_full_ports=[],
        max_oht_lots=int(_SIM_DEF.lot_count),
        process_time_priority=False,
    )
    snap_1: Dict[str, Any] = {}
    if n_ch <= 1:
        try:
            cap1 = _capture_per_screen_sim_settings(ext, 1)
        except Exception:
            cap1 = {}
        snap_1 = dict(cap1)
        try:
            timing, init_cfg = _timing_and_init_from_snapshot(ext, snap_1)
            _inject_lot_fix_proc_into_init(ext, init_cfg)
        except Exception:
            pass
        try:
            ep_count = int(getattr(init_cfg, "ep_count", 2) or 2)
        except Exception:
            ep_count = 2

    log_interval = 0.0
    log_cfg = SimulationLogConfig(
        progress_interval_sec=log_interval,
        input_status_interval_sec=log_interval,
    )
    lots: List[Lot] = []

    if not partial_startup:
        ext._sim_history_text.set_value("[SIM] 초기화")
        ext._sim_progress_text.set_value("[진행현황] 초기화 (시뮬레이션 시작 대기)")
        ext._sim_port_state_text.set_value("[포트상태] 초기화 (이벤트 대기)")
    elif 1 in target_screens:
        try:
            ext._sim_history_text.set_value("[SIM] 초기화")
            ext._sim_progress_text.set_value("[진행현황] 초기화 (시뮬레이션 시작 대기)")
            ext._sim_port_state_text.set_value("[포트상태] 초기화 (이벤트 대기)")
        except Exception:
            pass
    # EP 타임라인: 시작 버튼 누르는 순간부터(t=0) 빈 포트 상태로 표시/진행할 수 있도록 초기 스냅샷을 만든다.
    _per_scr_state_attrs = (
        "_sim_ep_occ_timeline_state_by_screen",
        "_sim_last_ports_occupancy_by_screen",
        "_sim_post_anim_port_applied_by_screen",
        "_sim_renewal_port_defer_by_screen",
        "_sim_pending_post_anim_port_by_screen",
        "_sim_post_anim_src_by_screen",
        "_sim_ep_timeline_virtual_time_by_screen",
    )
    if partial_startup:
        for attr in _per_scr_state_attrs:
            d = getattr(ext, attr, None)
            if not isinstance(d, dict):
                continue
            for sc in target_screens:
                d.pop(str(int(sc)), None)
    else:
        for attr in _per_scr_state_attrs:
            try:
                setattr(ext, attr, {})
            except Exception:
                pass
        try:
            sub = getattr(ext, "_sim_ep_timeline_ui_sub", None)
            if sub is not None:
                try:
                    sub.unsubscribe()
                except Exception:
                    pass
            ext._sim_ep_timeline_ui_sub = None
        except Exception:
            pass
        try:
            ext._sim_aux_anim_notice_screens = set()
        except Exception:
            pass
    chans0 = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans0, list) and len(chans0) > 1:
        for ch in chans0:
            if not isinstance(ch, dict):
                continue
            try:
                si = int(ch.get("screen", 1))
            except Exception:
                si = 1
            if partial_startup and si not in target_screens:
                continue
            ht = "[SIM] 초기화" if si == 1 else f"[SIM·화면{si}] 초기화"
            pt = "[진행현황] 초기화 (시뮬레이션 시작 대기)" if si == 1 else f"[진행현황·화면{si}] 초기화 (대기)"
            ph = f"[포트상태·화면{si}] 초기화 (이벤트 대기)"
            pl = ch.get("progress_label")
            phdr = ch.get("port_header")
            try:
                _set_channel_history_text(ch, ht)
            except Exception:
                pass
            if pl is not None:
                pl.text = pt
            if phdr is not None:
                phdr.text = ph
            cells = ch.get("port_cells") or {}
            boxes = ch.get("port_cell_boxes") or {}
            for port in ("INOUT", "BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
                if port in cells:
                    cells[port].text = "IN/OUT:-" if port == "INOUT" else f"{port}:-"
                try:
                    _set_port_box_style(ext, port, "-", boxes)
                except Exception:
                    pass
            ep3c = ch.get("port_ep3_cell")
            if ep3c is not None:
                ep3c.text = "EP3:-"
            try:
                _set_port_box_style(ext, "EP3", "-", boxes)
            except Exception:
                pass
            # EP 타임라인 초기 렌더(t=0.0) + 마지막 점유 스냅샷 저장
            try:
                occ0 = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
                ext._sim_last_ports_occupancy_by_screen[str(si)] = dict(occ0)  # type: ignore[index]
                _update_ep_timeline_under_port_state(ext, ch, occ0, "0.0")
            except Exception:
                pass
    else:
        if not partial_startup or 1 in target_screens:
            try:
                chans_s = getattr(ext, "_sim_monitor_channels", None)
                if isinstance(chans_s, list) and chans_s and isinstance(chans_s[0], dict):
                    _set_channel_history_text(chans_s[0], "[SIM] 초기화")
            except Exception:
                pass
            if getattr(ext, "_sim_history_label", None) is not None:
                try:
                    ext._sim_history_label.text = "[SIM] 초기화"
                except Exception:
                    pass
            if getattr(ext, "_sim_progress_label", None) is not None:
                ext._sim_progress_label.text = "[진행현황] 초기화 (시뮬레이션 시작 대기)"
            if getattr(ext, "_sim_port_state_label", None) is not None:
                ext._sim_port_state_label.text = "[포트상태] 초기화 (이벤트 대기)"
            if getattr(ext, "_sim_port_state_header_label", None) is not None:
                ext._sim_port_state_header_label.text = "[포트상태] 초기화 (이벤트 대기)"
            cells = getattr(ext, "_sim_port_cells", {}) or {}
            for port in ("INOUT", "BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
                if port in cells:
                    cells[port].text = "IN/OUT:-" if port == "INOUT" else f"{port}:-"
            if getattr(ext, "_sim_port_ep3_cell", None) is not None:
                ext._sim_port_ep3_cell.text = "EP3:-"
            # 단일(또는 모니터 채널 1개) 모드: 시작 직후 timeline_only 가 last_occ 없이 스킵되지 않도록 시드
            try:
                occ0 = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
                by = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                if not isinstance(by, dict):
                    by = {}
                    ext._sim_last_ports_occupancy_by_screen = by
                by["1"] = dict(occ0)
                chans_s = getattr(ext, "_sim_monitor_channels", None)
                if isinstance(chans_s, list) and len(chans_s) >= 1 and isinstance(chans_s[0], dict):
                    _update_ep_timeline_under_port_state(ext, chans_s[0], occ0, "0.0")
            except Exception:
                pass

    # -------------------------------------------------------------------
    # Start 직후(t=0) 초기 적재 포트(FULL) 즉시 반영
    # - 기존에는 occ0(전부 EMPTY)로 먼저 그린 뒤, 첫 공정/첫 이벤트까지 업데이트가 없어
    #   포트상태/막대가 "첫 공정 전까지 빨강"처럼 보일 수 있었다.
    # -------------------------------------------------------------------
    def _occ_from_snap(snap: Dict[str, Any], ep_cnt: int) -> Dict[str, str]:
        occ = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
        try:
            if bool(snap.get("init_inout")):
                occ["INOUT"] = "FULL"
            if bool(snap.get("init_bp1")):
                occ["BP1"] = "FULL"
            if bool(snap.get("init_bp2")):
                occ["BP2"] = "FULL"
            if bool(snap.get("init_bp3")):
                occ["BP3"] = "FULL"
            if bool(snap.get("init_bp4")):
                occ["BP4"] = "FULL"
            if bool(snap.get("init_ep1")):
                occ["EP1"] = "FULL"
            if bool(snap.get("init_ep2")):
                occ["EP2"] = "FULL"
            if bool(snap.get("init_ep3")):
                occ["EP3"] = "FULL"
        except Exception:
            pass
        if int(ep_cnt) < 3:
            occ["BP4"] = ""
            occ["EP3"] = ""
        return occ

    try:
        snaps_init = list(getattr(ext, "_sim_per_screen_snapshots", None) or [None, None, None, None])
    except Exception:
        snaps_init = [None, None, None, None]
    while len(snaps_init) < 4:
        snaps_init.append(None)
    snaps_init = snaps_init[:4]

    for scr0 in range(1, int(n_ch) + 1):
        if partial_startup and scr0 not in target_screens:
            continue
        s0 = snaps_init[scr0 - 1] if (scr0 - 1) < len(snaps_init) else None
        if not isinstance(s0, dict):
            try:
                s0 = _capture_per_screen_sim_settings(ext)
            except Exception:
                s0 = {}
        try:
            if "ep_count" in s0:
                ep_cnt0 = 3 if int(s0.get("ep_count", 2) or 2) >= 3 else 2
            else:
                ep_cnt0 = int(_SIM_DEF.ep_count())
        except Exception:
            ep_cnt0 = int(_SIM_DEF.ep_count())
        ep_idx0 = 0 if ep_cnt0 < 3 else 1
        occ_init = _occ_from_snap(s0, ep_cnt0)
        try:
            ext._sim_last_ports_occupancy_by_screen[str(scr0)] = dict(occ_init)  # type: ignore[index]
        except Exception:
            pass
        try:
            _update_port_occupancy_panel(ext, occ_init, sim_time="0.0", screen=int(scr0))
        except Exception:
            pass
        # 요구사항: 시뮬 시작 직후에도 포트 점유 상태를 바탕으로
        # visibility/위치 초기화를 함께 반영한 상태에서 시작해야 한다.
        try:
            _apply_sim_event_state_only(
                ext,
                {
                    "ports_occupancy": dict(occ_init),
                    "sim_time": "0.00",
                    "foup_proc_active_ep": "",
                },
                screen=int(scr0),
            )
        except Exception:
            pass
    if partial_startup:
        try:
            rows = getattr(ext, "_sim_progress_rows", None)
            if isinstance(rows, dict):
                for sc in target_screens:
                    rows.pop(str(int(sc)), None)
                    rows.pop(int(sc), None)
        except Exception:
            pass
    else:
        ext._sim_progress_rows = {}
        ext._sim_progress_history = []
        ext._sim_progress_start_times = {}
    ext._sim_log_queue = queue.SimpleQueue()
    _enqueue_sim_log(ext, "[SIM UI] 실시간 로그 큐 초기화")
    # 첫 공정 전에도 진행현황이 끊기지 않도록, 화면별 기본 progress payload를 1회 시드한다.
    try:
        for scr0 in range(1, int(n_ch) + 1):
            if partial_startup and scr0 not in target_screens:
                continue
            p0 = {
                "tbs_sim_screen": str(scr0),
                "sim_time": "0.00",
                "label": "대기",
                "detail": "",
                "status": "RUNNING",
                "elapsed": "0.0",
                "total": "0.0",
                "percent": "0",
            }
            _enqueue_sim_progress(ext, p0)
    except Exception:
        pass
    if partial_startup:
        try:
            active_by = getattr(ext, "_sim_anim_active_by_screen", None)
            pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
            for sc in target_screens:
                sk = str(int(sc))
                if isinstance(active_by, dict):
                    active_by[sk] = {}
                if isinstance(pending_by, dict):
                    pending_by[sk] = []
        except Exception:
            pass
    else:
        ext._sim_anim_active = {}
        ext._sim_anim_pending = []

    def _interrupt_anim_for_proc_priority(screen: Optional[str] = None) -> None:
        """
        공정설정 시간 우선 모드에서 '애니가 재생 중이어도 끊고 다음 단계로 진행'을 위해 호출.
        - 시뮬 tick pause 플래그를 해제하고
        - 시퀀스 러너/개별 애니메이션을 stop 한다.
        """
        try:
            from . import sim_multi_diag as _mdiag

            _mdiag.log_interrupt(ext, screen=screen, reason="proc_priority")
        except Exception:
            pass
        try:
            scr = str(screen or "").strip() or None
            # pending 큐 비우기(현재 모드에서는 애니 길이를 기다리지 않음)
            if scr is not None:
                pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
                if isinstance(pending_by, dict):
                    pending_by[str(scr)] = []
            else:
                ext._sim_anim_pending = []
        except Exception:
            pass
        try:
            # 멀티 화면: 화면별 runner가 있으면 전부 정리.
            runners = getattr(ext, "_sim_runners_by_screen", None)
            if isinstance(runners, dict) and runners:
                if scr is not None:
                    r = runners.get(str(scr))
                    try:
                        if r is not None:
                            r.pause()
                    except Exception:
                        pass
                else:
                    for r in list(runners.values()):
                        try:
                            if r is not None:
                                r.pause()
                        except Exception:
                            pass
            else:
                runner = getattr(ext, "_sim_runner", None)
                if runner is not None:
                    runner.pause()
        except Exception:
            pass
        try:
            from .sim_channel_scope import stop_channel_animations

            ctx_nm = None
            if scr is not None:
                try:
                    ctx_nm = _usd_context_name_for_sim_screen(ext, int(scr))
                except Exception:
                    ctx_nm = None
            stop_channel_animations(ctx_nm, preserve_foup_port_lot_prims=True, diag_reason="interrupt_proc_priority")
        except Exception:
            pass
        pe = getattr(ext, "_sim_tick_pause_event", None)
        if pe is not None:
            try:
                pe.clear()
            except Exception:
                pass
        try:
            ext._sim_tick_pause_until_wall = None
        except Exception:
            pass

    # tick 스레드 -> UI 스레드 마샬링: 중단 요청 플래그만 세팅하고, 실제 stop은 _drain_sim_log_queue에서 처리.
    try:
        ext._sim_interrupt_anim_event = threading.Event()
        ext._sim_interrupt_anim_apply_fn = lambda: _interrupt_anim_for_proc_priority(None)
        # 화면별 interrupt (멀티 시뮬에서 해당 화면만 중단)
        ext._sim_interrupt_anim_event_by_screen = {}
        ext._sim_interrupt_anim_apply_fn_by_screen = {}
        try:
            n_ch2 = int(getattr(ext, "_sim_split_channels", 1) or 1)
        except Exception:
            n_ch2 = 1
        n_ch2 = max(1, n_ch2)
        for i in range(n_ch2):
            scr = str(i + 1)
            ext._sim_interrupt_anim_event_by_screen[scr] = threading.Event()
            ext._sim_interrupt_anim_apply_fn_by_screen[scr] = (lambda s=scr: _interrupt_anim_for_proc_priority(s))
    except Exception:
        ext._sim_interrupt_anim_event = None
        ext._sim_interrupt_anim_apply_fn = None
        ext._sim_interrupt_anim_event_by_screen = None
        ext._sim_interrupt_anim_apply_fn_by_screen = None

    def _request_interrupt_anim_for_proc_priority(tags: Optional[Dict[str, Any]] = None) -> None:
        # 화면별 중단 지원: tbs_sim_screen이 있으면 해당 화면만 중단 요청.
        try:
            scr = "1"
            if isinstance(tags, dict):
                scr = str(tags.get("tbs_sim_screen", "1") or "1").strip() or "1"
            # per-screen event가 있으면 그쪽을 우선
            by = getattr(ext, "_sim_interrupt_anim_event_by_screen", None)
            if isinstance(by, dict) and scr in by and by[scr] is not None:
                by[scr].set()
                return
        except Exception:
            pass
        try:
            ie = getattr(ext, "_sim_interrupt_anim_event", None)
            if ie is not None:
                ie.set()
        except Exception:
            pass

    def _on_gate(payload: Dict[str, str]) -> float:
        # 요구사항: 공정시간보다 애니(JSON) 시간이 길면 다음 공정은 애니 종료까지 대기.
        # simulation_engine은 이 반환값(초)을 받아서 각 공정 timeout을 max(공정, 애니)로 확장한다.
        anim_est_sec = _estimate_anim_duration_for_gate_payload(ext, payload or {})
        # 공정확인 체크 시에는 "확인 클릭 전에는 애니/공정 시작 금지"가 목표이므로,
        # simulation_engine의 _request_gate() 시점에 UI 확인창을 띄우고 동기 블로킹한다.
        try:
            confirm_each = bool(
                getattr(ext, "_sim_confirm_each_step_model", None) is not None
                and ext._sim_confirm_each_step_model.get_value_as_bool()
            )
        except Exception:
            confirm_each = False
        if not confirm_each:
            return float(anim_est_sec)

        # 공정확인 중에는 sim tick thread도 멈춰야 "확인 전까지 완전 정지"가 된다.
        try:
            gp = getattr(ext, "_sim_gate_pause_event", None)
            if gp is not None:
                gp.set()
        except Exception:
            pass

        seq_raw = str(payload.get("seq", ""))
        seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
        lot = lot_id_from_payload(payload)
        est = str(payload.get("est_sec", ""))
        fr = str(payload.get("from_port_id", ""))
        to = str(payload.get("to_port_id", ""))
        port = str(payload.get("port_id", ""))
        try:
            # 게이트 다이얼로그의 XML도 실제 애니 매핑 파이프라인과 같은 규칙으로 생성한다.
            # (FROM/TO 시퀀스 vs PORT_ID 시퀀스 분기)
            if seq in xml_generator.FROM_TO_SEQS:
                fnum = _parse_port_num(fr, 1)
                tnum = _parse_port_num(to, 1)
                xml = xml_generator.build_xml_string(seq, from_port_id=fnum, to_port_id=tnum)
            elif seq in xml_generator.PORT_ID_ONLY_SEQS:
                pnum = _parse_port_num(port, 1)
                xml = xml_generator.build_xml_string(seq, port_id=pnum)
            else:
                xml = f"(XML 미생성: 비-XML 공정 seq={seq_raw}->{seq})"
        except Exception:
            xml = f"(XML 생성 실패: seq={seq})"

        xml_sequence_name = ""
        if isinstance(xml, str) and xml.strip().startswith("<"):
            try:
                _pd = xml_generator.parse_xml_string(xml) or {}
                xml_sequence_name = str(_pd.get("sequence_name", "") or "").strip().upper()
            except Exception:
                pass

        # Alert에서 "실행 대상 JSON 파일"과 존재 여부를 함께 안내한다.
        map_line = "JSON 매핑: 없음"
        try:
            mapping_payload = dict(payload or {})
            seq_for_mapping = seq
            if isinstance(xml, str) and xml.strip().startswith("<"):
                parsed = xml_generator.parse_xml_string(xml) or {}
                parsed_seq = str(parsed.get("sequence_name", "") or "").strip().upper()
                if parsed_seq:
                    seq_for_mapping = parsed_seq
                mapping_payload["seq"] = seq_for_mapping
                mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
            else:
                mapping_payload["seq"] = seq_for_mapping

            mapped_json, _meta, rule_name, source_name = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
            if mapped_json:
                jp = _normalize_json_path(mapped_json)
                exists_txt = "존재" if jp.is_file() else "없음"
                map_line = (
                    f"JSON 매핑: source={source_name or '-'} rule={rule_name or '-'} "
                    f"file={jp.name} ({exists_txt})"
                )
            else:
                map_line = f"JSON 매핑: 없음 (event={seq_for_mapping})"
        except Exception as e:
            map_line = f"JSON 매핑 확인 실패: {e}"
        done_evt = threading.Event()
        message = (
            f"공정: {payload.get('title','-')}\n"
            f"이벤트 sequence_name: 시뮬 seq={seq_raw or '-'}, 규격/별칭={seq or '-'}"
            + (f", XML SEQUENCE_NAME={xml_sequence_name}" if xml_sequence_name else "")
            + "\n"
            f"lot={lot} from={fr} to={to} port={port}\n"
            f"예상시간={est}s\n"
            f"애니예상={anim_est_sec:.2f}s (JSON 기준)\n\n"
            f"{map_line}\n\n"
            f"XML:\n{xml}"
        )
        _enqueue_gate_request(
            ext,
            {
                "title": payload.get("title", "공정 확인"),
                "message": message,
                "_done_event": done_evt,
                "gate_seq_raw": seq_raw,
                "gate_seq_canonical": seq,
                "gate_xml_sequence_name": xml_sequence_name,
            },
        )
        # 시뮬레이션 스레드는 사용자 확인 전까지 여기서 동기 대기한다.
        done_evt.wait()
        return float(anim_est_sec)

    if not partial_startup:
        try:
            ext._sim_engines = []
        except Exception:
            pass

    if n_ch <= 1:

        def _collect_faulty_ports_for_engine() -> Set[str]:
            # 단일도 스냅샷(dict) 기준으로 통일(= UI 모델/다른 경로 중복 제거)
            try:
                return set(_fault_ports_from_snapshot(snap_1, ep_count))
            except Exception:
                return set()

        # 프리런/재생 모드에서는 엔진 콜백을 UI로 직접 보내지 않는다(프리런 수집 → 재생 단계에서만 UI로 emit).
        engine = TBSSimulationEngine(
            lots=lots,
            timing=timing,
            log_config=log_cfg,
            init_config=init_cfg,
            on_log=lambda _line: None,
            on_event=lambda _payload: None,
            on_progress=lambda _payload: None,
            # 프리런은 사용자 확인 UI 없이 자동 통과 + 애니 길이만 반영
            on_gate=lambda payload: float(_estimate_anim_duration_for_gate_payload(ext, payload or {})),
            print_to_console=(not _is_progress_only_mode(ext)),
            event_tags={"tbs_sim_screen": "1"},
        )
        try:
            engine.set_runtime_hooks(
                faulty_ports_supplier=_collect_faulty_ports_for_engine,
                interrupt_anim_cb=_request_interrupt_anim_for_proc_priority,
            )
        except Exception:
            pass
        ext._sim_engine = engine
        ext._sim_engines = []
        # start() 직전에 스케일 주입: 시작 직후 이벤트가 큐에 들어가도 30초 폴백에 고정되지 않게 한다.
        try:
            te0 = float(getattr(engine, "_sim_total_est_sec", 0.0) or 0.0)
        except Exception:
            te0 = 0.0
        if te0 > 0.0:
            try:
                by = getattr(ext, "_sim_last_total_est_by_screen", None)
                if not isinstance(by, dict):
                    by = {}
                    ext._sim_last_total_est_by_screen = by
                by["1"] = float(te0)
            except Exception:
                pass
        if not engine.start():
            _append_sim_log(ext, "[SIM] 시작 실패")
            return
    else:
        from .ebs_case_models import snapshots_for_startup_channels

        channel_snaps = snapshots_for_startup_channels(ext, n_ch)
        engines: List[Any] = list(getattr(ext, "_sim_engines", None) or [])
        while len(engines) < n_ch:
            engines.append(None)
        engines = engines[:n_ch]

        def _make_fault_supplier(sf: Dict[str, Any], ec: int):
            def _sup() -> Set[str]:
                return set(_fault_ports_from_snapshot(sf, ec))

            return _sup

        for i in range(n_ch):
            screen = i + 1
            if partial_startup and screen not in target_screens:
                continue
            snap_i = copy.deepcopy(channel_snaps[i] if i < len(channel_snaps) else {})
            timing_i, init_i = _timing_and_init_from_snapshot(ext, snap_i)
            _inject_lot_fix_proc_into_init(ext, init_i)
            screen_tag = str(i + 1)
            snap_frozen = copy.deepcopy(snap_i)
            ep_i = int(getattr(init_i, "ep_count", 2) or 2)
            ep_i = 3 if ep_i >= 3 else 2
            # 프리런/재생 모드: 콜백은 프리런 수집 단계에서만 사용(여기선 노옵 주입).
            eng = TBSSimulationEngine(
                lots=lots,
                timing=timing_i,
                log_config=log_cfg,
                init_config=init_i,
                on_log=lambda _line: None,
                on_event=lambda _payload: None,
                on_progress=lambda _payload: None,
                on_gate=lambda payload: float(_estimate_anim_duration_for_gate_payload(ext, payload or {})),
                print_to_console=(not _is_progress_only_mode(ext)),
                event_tags={"tbs_sim_screen": screen_tag},
            )
            try:
                eng.set_runtime_hooks(
                    faulty_ports_supplier=_make_fault_supplier(snap_frozen, ep_i),
                    interrupt_anim_cb=_request_interrupt_anim_for_proc_priority,
                )
            except Exception:
                pass
            engines[i] = eng

        # 멀티: start() 전에 화면별 총 예상 시간을 넣어 두어 첫 타임라인 갱신이 30초 스케일에 묶이지 않게 한다.
        try:
            by_pre = getattr(ext, "_sim_last_total_est_by_screen", None)
            if not isinstance(by_pre, dict):
                by_pre = {}
                ext._sim_last_total_est_by_screen = by_pre
            for i_pre, eng_pre in enumerate(engines):
                if eng_pre is None:
                    continue
                try:
                    te_pre = float(getattr(eng_pre, "_sim_total_est_sec", 0.0) or 0.0)
                except Exception:
                    te_pre = 0.0
                if te_pre > 0.0:
                    by_pre[str(i_pre + 1)] = float(te_pre)
        except Exception:
            pass

        started: List[Any] = []
        for eng in engines:
            if eng is None:
                continue
            if not eng.start():
                for e2 in started:
                    try:
                        e2.stop()
                    except Exception:
                        pass
                try:
                    ext._sim_engine = None
                    ext._sim_engines = []
                except Exception:
                    pass
                _append_sim_log(ext, "[SIM] 멀티 채널 시작 실패")
                return
            started.append(eng)
        # 화면별 EP 타임라인 스케일(총 예상 시간) 선주입
        try:
            by = getattr(ext, "_sim_last_total_est_by_screen", None)
            if not isinstance(by, dict):
                by = {}
                ext._sim_last_total_est_by_screen = by
            for i, eng in enumerate(started):
                scr = str(i + 1)
                try:
                    te = float(getattr(eng, "_sim_total_est_sec", 0.0) or 0.0)
                except Exception:
                    te = 0.0
                if te > 0.0:
                    by[scr] = float(te)
        except Exception:
            pass
        try:
            ext._sim_engines = engines
            ext._sim_engine = engines[0] if engines else None
        except Exception:
            pass
        try:
            if partial_startup:
                _append_sim_log(
                    ext,
                    f"[SIM] 화면 {','.join(str(s) for s in target_screens)} 시뮼 시작 (CASE 실시간 설정)",
                )
            else:
                _append_sim_log(ext, f"[SIM] 멀티 시뮼 시작 (채널={n_ch}, CASE A/B 실시간 설정)")
        except Exception:
            pass
        try:
            from . import sim_multi_diag as _mdiag

            _mdiag.log_sim_start_multi(
                ext,
                n_ch=n_ch,
                run_gen=int(getattr(ext, "_sim_run_gen", 0) or 0),
                snaps=channel_snaps,
                engines=[e for e in engines if e is not None],
            )
        except Exception:
            pass

    # --- 프리런(오프라인) → 타임라인 재생 모드 ---
    # 요구사항:
    # - 사용자가 Start를 누르면 UI는 기존과 동일하게 "시뮬이 진행"되는 것처럼 보이되,
    #   내부적으로는 먼저 가능한 빠르게 시뮬을 끝까지 계산(prerun)하고 그 결과를 재생한다.
    # - 이 모드에서는 실시간 tick thread를 띄우지 않는다.
    tick_state = {"count": 0}
    stop_evt = threading.Event()
    ext._sim_thread_stop = stop_evt
    try:
        ext._sim_tick_threads = []
    except Exception:
        pass
    speed_value = max(0.1, ext._sim_speed_model.get_value_as_float())
    if partial_startup:
        _append_sim_log(
            ext,
            f"[SIM] 화면 {','.join(str(s) for s in target_screens)} 프리런 시작…",
        )
    else:
        _append_sim_log(ext, "[SIM] 프리런 시작: 내부적으로 전체 시뮬을 먼저 계산합니다...")
    try:
        _set_sim_prerun_ui_busy(ext, True)
    except Exception:
        pass
    # SIM 현황(이력) 영역은 타임테이블 전용으로 유지
    try:
        ext._sim_history_timetable_only = True
    except Exception:
        pass
    try:
        ext._sim_log_ui_sub = app.get_app().get_update_event_stream().create_subscription_to_pop(
            lambda e: _drain_sim_log_queue(ext),
            name="morph.tbs_control_2:sim_log_ui_drain",
        )
    except Exception as e:
        _append_sim_log(ext, f"[SIM UI] 로그 큐 드레인 구독 실패: {e}")

    # prerun 결과/플레이어 상태 초기화
    try:
        ext._sim_partial_prerun_screens = list(target_screens) if partial_startup else None
    except Exception:
        pass
    if partial_startup:
        try:
            ext._sim_prerun_done_evt = threading.Event()
            prev_res = getattr(ext, "_sim_prerun_results_by_screen", None)
            if isinstance(prev_res, dict):
                for sc in target_screens:
                    prev_res.pop(int(sc), None)
                    prev_res.pop(str(int(sc)), None)
        except Exception:
            pass
        try:
            for sc in target_screens:
                sk = str(int(sc))
                for attr in (
                    "_sim_playback_plan_by_screen",
                    "_sim_ep_bar_prerun_by_screen",
                    "_sim_playback_schedule_by_screen",
                    "_sim_playback_plan_initial_occ_by_screen",
                ):
                    d = getattr(ext, attr, None)
                    if isinstance(d, dict):
                        d.pop(sk, None)
                        try:
                            d.pop(int(sc), None)
                        except Exception:
                            pass
        except Exception:
            pass
    else:
        try:
            _clear_sim_timetable_storage(ext)
        except Exception:
            pass
        try:
            ext._sim_prerun_done_evt = threading.Event()
            ext._sim_prerun_results_by_screen = None
            ext._sim_playback_schedule_by_screen = None
            try:
                from .control_sim_playback_plan import clear_playback_plan_runtime_state

                clear_playback_plan_runtime_state(ext)
            except Exception:
                pass
            ext._sim_playback_player = None
            ext._sim_playback_players_by_screen = None
            ext._sim_playback_runtime = None
            ext._sim_playback_ui_sub = None
            ext._sim_prerun_timetable_printed = False
            set_sim_playback_active(ext, False)
            ext._sim_playback_done = False
            clear_proc_gates(ext)
        except Exception:
            pass

    def _prerun_thread_body(run_gen: int) -> None:
        try:
            engs = []
            # 단일은 ext._sim_engine, 멀티는 ext._sim_engines 를 사용
            eng_list_outer = getattr(ext, "_sim_engines", None)
            if isinstance(eng_list_outer, list) and len(eng_list_outer) > 0:
                engs = list(eng_list_outer)
            else:
                e0 = getattr(ext, "_sim_engine", None)
                engs = [e0] if e0 is not None else []

            results: Dict[int, SimPreRunResult] = {}
            stopped_scr = set()
            partial_targets: Any = None
            try:
                partial_targets = getattr(ext, "_sim_partial_prerun_screens", None)
            except Exception:
                partial_targets = None
            try:
                stopped_scr = {
                    int(x) for x in (getattr(ext, "_sim_stopped_screens", None) or set())
                }
            except Exception:
                stopped_scr = set()
            partial_set = set()
            if partial_targets is not None:
                try:
                    partial_set = {int(x) for x in partial_targets}
                except Exception:
                    partial_set = set()
            for idx, eng in enumerate(engs):
                if eng is None:
                    continue
                scr = idx + 1
                if scr in stopped_scr:
                    continue
                if partial_set and scr not in partial_set:
                    continue
                # 세대가 바뀌었으면 중단 (전역 시작만 해당)
                if partial_set:
                    pass
                else:
                    try:
                        if int(getattr(ext, "_sim_run_gen", 0) or 0) != int(run_gen):
                            return
                    except Exception:
                        pass
                try:
                    res = prerun_engine_to_timeline(screen=scr, engine=eng)
                    results[int(scr)] = res
                except Exception:
                    continue
            if not partial_set:
                try:
                    if int(getattr(ext, "_sim_run_gen", 0) or 0) != int(run_gen):
                        return
                except Exception:
                    pass
            try:
                if partial_set:
                    merged: Dict[int, SimPreRunResult] = {}
                    prev = getattr(ext, "_sim_prerun_results_by_screen", None)
                    if isinstance(prev, dict):
                        for k, v in prev.items():
                            try:
                                merged[int(k)] = v
                            except Exception:
                                continue
                    merged.update(results)
                    ext._sim_prerun_results_by_screen = merged
                else:
                    ext._sim_prerun_results_by_screen = results
            except Exception:
                pass
            try:
                from . import sim_multi_diag as _mdiag

                _mdiag.log_prerun_done(ext, results)
            except Exception:
                pass
        finally:
            try:
                ev = getattr(ext, "_sim_prerun_done_evt", None)
                if ev is not None:
                    ev.set()
            except Exception:
                pass

    try:
        run_gen = int(getattr(ext, "_sim_run_gen", 0) or 0)
    except Exception:
        run_gen = 0
    th_pr = threading.Thread(
        target=_prerun_thread_body,
        args=(run_gen,),
        name="morph.tbs_control_2.sim_prerun",
        daemon=True,
    )
    ext._sim_thread = th_pr
    th_pr.start()
    return

    def _tick_loop():
        """
        단일 스레드 시뮼 tick 루프(분할 1 또는 레거시 경로).

        - ``ext._sim_engines`` 가 비어 있지 않으면 모든 엔진에 동일 ``scaled`` 로 순차 ``tick``(구조상 한 스레드).
        - ``_sim_tick_pause_event`` 가 켜지면 **전역**으로 tick 을 멈출 수 있다(멀티 N>1 은 별도 worker 사용).
        - ``confirm_each`` + ``_sim_gate_pause_event`` 이면 게이트 확인 전까지 이 루프가 대기한다.
        """
        try:
            print("[SIM] tick thread 시작", flush=True)
            last = time.perf_counter()
            while not stop_evt.is_set():
                # 애니메이션이 재생 중이면 sim tick을 일시정지
                pause_evt = getattr(ext, "_sim_tick_pause_event", None)
                gate_pause_evt = getattr(ext, "_sim_gate_pause_event", None)
                try:
                    confirm_each = bool(getattr(ext, "_sim_confirm_each_step_model", None) is not None and ext._sim_confirm_each_step_model.get_value_as_bool())
                except Exception:
                    confirm_each = False
                if not confirm_each and gate_pause_evt is not None and gate_pause_evt.is_set():
                    try:
                        gate_pause_evt.clear()
                    except Exception:
                        pass
                if confirm_each and gate_pause_evt is not None and gate_pause_evt.is_set():
                    time.sleep(0.02)
                    continue
                if (not _is_multi_viewport_sim(ext)) and pause_evt is not None and pause_evt.is_set():
                    # 원칙: JSON 애니메이션이 실제로 진행 중이면(sim 모듈의 활성 상태가 있으면) 절대 tick 재개하지 않는다.
                    try:
                        anim_running = bool(
                            is_translate_animation_running()
                            or is_rotate_animation_running()
                            or is_curve_animation_running()
                            or (getattr(ext, "_sim_runner", None) is not None and getattr(ext._sim_runner, "is_running", lambda: False)())
                        )
                    except Exception:
                        anim_running = True
                    if anim_running:
                        time.sleep(0.02)
                        continue

                    # fail-safe: 추정 시간이 남아있으면 최소한 그동안은 pause 유지
                    until_wall = getattr(ext, "_sim_tick_pause_until_wall", None)
                    if isinstance(until_wall, (float, int)) and time.monotonic() < float(until_wall):
                        time.sleep(0.02)
                        continue

                    time.sleep(0.02)
                    continue
                eng_list = getattr(ext, "_sim_engines", None)
                if isinstance(eng_list, list) and len(eng_list) > 0:
                    now = time.perf_counter()
                    dt = now - last
                    last = now
                    dt = max(0.001, min(dt, 0.1))
                    scaled = dt * speed_value
                    for sim in eng_list:
                        if sim is not None:
                            try:
                                sim.tick(scaled)
                            except Exception:
                                pass
                    tick_state["count"] += 1
                    if tick_state["count"] == 1:
                        print("[SIM] tick 동작 확인 (first tick, multi)", flush=True)
                    if all(getattr(s, "is_done", False) for s in eng_list if s is not None):
                        print("[SIM] 멀티 종료 감지", flush=True)
                        _enqueue_control_action(ext, SimUiControlAction.EXPORT_XLSX.value)
                        break
                else:
                    sim = getattr(ext, "_sim_engine", None)
                    if sim is None:
                        break
                    now = time.perf_counter()
                    dt = now - last
                    last = now
                    dt = max(0.001, min(dt, 0.1))
                    sim.tick(dt * speed_value)
                    tick_state["count"] += 1
                    if tick_state["count"] == 1:
                        print("[SIM] tick 동작 확인 (first tick)", flush=True)
                    if sim.is_done:
                        print("[SIM] 종료 감지", flush=True)
                        _enqueue_control_action(ext, SimUiControlAction.EXPORT_XLSX.value)
                        break
                time.sleep(0.02)
        except Exception as err:
            # 원인 파악을 위해 traceback까지 출력한다.
            try:
                import traceback

                print(f"[SIM] tick thread 예외: {err}", flush=True)
                print(traceback.format_exc(), flush=True)
            except Exception:
                print(f"[SIM] tick thread 예외: {err}", flush=True)

    th = threading.Thread(target=_tick_loop, name="morph.tbs_control_2.sim_tick", daemon=True)
    ext._sim_thread = th
    th.start()


def _reset_sim_motion_before_json_run(
    ext: Any,
    job: Dict[str, Any],
    *,
    runner_obj: Any = None,
) -> None:
    """시뮬 중 **새 JSON** 직전 — 해당 화면만 애니 중지·TBS_OFFSET 초기화( evaluator replay 는 유지)."""
    try:
        scr_i = int(str((job or {}).get("tbs_sim_screen", "1") or "1").strip() or "1")
    except Exception:
        scr_i = 1
    scr_i = max(1, scr_i)
    ctx = _usd_context_name_for_sim_screen(ext, scr_i)
    extra = (job or {}).get("parsed") if isinstance((job or {}).get("parsed"), list) else []
    runner_was_running = False
    if runner_obj is not None:
        try:
            runner_was_running = bool(getattr(runner_obj, "is_running", lambda: False)())
        except Exception:
            runner_was_running = False
    if runner_obj is not None:
        try:
            if runner_was_running:
                runner_obj.pause(cancel_all_move_rotate=True)
                th = getattr(runner_obj, "_lam_thread", None)
                if th is not None and getattr(th, "is_alive", lambda: False)():
                    try:
                        th.join(timeout=3.0)
                    except Exception:
                        pass
        except Exception:
            try:
                if getattr(runner_obj, "is_running", lambda: False)():
                    runner_obj.pause(cancel_all_move_rotate=True)
            except Exception:
                pass
    try:
        from . import sim_multi_diag as _mdiag

        _mdiag.log_anim_reset(
            ext,
            screen=scr_i,
            ctx=ctx,
            motion_only=True,
            runner_was_running=runner_was_running,
            path_count=len(extra),
            reason="pre_json_run",
        )
    except Exception:
        pass
    active_ep = _resolve_foup_proc_active_ep(ext, scr_i, dict(job or {}))
    _restore_sim_prim_motion_to_initial(
        ext,
        extra_steps=extra if extra else None,
        usd_context_name=ctx,
        preserve_foup_offsets=True,
        foup_proc_active_ep=active_ep,
        motion_only=True,
        include_registry_paths=False,
    )


def _restore_all_sim_channels_prim_motion(ext: Any, **kwargs: Any) -> None:
    """멀티 분할 시 각 화면 스테이지를 독립적으로 초기 자세 복원."""
    try:
        n_ch = max(1, min(4, int(getattr(ext, "_sim_viewport_split_count", 1) or 1)))
    except Exception:
        n_ch = 1
    for si in range(1, n_ch + 1):
        ctx = _usd_context_name_for_sim_screen(ext, si)
        _restore_sim_prim_motion_to_initial(ext, usd_context_name=ctx, **kwargs)


def _restore_sim_prim_motion_to_initial(
    ext: Any,
    *,
    extra_steps: Optional[List[Dict[str, Any]]] = None,
    usd_context_name: Optional[str] = None,
    preserve_foup_offsets: bool = False,
    foup_proc_active_ep: str = "",
    motion_only: bool = False,
    include_registry_paths: bool = True,
) -> None:
    """시뮬 **시작·리셋** 및 **JSON 전환** 시 MOVE·ROTATE·FOUP·USD_TIMELINE prim 을 초기 자세로.

    포트 LOT 숨김/보임(visibility)은 건드리지 않는다 — transform·TBS_OFFSET·인스턴스 replay 만 복원.
    ``preserve_foup_offsets=True`` 이면 FOUP 공정 플래그를 지우지 않고 EP plateau 를 유지한다.
    ``motion_only=True`` 이면 JSON 직전용 — evaluator invalidate/replay 해제는 생략한다.
    """
    paths_seen: set[str] = set()
    paths: List[str] = []

    def _add(path: str) -> None:
        p = str(path or "").strip()
        if p.startswith("/") and p not in paths_seen:
            paths_seen.add(p)
            paths.append(p)

    try:
        from . import port_lot_visibility as _plv

        if not preserve_foup_offsets:
            for p in (_plv.load_port_lot_prim_paths() or {}).values():
                _add(str(p))
    except Exception:
        pass

    try:
        from .tbs_lam_sequence_engine import _collect_prim_paths_for_reset

        runners = getattr(ext, "_sim_runners_by_screen", None)
        scr_filter: Optional[int] = None
        if usd_context_name is not None:
            try:
                cn = str(usd_context_name or "").strip()
                names = list(getattr(ext, "_sim_multi_context_names", []) or [])
                for i, nm in enumerate(names):
                    if str(nm or "").strip() == cn:
                        scr_filter = i + 2
                        break
            except Exception:
                scr_filter = None
        if isinstance(runners, dict):
            if scr_filter is not None:
                r = runners.get(str(scr_filter))
                if r is not None:
                    for p in _collect_prim_paths_for_reset(
                        getattr(r, "_lam_last_steps", None) or []
                    ):
                        _add(p)
            else:
                for r in runners.values():
                    if r is not None:
                        for p in _collect_prim_paths_for_reset(
                            getattr(r, "_lam_last_steps", None) or []
                        ):
                            _add(p)
        if scr_filter is None or scr_filter == 1:
            r0 = getattr(ext, "_sim_runner", None)
            if r0 is not None:
                for p in _collect_prim_paths_for_reset(
                    getattr(r0, "_lam_last_steps", None) or []
                ):
                    _add(p)
        if extra_steps:
            for p in _collect_prim_paths_for_reset(extra_steps):
                _add(p)
    except Exception:
        pass

    try:
        from .tbs_split_composed_loader import get_split_runtime_for_usd_context

        if include_registry_paths:
            rt_ctx = get_split_runtime_for_usd_context(ext, usd_context_name)
            reg = rt_ctx.registry if rt_ctx is not None else getattr(ext, "_tbs_registry", None)
            if reg is not None and hasattr(reg, "all_instances"):
                for inst in reg.all_instances():
                    _add(str(getattr(inst, "prim_path", "") or ""))
    except Exception:
        if include_registry_paths:
            try:
                reg = getattr(ext, "_tbs_registry", None)
                if reg is not None and hasattr(reg, "all_instances"):
                    for inst in reg.all_instances():
                        _add(str(getattr(inst, "prim_path", "") or ""))
            except Exception:
                pass

    def _do_on_main() -> None:
        # USD write / stage 접근은 반드시 main thread 에서만 수행한다.
        try:
            from .sim_channel_scope import stop_channel_animations

            stop_channel_animations(
                usd_context_name,
                preserve_foup_port_lot_prims=bool(preserve_foup_offsets),
                diag_reason=f"restore_motion motion_only={motion_only}",
            )
        except Exception:
            if not preserve_foup_offsets:
                try:
                    from . import tbs_lam_rotate_animation as _lrx
                    from . import tbs_lam_translate_animation as _ltx

                    _ltx.stop_all_translate_animations()
                    _lrx.stop_all_rotate_animations()
                except Exception:
                    pass
            try:
                stop_all_translate_animations(preserve_foup_port_lot_prims=bool(preserve_foup_offsets))
                stop_all_rotate_animations()
                stop_all_curve_animations()
            except Exception:
                pass
        try:
            from . import port_lot_visibility as _plv

            if not preserve_foup_offsets:
                # 화면(USD 컨텍스트)별 독립: 이 화면의 FOUP 공정 상태만 비운다.
                # (전역 비우기를 하면 다른 화면에서 공정 중인 FOUP 의 lift 상태가 풀려 함께 내려간다.)
                _plv.clear_foup_in_progress(usd_context_name=usd_context_name)
                _plv.clear_foup_lifted(usd_context_name=usd_context_name)
                _plv.restore_port_lot_prims_to_authoring(
                    usd_context_name=usd_context_name,
                    foup_proc_active_ep=str(foup_proc_active_ep or ""),
                )
        except Exception:
            pass
        try:
            from .tbs_split_composed_loader import get_split_runtime_for_usd_context

            rt_sch = get_split_runtime_for_usd_context(ext, usd_context_name)
            sch = rt_sch.scheduler if rt_sch is not None else getattr(ext, "_tbs_scheduler", None)
            stop_fn = getattr(sch, "stop_all", None) if sch is not None else None
            if callable(stop_fn):
                try:
                    from . import sim_multi_diag as _mdiag

                    _mdiag.log_scheduler_stop(
                        ctx=usd_context_name,
                        scheduler=sch,
                        motion_only=bool(motion_only),
                        reason="restore_motion",
                    )
                except Exception:
                    pass
                stop_fn()
        except Exception:
            pass
        if paths:
            from . import port_lot_visibility as _plv_paths
            from .tbs_lam_sequence_engine import _reset_tbs_offset_ops_for_paths
            from .tbs_usd_stage_context import pop_usd_context_name, push_usd_context_name

            reset_paths = list(paths)
            if preserve_foup_offsets:
                try:
                    port_set = set(_plv_paths._iter_unique_mapped_prim_paths())
                    reset_paths = [p for p in reset_paths if p not in port_set]
                except Exception:
                    pass
            if reset_paths:
                prev_ctx = push_usd_context_name(usd_context_name)
                try:
                    _reset_tbs_offset_ops_for_paths(reset_paths, usd_context_name=usd_context_name)
                finally:
                    pop_usd_context_name(prev_ctx)

        if not motion_only:
            try:
                from .tbs_lam_sequence_editor import _range_start_seconds_for_instance
                from .tbs_split_composed_loader import get_split_runtime_for_usd_context

                rt_ctx = get_split_runtime_for_usd_context(ext, usd_context_name)
                reg = rt_ctx.registry if rt_ctx is not None else getattr(ext, "_tbs_registry", None)
                ev = rt_ctx.evaluator if rt_ctx is not None else getattr(ext, "_tbs_evaluator", None)
                sch = rt_ctx.scheduler if rt_ctx is not None else getattr(ext, "_tbs_scheduler", None)
                if reg is not None and hasattr(reg, "all_instances"):
                    for inst in reg.all_instances():
                        pp = str(getattr(inst, "prim_path", "") or "").strip()
                        if not pp.startswith("/"):
                            continue
                        try:
                            inst.virtual_time = _range_start_seconds_for_instance(inst)
                            inst.state = "stopped"
                        except Exception:
                            pass
                        if ev is not None:
                            for fn_name in (
                                "end_replay_mode",
                                "end_master_timeline_mode",
                                "invalidate_mapping",
                                "force_rebuild_attr_cache",
                            ):
                                fn = getattr(ev, fn_name, None)
                                if callable(fn):
                                    try:
                                        fn(pp)
                                    except Exception:
                                        pass
            except Exception:
                pass

        try:
            usd_animation_control.stop_usd_animation(usd_context_name)
            usd_animation_control.reset_timeline_to_zero(usd_context_name)
        except Exception:
            pass

    try:
        if threading.current_thread() is threading.main_thread():
            _do_on_main()
        else:
            from .tbs_lam_sequence_engine import _dispatch_main_wait

            _dispatch_main_wait(_do_on_main, timeout=20.0)
    except Exception as exc:
        print(f"[TBS/SIM] restore motion failed: {exc}", flush=True)


def on_sim_stop_clicked(
    ext: Any,
    *,
    freeze_ep_timeline: bool = True,
    clear_prerun_cache: bool = True,
) -> None:
    """
    시뮬레이션 중지(Stop).

    ``freeze_ep_timeline`` (기본 True):
    - True  → 정지 버튼/웹/종료 경로. EP 막대그래프를 **현 상태 그대로 동결**(위젯·렌더 상태 유지).
    - False → start/reset 내부 호출. 막대그래프 위젯·상태를 초기화(다음 실행을 깨끗이 시작).

    목표(요구사항):
    - 멀티 화면에서 화면별 runner/큐/인터럽트/일시정지(pause) 상태가 남아
      다음 실행이 깨지는 회귀(애니가 있는데 재생이 안 됨, 진행률 0% 교착 등)를 방지한다.

    주요 동작 요약(추적용):
    - (세대 토큰) ``ext._sim_run_gen`` 증가:
      - UI 큐/애니 이벤트 처리부가 이전 실행의 payload를 무시하도록 만드는 "세대" 구분자.
    - (엔진 stop) ``TBSSimulationEngine.stop()`` 호출:
      - 단일 엔진: ``ext._sim_engine.stop()``
      - 멀티 엔진: ``ext._sim_engines[*].stop()``
      - 엔진 내부에서는 simpy 프로세스 종료, 진행 emit 중단, 최종 상태 정리 등이 수행된다.
    - (UI 업데이트 detach) ``_detach_sim_update(ext)``:
      - 엔진/워커가 올리던 UI 업데이트 subscription(로그/진행현황 drain)을 해제한다.
    - (화면별 애니 상태 정리) ``_sim_runners_by_screen`` 및 아래 dict들을 전부 초기화/clear:
      - pending/active: ``_sim_anim_pending_by_screen``, ``_sim_anim_active_by_screen``
      - pause: ``_sim_tick_pause_events_by_screen``, ``_sim_tick_pause_until_wall_by_screen``
      - interrupt(공정시간우선): ``_sim_interrupt_anim_event_by_screen``
    - (게이트/일시정지 해제) ``_sim_tick_pause_event``, ``_sim_gate_pause_event`` clear +
      공정확인 다이얼로그(``_sim_gate_dialog``) 강제 destroy
    - (EP 타임라인 리셋) 그래프 상태/위젯 파기:
      - state dict: ``_sim_ep_timeline_state_by_screen``, ``_sim_ep_occ_timeline_state_by_screen``,
        ``_sim_last_ports_occupancy_by_screen``
      - (레거시) 혹시 남아있으면 ``_sim_ep_timeline_ui_sub`` unsubscribe
      - widget: 채널별 ``ep_timeline_widget`` destroy
    """
    # stop/reset 이후 남아있는 큐 아이템을 무시하기 위한 세대 토큰 증가
    try:
        ext._sim_run_gen = int(getattr(ext, "_sim_run_gen", 0) or 0) + 1
    except Exception:
        ext._sim_run_gen = 1
    try:
        from .sim_multi_diag import set_session_active

        set_session_active(False)
    except Exception:
        pass
    try:
        _set_sim_prerun_ui_busy(ext, False)
    except Exception:
        pass
    # FOUP 진행중 보호 표시 정리(안전망):
    # - END 후 1초 unmark 가 잡히기 전에 stop 이 들어오면 보호 플래그가 남을 수 있다.
    # - 보류 중인 unmark subscription 도 함께 해제하여 잔여 콜백을 차단한다.
    try:
        from . import port_lot_visibility as _plv  # type: ignore
        _plv.clear_foup_in_progress()
    except Exception:
        pass
    try:
        holders = getattr(ext, "_foup_unmark_subs", None)
        if isinstance(holders, list):
            for h in holders:
                try:
                    s = h.get("sub") if isinstance(h, dict) else None
                    if s is not None:
                        s.unsubscribe()
                except Exception:
                    pass
                try:
                    if isinstance(h, dict):
                        h["done"] = True
                        h["sub"] = None
                except Exception:
                    pass
            ext._foup_unmark_subs = []
    except Exception:
        pass
    # FOUP 라벨 idle 리셋 + 보류 중인 라벨 reset sub 정리(안전망):
    # - 진행 중이던 라벨이 'DONE/대기' 갱신 없이 stop 이 들어오면 텍스트가 잔여로 남을 수 있다.
    # - 보류된 1회성 update sub 도 모두 unsubscribe 해 잔여 콜백을 차단한다.
    try:
        holders2 = getattr(ext, "_foup_label_reset_subs", None)
        if isinstance(holders2, list):
            for h in holders2:
                try:
                    s = h.get("sub") if isinstance(h, dict) else None
                    if s is not None:
                        s.unsubscribe()
                except Exception:
                    pass
                try:
                    if isinstance(h, dict):
                        h["done"] = True
                        h["sub"] = None
                except Exception:
                    pass
            ext._foup_label_reset_subs = []
    except Exception:
        pass
    try:
        _reset_all_foup_labels(ext)
    except Exception:
        pass
    # FOUP material 안전망: 정상 종료가 아니어도 다음 실행 시작 상태가 깨끗하도록
    # 모든 매핑 prim 의 material 을 기본(phong1)으로 일괄 복원한다.
    # 분할화면(보조 USD 컨텍스트)도 함께 처리한다.
    try:
        from . import port_lot_visibility as _plv  # type: ignore
        # 기본 컨텍스트
        try:
            _plv.restore_port_lot_prims_to_default_material(None)
        except Exception:
            pass
        # 분할화면 보조 컨텍스트들
        try:
            ctx_names = list(getattr(ext, "_sim_multi_context_names", []) or [])
        except Exception:
            ctx_names = []
        for cn in ctx_names:
            try:
                _plv.restore_port_lot_prims_to_default_material(str(cn))
            except Exception:
                pass
    except Exception:
        pass
    # 프리런/재생 모드 정리
    try:
        stop_playback_runtime(ext)
    except Exception:
        pass
    try:
        clear_proc_gates(ext)
    except Exception:
        pass
    try:
        p = getattr(ext, "_sim_playback_player", None)
        if p is not None and hasattr(p, "stop"):
            try:
                p.stop()
            except Exception:
                pass
    except Exception:
        pass
    try:
        sub = getattr(ext, "_sim_playback_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_playback_ui_sub = None
    except Exception:
        pass
    try:
        set_sim_playback_active(ext, False)
    except Exception:
        pass
    try:
        ev = getattr(ext, "_sim_prerun_done_evt", None)
        if ev is not None and hasattr(ev, "clear"):
            ev.clear()
        if clear_prerun_cache:
            ext._sim_prerun_results_by_screen = None
            ext._sim_playback_schedule_by_screen = None
            try:
                from .control_sim_playback_plan import clear_playback_plan_runtime_state

                clear_playback_plan_runtime_state(ext)
            except Exception:
                pass
    except Exception:
        pass
    for eng in list(getattr(ext, "_sim_engines", None) or []):
        if eng is None:
            continue
        try:
            eng.stop()
        except Exception:
            pass
    try:
        ext._sim_engines = []
    except Exception:
        pass
    sim = getattr(ext, "_sim_engine", None)
    if sim is not None:
        try:
            sim.stop()
        except Exception:
            pass
    _detach_sim_update(ext)
    # 화면별 runner/큐/상태를 모두 정리해야 stop/reset 후 재시작에서
    # "애니가 있는데도 재생이 안 되는" 잔여 상태(구독/타이머/인터럽트) 회귀를 막을 수 있다.
    try:
        runners = getattr(ext, "_sim_runners_by_screen", None)
        if isinstance(runners, dict):
            for r in list(runners.values()):
                try:
                    if r is not None:
                        r.pause()
                except Exception:
                    pass
    except Exception:
        pass
    runner = getattr(ext, "_sim_runner", None)
    if runner is not None:
        try:
            runner.pause()
        except Exception:
            pass
    # pause 상태 해제
    pe = getattr(ext, "_sim_tick_pause_event", None)
    if pe is not None:
        try:
            pe.clear()
        except Exception:
            pass
    ge = getattr(ext, "_sim_gate_pause_event", None)
    if ge is not None:
        try:
            ge.clear()
        except Exception:
            pass
    # 공정확인 창이 열려있으면 강제 종료(리셋/중지 시 다음 실행에 pause가 남지 않게)
    try:
        w = getattr(ext, "_sim_gate_dialog", None)
        if w is not None:
            try:
                w.visible = False
            except Exception:
                pass
            try:
                w.destroy()
            except Exception:
                pass
        ext._sim_gate_dialog = None
    except Exception:
        pass
    try:
        ext._sim_tick_pause_until_wall = None
    except Exception:
        pass
    try:
        ext._sim_anim_pending = []
    except Exception:
        pass
    try:
        pending_by = getattr(ext, "_sim_anim_pending_by_screen", None)
        if isinstance(pending_by, dict):
            for k in list(pending_by.keys()):
                pending_by[k] = []
    except Exception:
        pass
    try:
        active_by = getattr(ext, "_sim_anim_active_by_screen", None)
        if isinstance(active_by, dict):
            for k in list(active_by.keys()):
                active_by[k] = {}
    except Exception:
        pass
    try:
        until_by = getattr(ext, "_sim_tick_pause_until_wall_by_screen", None)
        if isinstance(until_by, dict):
            for k in list(until_by.keys()):
                until_by[k] = None
    except Exception:
        pass
    try:
        pause_by = getattr(ext, "_sim_tick_pause_events_by_screen", None)
        if isinstance(pause_by, dict):
            for ev in list(pause_by.values()):
                try:
                    if ev is not None:
                        ev.clear()
                except Exception:
                    pass
    except Exception:
        pass
    try:
        ie_by = getattr(ext, "_sim_interrupt_anim_event_by_screen", None)
        if isinstance(ie_by, dict):
            for ev in list(ie_by.values()):
                try:
                    if ev is not None:
                        ev.clear()
                except Exception:
                    pass
    except Exception:
        pass
    # EP 타임라인 그래프 상태/위젯 초기화(리셋/정지 후 누적 잔상 방지)
    # freeze 시(정지 버튼)에는 막대그래프 렌더 상태를 그대로 두어 현 상태로 동결한다.
    if not freeze_ep_timeline:
        try:
            ext._sim_ep_timeline_state_by_screen = {}
        except Exception:
            pass
        try:
            ext._sim_ep_occ_timeline_state_by_screen = {}
        except Exception:
            pass
    try:
        ext._sim_last_ports_occupancy_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_post_anim_port_applied_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_renewal_port_defer_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_pending_post_anim_port_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_post_anim_src_by_screen = {}
    except Exception:
        pass
    try:
        sub = getattr(ext, "_sim_ep_timeline_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_ep_timeline_ui_sub = None
    except Exception:
        pass
    if not freeze_ep_timeline:
        try:
            ext._sim_ep_timeline_virtual_time_by_screen = {}
        except Exception:
            pass
    # freeze 시에는 EP 막대그래프 위젯을 destroy 하지 않고 현 상태로 둔다(빈 영역 방지).
    if not freeze_ep_timeline:
        try:
            chans = getattr(ext, "_sim_monitor_channels", None)
            if isinstance(chans, list):
                for ch in chans:
                    if not isinstance(ch, dict):
                        continue
                    w = ch.get("progress_ep_timeline_widget", None)
                    if w is not None:
                        try:
                            w.destroy()
                        except Exception:
                            pass
                        ch["progress_ep_timeline_widget"] = None
                    w2 = ch.get("ep_timeline_widget", None)
                    if w2 is not None:
                        try:
                            w2.destroy()
                        except Exception:
                            pass
                        ch["ep_timeline_widget"] = None
        except Exception:
            pass
    try:
        ext._sim_engines = []
    except Exception:
        pass


def on_sim_restart_clicked(ext: Any) -> None:
    """이전 프리런 결과로 재생만 다시 시작 (재프리런/설정 변경 없음)."""
    print("[SIM] restart clicked", flush=True)
    bundle = getattr(ext, "_sim_restart_prerun_bundle", None)
    if not isinstance(bundle, dict) or not isinstance(bundle.get("results"), dict):
        try:
            _append_sim_log(ext, "[SIM] 재시작 실패 — 보관된 프리런 결과가 없습니다. 먼저 시작하세요.")
        except Exception:
            pass
        print("[SIM] restart skipped — no prerun bundle", flush=True)
        return

    # 정지 후에도 잔여 애니/구독이 있을 수 있어 항상 정리 후 번들 복원
    try:
        on_sim_stop_clicked(
            ext, freeze_ep_timeline=False, clear_prerun_cache=True
        )
    except Exception as exc:
        print(f"[SIM] restart pre-stop: {exc}", flush=True)

    results = _restore_prerun_restart_bundle(ext)
    if not isinstance(results, dict) or not results:
        try:
            _append_sim_log(ext, "[SIM] 재시작 실패 — 프리런 번들 복원 실패")
        except Exception:
            pass
        print("[SIM] restart restore failed", flush=True)
        return

    try:
        # 화면별 stop 마크가 남아있으면 재생이 스킵될 수 있음
        ext._sim_stopped_screens = set()
    except Exception:
        pass

    try:
        _append_sim_log(ext, "[SIM] 재시작 — 이전 프리런으로 재생 재개")
    except Exception:
        pass
    print(
        f"[SIM] restart screens={sorted(int(s) for s in results.keys())}",
        flush=True,
    )

    try:
        set_sim_playback_active(ext, True)
        ext._sim_anim_pending = []
        ext._sim_anim_pending_by_screen = {}
        ext._sim_playback_done = False
        ext._sim_prerun_timetable_printed = True
    except Exception:
        pass

    try:
        by = getattr(ext, "_sim_last_total_est_by_screen", None)
        if not isinstance(by, dict):
            by = {}
            ext._sim_last_total_est_by_screen = by
        for scr, res in results.items():
            try:
                by[str(int(scr))] = float(res.final_sim_time)
            except Exception:
                continue
    except Exception:
        pass

    playback_engs: List[Any] = []
    try:
        max_scr = max(int(s) for s in results.keys())
    except Exception:
        max_scr = 1
    for i in range(1, max_scr + 1):
        rr = results.get(int(i))
        if rr is None:
            continue
        playback_engs.append(PlaybackEngine(final_sim_time=float(rr.final_sim_time)))
    try:
        ext._sim_engines = playback_engs
        ext._sim_engine = playback_engs[0] if playback_engs else None
    except Exception:
        pass

    def _speed() -> float:
        try:
            m = getattr(ext, "_sim_speed_model", None)
            return float(m.get_value_as_float()) if m is not None else 1.0
        except Exception:
            return 1.0

    def _timeline_event_gate(scr: int) -> bool:
        return can_emit_timeline_event(ext, int(scr))

    try:
        _prepare_playback_emit_environment(ext, results)
    except Exception as exc:
        print(f"[SIM] restart prepare emit: {exc}", flush=True)

    try:
        bootstrap_playback_after_prerun(
            ext,
            results,
            _make_playback_emit_fn(ext, results),
            _speed,
            gate_fn=_timeline_event_gate,
        )
    except Exception as exc:
        print(f"[SIM] restart bootstrap failed: {exc}", flush=True)
        try:
            _append_sim_log(ext, f"[SIM] 재시작 실패: {exc}")
        except Exception:
            pass
        return

    # 첫 이벤트 전에도 진행현황이 움직이도록 초기 payload
    try:
        for scr, rr in results.items():
            try:
                scr_i = int(scr)
            except Exception:
                scr_i = 1
            p0 = {
                "tbs_sim_screen": str(scr_i),
                "sim_time": "0.00",
                "sim_total_est_sec": f"{float(rr.final_sim_time):.2f}",
                "label": "대기",
                "detail": "",
                "status": "RUNNING",
                "elapsed": "0.0",
                "total": "0.0",
                "percent": "0",
            }
            _update_sim_progress(ext, p0)
    except Exception:
        pass

    # 시작(프리런 완료)과 동일 — UI update 구독이 있어야 재생 tick 이 돈다
    try:
        import omni.kit.app as app  # type: ignore

        sub = getattr(ext, "_sim_playback_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_playback_ui_sub = (
            app.get_app()
            .get_update_event_stream()
            .create_subscription_to_pop(
                lambda _e: _tick_playback_timeline(ext),
                name="morph.tbs_control_2:sim_playback_tick",
            )
        )
        print("[SIM] restart playback tick subscribed", flush=True)
    except Exception as exc:
        print(f"[SIM] restart tick subscribe failed: {exc}", flush=True)

    try:
        ev = getattr(ext, "_sim_prerun_done_evt", None)
        if ev is None:
            ext._sim_prerun_done_evt = threading.Event()
            ev = ext._sim_prerun_done_evt
        if hasattr(ev, "set"):
            ev.set()
    except Exception:
        pass

    try:
        _append_sim_log(ext, "[SIM] 재시작 완료: 보관된 타임라인을 재생합니다.")
    except Exception:
        pass
    print("[SIM] restart playback started", flush=True)

def on_sim_reset_clicked(ext: Any) -> None:
    """
    시뮬레이션 리셋(Reset).

    Stop과의 차이(추적용):
    - 먼저 ``on_sim_stop_clicked(ext)``를 호출하여 "실행 중인 것"을 완전히 끊는다.
    - 그 다음, UI 텍스트/포트 박스/그래프 등을 "초기 화면" 상태로 되돌린다.
    - 추가로, 포트 LOT authoring 캐시를 clear 하여 다음 실행에서 경로/프림/표시가
      이전 실행의 잔여 데이터에 영향을 받지 않게 한다.

    이 함수가 만지는 대표 상태:
    - 엔진/러너 참조: ``ext._sim_engine = None`` 및 ``ext._sim_engines = []``
    - UI 텍스트: history/progress/port_state 라벨 텍스트 초기화
    - 포트 박스: 각 포트를 '-'로 표기하고 스타일을 초기화
    - EP 타임라인: t=0.0 초기 렌더 + 관련 dict를 완전 초기화
    """
    try:
        unlock_timetable_rows(ext)
        _clear_sim_timetable_storage(ext)
    except Exception:
        pass
    try:
        _clear_prerun_restart_bundle(ext)
    except Exception:
        pass
    on_sim_stop_clicked(ext, freeze_ep_timeline=False)
    try:
        _restore_all_sim_channels_prim_motion(ext)
    except Exception:
        pass
    try:
        clear_port_lot_authoring_cache()
    except Exception:
        pass
    ext._sim_engine = None
    try:
        ext._sim_engines = []
    except Exception:
        pass
    if getattr(ext, "_sim_history_text", None):
        ext._sim_history_text.set_value("[SIM] 리셋 완료")
    if getattr(ext, "_sim_progress_text", None):
        ext._sim_progress_text.set_value("[진행현황] 없음")
    if getattr(ext, "_sim_port_state_text", None):
        ext._sim_port_state_text.set_value("[포트상태] 없음")
    if getattr(ext, "_sim_port_state_label", None) is not None:
        ext._sim_port_state_label.text = "[포트상태] 없음"
    chans_r = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans_r, list) and len(chans_r) > 1:
        for ch in chans_r:
            if not isinstance(ch, dict):
                continue
            try:
                si = int(ch.get("screen", 1))
            except Exception:
                si = 1
            ht = "[SIM] 리셋 완료" if si == 1 else f"[SIM·화면{si}] 리셋 완료"
            pt = "[진행현황] 없음" if si == 1 else f"[진행현황·화면{si}] 없음"
            ph = f"[포트상태·화면{si}] 없음"
            pl = ch.get("progress_label")
            phdr = ch.get("port_header")
            try:
                _set_channel_history_text(ch, ht)
            except Exception:
                pass
            if pl is not None:
                pl.text = pt
            if phdr is not None:
                phdr.text = ph
            cells = ch.get("port_cells") or {}
            boxes = ch.get("port_cell_boxes") or {}
            for port in ("INOUT", "BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
                if port in cells:
                    cells[port].text = "IN/OUT:-" if port == "INOUT" else f"{port}:-"
                try:
                    _set_port_box_style(ext, port, "-", boxes)
                except Exception:
                    pass
            ep3c = ch.get("port_ep3_cell")
            if ep3c is not None:
                ep3c.text = "EP3:-"
            try:
                _set_port_box_style(ext, "EP3", "-", boxes)
            except Exception:
                pass
            # 멀티 채널도 EP 타임라인을 즉시 t=0으로 재렌더(총시간 라벨 잔상 방지)
            try:
                occ0 = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
                by_occ = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
                if not isinstance(by_occ, dict):
                    by_occ = {}
                    ext._sim_last_ports_occupancy_by_screen = by_occ
                by_occ[str(si)] = dict(occ0)
                _update_ep_timeline_under_port_state(ext, ch, occ0, "0.0")
            except Exception:
                pass
    else:
        try:
            chans_s = getattr(ext, "_sim_monitor_channels", None)
            if isinstance(chans_s, list) and chans_s and isinstance(chans_s[0], dict):
                _set_channel_history_text(chans_s[0], "[SIM] 리셋 완료")
        except Exception:
            pass
        if getattr(ext, "_sim_history_label", None) is not None:
            try:
                ext._sim_history_label.text = "[SIM] 리셋 완료"
            except Exception:
                pass
        if getattr(ext, "_sim_progress_label", None) is not None:
            ext._sim_progress_label.text = "[진행현황] 없음"
        if getattr(ext, "_sim_port_state_header_label", None) is not None:
            ext._sim_port_state_header_label.text = "[포트상태] 없음"
        cells = getattr(ext, "_sim_port_cells", {}) or {}
        for port in ("INOUT", "BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
            if port in cells:
                cells[port].text = "IN/OUT:-" if port == "INOUT" else f"{port}:-"
            try:
                _set_port_box_style(ext, port, "-")
            except Exception:
                pass
        if getattr(ext, "_sim_port_ep3_cell", None) is not None:
            ext._sim_port_ep3_cell.text = "EP3:-"
        try:
            _set_port_box_style(ext, "EP3", "-")
        except Exception:
            pass
        # 단일 채널 EP 타임라인 초기 렌더(t=0.0)
        try:
            occ0 = {k: "" for k in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3")}
            ext._sim_last_ports_occupancy_by_screen["1"] = dict(occ0)
            # 단일 채널은 channels[0]이거나 ext 레거시 참조이므로, 가능하면 channels에서 가져온다.
            chans_s = getattr(ext, "_sim_monitor_channels", None)
            if isinstance(chans_s, list) and len(chans_s) >= 1 and isinstance(chans_s[0], dict):
                _update_ep_timeline_under_port_state(ext, chans_s[0], occ0, "0.0")
        except Exception:
            pass
    ext._sim_progress_rows = {}
    ext._sim_progress_history = []
    ext._sim_progress_start_times = {}
    # EP 타임라인도 완전 초기화
    try:
        ext._sim_ep_timeline_state_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_ep_occ_timeline_state_by_screen = {}
    except Exception:
        pass
    try:
        ext._sim_last_ports_occupancy_by_screen = {}
    except Exception:
        pass
    # 총시간(끝 라벨) 캐시도 초기화(리셋 후 이전 총시간이 남는 문제 방지)
    try:
        ext._sim_last_total_est_by_screen = {}
    except Exception:
        pass
    # 진행현황(텍스트용) 마지막 payload도 초기화(리셋 후 DONE/총시간 잔상 방지)
    try:
        ext._sim_progress_last_payload_by_screen = {}
    except Exception:
        pass
    try:
        clear_progress_step_state(ext)
    except Exception:
        pass
    # 프리런/재생 상태도 초기화(리셋 후 이전 결과 잔상 방지)
    try:
        ext._sim_prerun_results_by_screen = None
        ext._sim_playback_schedule_by_screen = None
        try:
            from .control_sim_playback_plan import clear_playback_plan_runtime_state

            clear_playback_plan_runtime_state(ext)
        except Exception:
            pass
        ext._sim_playback_player = None
        ext._sim_playback_players_by_screen = None
        ext._sim_playback_runtime = None
        set_sim_playback_active(ext, False)
        ext._sim_playback_done = False
        ext._sim_ep_bar_prerun_by_screen = {}
        ext._sim_prerun_export_json_by_screen = {}
    except Exception:
        pass
    try:
        stop_playback_runtime(ext)
    except Exception:
        pass
    try:
        clear_proc_gates(ext)
    except Exception:
        pass
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans, list):
            for ch in chans:
                if not isinstance(ch, dict):
                    continue
                w2 = ch.get("ep_timeline_widget", None)
                if w2 is not None:
                    try:
                        w2.destroy()
                    except Exception:
                        pass
                    ch["ep_timeline_widget"] = None
    except Exception:
        pass
    try:
        sub = getattr(ext, "_sim_ep_timeline_ui_sub", None)
        if sub is not None:
            try:
                sub.unsubscribe()
            except Exception:
                pass
        ext._sim_ep_timeline_ui_sub = None
    except Exception:
        pass
    try:
        ext._sim_ep_timeline_virtual_time_by_screen = {}
    except Exception:
        pass
    # 최근 요약/대기 토큰 초기화
    try:
        ext._sim_recent_story_blocks = []
    except Exception:
        pass
    try:
        _reset_all_channel_timetables_to_idle(
            ext,
            message="타임테이블 초기화됨 — Start 후 프리런 결과 표시",
        )
    except Exception:
        pass
    # (요청으로 제거) 점 표시 기능 비활성화


def on_sim_log_view_changed(ext: Any) -> None:
    # 표시모드 제거: 항상 둘다(진행현황+이력로그)
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and len(chans) > 0:
        for ch in chans:
            if not isinstance(ch, dict):
                continue
            pf = ch.get("progress_frame")
            hf = ch.get("history_frame")
            if pf is not None:
                try:
                    pf.visible = True
                except Exception:
                    pass
            if hf is not None:
                try:
                    hf.visible = True
                except Exception:
                    pass
        return
    if getattr(ext, "_sim_progress_frame", None) is not None:
        ext._sim_progress_frame.visible = True
    if getattr(ext, "_sim_history_frame", None) is not None:
        ext._sim_history_frame.visible = True


def on_copy_sim_progress(ext: Any) -> None:
    # 확장: 진행현황 + Sim 로그(이력) 함께 복사
    progress = ""
    history = ""
    chans = getattr(ext, "_sim_monitor_channels", None)
    if isinstance(chans, list) and len(chans) > 1:
        pb: List[str] = []
        hb: List[str] = []
        for ch in chans:
            if not isinstance(ch, dict):
                continue
            try:
                si = int(ch.get("screen", 0))
            except Exception:
                si = 0
            pl = ch.get("progress_label")
            hl = ch.get("history_label")
            if pl is not None:
                pt = str(pl.text or "").strip()
                if pt:
                    pb.append(f"=== 화면{si} 진행현황 ===\n{pt}")
            if hl is not None:
                ht = _get_channel_history_text(ch).strip()
                if ht:
                    hb.append(f"=== 화면{si} Sim 로그 ===\n{ht}")
        progress = "\n\n".join(pb).strip()
        history = "\n\n".join(hb).strip()
    else:
        if getattr(ext, "_sim_progress_label", None) is not None:
            progress = ext._sim_progress_label.text or ""
        if not progress.strip() and getattr(ext, "_sim_progress_text", None):
            progress = ext._sim_progress_text.as_string or ""
        if getattr(ext, "_sim_history_label", None) is not None:
            try:
                chans_s = getattr(ext, "_sim_monitor_channels", None)
                if isinstance(chans_s, list) and chans_s and isinstance(chans_s[0], dict):
                    history = _get_channel_history_text(chans_s[0])
                else:
                    history = ext._sim_history_label.text or ""
            except Exception:
                history = ext._sim_history_label.text or ""
        if not history.strip() and getattr(ext, "_sim_history_text", None):
            history = ext._sim_history_text.as_string or ""

    text = (progress or "").strip()
    if history.strip():
        text = (text + "\n\n" if text else "") + "[Sim 로그]\n" + history.strip()
    if not text.strip():
        _append_sim_log(ext, "[SIM UI] 복사할 진행현황/Sim 로그가 없습니다.")
        return
    try:
        import omni.kit.clipboard as cb  # type: ignore
        if hasattr(cb, "copy"):
            cb.copy(text)
        elif hasattr(cb, "set_text"):
            cb.set_text(text)
        else:
            raise RuntimeError("clipboard api not found")
        _append_sim_log(ext, "[SIM UI] 진행현황+Sim 로그 복사 완료")
    except Exception:
        print("[SIM UI] 클립보드 미지원: 텍스트를 콘솔에 출력합니다.", flush=True)
        print(text, flush=True)
        _append_sim_log(ext, "[SIM UI] 클립보드 미지원으로 콘솔 출력")


def _sim_stopped_screens_set(ext: Any) -> set:
    s = getattr(ext, "_sim_stopped_screens", None)
    if not isinstance(s, set):
        s = set()
        ext._sim_stopped_screens = s
    return s


def _stop_anim_screen_worker_for_screen(ext: Any, screen_idx: int) -> None:
    workers = getattr(ext, "_sim_anim_workers_by_screen", None)
    if not isinstance(workers, dict):
        return
    key = str(max(1, int(screen_idx)))
    ent = workers.get(key)
    if not isinstance(ent, dict):
        return
    lock = ent.get("lock")
    cond = ent.get("cond")
    queue = ent.get("queue")
    th = ent.get("thread")
    try:
        if lock is not None and cond is not None and queue is not None:
            with cond:
                queue.append(_ANIM_SCREEN_WORKER_STOP)
                cond.notify()
    except Exception:
        pass
    try:
        if th is not None:
            th.join(timeout=2.0)
    except Exception:
        pass
    try:
        workers.pop(key, None)
    except Exception:
        pass


def _stop_sim_screen_only(ext: Any, screen: int) -> None:
    """단일 화면 엔진·프리런 재생·애니만 정지(다른 화면 시뮬은 유지)."""
    try:
        sc = max(1, min(4, int(screen)))
    except Exception:
        return
    sk = str(sc)
    idx = sc - 1
    _sim_stopped_screens_set(ext).add(sc)

    try:
        stop_playback_for_screen(ext, sc)
    except Exception:
        pass

    try:
        engs = list(getattr(ext, "_sim_engines", None) or [])
        if idx < len(engs) and engs[idx] is not None:
            try:
                engs[idx].stop()
            except Exception:
                pass
            engs[idx] = None
            ext._sim_engines = engs
    except Exception:
        pass
    if sc == 1:
        sim = getattr(ext, "_sim_engine", None)
        if sim is not None:
            try:
                sim.stop()
            except Exception:
                pass
            try:
                ext._sim_engine = None
            except Exception:
                pass

    try:
        _halt_screen_json_anim(ext, sc, join_sec=2.0)
    except Exception:
        pass
    try:
        _stop_anim_screen_worker_for_screen(ext, sc)
    except Exception:
        pass
    try:
        runners = getattr(ext, "_sim_runners_by_screen", None)
        if isinstance(runners, dict):
            rr = runners.get(sk)
            if rr is not None:
                try:
                    rr.pause(cancel_all_move_rotate=True)
                except Exception:
                    pass
    except Exception:
        pass
    try:
        ie_by = getattr(ext, "_sim_interrupt_anim_event_by_screen", None)
        if isinstance(ie_by, dict) and sk in ie_by and ie_by[sk] is not None:
            try:
                ie_by[sk].clear()
            except Exception:
                pass
    except Exception:
        pass
    for attr, default in (
        ("_sim_anim_pending_by_screen", []),
        ("_sim_anim_active_by_screen", {}),
        ("_sim_tick_pause_events_by_screen", None),
        ("_sim_tick_pause_until_wall_by_screen", None),
        ("_sim_interrupt_anim_event_by_screen", None),
        ("_sim_post_anim_port_applied_by_screen", None),
        ("_sim_pending_post_anim_port_by_screen", None),
        ("_sim_renewal_port_defer_by_screen", None),
    ):
        try:
            d = getattr(ext, attr, None)
            if isinstance(d, dict) and sk in d:
                if attr == "_sim_anim_pending_by_screen":
                    d[sk] = []
                elif attr == "_sim_anim_active_by_screen":
                    d[sk] = {}
                elif attr == "_sim_interrupt_anim_event_by_screen":
                    try:
                        if d[sk] is not None:
                            d[sk].clear()
                    except Exception:
                        pass
                else:
                    d.pop(sk, None)
        except Exception:
            pass
    try:
        res = getattr(ext, "_sim_prerun_results_by_screen", None)
        if isinstance(res, dict):
            res.pop(sc, None)
            res.pop(sk, None)
    except Exception:
        pass
    try:
        by_occ = getattr(ext, "_sim_last_ports_occupancy_by_screen", None)
        if isinstance(by_occ, dict):
            by_occ.pop(sk, None)
    except Exception:
        pass
    try:
        _append_sim_log(ext, f"[SIM] 화면{sc} 정지")
    except Exception:
        pass


def on_sim_start_for_screen(ext: Any, screen: int) -> None:
    """CASE 창: 담당 화면만 시뮬 시작."""
    try:
        ext._sim_startup_target_screens = [int(screen)]
        on_sim_start_clicked(ext)
    finally:
        try:
            ext._sim_startup_target_screens = None
        except Exception:
            pass


def on_sim_stop_for_screen(ext: Any, screen: int) -> None:
    """CASE 창: 담당 화면만 정지."""
    _stop_sim_screen_only(ext, int(screen))


def on_sim_reset_for_screen(ext: Any, screen: int) -> None:
    """CASE 창: 담당 화면만 리셋."""
    try:
        sc = max(1, min(4, int(screen)))
    except Exception:
        return
    on_sim_stop_for_screen(ext, sc)
    try:
        ctx = _usd_context_name_for_sim_screen(ext, sc)
        _restore_sim_prim_motion_to_initial(ext, usd_context_name=ctx)
    except Exception:
        pass
    try:
        chans = getattr(ext, "_sim_monitor_channels", None)
        if isinstance(chans, list):
            for ch in chans:
                if not isinstance(ch, dict):
                    continue
                try:
                    if int(ch.get("screen", 0)) != sc:
                        continue
                except Exception:
                    continue
                ht = "[SIM] 초기화" if sc == 1 else f"[SIM·화면{sc}] 초기화"
                pt = "[진행현황] 초기화 (시뮬레이션 시작 대기)"
                ph = f"[포트상태·화면{sc}] 초기화 (이벤트 대기)"
                try:
                    _set_channel_history_text(ch, ht)
                except Exception:
                    pass
                pl = ch.get("progress_label")
                phdr = ch.get("port_header")
                if pl is not None:
                    pl.text = pt
                if phdr is not None:
                    phdr.text = ph
                cells = ch.get("port_cells") or {}
                boxes = ch.get("port_cell_boxes") or {}
                for port in ("INOUT", "BP1", "BP2", "BP3", "BP4", "EP1", "EP2", "EP3"):
                    if port in cells:
                        cells[port].text = "IN/OUT:-" if port == "INOUT" else f"{port}:-"
                    try:
                        _set_port_box_style(ext, port, "-", boxes)
                    except Exception:
                        pass
                try:
                    reset_timetable_channel_to_idle(ch, screen=sc, ext=ext)
                except Exception:
                    pass
    except Exception:
        pass
    try:
        _sim_stopped_screens_set(ext).discard(sc)
    except Exception:
        pass


def _apply_ep_port_layout_for_sim_screen(ext: Any, screen: int, *, reason: str = "") -> None:
    """화면별 CASE EP/EBS 설정을 해당 USD 컨텍스트에 반영한다."""
    from .ebs_case_models import case_from_screen, get_sim_ebs_enabled_for_case, get_sim_ep_count_idx_for_case
    from .tbs_ep_port_visibility import (
        apply_ep_port_layout_for_context,
        ep_count_from_combo_idx,
        schedule_apply_ep_port_layout,
    )

    try:
        s = max(1, int(screen))
    except Exception:
        s = 1
    cid = case_from_screen(s)
    try:
        idx = int(get_sim_ep_count_idx_for_case(ext, cid))
    except Exception:
        idx = 0
    ep_count = ep_count_from_combo_idx(idx)
    ebs_on = bool(get_sim_ebs_enabled_for_case(ext, cid))
    rs = str(reason or f"screen{s}_ep_ebs").strip() or f"screen{s}_ep_ebs"
    if s <= 1:
        schedule_apply_ep_port_layout(
            ext,
            ep_count,
            ebs_enabled=ebs_on,
            delay_frames=2,
            reason=rs,
        )
        return
    ctx_nm = _usd_context_name_for_sim_screen(ext, s)
    if ctx_nm:
        apply_ep_port_layout_for_context(ext, str(ctx_nm), s, reason=rs)


def _sync_ep3_port_cell_visibility_for_case(ext: Any, case_id: int) -> None:
    from .ebs_case_models import CASE_A, CASE_B, get_sim_ep_count_idx_for_case

    try:
        is_ep3 = int(get_sim_ep_count_idx_for_case(ext, int(case_id))) == 1
    except Exception:
        is_ep3 = False
    if int(case_id) == CASE_A:
        bp4_row = getattr(ext, "_sim_init_bp4_row", None)
        ep3_row = getattr(ext, "_sim_init_ep3_row", None)
        f_bp4 = getattr(ext, "_sim_fault_bp4_row", None)
        f_ep3 = getattr(ext, "_sim_fault_ep3_row", None)
        extra_rows = (
            list(getattr(ext, "_sim_init_bp4_rows", None) or [])
            + list(getattr(ext, "_sim_init_ep3_rows", None) or [])
            + list(getattr(ext, "_sim_fault_bp4_rows", None) or [])
            + list(getattr(ext, "_sim_fault_ep3_rows", None) or [])
        )
    else:
        bp4_row = getattr(ext, "_ebs_b_init_bp4_row", None)
        ep3_row = getattr(ext, "_ebs_b_init_ep3_row", None)
        f_bp4 = getattr(ext, "_ebs_b_fault_bp4_row", None)
        f_ep3 = getattr(ext, "_ebs_b_fault_ep3_row", None)
        extra_rows = (
            list(getattr(ext, "_ebs_b_init_bp4_rows", None) or [])
            + list(getattr(ext, "_ebs_b_init_ep3_rows", None) or [])
            + list(getattr(ext, "_ebs_b_fault_bp4_rows", None) or [])
            + list(getattr(ext, "_ebs_b_fault_ep3_rows", None) or [])
        )
    for row in (bp4_row, ep3_row, f_bp4, f_ep3):
        if row is None:
            continue
        try:
            row.visible = bool(is_ep3)
        except Exception:
            pass
    for row in extra_rows:
        try:
            row.visible = bool(is_ep3)
        except Exception:
            pass
    _sync_ep3_port_cell_visibility(ext)


def _sync_ebs_control_visibility_for_case(ext: Any, case_id: int) -> None:
    from .ebs_case_models import CASE_A, CASE_B, get_sim_ebs_enabled_for_case, get_sim_ep_count_idx_for_case

    try:
        ebs_on = bool(get_sim_ebs_enabled_for_case(ext, int(case_id)))
        is_ep3 = int(get_sim_ep_count_idx_for_case(ext, int(case_id))) == 1
    except Exception:
        ebs_on = True
        is_ep3 = False
    if int(case_id) == CASE_A:
        bp4_init = set(getattr(ext, "_sim_init_bp4_rows", None) or [])
        bp4_fault = set(getattr(ext, "_sim_fault_bp4_rows", None) or [])
        init_rows = list(getattr(ext, "_sim_init_ebs_rows", None) or [])
        fault_rows = list(getattr(ext, "_sim_fault_ebs_rows", None) or [])
        block = getattr(ext, "_sim_timing_inout_bp_block", None)
        bp_ep_row = getattr(ext, "_sim_timing_bp_ep_row", None)
        lbl = getattr(ext, "_sim_oht_timing_label", None)
        compact_rows = list(getattr(ext, "_sim_timing_ebs_compact_rows", None) or [])
    else:
        bp4_init = set(getattr(ext, "_ebs_b_init_bp4_rows", None) or [])
        bp4_fault = set(getattr(ext, "_ebs_b_fault_bp4_rows", None) or [])
        init_rows = list(getattr(ext, "_ebs_b_init_ebs_rows", None) or [])
        fault_rows = list(getattr(ext, "_ebs_b_fault_ebs_rows", None) or [])
        block = getattr(ext, "_ebs_b_timing_inout_bp_block", None)
        bp_ep_row = getattr(ext, "_ebs_b_timing_bp_ep_row", None)
        lbl = getattr(ext, "_ebs_b_oht_timing_label", None)
        compact_rows = []
    for row in init_rows:
        try:
            row.visible = bool(ebs_on and is_ep3) if row in bp4_init else bool(ebs_on)
        except Exception:
            pass
    for row in fault_rows:
        try:
            row.visible = bool(ebs_on and is_ep3) if row in bp4_fault else bool(ebs_on)
        except Exception:
            pass
    if block is not None:
        try:
            block.visible = ebs_on
        except Exception:
            pass
    if bp_ep_row is not None:
        try:
            bp_ep_row.visible = ebs_on
        except Exception:
            pass
    if lbl is not None:
        try:
            lbl.text = "OHT→IN/OUT/EP" if ebs_on else "OHT→EP"
        except Exception:
            pass
    for row in compact_rows:
        try:
            row.visible = ebs_on
        except Exception:
            pass
    _sync_ep3_port_cell_visibility_for_case(ext, int(case_id))


def on_sim_ep_count_changed_for_case(ext: Any, case_id: int) -> None:
    from .ebs_case_models import CASE_A, CASE_B, get_case_model, get_sim_ep_count_idx_for_case, screen_from_case

    if int(case_id) == CASE_A:
        on_sim_ep_count_changed(ext)
        return

    try:
        idx = int(get_sim_ep_count_idx_for_case(ext, CASE_B))
    except Exception:
        idx = 0
    is_ep3 = idx == 1

    if getattr(ext, "_ebs_b_init_bp4_row", None) is not None:
        ext._ebs_b_init_bp4_row.visible = is_ep3
    for row in list(getattr(ext, "_ebs_b_init_bp4_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3:
        m = get_case_model(ext, CASE_B, "init_bp4")
        if m is not None:
            try:
                m.set_value(False)
            except Exception:
                pass
    if getattr(ext, "_ebs_b_init_ep3_row", None) is not None:
        ext._ebs_b_init_ep3_row.visible = is_ep3
    for row in list(getattr(ext, "_ebs_b_init_ep3_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3:
        m = get_case_model(ext, CASE_B, "init_ep3")
        if m is not None:
            try:
                m.set_value(False)
            except Exception:
                pass
    if getattr(ext, "_ebs_b_fault_bp4_row", None) is not None:
        ext._ebs_b_fault_bp4_row.visible = is_ep3
    for row in list(getattr(ext, "_ebs_b_fault_bp4_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3:
        m = get_case_model(ext, CASE_B, "fault_bp4")
        if m is not None:
            try:
                m.set_value(False)
            except Exception:
                pass
    if getattr(ext, "_ebs_b_fault_ep3_row", None) is not None:
        ext._ebs_b_fault_ep3_row.visible = is_ep3
    for row in list(getattr(ext, "_ebs_b_fault_ep3_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3:
        m = get_case_model(ext, CASE_B, "fault_ep3")
        if m is not None:
            try:
                m.set_value(False)
            except Exception:
                pass
    _sync_ebs_control_visibility_for_case(ext, CASE_B)
    try:
        _apply_ep_port_layout_for_sim_screen(
            ext,
            screen_from_case(CASE_B),
            reason="ep_count_changed_case_b",
        )
    except Exception:
        pass


def on_sim_ebs_enabled_changed_for_case(ext: Any, case_id: int) -> None:
    from .ebs_case_models import screen_from_case

    cid = int(case_id)
    _sync_ebs_control_visibility_for_case(ext, cid)
    try:
        _apply_ep_port_layout_for_sim_screen(
            ext,
            screen_from_case(cid),
            reason=f"ebs_enabled_case{cid}",
        )
    except Exception:
        pass


def _sync_ebs_control_visibility(ext: Any) -> None:
    """EBS 적용여부에 따라 버퍼 관련 UI·포트상태·공정시간 필드를 표시/숨김."""
    try:
        from .ebs_control_panel_ui import get_sim_ebs_enabled, get_sim_ep_count_idx

        ebs_on = bool(get_sim_ebs_enabled(ext))
        is_ep3 = int(get_sim_ep_count_idx(ext)) == 1
    except Exception:
        ebs_on = True
        is_ep3 = False
    bp4_init = set(getattr(ext, "_sim_init_bp4_rows", None) or [])
    bp4_fault = set(getattr(ext, "_sim_fault_bp4_rows", None) or [])
    for row in list(getattr(ext, "_sim_init_ebs_rows", None) or []):
        try:
            row.visible = bool(ebs_on and is_ep3) if row in bp4_init else bool(ebs_on)
        except Exception:
            pass
    for row in list(getattr(ext, "_sim_fault_ebs_rows", None) or []):
        try:
            row.visible = bool(ebs_on and is_ep3) if row in bp4_fault else bool(ebs_on)
        except Exception:
            pass
    block = getattr(ext, "_sim_timing_inout_bp_block", None)
    if block is not None:
        try:
            block.visible = ebs_on
        except Exception:
            pass
    bp_ep_row = getattr(ext, "_sim_timing_bp_ep_row", None)
    if bp_ep_row is not None:
        try:
            bp_ep_row.visible = ebs_on
        except Exception:
            pass
    lbl = getattr(ext, "_sim_oht_timing_label", None)
    if lbl is not None:
        try:
            lbl.text = "OHT→IN/OUT/EP" if ebs_on else "OHT→EP"
        except Exception:
            pass
    for row in list(getattr(ext, "_sim_timing_ebs_compact_rows", None) or []):
        try:
            row.visible = ebs_on
        except Exception:
            pass
    _sync_ep3_port_cell_visibility(ext)


def on_sim_ep_count_changed(ext: Any) -> None:
    try:
        idx = int(get_sim_ep_count_idx(ext))
    except Exception:
        idx = 0
    is_ep3 = idx == 1

    if getattr(ext, "_sim_init_bp4_row", None) is not None:
        ext._sim_init_bp4_row.visible = is_ep3
    for row in list(getattr(ext, "_sim_init_bp4_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3 and getattr(ext, "_sim_init_bp4_model", None) is not None:
        ext._sim_init_bp4_model.set_value(False)
    if getattr(ext, "_sim_init_ep3_row", None) is not None:
        ext._sim_init_ep3_row.visible = is_ep3
    for row in list(getattr(ext, "_sim_init_ep3_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3 and getattr(ext, "_sim_init_ep3_model", None) is not None:
        ext._sim_init_ep3_model.set_value(False)
    # 고장 포트 행도 동일 규칙 적용
    if getattr(ext, "_sim_fault_bp4_row", None) is not None:
        ext._sim_fault_bp4_row.visible = is_ep3
    for row in list(getattr(ext, "_sim_fault_bp4_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3 and getattr(ext, "_sim_fault_bp4_model", None) is not None:
        ext._sim_fault_bp4_model.set_value(False)
    if getattr(ext, "_sim_fault_ep3_row", None) is not None:
        ext._sim_fault_ep3_row.visible = is_ep3
    for row in list(getattr(ext, "_sim_fault_ep3_rows", None) or []):
        try:
            row.visible = is_ep3
        except Exception:
            pass
    if not is_ep3 and getattr(ext, "_sim_fault_ep3_model", None) is not None:
        ext._sim_fault_ep3_model.set_value(False)
    _sync_ebs_control_visibility(ext)
    try:
        from .tbs_ep_port_visibility import on_sim_ep_count_combo_changed

        on_sim_ep_count_combo_changed(ext)
    except Exception:
        pass


def receive_signal_data(ext: Any, data: str, format: str = "json") -> bool:
    parsed = parse_signal(data, format)
    if not parsed:
        return False
    run_generator_from_parsed(ext, parsed)
    return True


def run_generator_from_parsed(ext: Any, parsed: dict) -> None:
    stage = get_stage()
    if not stage:
        return
    objects = parsed.get("objects") or []
    segments = parsed.get("segments") or []
    if not objects or not segments:
        return
    for name in objects:
        if not isinstance(name, str):
            continue
        paths = find_all_prim_paths_by_name(stage, name)
        for path in paths:
            if not path:
                continue
            stop_prim_translate_animation(path)
            stop_prim_curve_animation(path)
            run_prim_translate_animation(path, segments, loop=False)


def on_refresh_prim_list(ext: Any) -> None:
    stage = get_stage()
    if not stage:
        return
    ext._tracked_paths = collect_prim_paths_safe(stage)
    refresh_object_list(ext)


def refresh_object_list(ext: Any) -> None:
    if ext._object_list_frame is None:
        return
    ext._object_list_frame.clear()
    stage = get_stage()
    if not stage:
        with ext._object_list_frame:
            ui.Label("USD를 먼저 로드하세요.")
        return

    def _valid_path(p: str) -> bool:
        try:
            return stage.GetPrimAtPath(p).IsValid()
        except (UnicodeDecodeError, UnicodeEncodeError):
            return False

    valid_paths = [p for p in ext._tracked_paths if _valid_path(p)]
    total = len(valid_paths)
    priority_prefix = (ext._priority_prefix_model.get_value_as_string().strip() or "")

    if priority_prefix:
        priority_paths: List[str] = []
        rest_paths: List[str] = []
        for p in valid_paths:
            try:
                prim = stage.GetPrimAtPath(p)
                if not prim or not prim.IsValid():
                    rest_paths.append(p)
                    continue
                name = safe_str(prim.GetName())
                if name.startswith(priority_prefix):
                    priority_paths.append(p)
                else:
                    rest_paths.append(p)
            except Exception:
                rest_paths.append(p)
        need = max(0, MAX_PRIMS_DISPLAY - len(priority_paths))
        display_paths = priority_paths[:MAX_PRIMS_DISPLAY] + rest_paths[:need]
    else:
        display_paths = valid_paths[:MAX_PRIMS_DISPLAY]

    with ext._object_list_frame:
        if total > MAX_PRIMS_DISPLAY:
            ui.Label(f"총 {total}개 prim 중 {len(display_paths)}개만 표시됩니다.", height=0)
            ui.Spacer(height=4)
        if priority_prefix:
            n_priority = min(len(priority_paths), MAX_PRIMS_DISPLAY)
            n_rest = len(display_paths) - n_priority
            ui.Label(f"접두사 '{priority_prefix}' 우선: {n_priority}개, 나머지 순서대로 {n_rest}개", height=0)
            ui.Spacer(height=4)
        for idx, prim_path in enumerate(display_paths):
            build_object_panel(ext, ext._object_list_frame, prim_path, idx + 1)


def build_object_panel(ext: Any, parent: ui.VStack, prim_path: str, index: int) -> None:
    try:
        stage = get_stage()
        prim = stage.GetPrimAtPath(prim_path) if stage else None
        if not prim or not prim.IsValid():
            return
        title = get_prim_display_name(prim, index)
        local = get_prim_local_translate(prim)
        pos_models = [
            ui.SimpleFloatModel(local[0]),
            ui.SimpleFloatModel(local[1]),
            ui.SimpleFloatModel(local[2]),
        ]

        def update_prim_position():
            s = get_stage()
            p = s.GetPrimAtPath(prim_path) if s else None
            if p and p.IsValid():
                set_prim_translate_only(p, Gf.Vec3f(
                    pos_models[0].get_value_as_float(),
                    pos_models[1].get_value_as_float(),
                    pos_models[2].get_value_as_float(),
                ))

        with parent:
            with ui.CollapsableFrame(title, collapsed=False):
                with ui.VStack(spacing=6):
                    ui.Label("Position (X, Y, Z)", height=0)
                    with ui.HStack():
                        for i, label in enumerate(["X", "Y", "Z"]):
                            ui.Label(label, width=24)
                            ui.FloatField(model=pos_models[i])
                    for m in pos_models:
                        m.add_value_changed_fn(lambda _: update_prim_position())
                    ui.Spacer(height=4)
                    ui.Button("3D 정보 보기", height=24, clicked_fn=lambda p=prim_path: show_prim_info_in_viewport(ext, p))
                    ui.Spacer(height=4)
                    with ui.HStack(spacing=8):
                        ui.Button("button_0", width=0, clicked_fn=lambda p=prim_path: on_button_0(ext, p))
                        ui.Button("button_1", width=0, clicked_fn=lambda p=prim_path: on_button_1(ext, p))
                        ui.Button("button_2", width=0, clicked_fn=lambda p=prim_path: on_button_2(ext, p))
    except (UnicodeDecodeError, UnicodeEncodeError):
        return


def on_button_0(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    run_prim_translate_animation(
        prim_path,
        [
            {"duration": 1.0, "delta": (100.0, 0.0, 0.0)},
            {"duration": 1.0, "delta": (0.0, 0.0, 100.0)},
        ],
        loop=False,
    )


def on_button_1(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    stage = get_stage()
    prim = stage.GetPrimAtPath(prim_path) if stage else None
    if not prim or not prim.IsValid():
        return
    start = get_prim_local_translate(prim)
    start_t = (start[0], start[1], start[2])
    end_t = (start[0] + 100.0, start[1], start[2])
    path_points = make_parabolic_path(start=start_t, end=end_t, arc_height=30.0, num_points=24)
    run_prim_curve_animation(prim_path, path_points, duration_sec=1.0, loop=False)


def on_button_2(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    run_prim_rotate_animation(
        prim_path,
        [{"duration": 3.0, "delta": (0.0, 90.0, 0.0)}],
        loop=False,
    )
